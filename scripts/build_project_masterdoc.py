"""
Generates project_masterdoc.xlsx — comprehensive game design & implementation tracker for Farsword.
Run from repo root: python3 scripts/build_project_masterdoc.py

Sheets:
  1. Army        — units, weapons, armor, mil mods, combat, status effects
  2. Events      — event pipeline, all event chains and categories
  3. Ualani      — Ualani Carlisle systems, tile events, DC monthlies, secret arc
  4. Explicit    — adult content events, buttons, gating status
  5. Govs & Cmds — governor system, archetypes, arcs, loyalty, progression
  6. Game Systems — core world: resources, buildings, politics, factions, UI
  7. Win & Lose  — end-game conditions, collapse, game over
  8. Major Bugs  — confirmed code bugs and critical stubs
"""

import csv
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "project_masterdoc.xlsx"

HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
NORMAL_FG  = "1A1A1A"
EXPLICIT_BG = "FF6B35"

STATUS_COLORS = {
    "FULL PASS":   "00B050",
    "FIRST PASS":  "92D050",
    "FIRST DRAFT": "FFEB9C",
    "IDEA":        "C9B1E8",
    "STUB":        "FF6B35",
}
STATUS_FG = {
    "FULL PASS":   "FFFFFF",
    "FIRST PASS":  "1A3A00",
    "FIRST DRAFT": "7A5A00",
    "IDEA":        "3A1A6A",
    "STUB":        "FFFFFF",
}

SEV_COLORS = {
    "CRITICAL": ("FF0000", "FFFFFF"),
    "HIGH":     ("FF6B35", "FFFFFF"),
    "MEDIUM":   ("FFEB9C", "7A5A00"),
    "LOW":      ("D6E4F7", "1A3A6A"),
    "RESOLVED": ("00B050", "FFFFFF"),
}

thin   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

ROW_BG = {
    "FULL PASS":   "EDF5FF",
    "FIRST PASS":  "F2FAEA",
    "FIRST DRAFT": "FFFEF0",
    "IDEA":        "F7F2FF",
    "STUB":        "FFF0EA",
}


def mf(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def mfont(bold=False, color=NORMAL_FG, size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def hcell(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = mf(HEADER_BG)
    c.font = mfont(bold=True, color=HEADER_FG, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def pcell(ws, row, col, value, bg=None, fg=None, bold=False, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    c.font = mfont(bold=bold, color=fg or NORMAL_FG)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = BORDER
    if bg:
        c.fill = mf(bg)
    return c

def divider(ws, row, ncols, text, bg="2F4F6F"):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row, column=ci, value=text if ci == 1 else "")
        c.fill = mf(bg)
        c.font = mfont(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 16

# Standard 6-column layout used by most sheets
STD_COLS = [
    ("STATUS",           12),
    ("CATEGORY",         18),
    ("SYSTEM / FEATURE", 30),
    ("KEY FILE(S)",      34),
    ("DESCRIPTION",      60),
    ("GAPS / NOTES",     42),
]

def std_headers(ws):
    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(STD_COLS, 1):
        hcell(ws, 1, ci, h, w)

def std_row(ws, row, status, cat, name, files, desc, notes, height=40):
    sbg = STATUS_COLORS.get(status, "FFFFFF")
    sfg = STATUS_FG.get(status, NORMAL_FG)
    cbg = ROW_BG.get(status, "FAFAFA")
    pcell(ws, row, 1, status, bg=sbg,  fg=sfg,     bold=True, align="center")
    pcell(ws, row, 2, cat,    bg=cbg)
    pcell(ws, row, 3, name,   bg=cbg,  bold=True)
    pcell(ws, row, 4, files,  bg="FAFAFA")
    pcell(ws, row, 5, desc,   bg="FAFAFA")
    pcell(ws, row, 6, notes,  bg="FFFBE6")
    ws.row_dimensions[row].height = height

def std_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(sheet_name)
    std_headers(ws)
    current_cat = None
    row = 2
    CAT_DIVIDER_COLORS = {
        # Army
        "Unit Core":          "2E5C3A", "Weapons":          "5A3A1E",
        "Armor":              "5A2E6A", "Mil Mod Pipeline":  "2E5C8A",
        "Terrain Mods":       "6A4A1A", "Storm Mods":        "1A4A6A",
        "Cultural Mods":      "6A4A1A", "State Guard Mods":  "1A6A3A",
        "Protector Buffs":    "7A5A00", "Negative Status":   "8A1A1A",
        "Status System":      "4A1A7A", "Combat":            "8A2E2E",
        "Marine & Naval":     "1A4A6A", "Army Core":         "2E5C8A",
        # Events
        "Pipeline":           "2E5C3A", "Protector Events":  "7A5A00",
        "Commander Arc":      "2E5C8A", "Tile Crisis":       "5A3A1E",
        "Ualani Events":      "5A2E6A", "Calendar":          "1A4A6A",
        "City Events":        "6A4A1A", "State Events":      "6A4A1A",
        "Fort Events":        "5A5A5A", "Endgame Events":    "8A1A1A",
        "Misc Events":        "3A5A3A",
        # Ualani
        "Identity":           "5A2E6A", "Commander Mechanics":"2E5C8A",
        "Tile Events":        "6A4A1A", "DC Monthly Events": "8A2E2E",
        "Secret Arc":         "7A5A00", "Election System":   "2E5C3A",
        "Endgame":            "8A1A1A",
        # Govs
        "Governor Core":      "2E5C8A", "Named Governors":   "2E5C3A",
        "Proc Gen":           "6A4A1A", "Commander Arcs":    "5A2E6A",
        "Commander Progression": "7A5A00", "VP System":      "8A2E2E",
        # Game Systems
        "Core World":         "2E5C3A", "Factions":          "5A3A1E",
        "UI":                 "2E5C8A", "Spawn":             "6A4A1A",
        # Win/Lose
        "Lose: Collapse":     "8A1A1A", "Lose: Death":       "8A2E2E",
        "Lose: Secession":    "5A3A1E", "Lose: Occupation":  "5A5A5A",
        "Win: Unification":   "2E5C3A", "Win: Diplomatic":   "2E5C8A",
        "Win: Protectors":    "7A5A00", "Win: Screen":       "5A5A5A",
    }
    for (status, cat, name, files, desc, notes) in data:
        if cat != current_cat:
            bg = CAT_DIVIDER_COLORS.get(cat, "2F4F6F")
            divider(ws, row, len(STD_COLS), cat, bg=bg)
            row += 1
            current_cat = cat
        std_row(ws, row, status, cat, name, files, desc, notes)
        row += 1
    ws.freeze_panes = "A2"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ARMY
# ══════════════════════════════════════════════════════════════════════════════

ARMY_DATA = [
    # ── UNIT CORE ─────────────────────────────────────────────────────────────
    ("FULL PASS", "Unit Core", "Unit Class & Stats",
     "unit.gd",
     "unitType, unitLevel (1+), unitOffensiveScore, unitDefensiveScore, "
     "unitRangedOffence, unitRangedDefence, unitMagicDefence, unitMaxShield/unitShield. "
     "unitMaxManpower = 100 × level; unitMaxWeapons = 100 × level.",
     ""),

    ("FULL PASS", "Unit Core", "Unit Attribute Calculation",
     "unit.gd getUnitAttributes(), calculateGrossValues()",
     "effectMultiplier = (man% + weapons%) / 2; artillery uses weapons% only. "
     "meleePenalty = 0.5 for muskets, 1.0 for sabers. "
     "offScore = level × weaponOffence × meleePenalty × effectMultiplier.",
     ""),

    ("FULL PASS", "Unit Core", "Weapon / Ore / Armor Integration",
     "unit.gd calculateWeaponsOresArmor()",
     "Pulls stats from Weapon, Ore, Armor nodes; copies embedded milMods to "
     "unit.militaryModifierList. Shield = ore.oreMaxShield × unitLevel.",
     ""),

    ("FULL PASS", "Unit Core", "Damage Reception (takeLosses)",
     "unit.gd takeLosses(type, amount)",
     "Melee: blocked = amount × unitDefensiveScore; shield absorbs first; rest = manpower. "
     "Ranged: same with unitRangedDefence. Magic: blocked = amount × unitMagicDefence; bypasses shield.",
     ""),

    ("FULL PASS", "Unit Core", "Manpower Refill",
     "unit.gd refillManpower(RR)",
     "Adds armyReinforceRate to currentManpower each turn, capped at max. "
     "Skipped if army reinforcementBlocked flag is set.",
     ""),

    ("FULL PASS", "Unit Core", "Mil Mod Tick (tick_temp_mods)",
     "unit.gd tick_temp_mods()",
     "Decrements turnsRemaining for all non-permanent mods (turnsRemaining > −1). "
     "Removes expired mods. Returns true if any removed (caller re-runs surveySelf).",
     ""),

    # ── WEAPONS ────────────────────────────────────────────────────────────────
    ("FULL PASS", "Weapons", "Saber Class (4 types)",
     "weapon.gd",
     "Cutlass L1 (+3/+2), Cavalry Saber L2 (+5/+3), Light Saber L3 (+7/+4), Heavy Saber L4 (+9/+5). "
     "reloadTurns=0; chargeManpowerCost=10%/charge. SaberCharge on all. "
     "CavalryMorale on L2, HeavyCharge on L4.",
     "CavalryMorale and HeavyCharge mod effects declared but unimplemented."),

    ("FULL PASS", "Weapons", "Musket Class (4 types)",
     "weapon.gd",
     "Flintlock L1 (ranged +4, reload 2), Brown Bess L2 (+6, reload 2), "
     "Percussion Cap L3 (+8, reload 1), Lever Repeater L4 (+10, no reload, burns 2 ammo/lvl). "
     "All have Bayonet (melee at 50% penalty). VolleyFire on Brown Bess.",
     "VolleyFire declared but unimplemented."),

    ("FIRST PASS", "Weapons", "Artillery Class (4 types)",
     "weapon.gd",
     "Falconet L1 (ranged +12, reload 3), Field Gun L2 (+18), Howitzer L3 (+25, reload 2), "
     "Mortar L4 (+35, reload 2, Siege mod). No melee. CannonBlast+AreaDamage on L2+.",
     "AreaDamage and Siege effects unimplemented. Weapon cost calc uses legacy names → 0 cost shown."),

    ("FIRST PASS", "Weapons", "Mythic Weapons (9 event-unlock only)",
     "weapon.gd",
     "Baseball Bat, Trident, Mythic Atlatl, Sharps Carbine, Blackbeard's Pistols, "
     "Colt Revolver, Rocket Artillery, Trebuchet, Wright Flyer. "
     "Status effects on hit wired via battle.gd _apply_combat_status_effects(). "
     "is_mythic() helper defined. can_melee() excludes Rocket Artillery and Wright Flyer.",
     "PirateVolley, CylinderFire, TrebuchetLaunch, AerialBombing special effects are stubs. "
     "RocketBarrage tile-fire effect not implemented."),

    ("FIRST PASS", "Weapons", "Legacy Weapon Class (12 types)",
     "weapon.gd",
     "Atlatl, Club, Double Axe, Flail, Longsword, Mace, Machete, Macuahuitl, "
     "Pike, Shortsword, Tomahawk, Spear. DODK compat. AtlatlPierce + ClubBleed mods applied.",
     "Cost calc in surveySelf uses legacy names only — new weapon armies show 0 cost."),

    ("FULL PASS", "Weapons", "Reload System",
     "unit.gd, battle.gd",
     "reloadCounter counts down to 0 (ready). start_reload() sets counter=reloadTurns. "
     "GunCrewEfficiency reduces reloadTurns by 1 at start_reload(). "
     "get_effective_ranged_offence() returns 0 if reloading.",
     "Reload only ticks during ranged rounds. Melee rounds do not advance reload."),

    # ── ARMOR & UNIFORMS ───────────────────────────────────────────────────────
    ("FIRST PASS", "Armor", "Uniform Class (10 types)",
     "armor.gd",
     "Continental (DrillFormation), Redcoat (LineFormation), Militia, Light Infantry (Skirmish), "
     "Heavy Infantry (ShockTroop), Cavalry (MountedCharge+CavalryMorale), "
     "Artillery Corps (GunCrewEfficiency — WORKS), Minuteman (MinutemanSpirit), "
     "Tombstone Cap (QuickDraw), Hardee Hat (HardeeDisc). "
     "M/R/S block percentages defined on all.",
     "Formation mods (Drill/Line/Skirmish/ShockTroop/Minuteman) declared at 'army level' but no army-level check exists. "
     "QuickDraw (first-attack bonus) and HardeeDisc (adjacent-friend bonus) are FIRST DRAFT stubs."),

    ("FIRST PASS", "Armor", "Legacy Armor Class (12 types)",
     "armor.gd",
     "Canine, Cast, Chain, Archer, Plate, Padded, Scout, Scale, Shell, Point, Feline, Otter. "
     "DODK compat. Chain adds to rangedOffence (inverted disabled bug historically). "
     "Shell has embedded spell block mod.",
     ""),

    # ── MIL MOD PIPELINE ───────────────────────────────────────────────────────
    ("FULL PASS", "Mil Mod Pipeline", "MilMod Class (184 types defined)",
     "mil_mod.gd",
     "MilMod node: milModType, milModResource, disabled, turnsRemaining (−1=permanent), isNegative. "
     "Flag bools: infantryMod, rangedMod, siegeMod, commanderMod, civilianMod, resourceMod, "
     "marineMod, entrenchMod, terrainMod, toolMod, stormMod, culturalMod. "
     "buildSelf(Type) sets all flags from 184-entry match block.",
     ""),

    ("FULL PASS", "Mil Mod Pipeline", "Commander Mod Tiers",
     "governor.gd, army.gd commanderCheck()",
     "Level codes: 123=all tiers, 23=tiers 2+3, 3=tier 3 only. "
     "commanderCheck() picks govMilModsLvl1/2/3 by governorLevel. "
     "VP commanders: all three arrays each appended twice (all mods doubled).",
     ""),

    ("FULL PASS", "Mil Mod Pipeline", "Archetype Mod Assignment",
     "world.gd _apply_archetype_mods()",
     "25 archetypes × 4 mods each: 2 at T1(123), 1 at T2(23), 1 at T3(3). "
     "Called after governorLevel=1 in generateBarracksCommanders().",
     ""),

    ("FULL PASS", "Mil Mod Pipeline", "Resource Disable System",
     "army.gd surveySelf(), mil_mod.gd disableMilModType()",
     "surveySelf() resets all mods enabled, then calls disableMilModType(resource) for each depleted resource. "
     "Any mod whose milModResource matches depleted type gets disabled=true.",
     ""),

    ("FULL PASS", "Mil Mod Pipeline", "calculateMilMods()",
     "unit.gd calculateMilMods()",
     "Main stat modifier function. Early return if armyDemoralized (skips all mod bonuses). "
     "Full match block for terrain/storm/cultural/commander/weapon/armor mods. "
     "Post-match generic state guard check: "
     "culturalMod + culturalState != '' + currentState == culturalState → +2 atk +2 def/level.",
     ""),

    # ── TERRAIN / CULTURAL / STORM MODS ────────────────────────────────────────
    ("FIRST PASS", "Terrain Mods", "T1/T2/T3 Commander Terrain Mods (36 mods)",
     "mil_mod.gd, unit.gd calculateMilMods()",
     "All 36 commander mods have cases in calculateMilMods(). "
     "surveySelf() sets unit.currentTerrain + currentStorm before calling calculateMilMods(). "
     "Working: Woodsman, Swamp Legs, Hill Runner, Street Tough, Farmhand, Guerrilla Tactics, "
     "Backcountry Rider, Frontier Marksman, Everglades Tracker, Bayou Warrior, "
     "Continental Line, Last Stand, Saber Drill, Marksman, Steady Line.",
     "Pending subsystems: Marine (now FIRST PASS), Entrenched (no turn counter), "
     "Night Raider (no night cycle), Ghost March (no ZoC), The Long March (no MP hook)."),

    ("FIRST PASS", "Storm Mods", "Storm Counter Mods (12 mods)",
     "mil_mod.gd, unit.gd calculateMilMods()",
     "Fog-Born, Storm Rider, Thunder Proof, Blizzard March, Hurricane Eyes, Tornado Dancer, "
     "Nor'easter Veteran, Rain Reader, White Out Walker, Storm Chaser, Lightning Rod, Eye of the Storm. "
     "stormMod=true + stormType; calculateMilMods reads currentStorm for stat bonuses.",
     "Movement/immunity effects (Storm Rider, Blizzard March, Tornado Dancer, Storm Chaser) need army-level check."),

    ("FIRST PASS", "Cultural Mods", "Cultural / State-Specific Mods (12 mods)",
     "mil_mod.gd, unit.gd calculateMilMods()",
     "Country Musician (TN), Virginia Gentry (VA), Minuteman's Pride (MA), Quaker Steel (PA), "
     "Georgia Peach (GA), Backcountry Rider (SC), Harbor Watch (NY), Chesapeake Sailor (MD), "
     "Frontier Marksman (KY), River Runner (OH), Everglades Tracker (FL), Bayou Warrior (LA).",
     "Minuteman's Pride (first-3-rounds +5 ATK) and River Runner (+2 movement near water) IDEA-tier."),

    ("FULL PASS", "State Guard Mods", "26 State / Provincial Guard Mods",
     "mil_mod.gd, unit.gd calculateMilMods()",
     "PA, VA, NY, MA, MD, NC, SC, GA, CT, NJ, DE, NH, RI, VT, ME, TN, AL, FL, WV, DC, "
     "CA-QB, CA-OT, CA-NS, CA-NB, CA-PEI, BA. "
     "+2 attack +2 defence per level when in matching state tile. "
     "Generic post-match handler in calculateMilMods(): culturalMod + culturalState == currentState.",
     "tileContinent must match culturalState codes exactly. No gate preventing wrong-state assignment."),

    # ── PROTECTOR BUFFS ────────────────────────────────────────────────────────
    ("FIRST PASS", "Protector Buffs", "25 Protector Buff Mods (17 USA + 8 Canadian)",
     "mil_mod.gd, army.gd _apply_status_effects_to_stats()",
     "USA (17): Mothman Presence (+20 Rng/+15 Rng Def), Jersey Devil's Fury (+25 Atk), "
     "Bigfoot's Solidarity (+30 Block), Thunderbird's Sovereignty (+25 Ranged), "
     "Headless Terror (+20 Atk, attackers Terrified 2t), Chessie's Blessing (+20 Block), "
     "Bell Witch's Harassment (+15 Atk, attackers Demoralized 2t), Old Ironsides' Hull (+30 Shield), "
     "Valley Forge's Will (+10 Atk +25 Block +20 Def), Snallygaster's Claim (+20 Atk), "
     "Paul Revere's Ride (+15 Rng +3 Mvmt), Liberty Bell's Resonance (+25 Block), "
     "Green Mountain Haunting (+20 Block), Presidential Decree (+20 Atk +20 Block +15 Rng +15 Def), "
     "Skunk Ape's Domain (+20 Atk), Eternal Vigilance (+25 Block), Lincoln's Mandate (+15 Atk). "
     "Canadian (8): Le Wendigo (+30 Atk), Loup-Garou (+25 Atk), Feux Follets (+25 Rng Def), "
     "Mishepeshu (+20 Block/Rng Def), La Corriveau (+20 Rng), Le Carcajou (+20 Atk), "
     "La Chasse-Galerie (+15 Atk/Rng +4 Mvmt), Le Gougou (+15 Atk, attackers Terrified 3t). "
     "Applied as army status effects with magic upkeep; drain TotalMagic/turn; self-remove at 0.",
     "Wiring 'cast_protector_buff' outcomes to specific protector AGREE events still needed. "
     "No UI shows active protector buffs or remaining magic upkeep cost."),

    # ── NEGATIVE STATUS EFFECTS ────────────────────────────────────────────────
    ("FULL PASS", "Negative Status", "22 Negative Status Effects",
     "mil_mod.gd, army.gd _apply_status_effects_to_stats(), _apply_status_flags()",
     "Combat: Stunned (no melee atk), Suppressed (no ranged atk), "
     "Shaken (block ×0.5), Terrified (atk ×0.5 block ×0.7), "
     "Routed (atk=0, no attack, def ×0.5 — triggered by >40% loss). "
     "DoT: Burning (5×unitCount/turn), Diseased (3×/turn), Quarantined (1×/turn + no reinforce). "
     "Stat: Blinded (ranged ×0.5), Hexed (magic def=0), Demoralized (atk ×0.8, block ×0.8, all mil mods disabled), "
     "Waterlogged (atk ×0.8, storm), Frostbitten (atk ×0.7, storm). "
     "Movement: Exhausted (movement → min 1), Bogged Down (movement = 0). "
     "Supply: Pacified (no attacks), Supply Cut (no reinforce). "
     "Funny: Seduced (no attack, atk=0), Starstruck (all ×0.7), Hangover (all ×0.5), "
     "Love-Struck (atk ×0.3, no attack), Mutinous (atk ×0.6, 50% refuse orders).",
     "No UI displays active status effects on army. All 22 are mechanically wired."),

    # ── ARMY STATUS SYSTEM ─────────────────────────────────────────────────────
    ("FULL PASS", "Status System", "Army Status Effect System",
     "army.gd apply_status(), _tick_status_effects(), _apply_status_effects_to_stats(), _apply_status_flags()",
     "armyStatusEffects: Array of {type, turnsLeft, magicCostPerTurn}. "
     "apply_status(type, duration, magic_cost): refresh/replace if already present. "
     "_tick_status_effects() on onTurnEnd: magic upkeep drain for protector buffs (remove at 0); "
     "DoT for Burning/Diseased/Quarantined; tick and remove expired. "
     "_apply_status_effects_to_stats() from surveySelf(): all 47 stat modifiers (22 debuffs + 25 protector buffs). "
     "_apply_status_flags(): sets attackBlocked, reinforcementBlocked, movement from statuses.",
     "No UI widget shows active statuses on army panel."),

    ("FULL PASS", "Status System", "Corruption Disease Check",
     "army.gd onTurnEnd()",
     "If inTile.tileCorrution > 0: randf() < corruption/100.0 → Diseased for 2 turns. "
     "Park Ranger mod (civilianMod + toolMod) grants full immunity.",
     "Spelling 'tileCorrution' matches in-game variable."),

    ("FULL PASS", "Status System", "Storm Status Debuffs",
     "world.gd _apply_storm_debuffs(), _apply_storm_status_to_army()",
     "Applied at each turn start after _tick_storms(). "
     "Thunderstorm/Hurricane → Waterlogged 2t. Blizzard/Nor'easter → Frostbitten 2t. "
     "Fog → Blinded 1t. Tornado → Bogged Down 1t + Shaken 1t.",
     "Armies moving into storm tiles mid-turn miss debuff until next turn start."),

    # ── COMBAT ────────────────────────────────────────────────────────────────
    ("FULL PASS", "Combat", "Battle Class",
     "battle.gd",
     "Instantiated by army.calculateBattle(). Snapshots both army states. "
     "Calculates projected damage for display. Player clicks Attack to apply.",
     ""),

    ("FULL PASS", "Combat", "Melee Damage Calculation",
     "battle.gd _calculate_melee_damage()",
     "raw_attack = attacker.armyPunch + _army_naval_supremacy_bonus(). "
     "block_ratio = armyBlock / unitCount (capped 0–0.9). "
     "Shield absorbs first; remainder hits manpower. Defender counter-attacks simultaneously.",
     ""),

    ("FULL PASS", "Combat", "Ranged Damage Calculation",
     "battle.gd _calculate_ranged_damage()",
     "effective_launch = sum get_effective_ranged_offence() (0 if reloading). "
     "ranged_block_ratio = armyDefence / unitCount. Defender can counter-fire if ready. "
     "Double Shot / Double Cannonade second volley applied against remaining shield.",
     ""),

    ("FULL PASS", "Combat", "Combat Status Effects",
     "battle.gd _apply_combat_status_effects()",
     "Routed: >40% manpower lost in one battle → Routed 2t. "
     "Weapon-based hits: Baseball Bat→Shaken, Trident→Terrified 2t, Mythic Atlatl→Stunned, "
     "Sharps Carbine→Blinded, Blackbeard's Pistols→Terrified, Colt Revolver→Shaken, "
     "Rocket Artillery→Burning 2t+Suppressed, Trebuchet→Stunned (ranged only), Wright Flyer→Suppressed. "
     "Protector retaliations: Bell Witch→Demoralize attacker 2t, "
     "Headless Terror→Terrify attacker 2t, Le Gougou→Terrify attacker 3t.",
     "Weapon statuses fire on every attack, not just crits."),

    ("FULL PASS", "Combat", "Saber Charge Cost",
     "battle.gd _apply_charge_costs(), unit.gd apply_charge_cost()",
     "Each saber unit pays 10% × currentManpower on every melee attack. Applied after resolve.",
     ""),

    ("FIRST PASS", "Combat", "Double Shot / Double Cannonade",
     "battle.gd _calculate_ranged_damage()",
     "Double Shot (T2): second volley at 50% power for artillery. "
     "Double Cannonade (T3): 100% + 3 bonus per artillery unit. "
     "Second shot calculated against remaining shield after first.",
     "Only applies to artillery weapon class."),

    ("FULL PASS", "Combat", "Shield System",
     "unit.gd, battle.gd",
     "unitMaxShield = unitLevel × ore.oreMaxShield. Shield absorbs before manpower. "
     "Magic damage bypasses shield. restore_shield() resets at new engagement.",
     ""),

    ("FULL PASS", "Combat", "Morale Multiplier",
     "army.gd surveySelf()",
     "mm = 1.0 + (commander.morale / 100) × 0.25. Applied to armyPunch + armyDefence. "
     "Range: ×1.0 (morale 0) to ×1.25 (morale 100).",
     "governor.loyalty is −20 to +10; morale var may not be correctly initialized — "
     "multiplier may always be 1.0. See Major Bugs."),

    ("FULL PASS", "Combat", "Army Death (Civ-style)",
     "army.gd, country.gd _on_army_destroyed()",
     "manpowerInArmy ≤ 0 → emit armyDestroyed signal. "
     "country._on_army_destroyed(): erases from countryArmyList, nulls tile.stationedArmy, queue_free().",
     ""),

    ("STUB", "Combat", "_apply_army_buff() — Army Buff Outcome",
     "world.gd _apply_army_buff() ~line 3351",
     "Function called from event button outcomes but body is a single `pass` statement. "
     "All army_buff outcome types in events are silently ignored.",
     "CRITICAL — implement: iterate countryArmyList, call army.apply_status(buff_type, duration)."),

    # ── MARINE & NAVAL ─────────────────────────────────────────────────────────
    ("FIRST PASS", "Marine & Naval", "Marine Mechanic",
     "mil_mod.gd (marineMod), world.gd meleePressed(), path_point_button.gd navalPathPoints",
     "Army with Marine mod may melee-attack into adjacent navalPathPoints (occupied PPBs). "
     "_army_has_active_marine() checks units for enabled Marine mod. "
     "meleePressed() loops inTile.navalPathPoints and builds battles on occupied ones. "
     "path_point_button.buildSelf() populates navalPathPoints from navalPathPointsEXP.",
     "PPBs must be pre-wired in editor navalPathPointsEXP. No separate naval-target UI."),

    ("FIRST PASS", "Marine & Naval", "Naval Supremacy",
     "battle.gd _army_naval_supremacy_bonus()",
     "+5 × unitLevel added to raw_attack in _calculate_melee_damage() for armies with Naval Supremacy mod.",
     ""),

    # ── ARMY CORE ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Army Core", "Army Survey (surveySelf)",
     "army.gd surveySelf()",
     "Rebuilds all stats from units + milMods; applies morale multiplier; calculates resource costs; "
     "disables mods for depleted resources; sets unit.currentState = inTile.tileContinent; "
     "sets unit.armyDemoralized from _has_status(); calls _apply_status_effects_to_stats() at tail.",
     ""),

    ("FULL PASS", "Army Core", "Turn-End Reset (onTurnEnd)",
     "army.gd onTurnEnd()",
     "currentMovementPoints = maxMovementPoints + _commander_movement_bonus(). "
     "Resets attackBlocked/reinforcementBlocked. Ticks spy timers. "
     "tick_temp_mods() per unit. _tick_status_effects(). _apply_status_flags(). "
     "Refills manpower if not reinforcementBlocked. Runs corruption disease check.",
     ""),

    ("FULL PASS", "Army Core", "Commander Movement Bonus",
     "army.gd _commander_movement_bonus()",
     "Sums +3 for President mod and +3 for Election Season mod from govMilModsLvl arrays. "
     "Applied before _apply_status_flags() (Exhausted→min 1, Bogged Down→0).",
     ""),

    ("FULL PASS", "Army Core", "Spy Effects (Sabotage / Recon)",
     "army.gd",
     "sabotaged bool + sabotageTimer: blocks movement/action for N turns. "
     "reconDebuffed bool + reconDebuffTimer: reduces armyDefence. Both tick on onTurnEnd().",
     "Wiring from espionage events to set flags is partial."),

    ("FIRST DRAFT", "Army Core", "Entrenched Mechanic",
     "mil_mod.gd (entrenchMod flag)",
     "entrenchMod bool defined. Design: after 3 stationary turns, +5 def/level; lost on movement.",
     "No stationary turn counter. Nothing reads entrenchMod."),

    ("IDEA", "Army Core", "Unit Promotion / Experience",
     "army.gd (averageExperience var)",
     "averageExperience and armyLevel vars declared. No experience gain logic anywhere.",
     ""),

    ("IDEA", "Army Core", "Formation Bonuses (Army-Level)",
     "(DrillFormation, LineFormation, VolleyFire, ShockTroop, Skirmish mods)",
     "Multiple weapon/armor mods reference formation bonuses. No army-level formation check exists.",
     ""),
]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — EVENTS
# ══════════════════════════════════════════════════════════════════════════════

EVENTS_DATA = [
    # ── PIPELINE ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Pipeline", "Event CSV Pipeline",
     "EventDatabase.gd, world.gd, event_scene.gd",
     "events.csv (260 events) + event_buttons.csv → evaluate_date_triggers() → "
     "createNewEvent() → event_scene.build_from_csv(). "
     "Events: event_id, headline, body, image_tag, content_flag. "
     "Buttons: event_id, button_id, text, outcome_type, outcome_value, outcome_amount, "
     "button_type, chain_event_id.",
     ""),

    ("FULL PASS", "Pipeline", "Event Cooldown System",
     "world.gd _event_cooldowns dict",
     "_event_cooldowns dict; _event_on_cooldown() / _start_cooldown() / _tick_event_cooldowns(). "
     "Prevents same event firing again within cooldown window.",
     ""),

    ("FULL PASS", "Pipeline", "Mission Flag System",
     "world.gd mission_<id>_<tile> flags",
     "Persistent boolean flags on world: set_mission_flag / set_mission_flag_own outcomes. "
     "checkPendingMissions() evaluates conditions each turn.",
     ""),

    ("FULL PASS", "Pipeline", "Button Content Filtering",
     "event_scene.gd _build_buttons()",
     "button_type (standard/explicit/kinky_lewd/sensual) checked against player Settings. "
     "Explicit buttons hidden if player has content disabled.",
     "No global Settings.enableExplicit toggle found — explicit always visible. See Major Bugs."),

    ("FULL PASS", "Pipeline", "Outcome Execution Engine",
     "world.gd executeOutcome()",
     "Handles 40+ outcome types: resource_change, loyalty_change, building_enable/disable, "
     "army_buff (STUB), promote_commander, morale_boost, add_tech, cast_protector_buff, "
     "set_mission_flag, add_governor, chain_event, tile_moral_decay_change, "
     "faction_loyalty_change, and more.",
     "_apply_army_buff() is a critical stub (see Major Bugs)."),

    # ── PROTECTOR CHAINS ───────────────────────────────────────────────────────
    ("FULL PASS", "Protector Events", "USA Protector 3-Stage Chain (17 × 3 = 51 events)",
     "events.csv, world.gd _check_protector_summons()",
     "PROT_01–17: each has PROT_xx_SUMMON → PROT_xx_TAME → PROT_xx_AGREE. "
     "SUMMON fires via _check_protector_summons() with wild corruption risk 25%/turn. "
     "All 17 protectors: Mothman, Jersey Devil, Bigfoot, Thunderbird, Headless Horseman, "
     "Chessie, Bell Witch, Old Ironsides, Valley Forge Spirit, Snallygaster, Paul Revere, "
     "Liberty Bell, Green Mountain Boys, Mount Rushmore Spirits, Skunk Ape, "
     "Eternal Minuteman, Lincoln's Ghost. "
     "Staggered turn thresholds per protector.",
     ""),

    ("FULL PASS", "Protector Events", "Canadian Protector 3-Stage Chain (8 × 3 = 24 events)",
     "events.csv, world.gd",
     "CA_PROT_01–08: SUMMON → TAME → AGREE per protector. "
     "Le Wendigo, Loup-Garou, Feux Follets, Mishepeshu, La Corriveau, "
     "Le Carcajou, La Chasse-Galerie, Le Gougou.",
     ""),

    # ── COMMANDER ARC ──────────────────────────────────────────────────────────
    ("FULL PASS", "Commander Arc", "Commander Arc Events (25 arcs × 2 = 50 events)",
     "events.csv, world.gd, WarRoomPanel.gd",
     "ARC_01_FIRST–ARC_25_FIRST: fire when objective 1 complete. "
     "ARC_01_DONE–ARC_25_DONE: fire when all 3 objectives done. "
     "Fired via arcObjectiveCompleted signal from WarRoomPanel.",
     ""),

    ("FULL PASS", "Commander Arc", "Governor Loyalty Arc Events (25 events)",
     "events.csv, world.gd",
     "GOV_LOYAL_ARC_01–25: fire at governor loyalty thresholds. One per archetype.",
     ""),

    # ── TILE CRISIS ────────────────────────────────────────────────────────────
    ("FULL PASS", "Tile Crisis", "9 Tile Crisis Event Chains",
     "world.gd, events.csv",
     "Harvest Crisis, Harbor Crisis, Forge Crisis, Corrupt Event, Border Tension, "
     "Garrison Event, Legitimacy Crisis, Turncoat Warning, Memorial Restoration. "
     "All 9 wired with conditions + cooldowns. Fire on qualifying tile state.",
     ""),

    # ── UALANI EVENTS ───────────────────────────────────────────────────────────
    ("FULL PASS", "Ualani Events", "9 Ualani Tile Events",
     "world.gd _check_ualani_*(), events.csv",
     "UALANI_AMBUSH_01, UALANI_DIGNITARY_01, UALANI_MEMORIAL_01, UALANI_WOUNDED_01, "
     "UALANI_WINTER_01, UALANI_FORGE_01, UALANI_CULPER_01, UALANI_ALLIANCE_01, UALANI_FRONTIER_01. "
     "All have cooldowns + location/condition checks. See Ualani tab for full breakdown.",
     ""),

    ("FULL PASS", "Ualani Events", "12 DC Monthly Events (WH_SECRET_01–12)",
     "world.gd, events.csv",
     "Trigger: Ualani at DC (tile 188) + matching calendar month (1–12). "
     "One-shot per month. Sensual/intimate tone. Several have explicit BTN2.",
     "No UI indicator that a monthly event has already fired for the current month."),

    ("FULL PASS", "Ualani Events", "Chalchiuhtotolin Secret Arc (5 events)",
     "world.gd, events.csv",
     "CHALCH_SUMMON (Ualani at Plymouth tile 66 on month 11) → "
     "CHALCH_Q1 (build 3 farms) → CHALCH_Q2 (provide 150 food) → "
     "CHALCH_Q3 (build 5 farms) → CHALCH_AGREE (reward: +220 food total).",
     "Ultra-rare trigger — player must know to position Ualani at Plymouth on month 11."),

    # ── CALENDAR EVENTS ─────────────────────────────────────────────────────────
    ("FULL PASS", "Calendar", "Date / Calendar Events (DE_001–005)",
     "EventDatabase.gd, event_triggers.csv",
     "DE_001 New Year (month 1), DE_002 Valentine's Day (month 2), "
     "DE_003 Independence Day (month 7), DE_004 Halloween (month 10), DE_005 Mothman Day (month 11). "
     "Annual; evaluate_date_triggers() handles cooldowns.",
     ""),

    # ── CITY / STATE EVENTS ─────────────────────────────────────────────────────
    ("FULL PASS", "City Events", "City Liberated / Lost Events",
     "events.csv",
     "CL_001 (NYC), CL_006 (Gettysburg), CL_007 (Kennett Sq.), CL_009 (Brattleboro), CL_182 (Nassau). "
     "CX_001 (DC falls), CX_005 (Gettysburg falls). Non-repeatable tile capture/loss events.",
     ""),

    ("FULL PASS", "State Events", "State Liberated + Secession Events",
     "events.csv, world.gd checkStateSecessionConditions()",
     "SL_001 (Pennsylvania), SL_004 (Vermont). "
     "STATE_REBEL_01 (state secedes), STATE_REINTEGRATED_01 (rejoins). "
     "rebel_ flags per state drive the secession cascade.",
     ""),

    # ── FORT CHAIN ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Fort Events", "Fort Disrepair Chain (FORT_001–005)",
     "world.gd, events.csv",
     "FORT_001 → FORT_003 (public resolution) OR FORT_004 → FORT_005 (private/explicit path). "
     "FORT_005 is explicit-flagged.",
     ""),

    # ── ENDGAME EVENTS ──────────────────────────────────────────────────────────
    ("FIRST PASS", "Endgame Events", "Republic Collapse Events (COLLAPSE_01–02)",
     "events.csv, world.gd _execute_republic_collapse()",
     "COLLAPSE_01: 'THE REPUBLIC HAS FALLEN' — all metros lost, surrender terms. "
     "COLLAPSE_02: 'WASHINGTON STANDS ALONE' — fired directly from _execute_republic_collapse(). "
     "Both have explicit content buttons.",
     "No game-over screen wired. Game continues after collapse events fire."),

    # ── MISC EVENTS ─────────────────────────────────────────────────────────────
    ("FULL PASS", "Misc Events", "Commander Turn Events (CMD_MERIT / CMD_RECOGNITION / CMD_THANKS)",
     "world.gd _commander_turns, events.csv",
     "Fire at 5 / 20 / 50 turns_served per governor. "
     "promote_commander outcome wired. CMD_THANKS has explicit button.",
     ""),

    ("FULL PASS", "Misc Events", "VP / Canadian / UK / Peace Events",
     "events.csv",
     "VP_DOUBT, VP_BATTLEFIELD (VP relationship). "
     "CAN_PEACE_01, CAN_CLEARWATER, CAN_ELECTION, CAN_PEACE (Canadian diplomacy). "
     "PEACE_ALLIED, PEACE_USA (peace treaties). UK faction interaction events.",
     ""),

    ("FULL PASS", "Misc Events", "Memorial, Stump Speech, Harvest, Harbor, Garrison",
     "events.csv",
     "MEMORIAL / MEMORIAL_RESTORE: specific tile memorial events. "
     "STUMP_SPEECH: political rally events. "
     "HARVEST / HARBOR / FORGE / GARRISON / BORDER: tile crisis single-instance events.",
     ""),
]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — UALANI
# ══════════════════════════════════════════════════════════════════════════════

UALANI_DATA = [
    # ── IDENTITY ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Identity", "Ualani Carlisle — President of the Republic",
     "governor.gd buildSelf()",
     "governorType='Ualani Carlisle'; isLeader=true; isVicePresident=false (she IS president). "
     "Base mods: Visionary×123, Champion of the Sun×23, Healer×3, President×123. "
     "_find_ualani_tile() locates her by governorType == 'Ualani Carlisle'.",
     "Named governor mods (Visionary, Champion of the Sun) have no cases in calculateMilMods. "
     "Only President mod actively effects gameplay (movement bonus)."),

    ("FULL PASS", "Identity", "isLeader Flag",
     "governor.gd, world.gd",
     "isLeader=true on Ualani. Used by world.gd to check head-of-state death condition. "
     "_game_ended checks isLeader to identify if protagonist is dead.",
     ""),

    # ── COMMANDER MECHANICS ────────────────────────────────────────────────────
    ("FULL PASS", "Commander Mechanics", "President Mod (+3 Movement)",
     "mil_mod.gd, army.gd _commander_movement_bonus()",
     "commanderMod; +3 movement/turn. "
     "_commander_movement_bonus() sums +3 for President mod from govMilModsLvl arrays. "
     "Applied in onTurnEnd(): currentMovementPoints = maxMovementPoints + bonus, then status flags clamp.",
     ""),

    ("FULL PASS", "Commander Mechanics", "Election Season Mod (+3 Movement, Permanent)",
     "mil_mod.gd, world.gd _grant_election_season_mods(), _check_election_season()",
     "_check_election_season() → _grant_election_season_mods(): "
     "grants 'Election Season' commanderMod to Ualani and any governor with isVicePresident=true. "
     "Permanent once granted. _commander_movement_bonus() sums +3 for this mod.",
     ""),

    ("FULL PASS", "Commander Mechanics", "VP: All Mods Doubled",
     "army.gd addUnitCommander()",
     "isVicePresident=true governors get govMilModsLvl1/2/3 each appended twice (all mods doubled). "
     "VP also gains Election Season mod. Loyalty doubled during election season.",
     "No dedicated VP appointment UI — isVicePresident set via event outcomes."),

    # ── TILE EVENTS ─────────────────────────────────────────────────────────────
    ("FULL PASS", "Tile Events", "UALANI_AMBUSH_01 — Ambush",
     "world.gd _check_ualani_ambush()",
     "Trigger: Ualani in Wetlands tile adjacent to UK-controlled tile. Cooldown 15t. "
     "Standard btn: +weapons 40. BTN3: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_DIGNITARY_01 — Dignitary Reception",
     "world.gd _check_ualani_dignitary()",
     "Trigger: Ualani in Metro/Suburbs tile with Courthouse building. Cooldown 15t. "
     "Standard: −20 moral decay. BTN3: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_MEMORIAL_01 — Memorial Address",
     "world.gd _check_ualani_memorial()",
     "Trigger: Ualani in tile with specialFeatures. Cooldown 15t. "
     "Standard: +morale 50. BTN3: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_WOUNDED_01 — Field Hospital",
     "world.gd _check_ualani_wounded()",
     "Trigger: Ualani army < 80% manpower. Cooldown 15t. "
     "Standard: +manpower 40. BTN2: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_WINTER_01 — Winter March",
     "world.gd _check_ualani_winter()",
     "Trigger: Ualani in Foothills/Woods tile, months 11–2. Cooldown 20t. "
     "No explicit button. Winter perseverance narrative.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_FORGE_01 — Forge Inspection",
     "world.gd _check_ualani_forge()",
     "Trigger: Ualani in tile with enabled Forge building. Cooldown 15t. "
     "Standard: +weapons 60. BTN3: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_CULPER_01 + UALANI_CULPER_BRIEF_01 — Culper Ring",
     "world.gd _check_ualani_culper()",
     "Trigger: Ualani in tile with espionage-active neighboring tile. Cooldown 15t. "
     "Standard: +gold 40 → chains to UALANI_CULPER_BRIEF_01; sets culper_active flag. "
     "BTN3: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_ALLIANCE_01 — Alliance Council",
     "world.gd _check_ualani_alliance()",
     "Trigger: Ualani in tile with governor-occupied neighbor. Cooldown 15t. "
     "Standard: +loyalty 12. BTN3: explicit.",
     ""),

    ("FULL PASS", "Tile Events", "UALANI_FRONTIER_01 — Frontier",
     "world.gd _check_ualani_frontier()",
     "Trigger: Ualani in Woods/Foothills tile bordering Canadian territory. Cooldown 15t. "
     "Standard: +morale 30. BTN3: explicit.",
     ""),

    # ── DC MONTHLY EVENTS ──────────────────────────────────────────────────────
    ("FULL PASS", "DC Monthly Events", "WH_SECRET_01–12 (one per calendar month)",
     "world.gd _check_white_house_monthly(), events.csv",
     "Trigger: Ualani stationed at DC (tile 188) + current month matches (1–12). "
     "One-shot per month; cannot repeat until next year. "
     "Sensual/intimate narrative tone throughout. Several months have explicit BTN2. "
     "All 12 months fully authored.",
     "No UI indicator that a monthly event has already fired for the current month."),

    # ── CHALCHIUHTOTOLIN SECRET ARC ────────────────────────────────────────────
    ("FULL PASS", "Secret Arc", "CHALCH_SUMMON — Chalchiuhtotolin Appears",
     "world.gd, events.csv",
     "Trigger: Ualani at Plymouth (tile 66) during month 11 (Thanksgiving). One-shot. "
     "Sets chalch_summoned flag. First contact with Aztec pestilence deity.",
     "Ultra-rare — player must know to position Ualani at Plymouth in November."),

    ("FULL PASS", "Secret Arc", "CHALCH_Q1 / Q2 / Q3 — Three Tasks",
     "world.gd, events.csv",
     "Q1: Build 3 farms in republic territory. "
     "Q2: Provide 150 food as tribute. "
     "Q3: Build 5 farms total. "
     "Each task unlocks after prior is complete. Progressive resource rewards at each stage.",
     ""),

    ("FULL PASS", "Secret Arc", "CHALCH_AGREE — Alliance Sealed",
     "world.gd, events.csv",
     "Final arc event after all 3 tasks complete. Total reward: +220 food. "
     "Chalchiuhtotolin pledges support to the Republic narrative.",
     ""),

    # ── ELECTION SYSTEM ────────────────────────────────────────────────────────
    ("FULL PASS", "Election System", "Election Season + VP Succession",
     "world.gd _check_election_season(), _grant_election_season_mods()",
     "_check_election_season() fires UALANI_ELECTION_01 event and calls _grant_election_season_mods(). "
     "VP governor gets isVicePresident=true, Election Season mod, doubled mods, doubled loyalty. "
     "_game_ended flag checked to prevent election loop during game over.",
     "No VP appointment UI. VP status set by event outcomes."),

    # ── ENDGAME ROLE ───────────────────────────────────────────────────────────
    ("FIRST PASS", "Endgame", "Republic Collapse — Ualani's Last Stand",
     "world.gd _execute_republic_collapse()",
     "On collapse: UK receives coastal tiles. Interior tiles spawn state nations. "
     "USA retains only DC (tile 188) — Ualani's final stronghold. "
     "COLLAPSE_02 'Washington Stands Alone' fires from this function. Explicit content option.",
     "Game continues after collapse. No rebuild-from-DC win path currently defined."),

    ("FIRST DRAFT", "Endgame", "Head of State Death → Game Over (STUB)",
     "world.gd ~line 4001",
     "_game_ended bool flag exists and is checked in election logic. "
     "Priority comment in code: 'head of state death → game over (Ualani for USA)'. "
     "TODO comment: '# call _trigger_game_over() once game over screen exists'.",
     "Game over screen not built. Death has no mechanical consequence yet. High priority for ship."),
]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — EXPLICIT CONTENT (CSV-driven + system overview rows)
# ══════════════════════════════════════════════════════════════════════════════

EXPLICIT_SYSTEM_ROWS = [
    # (label, status, description)
    ("Content Flag System",          "FULL PASS",
     "events.csv content_flag column + event_buttons.csv button_type column. "
     "Values: standard, explicit, kinky_lewd, sensual. "
     "event_scene.gd _build_buttons() checks button_type before rendering."),
    ("Explicit Content Gating",      "STUB",
     "button_type filtering exists in event_scene.gd but no global Settings.enableExplicit flag found. "
     "Explicit buttons always displayed — no parental control or age-gate. "
     "HIGH PRIORITY before public release."),
    ("Total Explicit Buttons",       "FULL PASS",
     "~29 explicit-flagged buttons + ~15 sensual buttons + ~2 kinky_lewd buttons = ~44 adult-content buttons. "
     "Spread across protector events, commander events, Ualani tile events, DC monthlies, fort chain, "
     "collapse events, VP events, Canadian events."),
    ("Explicit Events in events.csv","FULL PASS",
     "CMD_THANKS (explicit), FORT_005 (explicit), PROT_01_SUMMON (kinky_lewd), "
     "CAN_PEACE_01 (explicit), WH_SECRET_07 (explicit). "
     "PROT_01_AGREE, PROT_14_AGREE, PROT_14_SUMMON also kinky_lewd flagged."),
]


def build_explicit_sheet(wb):
    ws = wb.create_sheet("Explicit Content")
    ws.row_dimensions[1].height = 28

    # System overview section
    sys_cols = [("SYSTEM",30), ("STATUS",14), ("DESCRIPTION",88)]
    for ci, (h, w) in enumerate(sys_cols, 1):
        hcell(ws, 1, ci, h, w)

    row = 2
    divider(ws, row, 3, "EXPLICIT CONTENT SYSTEM OVERVIEW", bg="8A1A1A")
    row += 1
    for (label, status, desc) in EXPLICIT_SYSTEM_ROWS:
        sbg = STATUS_COLORS.get(status, "FFFFFF")
        sfg = STATUS_FG.get(status, NORMAL_FG)
        pcell(ws, row, 1, label,  bg="FAFAFA", bold=True)
        pcell(ws, row, 2, status, bg=sbg, fg=sfg, bold=True, align="center")
        pcell(ws, row, 3, desc,   bg="FAFAFA")
        ws.row_dimensions[row].height = 36
        row += 1

    # Spacer
    row += 1
    divider(ws, row, 3, "EXPLICIT / KINKY-LEWD / SENSUAL BUTTONS (from event_buttons.csv)", bg="8A2E2E")
    row += 1

    # Column headers for button list
    btn_cols = [
        ("EVENT ID",       14),
        ("HEADLINE",       36),
        ("CONTENT FLAG",   13),
        ("BTN TYPE",       13),
        ("BUTTON TEXT",    30),
        ("OUTCOME",        26),
        ("BTN ID",         16),
    ]
    for ci, (h, w) in enumerate(btn_cols, 1):
        hcell(ws, row, ci, h, w)
    row += 1

    events_by_id = {}
    events_path = os.path.join("data", "events.csv")
    try:
        with open(events_path, newline="", encoding="utf-8") as f:
            for rec in csv.DictReader(f):
                events_by_id[rec.get("event_id","").strip()] = rec
    except FileNotFoundError:
        pass

    buttons_path = os.path.join("data", "event_buttons.csv")
    try:
        with open(buttons_path, newline="", encoding="utf-8") as f:
            for rec in csv.DictReader(f):
                bt = rec.get("button_type","").strip().lower()
                if bt not in ("explicit", "kinky_lewd", "sensual"):
                    continue
                eid  = rec.get("event_id","").strip()
                ev   = events_by_id.get(eid, {})
                headline = ev.get("headline", "")
                cflag    = ev.get("content_flag","").strip() or "standard"

                is_kinky    = bt == "kinky_lewd"
                is_explicit = bt == "explicit"
                bg = "FF4444" if is_kinky else (EXPLICIT_BG if is_explicit else "FFF3CC")
                fg = "FFFFFF" if (is_kinky or is_explicit) else NORMAL_FG

                outcome_str = (rec.get("outcome_type","") + " → " +
                               str(rec.get("outcome_value",""))).strip(" →")

                pcell(ws, row, 1, eid,                          bg=bg, fg=fg, bold=True)
                pcell(ws, row, 2, headline,                     bg="FAFAFA")
                pcell(ws, row, 3, cflag,                        bg=bg, fg=fg, align="center")
                pcell(ws, row, 4, bt,                           bg=bg, fg=fg, bold=True, align="center")
                pcell(ws, row, 5, rec.get("button_text",""),    bg="FAFAFA")
                pcell(ws, row, 6, outcome_str,                  bg="FAFAFA")
                pcell(ws, row, 7, rec.get("button_id",""),      bg="FAFAFA")
                ws.row_dimensions[row].height = 20
                row += 1
    except FileNotFoundError:
        pcell(ws, row, 1, "event_buttons.csv NOT FOUND", bg="FCE0D6")

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — GOVERNORS & COMMANDERS
# ══════════════════════════════════════════════════════════════════════════════

GOVS_DATA = [
    # ── GOVERNOR CORE ──────────────────────────────────────────────────────────
    ("FULL PASS", "Governor Core", "Governor Class",
     "governor.gd",
     "governorType (name string), governorLevel 1–3, governorArchetypeId (ARC_01–25 / CA_ARC_01+), "
     "loyalty −20 to +20, morale 0–100, isVicePresident bool, isLeader bool, questComplete bool. "
     "govMilModsLvl1/2/3 arrays hold tier-tiered mil mods.",
     "morale var (0–100) may not be correctly initialized — see Major Bugs."),

    ("FULL PASS", "Governor Core", "Governor Loyalty",
     "governor.gd, world.gd",
     "loyalty −20 to +20 per turn via update_loyalty(). "
     "questComplete uncaps ceiling to +20. TURNCOAT event fires at loyalty ≤ −8. "
     "governor_loyalty_change outcome clamped ±20.",
     ""),

    ("FULL PASS", "Governor Core", "tileMoralDecay",
     "tile.gd, world.gd",
     "0–100 political corruption per tile. CORRUPT events fire at ≥50. "
     "tile_moral_decay_change outcome wired. "
     "Cleaner commander mod designed to reduce decay/turn (unimplemented).",
     ""),

    ("FULL PASS", "Governor Core", "addMilMod() Pipeline",
     "governor.gd addMilMod(type, levelCode)",
     "Creates MilMod.new(), calls buildSelf(type) with Sprite2D guard, "
     "adds to govMilModsLvl1/2/3 based on levelCode (123/23/3). "
     "Programmatic mods work without scene node due to guard.",
     ""),

    # ── NAMED GOVERNORS ────────────────────────────────────────────────────────
    ("FULL PASS", "Named Governors", "Ualani Carlisle (President)",
     "governor.gd",
     "isLeader=true. Mods: Visionary×123, Champion of the Sun×23, Healer×3, President×123. "
     "Protagonist; drives win/lose. Full detail in Ualani tab.",
     "Visionary/Champion/Healer mods have no calculateMilMods cases."),

    ("FULL PASS", "Named Governors", "Benjamin Tallmadge",
     "governor.gd",
     "Translator×123 (civilianMod). Espionage specialist.",
     ""),

    ("FULL PASS", "Named Governors", "Patrick Henry / Abigail Adams / Thomas Paine / Joseph Warren",
     "governor.gd",
     "Historical named governors with unique mod assignments. "
     "Patrick Henry (Virginia orator), Abigail Adams (Massachusetts), "
     "Thomas Paine (radical pamphleteer), Joseph Warren (MA doctor/soldier).",
     ""),

    ("FULL PASS", "Named Governors", "Daniel Shays / Phillis Wheatley / Francis Asbury",
     "governor.gd",
     "Daniel Shays (Shays' Rebellion leader), "
     "Phillis Wheatley (historical poet), Francis Asbury (Methodist bishop). "
     "All have unique mod sets per buildSelf().",
     ""),

    # ── PROCEDURAL GENERATION ──────────────────────────────────────────────────
    ("FULL PASS", "Proc Gen", "25 USA Archetypes + 10 Canadian Archetypes",
     "world.gd generateBarracksCommanders(), _apply_archetype_mods()",
     "ARC_01–25: Wetlands Fisher, Appalachian Miner, Ivy League Dropout, Seminole Fighter, "
     "Green Mountain Farmer, Chesapeake Shipwright, Loyalist Turncoat, Tobacco Belt Drifter, "
     "War Widow, Indigenous Scout, Boston Rabble-Rouser, Continental Surgeon, Nantucket Sailor, "
     "Frontier Preacher, DC Bureaucrat, Rust Belt Steelworker, Plantation Deserter, "
     "Swamp Witch, Caribbean Privateer, Hawaiian Refugee, Border Mercenary, "
     "Acadian Forest Ranger, Gettysburg Descendant, LGBTQ+ Organizer, Carnival Barker. "
     "CA_ARC_01–10: French-Canadian equivalents with NP_06 name pool. "
     "Each archetype gets 4 mods: T1a×123, T1b×123, T2×23, T3×3.",
     ""),

    ("FULL PASS", "Proc Gen", "Name Pool System (NP_01–06)",
     "world.gd generateBarracksCommanders()",
     "NP_01–05: USA cultural name pools by terrain/region affinity. "
     "NP_06: French-Canadian name pool. Pool selected by barracks tile terrain + cultural context.",
     ""),

    # ── COMMANDER ARC OBJECTIVES ────────────────────────────────────────────────
    ("FULL PASS", "Commander Arcs", "Arc Registration",
     "WarRoomPanel.gd registerCommanderArc()",
     "Called when governor hired. Builds arcData dict: "
     "governor, archetype_id, home_tile, objectives[3], objectives_complete[3], "
     "arc_active, arc_complete flags.",
     ""),

    ("FULL PASS", "Commander Arcs", "19 Objective Condition Evaluators",
     "WarRoomPanel.gd checkObjectives(), evaluateCondition()",
     "liberate_tile_terrain, liberate_specific_tile, liberate_tile_count, "
     "hold_tile_turns, commander_present_in_tile, liberate_tile_with_feature, "
     "turns_survived, build_in_home_tile, home_tile_corruption_below, "
     "home_tile_moral_decay_below, resource_threshold, liberate_tiles_in_state, "
     "home_tile_has_building, liberate_any_with_building, liberate_state_count + 4 more. "
     "Called each turn from world.gd _on_next_turn_pressed().",
     ""),

    ("FULL PASS", "Commander Arcs", "Arc Event Firing",
     "WarRoomPanel.gd → EventDatabase",
     "arcObjectiveCompleted signal fires ARC_xx_FIRST (obj 1 done) and ARC_xx_DONE (all done). "
     "50 total events across all 25 archetypes.",
     ""),

    # ── COMMANDER PROGRESSION ──────────────────────────────────────────────────
    ("FULL PASS", "Commander Progression", "Turn-Count Progression (CMD Events)",
     "world.gd _commander_turns dict",
     "turns_served tracked per governor. "
     "CMD_MERIT at 5t (level→1), CMD_RECOGNITION at 20t (→2), CMD_THANKS at 50t (→3). "
     "promote_commander outcome bumps governorLevel → activates higher-tier mods. "
     "CMD_THANKS has explicit button.",
     ""),

    ("FULL PASS", "Commander Progression", "Governor Level + Mod Tier Activation",
     "army.gd commanderCheck()",
     "governorLevel 1: T1 mods only; level 2: T1+T2; level 3: T1+T2+T3. "
     "commanderCheck() selects appropriate govMilModsLvl and distributes to units.",
     ""),

    # ── VP SYSTEM ──────────────────────────────────────────────────────────────
    ("FULL PASS", "VP System", "Vice President (isVicePresident)",
     "army.gd addUnitCommander(), governor.gd",
     "isVicePresident=true: all mods doubled (govMilModsLvl arrays appended twice). "
     "Gains Election Season mod. Loyalty doubled during election season. "
     "Can be any named or procedural governor.",
     "VP status set via event outcome; no dedicated appointment UI."),
]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — GAME SYSTEMS
# ══════════════════════════════════════════════════════════════════════════════

GAMESYS_DATA = [
    # ── CORE WORLD ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Core World", "Turn / Date System",
     "world.gd, tile.gd",
     "currentWorldTurn + month (1–12). evaluateDateEvents() called each turn. "
     "Month wraps → seasonal changes, storm spawns, calendar events. "
     "_on_next_turn_pressed() is the main turn-end loop.",
     ""),

    ("FULL PASS", "Core World", "Tile Ownership",
     "world.gd, tile.gd",
     "tileOwner string per tile. OwnedTileList on country. "
     "liberate_tile / lose_tile outcomes update ownership and fire CL/CX events. "
     "Presidential claim calculated from owned tiles.",
     ""),

    ("FULL PASS", "Core World", "Resource System",
     "world.gd, country.gd",
     "12 resources: food, gold, weapons, manpower, wood, metal, magic, science, "
     "culture, influence, harmony, faith. _apply_resource_change(resource, amount) on country. "
     "Army costs deducted each turn via surveySelf() aggregation.",
     "Weapon cost calc uses legacy names — new weapon armies report 0 cost."),

    ("FULL PASS", "Core World", "Building System",
     "tile.gd, building.gd",
     "tileBuildingsList per tile. disabled_buildings dict. "
     "disable_building / enable_building outcomes. "
     "Barracks, Forge, Courthouse, Fortress etc. checked in event conditions.",
     ""),

    ("FULL PASS", "Core World", "Presidential Claim System",
     "world.gd _calculate_presidential_claim()",
     "Calculates republic legitimacy score from owned tiles + loyalty + morale. "
     "Drives secession + legitimacy crisis events.",
     ""),

    ("FULL PASS", "Core World", "State Secession",
     "world.gd checkStateSecessionConditions()",
     "rebel_ flag per state. STATE_REBEL_01 → STATE_REINTEGRATED_01 chain. "
     "Mass secession cascades toward _execute_republic_collapse().",
     "Individual cascade logic to full collapse is partial."),

    ("FULL PASS", "Core World", "Republic Collapse",
     "world.gd _execute_republic_collapse()",
     "UK receives coastal tiles (182, 65-67, 151, 114, 123, 195). "
     "Interior tiles grouped by state code → spawn state country objects. "
     "USA retains DC (tile 188). COLLAPSE_02 fires. _republic_collapsed flag set.",
     "No game-over sequence wired post-collapse."),

    ("FULL PASS", "Core World", "Weather / Storm System",
     "tile.gd, world.gd _tick_storms(), _spawn_storm()",
     "Storms spawn 3%/turn on random tile. Type by winterScore+terrain. "
     "Spreads to TileNeighbors; duration 2–6 turns; intensity 1–3. "
     "Types: Tornado, Hurricane, Thunderstorm, Blizzard, Nor'easter, Fog. "
     "_apply_storm_debuffs() applies status effects to armies in storm tiles each turn start.",
     ""),

    ("FULL PASS", "Core World", "Winter Score + Supply Drain",
     "tile.gd winterScore, world.gd _apply_winter_army_drain()",
     "winterScore per tile (−80 to 80+). Categories: Extreme Hurricane → Blizzard. "
     "Months 11–2: food drain per army by cold category. Cold Weather armyTag exempts.",
     ""),

    # ── FACTIONS ───────────────────────────────────────────────────────────────
    ("FULL PASS", "Factions", "Faction System",
     "faction_control.gd, world.gd",
     "Loyalty scores per faction; changeFactionLoyalty(). 9 USA factions defined. "
     "faction_loyalty_change outcome wired in executeOutcome().",
     ""),

    ("PARTIAL", "Factions", "Espionage System",
     "tile.gd, world.gd",
     "espionageActive flag on tile. has_neighbor_with_espionage() helper. "
     "Culper Ring flag. sabotaged/reconDebuffed army flags set by spy outcomes. "
     "Wired to UALANI_CULPER_01 event condition.",
     "No spy unit type. No full spy action loop. Army flag wiring from events is partial."),

    ("PARTIAL", "Factions", "Codebook / Spell Schools",
     "world.gd, spell_schools_control.gd",
     "levelUpSchool(); spell school tiers defined; protector spell unlocks referenced. "
     "TotalMagic on country drained by protector buff upkeep.",
     "Full casting loop (player selecting and casting spells) TBD."),

    # ── UI & INFRASTRUCTURE ──────────────────────────────────────────────────────
    ("FULL PASS", "UI", "Camera / Map Navigation",
     "camera_movement_controller.gd",
     "Pan/zoom; tile click-to-select; path control. Army drag/movement.",
     ""),

    ("FULL PASS", "UI", "Resource Bar",
     "ResourceBar.gd",
     "Displays all 12 resource totals. Updates on turn end.",
     ""),

    ("FULL PASS", "UI", "Military Panel",
     "military_panel_control.gd",
     "Army roster; unit list; commander picker; Melee/Ranged attack buttons. "
     "attackBlocked + reinforcementBlocked flags checked before triggering battles.",
     "Visual category separator labels TODO (minor cosmetic)."),

    ("FULL PASS", "UI", "Governor Selection Panel",
     "governor_selection.gd",
     "Shows governor level + loyalty; assign to tile; pick from available pool.",
     ""),

    ("FULL PASS", "UI", "War Room Panel",
     "WarRoomPanel.gd",
     "Commander arc tracking; 3-objective display per active governor. "
     "registerCommanderArc() + checkObjectives() evaluated each turn.",
     "Checkmark/circle art assets not yet loaded (TODO in WarRoomPanel.gd:33)."),

    ("PARTIAL", "UI", "Tech Tree",
     "tech_tree.gd",
     "UI exists; add_tech outcome wired in executeOutcome(). Research nodes defined.",
     "Full research action loop (player spending resources to research) TBD."),

    ("PARTIAL", "UI", "Civilian Control",
     "civilian_control.gd, civilian_action_buttons.gd",
     "Civilian unit action buttons; some wired. Park Ranger, Cartographer etc. defined.",
     "Most civilian tool mod gameplay effects are IDEA-tier — no per-turn hooks."),

    ("PARTIAL", "UI", "Path / Movement Control",
     "path_control.gd, army.gd",
     "pathLine and targetTile vars; movement across tiles via tile cost system. "
     "Road level reduces cost. Cold Weather exemptions.",
     "'Not enough movement points' indicator TODO (path_control.gd:204)."),

    # ── SPAWN ───────────────────────────────────────────────────────────────────
    ("FULL PASS", "Spawn", "Barracks Army Generation (game start)",
     "world.gd generateBarracksCommanders()",
     "On game start: scans all player barracks tiles, selects archetype by terrain, "
     "generates commander, calls _apply_archetype_mods(), assigns to tile army. "
     "Canadian barracks use CA_ARC pool.",
     ""),

    ("FULL PASS", "Spawn", "Army Template from CSV",
     "army_database.gd, data/armies/army_templates.csv",
     "army_templates.csv: ArmyName, country, tile, armyMods (pipe-delimited tags). "
     "armyTags populated from armyMods column at spawn.",
     ""),
]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — WIN & LOSE CONDITIONS
# ══════════════════════════════════════════════════════════════════════════════

WINLOSE_DATA = [
    # ── LOSE CONDITIONS ─────────────────────────────────────────────────────────
    ("FIRST PASS", "Lose: Collapse", "_execute_republic_collapse() — Full Fragmentation",
     "world.gd ~line 1925",
     "Triggered when republic morale/legitimacy reaches critical threshold. "
     "UK receives coastal tiles (182, 65-67, 151, 114, 123, 195). "
     "Interior tiles grouped by state code → spawn independent state country objects. "
     "USA retains only DC (tile 188) — Ualani's last stronghold. "
     "_republic_collapsed flag set. COLLAPSE_02 fires.",
     "Game continues after collapse — no automatic lose sequence. "
     "Player can theoretically rebuild from DC. Rebuild win path is IDEA."),

    ("FIRST PASS", "Lose: Collapse", "COLLAPSE_01 — 'The Republic Has Fallen'",
     "events.csv",
     "Fires when all major metropolitan tiles are lost / surrender terms triggered. "
     "All major city falls consolidated into this narrative event. "
     "Explicit content buttons available.",
     "No game-over screen wired to this event."),

    ("FIRST PASS", "Lose: Collapse", "COLLAPSE_02 — 'Washington Stands Alone'",
     "events.csv, world.gd _execute_republic_collapse()",
     "Fired directly from _execute_republic_collapse(). "
     "Ualani's final stand narrative. DC as last holdout. "
     "Explicit content option. Represents deepest lose state currently modeled.",
     "No game-over trigger wired after this event."),

    ("FIRST DRAFT", "Lose: Death", "Head of State Death → Game Over",
     "world.gd ~line 4001",
     "_game_ended bool flag exists and is checked in election logic. "
     "Code priority comment: 'head of state death → game over (Ualani for USA)'. "
     "TODO: '# call _trigger_game_over() once game over screen exists'.",
     "Game over screen scene not built. Ualani death has no mechanical consequence yet. "
     "HIGH PRIORITY for ship."),

    ("FIRST PASS", "Lose: Secession", "State Secession Cascade",
     "world.gd checkStateSecessionConditions()",
     "rebel_ flags per state fire STATE_REBEL_01 and remove tiles from player control. "
     "Reintegration via STATE_REINTEGRATED_01 possible. "
     "Mass secession feeds into _execute_republic_collapse() path.",
     "Individual cascade logic from multi-state secession to full collapse is partial."),

    ("IDEA", "Lose: Occupation", "Full Territory Loss",
     "(no implementation)",
     "Design intent: losing all player-controlled tiles → game over. "
     "No zero-tiles check exists. No loss-on-zero-territories logic.",
     ""),

    # ── WIN CONDITIONS ─────────────────────────────────────────────────────────
    ("IDEA", "Win: Unification", "Full Continental Unification",
     "(no implementation)",
     "Design intent: liberate all US + Canadian territories → victory. "
     "No all-tiles-owned check. No victory screen.",
     "Most important win condition to define and implement."),

    ("IDEA", "Win: Diplomatic", "Canadian Alliance + UK Peace Treaty",
     "events.csv (CAN_PEACE_01, PEACE_ALLIED, PEACE_USA)",
     "Peace treaty events are authored (CAN_PEACE_01, PEACE_ALLIED, PEACE_USA). "
     "Design intent: achieving diplomatic resolution = win state or major milestone. "
     "No win-condition check fires from these event outcomes.",
     ""),

    ("IDEA", "Win: Protectors", "All 25 Protectors Agree",
     "events.csv PROT_xx_AGREE + CA_PROT_xx_AGREE",
     "All 17 USA + 8 Canadian protectors have AGREE stage events fully authored. "
     "Design intent: recruiting all 25 = win trigger or major milestone. "
     "No all-protectors-agreed win check exists.",
     ""),

    ("IDEA", "Win: Screen", "Victory + Game Over Screens",
     "(no implementation)",
     "No victory screen UI. No game over screen UI. "
     "_game_ended flag exists but has no visual/audio/mechanical consequence. "
     "Both screens required for any playable build.",
     "HIGHEST PRIORITY for any public demo or ship candidate."),
]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — MAJOR BUGS
# ══════════════════════════════════════════════════════════════════════════════

BUGS_DATA = [
    ("CRITICAL", "_apply_army_buff() is a complete stub",
     "world.gd _apply_army_buff() ~line 3351",
     "Function is called by event button outcomes with army_buff outcome type but "
     "the entire body is a single `pass` statement — does nothing. "
     "Every army_buff event outcome in the entire game is silently ignored.",
     "Implement: iterate countryArmyList, filter by tile if provided, "
     "call army.apply_status(buff_type, duration) for matching armies."),

    ("HIGH", "Game over / victory screens not implemented",
     "world.gd ~line 4001; (no scene file)",
     "_game_ended bool and _game_won bool exist. Head of state death check written "
     "with TODO comment. No GameOver scene. No Victory scene. "
     "_trigger_game_over() is referenced but not defined.",
     "Build GameOver and Victory scenes. Define _trigger_game_over() and "
     "_trigger_game_victory(). Wire to head-of-state death and win condition checks."),

    ("HIGH", "Explicit content has no global settings toggle",
     "event_scene.gd _build_buttons(); settings system",
     "button_type filtering code exists in event_scene.gd but no global "
     "Settings.enableExplicit flag was found. Explicit buttons render unconditionally "
     "for all players regardless of preference or age.",
     "Add Settings.enableExplicit bool (default false). Check it in _build_buttons() "
     "before rendering explicit/kinky_lewd buttons. Add toggle to settings menu."),

    ("MEDIUM", "morale vs loyalty: commander bonus may always be 1.0×",
     "army.gd surveySelf() ~line 371",
     "Morale multiplier uses commander.morale (0–100 field) but governor.gd tracks "
     "loyalty (−20 to +10). If morale is never set/initialized it stays 0, "
     "so armyPunch multiplier is always ×1.0 — no commander ever gives a bonus.",
     "Audit governor morale initialization. Map loyalty to 0–100 for morale input, "
     "or verify a separate morale var is correctly set."),

    ("MEDIUM", "Weapon cost calc uses legacy weapon names only",
     "army.gd surveySelf() lines ~357-365",
     "Match block calculating armyWeaponsCost checks old DODK names (Spear, Club, Atlatl). "
     "No case for Cutlass, Flintlock, Field Gun, etc. "
     "All new-weapon armies show 0 weapons cost.",
     "Add Saber/Musket/Artillery weapon cost cases to the match block."),

    ("MEDIUM", "commanderCheck() accumulates duplicate mods on repeated updateArmyUI calls",
     "army.gd commanderCheck()",
     "commanderCheck() appends mods to each unit on every call. "
     "Units clear militaryModifierList in getUnitAttributes() but commanderCheck runs after. "
     "Rapid successive updateArmyUI calls can double/triple mods on units.",
     "Verify getUnitAttributes() fully clears militaryModifierList before commanderCheck. "
     "Audit call order in updateArmyUI()."),

    ("MEDIUM", "No UI for active army status effects",
     "army.gd armyStatusEffects; military panel",
     "Army can have Diseased, Routed, Terrified, Waterlogged etc. active with no visual feedback. "
     "Player cannot see what statuses are active or how many turns remain.",
     "Add status effect icon row to military panel or army tooltip. "
     "Iterate armyStatusEffects and display type + turnsLeft."),

    ("LOW", "armyShield vs armyMaxShield display inconsistency",
     "army.gd updateFinalTotals()",
     "updateFinalTotals() passes armyShield for both current and max in armyCostUI3. "
     "Shield depletion bar never shows correctly.",
     "Change: armyCostUI3.updateSelf(armyShield, armyMaxShield)."),

    ("LOW", "Nassau/BA faction logic stub",
     "country.gd ~line 1180",
     "Comment: 'TODO: BA (Nassau) is opportunist — logic stub'. "
     "Bahamas has no behavioral AI logic implemented.",
     "Implement opportunist faction behavior for BA country."),

    ("LOW", "~25 debug print statements throughout codebase",
     "Various .gd files",
     "print() debug statements: 'DEBUG CLICK', 'RETURN DEBUG', 'ARMYARRIVEDDEBUG', "
     "'ControlDEBUG' etc. Non-critical but noisy in console.",
     "Search for print( and remove/comment all debug prints before any public build."),

    ("LOW", "No 'not enough movement points' UI feedback",
     "path_control.gd ~line 204",
     "TODO comment in path_control.gd. Player gets no visual indication when "
     "they try to move an army without sufficient movement points.",
     "Add tooltip or panel flash when movement fails due to insufficient points."),

    ("LOW", "War Room arc checkmark/circle art assets missing",
     "WarRoomPanel.gd:33-34, CommanderArcEntry.gd:33-34",
     "TODO comments: load checkmark and circle assets for objective display icons. "
     "Currently no visual state indicator for completed vs incomplete objectives.",
     "Load and wire checkmark/circle textures for arc objective UI nodes."),
]


def build_bugs_sheet(wb):
    ws = wb.create_sheet("Major Bugs")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("SEVERITY",    12),
        ("BUG",         34),
        ("FILE / LINE", 34),
        ("DESCRIPTION", 58),
        ("SUGGESTED FIX", 44),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    row = 2
    for (sev, bug, loc, desc, fix) in BUGS_DATA:
        sbg, sfg = SEV_COLORS.get(sev, ("FFFFFF", NORMAL_FG))
        pcell(ws, row, 1, sev,  bg=sbg, fg=sfg, bold=True, align="center")
        pcell(ws, row, 2, bug,  bg="FAFAFA", bold=True)
        pcell(ws, row, 3, loc,  bg="FAFAFA")
        pcell(ws, row, 4, desc, bg="FAFAFA")
        pcell(ws, row, 5, fix,  bg="FFFBE6")
        ws.row_dimensions[row].height = 54
        row += 1

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    std_sheet(wb, "Army",           ARMY_DATA)
    std_sheet(wb, "Events",         EVENTS_DATA)
    std_sheet(wb, "Ualani",         UALANI_DATA)
    build_explicit_sheet(wb)
    std_sheet(wb, "Govs & Cmds",    GOVS_DATA)
    std_sheet(wb, "Game Systems",   GAMESYS_DATA)
    std_sheet(wb, "Win & Lose",     WINLOSE_DATA)
    build_bugs_sheet(wb)

    wb.save(OUT_PATH)

    army_fp   = sum(1 for r in ARMY_DATA    if r[0] == "FULL PASS")
    army_pp   = sum(1 for r in ARMY_DATA    if r[0] == "FIRST PASS")
    army_fd   = sum(1 for r in ARMY_DATA    if r[0] == "FIRST DRAFT")
    army_idea = sum(1 for r in ARMY_DATA    if r[0] == "IDEA")
    army_stub = sum(1 for r in ARMY_DATA    if r[0] == "STUB")
    ev_total  = len(EVENTS_DATA)
    win_idea  = sum(1 for r in WINLOSE_DATA if r[0] == "IDEA")
    win_imp   = sum(1 for r in WINLOSE_DATA if r[0] in ("FULL PASS","FIRST PASS","FIRST DRAFT"))
    crit_bugs = sum(1 for b in BUGS_DATA    if b[0] == "CRITICAL")
    high_bugs = sum(1 for b in BUGS_DATA    if b[0] == "HIGH")

    print(f"Saved: {OUT_PATH}")
    print(f"  Army       : {army_fp} FULL PASS / {army_pp} FIRST PASS / "
          f"{army_fd} FIRST DRAFT / {army_idea} IDEA / {army_stub} STUB")
    print(f"  Events     : {ev_total} categories documented")
    print(f"  Ualani     : {len(UALANI_DATA)} systems/events documented")
    print(f"  Win & Lose : {win_imp} implemented conditions / {win_idea} IDEA")
    print(f"  Bugs       : {crit_bugs} CRITICAL / {high_bugs} HIGH / "
          f"{len(BUGS_DATA) - crit_bugs - high_bugs} MEDIUM-LOW")


if __name__ == "__main__":
    main()
