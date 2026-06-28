"""
Generates event_audit.xlsx — every triggerable event in the game, joined from
the source-of-truth CSVs the runtime actually loads:
  data/events.csv          (325 events: headline / writing / image_tag / flags)
  data/event_buttons.csv   (540 buttons, joined by event_id)
  data/event_triggers.csv  (89 triggers, joined by event_id)

One row per event. Columns: id, type, country, headline, subtitle, writing,
image tag + resolved EventArt png (or MISSING), buttons, trigger(s), content
flag, repeat/cooldown, and an empty Notes column for the audit pass.

Run from repo root:  python3 scripts/build_event_audit.py
"""
import csv, os
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EVENTS   = os.path.join(ROOT, "data/events.csv")
BUTTONS  = os.path.join(ROOT, "data/event_buttons.csv")
TRIGGERS = os.path.join(ROOT, "data/event_triggers.csv")
EVENTART = os.path.join(ROOT, "art assets/AmericanRevolutionArt/EventArt")
OUT      = os.path.join(ROOT, "event_audit.xlsx")

def rows(path):
    with open(path, encoding="utf-8") as f:
        return [{k: (v if v is not None else "") for k, v in row.items()} for row in csv.DictReader(f)]

events = rows(EVENTS)
buttons_by = defaultdict(list)
for b in rows(BUTTONS):
    buttons_by[b["event_id"]].append(b)
trigs_by = defaultdict(list)
for t in rows(TRIGGERS):
    trigs_by[t["event_id"]].append(t)

# event_type -> broad category (for colour grouping)
def category(t):
    t = t or ""
    if t.startswith("commander"): return "Commander"
    if t.startswith("ca_protector") or t.startswith("protector"): return "Protector"
    if t == "loyal_governor": return "Governor"
    if t.startswith("ualani"): return "Ualani"
    if t == "white_house_secret": return "White House"
    if t in ("vp_event","ca_vp","stump_speech") or t.startswith("election"): return "VP / Election"
    if t.startswith("canada") or t.startswith("border") or t.startswith("ca_"): return "Canada"
    if t in ("date","peace","war_buildup","war_declaration"): return "Date / War"
    if t in ("city_liberated","city_lost","state_liberated","secession","reintegration",
             "collapse","fort_reward","fort_disrepair","fort_confrontation","fort_private",
             "harbor_threat","harbor_secured","harbor_burned","forge_threat","forge_reinforced",
             "forge_evacuated","corruption_crisis","corrupt_trial","corrupt_reform","corrupt_deal",
             "harvest_crisis","harvest_resolution","garrison_starving","legitimacy_crisis",
             "turncoat_discovered","memorial_lost","memorial_restored","tile_crisis"): return "Crisis / Tile"
    return "Other"

CAT_FILL = {
    "Commander":"D6E1F7", "Protector":"EDD6F7", "Governor":"F5ECD5", "Ualani":"D6F5EA",
    "White House":"F5E6F0", "VP / Election":"F5E6D6", "Canada":"D6EAF8",
    "Date / War":"FFF3CC", "Crisis / Tile":"FDEBD0", "Other":"ECECEC",
}

def resolve_art(tag):
    tag = (tag or "").strip()
    if not tag:
        return ("", "— no image_tag —")
    p = os.path.join(EVENTART, tag + ".png")
    if os.path.exists(p):
        return (tag, "EventArt/" + tag + ".png")
    return (tag, "✗ MISSING — needs art (placeholder shown in-game)")

def fmt_buttons(bs):
    out = []
    for i, b in enumerate(bs, 1):
        parts = [f'{i}. "{b["button_text"]}"']
        oc = b.get("outcome_type","").strip()
        if oc:
            ov = b.get("outcome_value","").strip(); oa = b.get("outcome_amount","").strip()
            seg = oc + (f"={ov}" if ov else "") + (f" ({oa})" if oa else "")
            parts.append("→ " + seg)
        lc = b.get("loyalty_change","").strip()
        if lc and lc != "0": parts.append(f"loyalty {lc}")
        nx = b.get("next_event_id","").strip()
        if nx: parts.append("next: " + nx)
        pr = b.get("prerequisite_flag","").strip()
        if pr: parts.append("needs: " + pr)
        out.append("  ".join(parts))
    return "\n".join(out) if out else "— no buttons —"

def fmt_triggers(ts):
    if not ts:
        return "— no direct trigger row (fires as a chain via a button's next_event_id, or from a world.gd check function) —"
    out = []
    for t in ts:
        seg = [t["trigger_type"]]
        tmin, tmax = t.get("turn_min","").strip(), t.get("turn_max","").strip()
        if tmin or tmax: seg.append(f"turn {tmin or '?'}–{tmax or '?'}")
        for key, lbl in (("tile_id","tile"),("tile_owner","owner"),("state_code","state"),
                         ("commander_archetype","cmd"),("protector_id","prot"),
                         ("min_liberty_score","minLib"),("min_corruption","minCorr"),
                         ("trigger_condition","cond"),("priority","prio")):
            v = t.get(key,"").strip()
            if v and v not in ("0",""): seg.append(f"{lbl}={v}")
        out.append(" | ".join(seg))
    return "\n".join(out)

# ── Build workbook ───────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Events"

HEADERS = ["Event ID","Type","Country","Headline","Subtitle","Writing (long_desc)",
           "Image Tag","Art File","Buttons (choices → outcome)","Trigger(s)",
           "Content","Repeat / CD","Notes (audit)"]
WIDTHS  = [20, 20, 9, 30, 26, 60, 20, 30, 46, 44, 11, 11, 26]

HEAD_FILL = PatternFill("solid", fgColor="33415C")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BASE_FONT = Font(name="Arial", size=10)
MONO_FONT = Font(name="Consolas", size=9)
TOP = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MISS_FILL = PatternFill("solid", fgColor="FFD9CC")
NSFW_FILL = PatternFill("solid", fgColor="F5D6E5")

for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    cell = ws.cell(1, c, h); cell.fill = HEAD_FILL; cell.font = HEAD_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = BORDER
    ws.column_dimensions[chr(64+c)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

events.sort(key=lambda e: (category(e["event_type"]), e["event_type"], e["event_id"]))
r = 2
for e in events:
    tag, art = resolve_art(e["image_tag"])
    content = e.get("content_flag","").strip()
    content_disp = content if content and content != "false" else "SFW"
    rep = "yes" if e.get("repeatable") in ("true","True") else "no"
    cd  = e.get("cooldown_turns","0").strip() or "0"
    vals = [e["event_id"], e["event_type"], e["country_cid"], e["headline"], e["short_desc"],
            e["long_desc"], tag, art, fmt_buttons(buttons_by.get(e["event_id"], [])),
            fmt_triggers(trigs_by.get(e["event_id"], [])), content_disp,
            f"{rep} / {cd}", ""]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.alignment = TOP; cell.border = BORDER
        cell.font = MONO_FONT if c in (9, 10) else BASE_FONT
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=CAT_FILL[category(e["event_type"])])
    if art.startswith("✗"): ws.cell(r, 8).fill = MISS_FILL
    if content in ("sensual","kinky","explicit"): ws.cell(r, 11).fill = NSFW_FILL
    r += 1

# ── Summary sheet ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
miss = sum(1 for e in events if resolve_art(e["image_tag"])[1].startswith("✗"))
has_art = sum(1 for e in events if resolve_art(e["image_tag"])[1].startswith("EventArt"))
no_tag = sum(1 for e in events if not (e["image_tag"] or "").strip())
with_trig = len(set(t["event_id"] for t in rows(TRIGGERS)))
nsfw = sum(1 for e in events if e.get("content_flag","").strip() in ("sensual","kinky","explicit"))
summary = [
    ("Total events", len(events)),
    ("Events WITH a direct trigger row", with_trig),
    ("Events with NO trigger row (chain / world.gd driven)", len(events) - with_trig),
    ("Events with real art (EventArt png exists)", has_art),
    ("Events referencing MISSING art (tag set, no png)", miss),
    ("Events with NO image_tag (default placeholder)", no_tag),
    ("NSFW-flagged events (sensual/kinky/explicit)", nsfw),
    ("Total buttons", sum(len(v) for v in buttons_by.values())),
    ("Total trigger rows", sum(len(v) for v in trigs_by.values())),
]
ws2.cell(1,1,"Event Audit — Summary").font = Font(name="Arial", bold=True, size=14)
ws2.column_dimensions["A"].width = 52; ws2.column_dimensions["B"].width = 12
r = 3
for label, n in summary:
    ws2.cell(r,1,label).font = BASE_FONT
    ws2.cell(r,2,n).font = Font(name="Arial", bold=True, size=10); r += 1
r += 1
ws2.cell(r,1,"Events by type").font = Font(name="Arial", bold=True, size=11); r += 1
for t, n in Counter(e["event_type"] for e in events).most_common():
    ws2.cell(r,1,t).font = BASE_FONT; ws2.cell(r,2,n).font = BASE_FONT; r += 1

wb.save(OUT)
print(f"Wrote {OUT}: {len(events)} events, {sum(len(v) for v in buttons_by.values())} buttons, {sum(len(v) for v in trigs_by.values())} triggers")
print(f"Missing art: {miss} | With trigger: {with_trig} | NSFW: {nsfw}")
