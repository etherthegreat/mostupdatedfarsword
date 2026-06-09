#!/usr/bin/env python3
"""
Generate editortasks_masterdoc.xlsx — every deferred Godot editor task.

HOW TO ADD NEW TASKS:
  Append entries to the TASKS list below.
  Each entry is a dict with these keys:
    id          – unique short ID, e.g. "WAR-003"
    status      – "OPEN" | "IN PROGRESS" | "DONE"
    scene       – .tscn or scene name, e.g. "WarRoomPanel.tscn"
    category    – "New Scene" | "Node Connection" | "Asset" | "Signal Wire"
                  | "Layout/UI" | "Game Flow" | "Bug Fix"
    description – one-sentence description of the change
    detail      – full step-by-step notes for doing the work in the editor
    added_in    – short label for the feature/PR that created this task,
                  e.g. "Commander Arc system"
    priority    – "HIGH" | "MEDIUM" | "LOW"
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette ───────────────────────────────────────────────────────────────────
HDR_BG   = "1F2D3D"
HDR_FG   = "FFFFFF"
CAT_BG   = "2E4057"
CAT_FG   = "FFFFFF"
OPEN_BG  = "FFD6D6"   # red-tint  — OPEN
WIP_BG   = "FFF3CD"   # amber     — IN PROGRESS
DONE_BG  = "D6F5D6"   # green     — DONE
ALT_BG   = "EEF2F7"
WHITE    = "FFFFFF"

def _fill(h): return PatternFill("solid", fgColor=h)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _status_fill(s):
    return _fill(OPEN_BG if s == "OPEN" else WIP_BG if s == "IN PROGRESS" else DONE_BG)

# ── columns ───────────────────────────────────────────────────────────────────
COLUMNS = [
    ("ID",           10),
    ("Status",       14),
    ("Scene / File", 32),
    ("Category",     18),
    ("Description",  46),
    ("Detail / Steps", 60),
    ("Added In",     28),
    ("Priority",     12),
]

# ── task list ─────────────────────────────────────────────────────────────────
TASKS = [

    # ── War Room Panel ────────────────────────────────────────────────────────
    {
        "id":          "WAR-001",
        "status":      "OPEN",
        "scene":       "WarRoomPanel.tscn",
        "category":    "Node Connection",
        "description": "Add CloseButton node and wire its pressed signal",
        "detail":      (
            "1. Open WarRoomPanel.tscn in the editor.\n"
            "2. Add a Button node named 'CloseButton' as a child of PanelBackground.\n"
            "3. Position it at the top-right corner of the panel.\n"
            "4. In the Node panel, connect its 'pressed' signal to "
            "_on_close_button_pressed() in WarRoomPanel.gd (line 814)."
        ),
        "added_in":    "Commander / Protector Arc system",
        "priority":    "HIGH",
    },
    {
        "id":          "WAR-002",
        "status":      "OPEN",
        "scene":       "WarRoomPanel.tscn",
        "category":    "Bug Fix",
        "description": "Fix tab name typo: 'DEPTARTMENT' → 'DEPARTMENT' (remove double-space too)",
        "detail":      (
            "1. Open WarRoomPanel.tscn.\n"
            "2. Select the tab node currently named "
            "'DEPTARTMENT OF MYTHOLOGICAL  AFFAIRS' (note: typo + double-space).\n"
            "3. Rename it to 'DEPARTMENT OF MYTHOLOGICAL AFFAIRS'.\n"
            "4. Update the matching string in WarRoomPanel.gd line 57 to match exactly."
        ),
        "added_in":    "Commander / Protector Arc system",
        "priority":    "MEDIUM",
    },

    # ── CommanderArcEntry ─────────────────────────────────────────────────────
    {
        "id":          "CMD-001",
        "status":      "OPEN",
        "scene":       "CommanderArcEntry.tscn",
        "category":    "Asset",
        "description": "Set checkTexture and circleTexture for objective completion indicators",
        "detail":      (
            "CommanderArcEntry.gd lines 33-34 have:\n"
            "  var checkTexture = null   # TODO: load your checkmark asset\n"
            "  var circleTexture = null  # TODO: load your empty circle asset\n\n"
            "1. Open CommanderArcEntry.gd.\n"
            "2. Set checkTexture to a green checkmark — zGreen.png is already "
            "imported (res://art assets/finishedAssets/armyicons/finished/zGreen.png).\n"
            "3. Set circleTexture to an empty circle/hollow-dot texture "
            "(create or import one if needed).\n"
            "These drive the filled/empty objective row indicators."
        ),
        "added_in":    "Commander Arc system",
        "priority":    "MEDIUM",
    },

    # ── ProtectorArcEntry ─────────────────────────────────────────────────────
    {
        "id":          "PROT-001",
        "status":      "OPEN",
        "scene":       "ProtectorArcEntry.tscn",
        "category":    "Asset",
        "description": "Set checkTexture and circleTexture for prayer/devotion indicators",
        "detail":      (
            "ProtectorArcEntry.gd lines 33-34 have:\n"
            "  var checkTexture = null   # TODO: load your checkmark asset\n"
            "  var circleTexture = null  # TODO: load your empty circle asset\n\n"
            "1. Open ProtectorArcEntry.gd.\n"
            "2. Set checkTexture to a green checkmark (match CMD-001 asset).\n"
            "3. Set circleTexture to an empty circle texture (match CMD-001 asset).\n"
            "These control the prayer completion display in the Dept. of Myth. Affairs tab."
        ),
        "added_in":    "Protector Arc system",
        "priority":    "MEDIUM",
    },

    # ── MilitaryPanelControl ──────────────────────────────────────────────────
    {
        "id":          "MIL-001",
        "status":      "OPEN",
        "scene":       "MilitaryPanelControl scene",
        "category":    "Layout/UI",
        "description": "Add visual separator labels between barracks groups in the military panel",
        "detail":      (
            "military_panel_control.gd line 30:\n"
            "  # TODO: add visual category separator labels between groups once UI nodes exist.\n\n"
            "The barracks list sorts into 3 groups:\n"
            "  1. Ualani's barracks (always pinned top)\n"
            "  2. Occupied barracks (army present)\n"
            "  3. Unoccupied barracks\n\n"
            "Add Label nodes or HSeparator+Label pairs as dividers in the "
            "ScrollContainer/GridContainer hierarchy. The code will inject them "
            "dynamically once the nodes exist. "
            "Coordinate with the dev to wire the insertion logic."
        ),
        "added_in":    "Military Panel redesign",
        "priority":    "LOW",
    },

    # ── Game Over Screen ──────────────────────────────────────────────────────
    {
        "id":          "GOV-001",
        "status":      "OPEN",
        "scene":       "world.tscn  →  new GameOverPanel",
        "category":    "New Scene",
        "description": "Create GameOverPanel scene and add it to CanvasLayer in world.tscn",
        "detail":      (
            "world.gd line 4104:\n"
            "  # TODO: call _trigger_game_over() here once game over screen exists\n\n"
            "1. Create a new scene: GameOverPanel.tscn (Control root).\n"
            "2. Add: background overlay, title label ('GAME OVER' or win/loss variant), "
            "subtitle/reason label, a Return to Menu button.\n"
            "3. Add GameOverPanel as a child of CanvasLayer in world.tscn, hidden by default.\n"
            "4. Connect its visibility and the Return button to _trigger_game_over() "
            "and _on_game_over_return_pressed() in world.gd.\n"
            "This unblocks the president-death → game-over flow."
        ),
        "added_in":    "President death / game-over scaffold",
        "priority":    "HIGH",
    },

    # ── Resource Bar ──────────────────────────────────────────────────────────
    {
        "id":          "RES-001",
        "status":      "OPEN",
        "scene":       "Resource Bar (TOP) scene",
        "category":    "Layout/UI",
        "description": "Add Manpower, Influence, Mandate, Happiness, and Boats labels to top resource bar",
        "detail":      (
            "world.gd lines 162-164 reference labels that may not exist in the bar yet:\n"
            "  $MandateLabel/Label, $HarmonyLabel/Label, $InfluenceLabel/Label, "
            "$ManpowerLabel/Label\n\n"
            "The tile census now emits all 13 resource types (Food, Dollars, Wood, Metal, "
            "Magic, Culture, Weapons, Science, Mandate, Happiness, Manpower, Influence, Boats).\n\n"
            "Check which label nodes are missing from the Resource Bar scene and add them. "
            "Each follows the pattern: HBoxContainer child → icon Sprite + Label child.\n"
            "Match existing nodes (FoodLabel, GoldLabel, etc.) for style."
        ),
        "added_in":    "Tile census all-resources expansion",
        "priority":    "HIGH",
    },

    # ── Tile Info Panel ───────────────────────────────────────────────────────
    {
        "id":          "TIP-001",
        "status":      "OPEN",
        "scene":       "tile_info_panel.tscn",
        "category":    "Layout/UI",
        "description": "Verify ManaPanelContainer has enough space to show all 13 resource mana panels",
        "detail":      (
            "tile_info_panel.gd now calls buildTileOutput() for up to 13 resource types "
            "(Food, Dollars, Wood, Metal, Magic, Culture, Weapons, Science, Mandate, "
            "Happiness, Manpower, Influence, Boats).\n\n"
            "1. Open tile_info_panel.tscn.\n"
            "2. Check $ManaPanelContainer — it likely uses an HBoxContainer or GridContainer.\n"
            "3. Ensure it wraps or scrolls so all active resource panels are visible.\n"
            "4. Consider switching to a GridContainer (3–4 columns) or a ScrollContainer "
            "wrapper if the panel overflows."
        ),
        "added_in":    "Tile census all-resources expansion",
        "priority":    "MEDIUM",
    },

    # ── Governor Selection ────────────────────────────────────────────────────
    {
        "id":          "GOV-002",
        "status":      "OPEN",
        "scene":       "governor_selection.tscn",
        "category":    "Layout/UI",
        "description": "Add 'Wizard' tab/section to GovernorSelection panel (visible when tile has wizard)",
        "detail":      (
            "tile_info_panel.gd has a $WizardButton that shows the tile wizard's name, "
            "but there is no full wizard info view yet.\n\n"
            "1. Open governor_selection.tscn.\n"
            "2. Add a 'Wizard' tab or expandable section that shows:\n"
            "   - Wizard name (wizardType)\n"
            "   - Magic school\n"
            "   - Active spells / buffs granted\n"
            "3. Wire it to the changePanel('wizard') call path in tile_info_panel.gd."
        ),
        "added_in":    "Protector agree → wizard assignment",
        "priority":    "LOW",
    },

    # ── Mana Panel ────────────────────────────────────────────────────────────
    {
        "id":          "MAN-001",
        "status":      "OPEN",
        "scene":       "mana_panel.tscn",
        "category":    "Asset",
        "description": "Add a proper Boats icon — currently reuses the Manpower icon",
        "detail":      (
            "mana_panel.gd line 47-48:\n"
            "  'Boats':\n"
            "      $ManaIcon.texture = load('.../manpower.png')  # placeholder\n\n"
            "1. Create or import a boat/ship icon asset.\n"
            "2. Save it to res://art assets/finishedAssets/manaicons/boats.png "
            "(or similar path).\n"
            "3. Update mana_panel.gd line 48 to reference the new asset."
        ),
        "added_in":    "Tile census all-resources expansion",
        "priority":    "LOW",
    },

    # ── Sound System ─────────────────────────────────────────────────────────
    {
        "id":          "SFX-001",
        "status":      "OPEN",
        "scene":       "world.tscn  /  AutoLoad",
        "category":    "New Scene",
        "description": "Create AudioManager autoload and add AudioStreamPlayer nodes for SFX bus",
        "detail":      (
            "See sound_masterdoc.xlsx for the full 129-SFX catalog.\n\n"
            "1. Create an AudioManager.gd singleton (autoload).\n"
            "2. Add one AudioStreamPlayer node per concurrent channel needed "
            "(UI bus, combat bus, ambient bus, event bus).\n"
            "3. In Project → Project Settings → Audio, add buses: "
            "Master / Music / SFX / Ambient.\n"
            "4. Wire each AudioStreamPlayer to the SFX bus.\n"
            "5. Expose play_sfx(sound_id: String) and stop_sfx(sound_id) methods.\n"
            "6. Register AudioManager in Project → AutoLoad."
        ),
        "added_in":    "Sound effects planning (sound_masterdoc)",
        "priority":    "MEDIUM",
    },
    {
        "id":          "SFX-002",
        "status":      "OPEN",
        "scene":       "world.tscn  /  all panels",
        "category":    "Signal Wire",
        "description": "Wire SFX calls to all HIGH-priority sound trigger points",
        "detail":      (
            "See sound_masterdoc.xlsx — filter Priority = HIGH for the full list.\n"
            "Key wiring points:\n"
            "  • End Turn button → ui_end_turn.wav\n"
            "  • Event panel shown → event_standard / event_crisis / event_ualani\n"
            "  • Building constructed → building_construct.wav\n"
            "  • Battle start/win/loss → army_battle_*.wav\n"
            "  • Season change → event_season_<season>.wav\n"
            "  • Protector agree → protector_agree.wav + protector_tower_built.wav\n"
            "  • Resource drops below 0 → resource_critical.wav\n\n"
            "Depends on SFX-001 (AudioManager) being in place first."
        ),
        "added_in":    "Sound effects planning (sound_masterdoc)",
        "priority":    "MEDIUM",
    },

    # ── MAIN MENU ─────────────────────────────────────────────────────────────
    {
        "id":          "MENU-001",
        "status":      "OPEN",
        "scene":       "main_menu.tscn",
        "category":    "Layout/UI",
        "description": "Rebuild main_menu.tscn: add VideoStreamPlayer background + 5-button left panel",
        "detail":      (
            "See SCENE STRUCTURE comment in main_menu.gd for the full node tree.\n\n"
            "1. Add a VideoStreamPlayer node (full screen, autoplay=true).\n"
            "   Set stream to your menu .ogv file.\n"
            "2. Add LeftPanel (Panel, ~360px wide, left-anchored, translucent dark).\n"
            "3. Inside LeftPanel add LogoSprite + ButtonsVBox (VBoxContainer).\n"
            "4. In ButtonsVBox add 5 buttons: NewGameButton, ContinueButton, "
            "LibraryButton, SettingsButton, ExitButton.\n"
            "5. Add SaveInfoLabel (Label, italic, small) as child of ContinueButton.\n"
            "6. Add VersionLabel at the bottom of LeftPanel.\n"
            "7. Wire all button pressed signals to main_menu.gd handlers.\n"
            "8. Move existing LanguageSelection / GridContainer nodes to stay as children.\n"
            "9. Instance PresidentialLibraryPanel.tscn and SettingsPanel.tscn as children "
            "(both hidden by default)."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-002",
        "status":      "OPEN",
        "scene":       "PresidentialLibraryPanel.tscn",
        "category":    "New Scene",
        "description": "Create PresidentialLibraryPanel.tscn: full-screen dark overlay with 3-tab panel",
        "detail":      (
            "See SCENE STRUCTURE comment in PresidentialLibraryPanel.gd.\n\n"
            "1. Root: Control (full screen).\n"
            "2. Add DarkOverlay (ColorRect, full screen, rgba 0,0,0,0.88).\n"
            "3. Add LibraryContainer (Panel, ~1200x800, centered).\n"
            "4. Inside LibraryContainer:\n"
            "   - SealSprite (TextureRect, presidential seal art, top-center)\n"
            "   - TitleLabel: 'PRESIDENTIAL LIBRARY' (bold, large)\n"
            "   - TabContainer with 3 tabs named: GALLERY / RECORDS / JOURNAL\n"
            "   - CloseButton (top-right)\n"
            "5. Inside each tab Control, instance the matching tab .tscn:\n"
            "   - GALLERY → GalleryTab.tscn\n"
            "   - RECORDS → RecordsTab.tscn\n"
            "   - JOURNAL → JournalTab.tscn\n"
            "6. Wire CloseButton → _on_close_button_pressed().\n"
            "7. Assign PresidentialLibraryPanel.gd as script."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-003",
        "status":      "OPEN",
        "scene":       "GalleryTab.tscn",
        "category":    "New Scene",
        "description": "Create GalleryTab.tscn: 4-column grid + lightbox overlay",
        "detail":      (
            "See SCENE STRUCTURE comment in GalleryTab.gd.\n\n"
            "1. Root: Control.\n"
            "2. Add ScrollContainer → GalleryGrid (GridContainer, columns=4, "
            "h_separation=8, v_separation=8).\n"
            "3. Add Lightbox (Control, full screen, hidden):\n"
            "   - DarkBG (ColorRect, rgba 0,0,0,0.9)\n"
            "   - FullArtRect (TextureRect, large centered, expand_mode=fit)\n"
            "   - LightboxTitle (Label, bold, bottom center)\n"
            "   - LightboxFlavor (Label, italic, smaller)\n"
            "   - PrevButton ('<', left edge), NextButton ('>', right edge)\n"
            "   - CloseButton ('✕', top-right)\n"
            "4. Wire all lightbox buttons to GalleryTab.gd handlers.\n"
            "5. Assign GalleryTab.gd as script."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-004",
        "status":      "OPEN",
        "scene":       "GalleryTile.tscn",
        "category":    "New Scene",
        "description": "Create GalleryTile.tscn: individual gallery entry tile (locked/unlocked states)",
        "detail":      (
            "See SCENE STRUCTURE comment in GalleryTile.gd.\n\n"
            "1. Root: Control (custom_minimum_size 160x200).\n"
            "2. Add Panel (NinePatchRect or Panel).\n"
            "3. Inside Panel:\n"
            "   - ArtRect (TextureRect, expand/fill, top portion)\n"
            "   - LockedOverlay (ColorRect, full size, rgba 0,0,0,0.85):\n"
            "       + LockIcon (TextureRect, 32x32, centered)\n"
            "       + LockedTitle (Label, '???', bottom center)\n"
            "   - UnlockedTitle (Label, bottom strip, visible when unlocked)\n"
            "   - HoverHint (PanelContainer, hidden by default):\n"
            "       + HintLabel (Label, autowrap)\n"
            "4. Add Area2D or mouse_entered/exited signals to root Control.\n"
            "5. Wire mouse_entered → _on_mouse_entered(), mouse_exited → _on_mouse_exited().\n"
            "6. Wire gui_input → _on_gui_input().\n"
            "7. Assign GalleryTile.gd as script."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-005",
        "status":      "OPEN",
        "scene":       "RecordsTab.tscn",
        "category":    "New Scene",
        "description": "Create RecordsTab.tscn: category sidebar + entry panel (HSplitContainer)",
        "detail":      (
            "See SCENE STRUCTURE comment in RecordsTab.gd.\n\n"
            "1. Root: Control.\n"
            "2. HSplitContainer (~25% / 75% split):\n"
            "   Left: CategoryScroll (ScrollContainer) → CategoryList (VBoxContainer)\n"
            "   Right: ScrollContainer → EntryPanel (VBoxContainer, padding 12px):\n"
            "     - EntryHeader (HBoxContainer):\n"
            "         + EntryIcon (TextureRect, 64x64)\n"
            "         + EntryMeta (VBoxContainer):\n"
            "             EntryName (Label, bold, large)\n"
            "             EntryCategory (Label, italic, smaller)\n"
            "     - Divider (HSeparator)\n"
            "     - EntryBody (RichTextLabel, bbcode=on, fit_content=true)\n"
            "     - SeeAlsoBox (HBoxContainer, hidden when empty):\n"
            "         + SeeAlsoLabel (Label, 'See also:')\n"
            "         + SeeAlsoLinks (HBoxContainer)\n"
            "3. Assign RecordsTab.gd as script."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-006",
        "status":      "OPEN",
        "scene":       "JournalTab.tscn",
        "category":    "New Scene",
        "description": "Create JournalTab.tscn: entry list sidebar + memo document panel",
        "detail":      (
            "See SCENE STRUCTURE comment in JournalTab.gd.\n\n"
            "1. Root: Control.\n"
            "2. HSplitContainer (~30% / 70% split):\n"
            "   Left: EntryListScroll (ScrollContainer) → EntryList (VBoxContainer)\n"
            "   Right: MemoPanel (Panel, parchment/paper texture recommended):\n"
            "     - ClassificationLabel (Label, bold caps, red color)\n"
            "     - DateLabel (Label, italic, small)\n"
            "     - SubjectLabel (Label, bold, 'RE: ...')\n"
            "     - Divider (HSeparator)\n"
            "     - MemoBody (RichTextLabel, bbcode=on, scroll_active=true)\n"
            "3. Assign JournalTab.gd as script."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-007",
        "status":      "OPEN",
        "scene":       "SettingsPanel.tscn",
        "category":    "New Scene",
        "description": "Create SettingsPanel.tscn: full-screen overlay with 6-category sidebar",
        "detail":      (
            "See SCENE STRUCTURE comment in SettingsPanel.gd.\n\n"
            "1. Root: Control (full screen).\n"
            "2. DarkOverlay (ColorRect, full screen, rgba 0,0,0,0.85).\n"
            "3. SettingsContainer (Panel, ~900x700, centered):\n"
            "   - TitleLabel: 'SETTINGS'\n"
            "   - HSplitContainer:\n"
            "       Left (SidebarScroll → SidebarList VBoxContainer, ~200px)\n"
            "       Right (ContentScroll → ContentVBox VBoxContainer):\n"
            "           AudioSection, DisplaySection, GameplaySection,\n"
            "           AccessibilitySection, LanguageSection, DataSection\n"
            "           (all VBoxContainers, hidden by default except first)\n"
            "   - CloseButton\n"
            "   - ResetConfirmDialog (ConfirmationDialog node)\n"
            "4. Populate each Section with appropriate controls:\n"
            "   Audio: 4 HSliders (Master/Music/SFX/Ambient) with labels\n"
            "   Display: CheckButton (Fullscreen), OptionButton (Resolution), "
            "HSlider (UI Scale)\n"
            "   Gameplay: OptionButton (Autosave), CheckButton (Tooltips), "
            "OptionButton (Notif Speed)\n"
            "   Accessibility: OptionButton (Text Size), OptionButton (Colorblind), "
            "CheckButton (High Contrast)\n"
            "   Language: Existing flag-picker GridContainer\n"
            "   Data: Label (description), Button ('RESET GAME DATA', red tint)\n"
            "5. Wire all controls to SettingsPanel.gd handlers.\n"
            "6. Wire CloseButton → _on_close_button_pressed().\n"
            "7. Wire ResetConfirmDialog confirmed → _on_reset_confirmed().\n"
            "8. Wire DataSection reset button → _on_reset_data_pressed().\n"
            "9. Assign SettingsPanel.gd as script."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
    {
        "id":          "MENU-008",
        "status":      "OPEN",
        "scene":       "Project Settings → AutoLoad",
        "category":    "Game Flow",
        "description": "Register LibraryData and RecordsDatabase as AutoLoads",
        "detail":      (
            "1. Open Project → Project Settings → AutoLoad tab.\n"
            "2. Add LibraryData.gd with node name 'LibraryData'.\n"
            "3. Add RecordsDatabase.gd with node name 'RecordsDatabase'.\n"
            "4. Ensure LibraryData loads BEFORE RecordsDatabase in the list "
            "(LibraryData must be ready first since RecordsDatabase reads it for discovery state).\n"
            "5. The existing Settings.gd autoload should remain — LibraryData.gd "
            "syncs to it automatically."
        ),
        "added_in":    "Main Menu rebuild",
        "priority":    "HIGH",
    },
]


# ── workbook builder ──────────────────────────────────────────────────────────

def _hdr_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=HDR_FG)
    c.fill      = _fill(HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()
    return c


def build_main_sheet(wb):
    ws = wb.active
    ws.title = "Editor Tasks"

    # header
    ws.row_dimensions[1].height = 28
    for ci, (name, width) in enumerate(COLUMNS, 1):
        _hdr_cell(ws, 1, ci, name)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    # sort: OPEN first, then IN PROGRESS, then DONE; within each by priority
    prio_order  = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    stat_order  = {"OPEN": 0, "IN PROGRESS": 1, "DONE": 2}
    sorted_tasks = sorted(TASKS,
        key=lambda t: (stat_order.get(t["status"], 9),
                       prio_order.get(t["priority"], 9),
                       t["id"]))

    for ri, task in enumerate(sorted_tasks, start=2):
        alt  = (ri % 2 == 0)
        base = _fill(ALT_BG) if alt else _fill(WHITE)

        vals = [
            task["id"],
            task["status"],
            task["scene"],
            task["category"],
            task["description"],
            task["detail"],
            task["added_in"],
            task["priority"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = _border()
            if ci == 2:   # Status
                cell.fill = _status_fill(task["status"])
                cell.font = Font(bold=True)
            elif ci == 8: # Priority
                pf = _fill("FFD6D6") if task["priority"] == "HIGH" \
                     else _fill("FFF3CD") if task["priority"] == "MEDIUM" \
                     else _fill("D6F5D6")
                cell.fill = pf
            else:
                cell.fill = base

        ws.row_dimensions[ri].height = max(60, task["detail"].count("\n") * 14 + 20)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(sorted_tasks) + 1}"


def build_summary_sheet(wb):
    from collections import Counter
    ws = wb.create_sheet("Summary")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12

    def hdr(cell, val):
        ws[cell].value = val
        ws[cell].font  = Font(bold=True, color=HDR_FG)
        ws[cell].fill  = _fill(HDR_BG)

    hdr("A1", "Status");   hdr("B1", "Count")
    hdr("D1", "Category"); hdr("E1", "Count")

    stat_counts = Counter(t["status"]   for t in TASKS)
    cat_counts  = Counter(t["category"] for t in TASKS)

    for i, (s, c) in enumerate(sorted(stat_counts.items()), 2):
        ws.cell(row=i, column=1, value=s).fill = _status_fill(s)
        ws.cell(row=i, column=2, value=c).fill = _status_fill(s)

    tr = len(stat_counts) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=2, value=len(TASKS)).font = Font(bold=True)

    open_count = sum(1 for t in TASKS if t["status"] == "OPEN")
    ws.cell(row=tr+1, column=1, value="Open").font = Font(bold=True)
    ws.cell(row=tr+1, column=2, value=open_count).font = Font(bold=True)

    for i, (cat, cnt) in enumerate(sorted(cat_counts.items()), 2):
        ws.cell(row=i, column=4, value=cat).fill = _fill(ALT_BG)
        ws.cell(row=i, column=5, value=cnt).fill = _fill(ALT_BG)


def main():
    wb = openpyxl.Workbook()
    build_main_sheet(wb)
    build_summary_sheet(wb)
    out = "editortasks_masterdoc.xlsx"
    wb.save(out)
    open_count = sum(1 for t in TASKS if t["status"] == "OPEN")
    done_count = sum(1 for t in TASKS if t["status"] == "DONE")
    print(f"Saved {out}  —  {len(TASKS)} tasks total  |  {open_count} OPEN  |  {done_count} DONE")


if __name__ == "__main__":
    main()
