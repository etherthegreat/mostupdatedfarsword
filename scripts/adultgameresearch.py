#!/usr/bin/env python3
"""
/adultgameresearch — SpaceDesk Adult Game Research Command

Searches internationally for adult strategy/city builder games using Claude AI
with live web search. Filters out incest content. Prioritizes strategy, city
builder, management, and political simulation games from underrepresented regions.
Rebuilds the SpaceDesk adult games database Excel spreadsheet.

Setup:
    pip install anthropic openpyxl requests --break-system-packages
    export ANTHROPIC_API_KEY="your-key-here"

Usage:
    python adultgameresearch.py                  # Full rebuild
    python adultgameresearch.py --append         # Add to existing DB
    python adultgameresearch.py --quick          # Fewer queries, faster run
    python adultgameresearch.py --output mydb.xlsx

Install as command:
    chmod +x adultgameresearch.py
    sudo cp adultgameresearch.py /usr/local/bin/adultgameresearch
"""

import os, sys, json, time, argparse, re
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import anthropic
except ImportError:
    print("Missing: pip install anthropic openpyxl requests --break-system-packages")
    sys.exit(1)

# ─── Content filter ───────────────────────────────────────────────────────────
BLOCKED_TAGS = {
    "incest","step-family","step-sister","step-brother","step-mom",
    "step-dad","step-parent","step-son","step-daughter","stepsister",
    "stepbrother","stepmother","stepfather","rape","non-con","noncon",
    "non_con","forced","sexual assault","sleep","sleeping"
}

PRIORITY_TAGS = {
    "strategy","grand-strategy","grand strategy","city-builder","city builder",
    "management","nation-building","nation building","political","political-simulator",
    "turn-based","turn based","tactical","simulation","4x","rts","wargame",
    "conquest","empire","colony","civilization","resource management",
    "kingdom","dungeon management","tycoon","sandbox","war"
}

# ─── Research queries — ordered by discovery value ───────────────────────────
RESEARCH_QUERIES_FULL = [
    ("Robert Yang gay games strategy tactics 2014 2025 queer adult itch.io",
     "USA — Queer Art Games"),
    ("Brazilian Portuguese adult strategy game itch.io F95Zone indie developer 2023 2024 2025",
     "Brazil / Latin America"),
    ("Spanish language adult strategy game city builder itch.io developer Argentina Chile Mexico",
     "Latin America — Spanish"),
    ("Russian adult strategy game NSFW itch.io 2023 2024 стратегия взрослый",
     "Russia / Eastern Europe"),
    ("Polish Czech Romanian adult indie strategy game NSFW 2023 2024 2025",
     "Eastern Europe"),
    ("Southeast Asia Philippines Indonesia Thailand Vietnam adult NSFW strategy game 2024 2025",
     "Southeast Asia"),
    ("Chinese adult strategy game English patch DLsite Steam 2023 2024 2025",
     "China / Taiwan"),
    ("Korean adult strategy game DLsite 한국 성인 전략 2023 2024",
     "Korea"),
    ("site:itch.io adult strategy NSFW grand strategy city builder management",
     "itch.io — Strategy + NSFW direct"),
    ("site:itch.io adult city builder management simulation NSFW political",
     "itch.io — City Builder + NSFW direct"),
    ("F95Zone grand strategy adult game thread 2024 2025 new release management",
     "F95Zone — Strategy threads"),
    ("F95Zone city builder adult management game 2024 2025 thread",
     "F95Zone — City Builder threads"),
    ("obscure unknown adult strategy game indie developer 2024 2025 F95Zone itch",
     "Obscure Discoveries"),
    ("adult game nation building political simulator indie 2024 2025",
     "Political Sim Hunting"),
    ("adult game developer Africa Nigeria Kenya South Africa strategy indie",
     "Africa — Emerging Scene"),
    ("adult game developer India strategy management NSFW anonymous developer",
     "India — Hidden Devs"),
    ("adult game developer Turkey Middle East NSFW strategy 2024 2025",
     "Middle East / Turkey"),
    ("Australian adult strategy game developer NSFW indie 2023 2024 2025",
     "Australia / NZ"),
    ("DLsite western developer adult strategy English 2023 2024 2025",
     "DLsite — Western Devs"),
    ("adult strategy RPG tactical management Patreon SubscribeStar 2024 2025 indie developer",
     "Patreon/SubscribeStar — Strategy Devs"),
]

RESEARCH_QUERIES_QUICK = RESEARCH_QUERIES_FULL[:8]

# ─── Claude API call with web search ─────────────────────────────────────────
SYSTEM_PROMPT = """You are a research assistant for SpaceDesk, an adult gaming platform.
Your job: find real, existing adult video games that have strategy, city-building,
management, nation-building, or political simulation elements.

CRITICAL CONTENT FILTER — RETURN NOTHING THAT CONTAINS:
- Incest, step-family, step-relative content of any kind
- Non-consensual sexual content (rape, forced)
- Content that sexualises minors in any form

PRIORITY — RETURN GAMES WITH:
- Grand strategy, 4X, city builder, management, political simulation, turn-based strategy
- Games from non-US/non-Japanese developers especially
- Early builds, alpha, beta — include these
- De-listed or hard to find games
- Games in non-English languages

OUTPUT FORMAT — respond ONLY with a JSON array. No text before or after.
Each game object must have exactly these fields:
{
  "title": "exact game title",
  "developer": "developer name or 'Unknown'",
  "country": "country of origin if known, else 'Unknown'",
  "year": "release year or 'In dev' or 'Unknown'",
  "genre": "specific genre description",
  "strategy_depth": 1-5,
  "adult_level": 1-5,
  "platform": "platforms",
  "status": "Released/Early Access/In dev/Alpha/Beta/Abandoned",
  "english_available": "Yes/No/Partial/Unknown",
  "where_to_play": "URL or platform name",
  "lgbtq_content": "Yes/No/Some/Unknown",
  "content_warnings": "specific warnings, or 'None noted'",
  "notes": "why this is interesting, historical significance, what makes it notable",
  "incest_flag": false
}

If you find NO games matching criteria for a query, return: []
Return 3-8 games maximum per query. Quality over quantity.
Be honest — only include games you found evidence actually exists."""


def research_query(client, query, region_label, verbose=True):
    """Run one research query with web search and return parsed games."""
    if verbose:
        print(f"  🔍 {region_label}...")

    prompt = f"""Search for adult strategy/management/city-builder games using this research direction:
"{query}"

Region focus: {region_label}

Search thoroughly. Include:
- Games that may be de-listed from itch.io main search but still accessible
- Games on F95Zone, DLsite, Patreon, SubscribeStar, personal dev sites
- Games in non-English languages with or without English versions
- Early builds and alpha releases from indie developers
- Robert Yang's catalog if relevant to strategy/management
- ANY developer from the specified region making adult strategy content

Return your findings as a JSON array following the exact format specified."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            system=SYSTEM_PROMPT,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}]
        )

        full_text = ""
        for block in response.content:
            if hasattr(block, "text"):
                full_text += block.text

        json_match = re.search(r'\[[\s\S]*\]', full_text)
        if not json_match:
            return []

        games = json.loads(json_match.group())
        if not isinstance(games, list):
            return []

        clean = []
        for g in games:
            if not isinstance(g, dict):
                continue
            if g.get("incest_flag", False):
                continue
            combined = " ".join([
                str(g.get("content_warnings", "")),
                str(g.get("notes", "")),
                str(g.get("genre", "")),
                str(g.get("title", ""))
            ]).lower()
            if any(tag in combined for tag in BLOCKED_TAGS):
                continue
            g["region_found"] = region_label
            clean.append(g)

        if verbose and clean:
            print(f"    ✓ Found {len(clean)} game(s)")
        return clean

    except json.JSONDecodeError:
        return []
    except Exception as e:
        if verbose:
            print(f"    ⚠ Error: {str(e)[:60]}")
        return []


def deduplicate(games):
    seen = set()
    unique = []
    for g in games:
        key = g.get("title", "").lower().strip()
        if key and key not in seen:
            seen.add(key)
            unique.append(g)
    return unique


def strategy_score(game):
    text = " ".join([game.get("genre", ""), game.get("notes", "")]).lower()
    score = sum(1 for tag in PRIORITY_TAGS if tag in text)
    score += game.get("strategy_depth", 0)
    return -score


# ─── Excel palette ────────────────────────────────────────────────────────────
C = {
    "void":   "070710", "violet": "5A1F8A", "teal":   "0A7C6E",
    "gold":   "C9A84C", "red":    "C0392B", "white":  "FFFFFF",
    "ink":    "1A1410", "grey":   "F0EEE8", "gold_l": "FFF3C4",
    "red_l":  "FCE8E6", "teal_l": "E0F5F0", "viol_l": "EDE0FF",
    "blue_l": "E0EEFF", "grn_l":  "E0FFE8",
}


def fill(c): return PatternFill("solid", fgColor=c)
def af(bold=False, sz=10, color="1A1410", italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
def al(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def bd():
    s = Side(style="thin", color="D0C8B0")
    return Border(left=s, right=s, top=s, bottom=s)


def wcell(ws, r, col, v="", bg=None, fg="1A1410", bold=False,
          sz=10, align="left", italic=False):
    cell = ws.cell(row=r, column=col, value=v)
    if bg:
        cell.fill = fill(bg)
    cell.font = af(bold=bold, sz=sz, color=fg, italic=italic)
    cell.alignment = al(align)
    cell.border = bd()
    return cell


def mtitle(ws, r, c1, c2, v, bg="070710", fg="C9A84C", sz=13):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=v)
    cell.fill = fill(bg)
    cell.font = af(bold=True, sz=sz, color=fg)
    cell.alignment = al("center")


def tier_color(game):
    depth = game.get("strategy_depth", 1)
    region = game.get("region_found", "")
    if depth >= 4:
        return C["gold_l"]
    if "itch" in region.lower() or "f95" in region.lower():
        return C["teal_l"]
    if any(r in region for r in ["Brazil","Latin","Russia","Eastern","Asia","Africa","India","Korea","China","Turkey","Middle"]):
        return C["viol_l"]
    return C["blue_l"]


COLS = [
    "#", "Title", "Developer", "Country", "Year",
    "Genre", "Strategy\nDepth", "Adult\nLevel",
    "Platform", "Status", "English?", "Where to Play",
    "LGBTQ+", "Content Warnings", "Notes", "Found Via"
]
COL_W = [4, 28, 22, 12, 6, 22, 9, 9, 16, 14, 10, 28, 10, 22, 50, 24]


def build_spreadsheet(games, output_path):
    wb = Workbook()
    today = date.today().isoformat()

    # Sheet 1: Full database
    ws = wb.active
    ws.title = "Research Database"
    mtitle(ws, 1, 1, len(COLS),
           f"SPACEDESK — ADULT STRATEGY GAME RESEARCH  ·  Last updated: {today}")
    mtitle(ws, 2, 1, len(COLS),
           f"{len(games)} games  ·  No incest content  ·  Strategy/City Builder priority  ·  International focus",
           bg=C["violet"], fg=C["viol_l"], sz=10)

    for i, h in enumerate(COLS):
        wcell(ws, 4, i+1, h, bg=C["void"], fg=C["gold"], bold=True, sz=10, align="center")

    for idx, game in enumerate(games):
        r = 5 + idx
        bg = tier_color(game)
        vals = [
            str(idx + 1),
            game.get("title", ""),
            game.get("developer", ""),
            game.get("country", ""),
            str(game.get("year", "")),
            game.get("genre", ""),
            game.get("strategy_depth", ""),
            game.get("adult_level", ""),
            game.get("platform", ""),
            game.get("status", ""),
            game.get("english_available", ""),
            game.get("where_to_play", ""),
            game.get("lgbtq_content", ""),
            game.get("content_warnings", ""),
            game.get("notes", ""),
            game.get("region_found", ""),
        ]
        for j, v in enumerate(vals):
            wcell(ws, r, j+1, v, bg=bg, fg=C["ink"], sz=9)
        ws.row_dimensions[r].height = 50

    for i, w in enumerate(COL_W):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.freeze_panes = "A5"

    # Sheet 2: High priority
    ws2 = wb.create_sheet("High Priority")
    mtitle(ws2, 1, 1, 6, "HIGH PRIORITY STREAM LIST — Strategy Depth 4-5",
           bg=C["void"], fg=C["gold"])
    priority = [g for g in games if g.get("strategy_depth", 0) >= 4]
    for i, h in enumerate(["Title", "Developer", "Country", "Genre", "Where to Play", "Notes"]):
        wcell(ws2, 3, i+1, h, bg=C["void"], fg=C["gold"], bold=True)
    for idx, game in enumerate(priority):
        r = 4 + idx
        for j, v in enumerate([game.get("title",""), game.get("developer",""),
                                game.get("country",""), game.get("genre",""),
                                game.get("where_to_play",""), game.get("notes","")]):
            wcell(ws2, r, j+1, v, bg=C["gold_l"], fg=C["ink"], sz=9)
        ws2.row_dimensions[r].height = 45
    for i, w in enumerate([28, 22, 14, 22, 28, 50]):
        ws2.column_dimensions[get_column_letter(i+1)].width = w

    # Sheet 3: By region
    ws3 = wb.create_sheet("By Region")
    mtitle(ws3, 1, 1, 4, "GAMES BY REGION — International Discovery Map",
           bg=C["teal"], fg=C["white"])
    regions = {}
    for g in games:
        regions.setdefault(g.get("region_found", "Unknown"), []).append(g)
    r = 3
    for region, rg in sorted(regions.items()):
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        hc = ws3.cell(row=r, column=1, value=f"  {region}  ({len(rg)} games)")
        hc.fill = fill(C["void"]); hc.font = af(bold=True, sz=11, color=C["gold"])
        hc.alignment = al("left"); r += 1
        for game in rg:
            wcell(ws3, r, 1, game.get("title",""), bg=C["teal_l"], fg=C["ink"], sz=9, bold=True)
            wcell(ws3, r, 2, game.get("country",""), bg=C["teal_l"], fg=C["ink"], sz=9)
            ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            wcell(ws3, r, 3, game.get("genre",""), bg=C["teal_l"], fg=C["ink"], sz=9)
            ws3.row_dimensions[r].height = 35; r += 1
        r += 1
    for i, w in enumerate([32, 14, 30, 30]):
        ws3.column_dimensions[get_column_letter(i+1)].width = w

    # Sheet 4: Research log
    ws4 = wb.create_sheet("Research Log")
    mtitle(ws4, 1, 1, 3,
           f"RESEARCH LOG  ·  Run: {today}  ·  {len(games)} games discovered",
           bg=C["void"], fg=C["gold"])
    wcell(ws4, 3, 1, "Filter Rules Applied", bg=C["void"], fg=C["gold"], bold=True)
    ws4.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    rules = [
        ("BLOCKED tags", ", ".join(sorted(BLOCKED_TAGS))),
        ("PRIORITY tags", ", ".join(sorted(PRIORITY_TAGS))),
        ("Content filter", "Any game where title, notes, warnings, or genre contains blocked tags is excluded"),
        ("Incest flag", "Games where Claude's research flagged incest_flag=true are excluded"),
        ("Source", "Claude claude-sonnet-4-6 with live web_search tool"),
        ("Deduplication", "Titles normalized to lowercase, exact duplicates removed"),
    ]
    for i, (k, v) in enumerate(rules):
        r2 = 4 + i
        wcell(ws4, r2, 1, k, bg=C["grey"], fg=C["ink"], bold=True, sz=10)
        c2 = ws4.cell(row=r2, column=2, value=v)
        ws4.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
        c2.fill = fill(C["grey"]); c2.font = af(sz=9, color=C["ink"])
        c2.alignment = al("left"); c2.border = bd()
        ws4.row_dimensions[r2].height = 40
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 80
    ws4.column_dimensions["C"].width = 10

    wb.save(output_path)
    return len(games)


def main():
    parser = argparse.ArgumentParser(
        description="/adultgameresearch — SpaceDesk adult strategy game research"
    )
    parser.add_argument("--append", action="store_true",
                        help="Add to existing database instead of rebuilding")
    parser.add_argument("--quick", action="store_true",
                        help="Run fewer queries (faster, less thorough)")
    parser.add_argument("--output", default=None,
                        help="Output filename (default: SpaceDesk_GameDB_Research_DATE.xlsx)")
    parser.add_argument("--quiet", action="store_true",
                        help="Minimal output")
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: Set ANTHROPIC_API_KEY environment variable first.")
        print("    export ANTHROPIC_API_KEY='your-key-here'")
        sys.exit(1)

    output_path = args.output or f"SpaceDesk_GameDB_Research_{date.today().isoformat()}.xlsx"
    queries = RESEARCH_QUERIES_QUICK if args.quick else RESEARCH_QUERIES_FULL
    verbose = not args.quiet

    if verbose:
        print(f"\n  SpaceDesk /adultgameresearch")
        print(f"  Mode:    {'Quick' if args.quick else 'Full'} ({len(queries)} queries)")
        print(f"  Output:  {output_path}")
        print(f"  Filter:  No incest · Strategy/City Builder priority")
        print(f"  Focus:   International / underrepresented regions\n")

    client = anthropic.Anthropic(api_key=api_key)
    all_games = []

    for query_text, region_label in queries:
        games = research_query(client, query_text, region_label, verbose=verbose)
        all_games.extend(games)
        time.sleep(2)

    if verbose:
        print(f"\n  Raw results: {len(all_games)} games")

    all_games.sort(key=strategy_score)
    all_games = deduplicate(all_games)

    if verbose:
        print(f"  After dedup: {len(all_games)} unique games")

    existing_games = []
    if args.append and os.path.exists(output_path):
        try:
            ewb = load_workbook(output_path)
            ews = ewb["Research Database"]
            for row in ews.iter_rows(min_row=5, values_only=True):
                if row[1]:
                    existing_games.append({
                        "title": row[1] or "", "developer": row[2] or "",
                        "country": row[3] or "", "year": row[4] or "",
                        "genre": row[5] or "", "strategy_depth": row[6] or 1,
                        "adult_level": row[7] or 1, "platform": row[8] or "",
                        "status": row[9] or "", "english_available": row[10] or "",
                        "where_to_play": row[11] or "", "lgbtq_content": row[12] or "",
                        "content_warnings": row[13] or "", "notes": row[14] or "",
                        "region_found": row[15] or ""
                    })
            if verbose:
                print(f"  Loaded {len(existing_games)} existing games for append")
        except Exception as e:
            if verbose:
                print(f"  Could not load existing DB: {e}")

    final_games = deduplicate(existing_games + all_games)

    if verbose:
        print(f"\n  Building spreadsheet: {output_path}")

    count = build_spreadsheet(final_games, output_path)

    if verbose:
        print(f"  Done — {count} games in database")
        print(f"  File: {output_path}\n")
        regions = {}
        for g in final_games:
            r = g.get("region_found", "Unknown")
            regions[r] = regions.get(r, 0) + 1
        print("  Games by region:")
        for region, n in sorted(regions.items(), key=lambda x: -x[1]):
            print(f"    {region:<40} {n}")

    return output_path


if __name__ == "__main__":
    main()
