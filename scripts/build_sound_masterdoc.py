#!/usr/bin/env python3
"""Generate sound_masterdoc.xlsx — sound design asset catalog for the game."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── colour palette ─────────────────────────────────────────────────────────────
HDR_BG  = "1F2D3D"
HDR_FG  = "FFFFFF"
CAT_BG  = "2E4057"
CAT_FG  = "FFFFFF"
HIGH_BG = "FFD6D6"   # light red
MED_BG  = "FFF3CD"   # amber
LOW_BG  = "D6F5D6"   # light green
ALT_BG  = "EEF2F7"
WHITE   = "FFFFFF"

def hex_fill(h): return PatternFill("solid", fgColor=h)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── columns ───────────────────────────────────────────────────────────────────
COLUMNS = [
    ("Category",        18),
    ("Asset Name",      30),
    ("When It Plays",   44),
    ("Priority",        12),
    ("Duration",        12),
    ("Character",       38),
    ("Variants Needed", 14),
    ("Notes",           40),
]

# ── data ──────────────────────────────────────────────────────────────────────
# (category, asset_name, when_it_plays, priority, duration, character, variants, notes)
SOUNDS = [

    # ── UI BUTTONS ────────────────────────────────────────────────────────────
    ("UI Buttons", "ui_click_generic",
     "Any standard button pressed",
     "HIGH", "< 100ms",
     "Dry, neutral tap — like pressing a key on a wooden keyboard",
     "3–4", "Randomise to avoid repetition"),

    ("UI Buttons", "ui_click_confirm",
     "Confirm / build / hire — affirmative action",
     "HIGH", "150–250ms",
     "Confident thud + short sparkle tail",
     "1–2", "Slightly more satisfying than generic click"),

    ("UI Buttons", "ui_click_cancel",
     "Cancel, close, dismiss a panel",
     "MEDIUM", "< 100ms",
     "Soft negative click — quiet paper slide",
     "2", "Lower volume than confirm"),

    ("UI Buttons", "ui_end_turn",
     "End Turn button pressed",
     "HIGH", "400–600ms",
     "Deep 'thunk' + 2-note orchestral sting — signals real time passing",
     "1", "Should feel weighty and decisive"),

    ("UI Buttons", "ui_tab_switch",
     "Switch tabs inside a multi-tab panel",
     "LOW", "< 80ms",
     "Paper page-turn, very subtle",
     "2", "Keep quiet — happens a lot"),

    ("UI Buttons", "ui_scroll",
     "Scroll list or slider moved",
     "LOW", "< 60ms",
     "Faint paper rustle",
     "1", "Very low volume"),

    ("UI Buttons", "ui_hover",
     "Mouse enters a button or panel element",
     "LOW", "< 60ms",
     "Almost inaudible tick / breath",
     "1", "Optional — consider disabling by default"),

    # ── PANELS ────────────────────────────────────────────────────────────────
    ("Panels", "panel_open",
     "Any info panel slides into view",
     "MEDIUM", "150–300ms",
     "Smooth whoosh right-to-left with soft landing thud",
     "2–3", "Generic fallback; specific panels override"),

    ("Panels", "panel_close",
     "Panel dismissed / hidden",
     "LOW", "100–200ms",
     "Reverse whoosh, shorter than open",
     "2", ""),

    ("Panels", "panel_open_warroom",
     "War Room panel opens",
     "MEDIUM", "300ms",
     "Heavy iron door creak + clank",
     "1", "Distinct character for war context"),

    ("Panels", "panel_open_event",
     "Event card appears on screen",
     "HIGH", "300–500ms",
     "Dramatic paper drop + wax seal crack — grabs attention",
     "1", "Central gameplay moment; should command attention"),

    ("Panels", "panel_open_spellbook",
     "Spellbook / magic panel opens",
     "MEDIUM", "250ms",
     "Arcane pages fluttering + faint shimmer",
     "1", ""),

    ("Panels", "panel_open_laws",
     "Laws / legislature panel opens",
     "MEDIUM", "200ms",
     "Gavel light tap on wood",
     "1", ""),

    ("Panels", "panel_open_techTree",
     "Research / tech tree panel opens",
     "MEDIUM", "250ms",
     "Electrical startup hum + subtle beep sequence",
     "1", ""),

    ("Panels", "panel_open_governors",
     "Governor selection panel opens",
     "MEDIUM", "200ms",
     "Parchment unrolling",
     "1", ""),

    # ── EVENTS ────────────────────────────────────────────────────────────────
    ("Events", "event_standard",
     "A standard random event fires",
     "HIGH", "400ms",
     "Notification bell + paper rustle — draws attention without alarm",
     "3", "3 variants rotate to keep variety"),

    ("Events", "event_crisis",
     "A crisis or major negative event fires",
     "HIGH", "500–700ms",
     "Low brass stab + alarm bell — urgent and ominous",
     "2", ""),

    ("Events", "event_ualani",
     "A personal Ualani event fires",
     "HIGH", "400–600ms",
     "Mystical Hawaiian-inspired chime sequence — her personal theme",
     "2", "Ualani's sound identity should be consistent throughout"),

    ("Events", "event_election",
     "Election event / results fire",
     "HIGH", "600ms",
     "TV news intro jingle + crowd murmur",
     "1", ""),

    ("Events", "event_good_outcome",
     "Player picks an outcome that grants resources or bonuses",
     "MEDIUM", "250ms",
     "Warm ascending chime",
     "2", ""),

    ("Events", "event_bad_outcome",
     "Player picks an outcome that costs resources or triggers penalties",
     "MEDIUM", "250ms",
     "Descending tone + soft thud",
     "2", ""),

    ("Events", "event_death",
     "A named character (governor, commander) dies in an event",
     "HIGH", "800ms–1s",
     "Solitary bell toll — somber and resonant",
     "1", ""),

    ("Events", "event_mission_complete",
     "A mission / quest objective completed",
     "HIGH", "500ms",
     "Short victory fanfare — uplifting but not too long",
     "1", ""),

    ("Events", "event_mission_expire",
     "A mission expires without completion",
     "HIGH", "500ms",
     "Failure buzzer + low string drop",
     "1", ""),

    ("Events", "event_season_spring",
     "Season changes to Spring",
     "HIGH", "2–3s",
     "Birds chirping + light flute sting",
     "1", "Play once at season transition"),

    ("Events", "event_season_summer",
     "Season changes to Summer",
     "HIGH", "2–3s",
     "Crickets + warm brass swell",
     "1", ""),

    ("Events", "event_season_autumn",
     "Season changes to Autumn",
     "HIGH", "2–3s",
     "Wind through leaves + melancholy oboe",
     "1", ""),

    ("Events", "event_season_winter",
     "Season changes to Winter",
     "HIGH", "2–3s",
     "Howling wind + low string + distant chime",
     "1", ""),

    ("Events", "event_year_change",
     "New calendar year begins",
     "MEDIUM", "1s",
     "Clock chiming + brief orchestral swell",
     "1", ""),

    # ── RESOURCES ────────────────────────────────────────────────────────────
    ("Resources", "resource_gain_currency",
     "Currency / dollars gained",
     "MEDIUM", "200ms",
     "Coins clinking / cash register ding",
     "3", "Pitch-randomise for variety"),

    ("Resources", "resource_gain_food",
     "Food gained",
     "LOW", "200ms",
     "Grain pour / harvest rustle",
     "2", ""),

    ("Resources", "resource_gain_wood",
     "Wood gained",
     "LOW", "200ms",
     "Logs stacking / wood creak",
     "2", ""),

    ("Resources", "resource_gain_metal",
     "Metal gained",
     "LOW", "200ms",
     "Metal clank / forge ring",
     "2", ""),

    ("Resources", "resource_gain_manpower",
     "Manpower gained",
     "LOW", "200ms",
     "Marching footstep burst (2 steps)",
     "2", ""),

    ("Resources", "resource_gain_science",
     "Science gained",
     "LOW", "200ms",
     "Electric beep / computer chirp",
     "2", ""),

    ("Resources", "resource_gain_magic",
     "Magic gained",
     "LOW", "200ms",
     "Soft arcane shimmer",
     "2", ""),

    ("Resources", "resource_gain_faith",
     "Faith gained",
     "LOW", "200ms",
     "Soft choir hum (single note)",
     "2", ""),

    ("Resources", "resource_gain_culture",
     "Culture gained",
     "LOW", "200ms",
     "Brush on canvas / harp pluck",
     "2", ""),

    ("Resources", "resource_gain_mandate",
     "Mandate gained",
     "LOW", "200ms",
     "Stamp + crowd murmur",
     "2", ""),

    ("Resources", "resource_gain_influence",
     "Influence gained",
     "LOW", "200ms",
     "Whisper + smooth tone",
     "2", ""),

    ("Resources", "resource_gain_weapons",
     "Weapons gained",
     "LOW", "200ms",
     "Sword slide into sheath",
     "2", ""),

    ("Resources", "resource_gain_boats",
     "Boats gained",
     "LOW", "200ms",
     "Ship bell + water lap",
     "2", ""),

    ("Resources", "resource_lost",
     "Any resource lost",
     "MEDIUM", "200ms",
     "Coin drop / heavy thud — generic negative",
     "2", "Use when resource decreases; can layer over specific resource sounds"),

    ("Resources", "resource_critical",
     "A resource hits zero or goes negative (shown in red)",
     "HIGH", "400ms",
     "Low alarm tone + discordant sting",
     "1", "Should feel alarming — player needs to act"),

    # ── BUILDINGS ────────────────────────────────────────────────────────────
    ("Buildings", "building_construct",
     "A new building is placed on a tile",
     "HIGH", "400–600ms",
     "Construction sequence: hammer strikes + wood settling + completion chime",
     "1", ""),

    ("Buildings", "building_upgrade",
     "Existing building upgraded to next level",
     "HIGH", "400ms",
     "Upgrade sparkle + hammer + ascending tone",
     "1", "More positive/sparkly than construct"),

    ("Buildings", "building_downgrade",
     "Building level reduced by damage or event",
     "MEDIUM", "400ms",
     "Crumbling stone / structural groan",
     "1", ""),

    ("Buildings", "building_destroyed",
     "Building fully destroyed and removed",
     "HIGH", "600ms–1s",
     "Collapse crash + dust burst",
     "2", "More dramatic than downgrade"),

    ("Buildings", "building_farm",
     "Farm produces output (turn resolution)",
     "LOW", "300ms",
     "Harvest / scythe swish",
     "2", "Play once per turn if farm present"),

    ("Buildings", "building_mine",
     "Mine produces metal output",
     "LOW", "300ms",
     "Pickaxe clink + echo",
     "2", ""),

    ("Buildings", "building_barracks",
     "Barracks produces manpower",
     "LOW", "300ms",
     "Marching cadence (2–3 steps)",
     "2", ""),

    ("Buildings", "building_market",
     "Market produces currency",
     "LOW", "300ms",
     "Coins on counter / busy market chatter",
     "2", ""),

    ("Buildings", "building_granary",
     "Granary produces food",
     "LOW", "300ms",
     "Grain pour",
     "2", ""),

    ("Buildings", "building_dock",
     "Dock produces boats",
     "LOW", "300ms",
     "Ship bell + gentle water",
     "2", ""),

    ("Buildings", "building_library",
     "Library produces science",
     "LOW", "300ms",
     "Page turn + soft chime",
     "2", ""),

    ("Buildings", "building_temple",
     "Temple / monastery produces faith",
     "LOW", "300ms",
     "Soft choir note / bell",
     "2", ""),

    ("Buildings", "building_forge",
     "Forge produces weapons",
     "LOW", "300ms",
     "Hammer on anvil",
     "2", ""),

    ("Buildings", "building_tower",
     "Tower produces magic or fires a spell",
     "MEDIUM", "400ms",
     "Arcane resonance pulse + shimmer",
     "2", "Distinct from generic magic"),

    # ── TILES / MAP ──────────────────────────────────────────────────────────
    ("Tiles/Map", "tile_click",
     "Player clicks on a map tile",
     "HIGH", "< 100ms",
     "Soft map-paper tap — subtle, non-intrusive",
     "3", "Happens constantly; must be quiet"),

    ("Tiles/Map", "tile_colonize_complete",
     "Tile fully colonized and claimed",
     "HIGH", "600ms",
     "Flag planted + small crowd cheer",
     "1", ""),

    ("Tiles/Map", "tile_owner_change",
     "Tile ownership changes (conquest or event)",
     "HIGH", "400ms",
     "Flag captured / swift sword sting",
     "1", ""),

    ("Tiles/Map", "tile_corruption_high",
     "Tile corruption crosses a high threshold",
     "MEDIUM", "400ms",
     "Dark hum / sinister low crackle",
     "2", "Play at threshold crossings: 40, 60, 80"),

    ("Tiles/Map", "tile_discover",
     "New tile revealed / discovered on map",
     "MEDIUM", "300ms",
     "Exploration ping + soft map-paper sound",
     "2", ""),

    ("Tiles/Map", "tile_storm",
     "Storm or disaster hits a tile",
     "HIGH", "800ms–1s",
     "Thunder crack + howling wind",
     "3", "Randomise thunder for variety"),

    ("Tiles/Map", "tile_siege_start",
     "Siege of a tile begins",
     "HIGH", "600ms",
     "Battering ram impact + crowd war cry",
     "1", ""),

    ("Tiles/Map", "tile_siege_end_taken",
     "Besieged tile falls to attacker",
     "HIGH", "800ms",
     "Gate collapse + brief cheer",
     "1", ""),

    ("Tiles/Map", "tile_siege_end_repelled",
     "Siege repelled — defenders hold",
     "HIGH", "600ms",
     "Defense horn + cheer",
     "1", ""),

    ("Tiles/Map", "tile_atmosphere_jungle",
     "Playing near / hovering over jungle tiles",
     "LOW", "3–5s loop",
     "Birds, insects, rustling leaves",
     "1", "Soft ambient loop; fade in/out"),

    ("Tiles/Map", "tile_atmosphere_desert",
     "Playing near desert tiles",
     "LOW", "3–5s loop",
     "Wind, distant sand hiss",
     "1", ""),

    ("Tiles/Map", "tile_atmosphere_coast",
     "Playing near coastal tiles",
     "LOW", "3–5s loop",
     "Ocean waves breaking",
     "1", ""),

    ("Tiles/Map", "tile_atmosphere_mountain",
     "Playing near mountain tiles",
     "LOW", "3–5s loop",
     "High wind, distant eagle cry",
     "1", ""),

    ("Tiles/Map", "tile_atmosphere_forest",
     "Playing near forest / taiga tiles",
     "LOW", "3–5s loop",
     "Forest ambience, light wind through trees",
     "1", ""),

    # ── ARMY / COMBAT ────────────────────────────────────────────────────────
    ("Army/Combat", "army_spawn",
     "New army unit created and placed on map",
     "HIGH", "400ms",
     "Drum roll burst + metallic boot clank",
     "2", ""),

    ("Army/Combat", "army_move",
     "Army unit moves across tiles",
     "MEDIUM", "500ms–1s loop",
     "Marching footsteps (boots on varied terrain)",
     "4", "4 terrain variants: normal/forest/snow/road"),

    ("Army/Combat", "army_destroy",
     "Army unit destroyed / removed",
     "HIGH", "700ms",
     "Explosion + defeat sting (low brass)",
     "2", ""),

    ("Army/Combat", "army_battle_start",
     "Two armies engage — combat begins",
     "HIGH", "600ms",
     "Clash of shields + battle cry + drum burst",
     "1", ""),

    ("Army/Combat", "army_battle_win",
     "Player wins a battle",
     "HIGH", "1–2s",
     "Victory shout + triumphant fanfare (short)",
     "1", ""),

    ("Army/Combat", "army_battle_loss",
     "Player loses a battle",
     "HIGH", "1–2s",
     "Retreat horn + somber string drop",
     "1", ""),

    ("Army/Combat", "army_melee_hit",
     "Melee attack lands on target",
     "HIGH", "200ms",
     "Sword clang / axe crack on armor",
     "4", "Randomise 4 variants; avoid machine-gun effect"),

    ("Army/Combat", "army_melee_miss",
     "Melee attack misses",
     "MEDIUM", "150ms",
     "Sword whoosh past — no impact",
     "3", ""),

    ("Army/Combat", "army_ranged_fire",
     "Ranged unit fires (arrow / bullet)",
     "MEDIUM", "200ms",
     "Bow twang or musket crack",
     "3", ""),

    ("Army/Combat", "army_ranged_hit",
     "Ranged projectile hits",
     "HIGH", "200ms",
     "Arrow thud + impact",
     "3", ""),

    ("Army/Combat", "army_ranged_miss",
     "Ranged projectile misses",
     "LOW", "150ms",
     "Whoosh only — no impact",
     "2", ""),

    ("Army/Combat", "army_rout",
     "Army unit routes / flees combat",
     "HIGH", "600ms",
     "Panicked shout + retreat horn + running footsteps",
     "1", ""),

    ("Army/Combat", "army_status_buff",
     "Positive status applied to army (inspired, fortified, etc.)",
     "MEDIUM", "300ms",
     "Ascending chime + shimmer",
     "2", ""),

    ("Army/Combat", "army_status_debuff",
     "Negative status applied (diseased, exhausted, etc.)",
     "MEDIUM", "300ms",
     "Low drone + cough/groan",
     "2", ""),

    ("Army/Combat", "army_level_up",
     "Army unit gains a promotion / level",
     "HIGH", "500ms",
     "Ascending trumpet + medal ding + crowd cheer (small)",
     "1", ""),

    ("Army/Combat", "army_magic_cast",
     "Wizard / caster attacks in battle",
     "HIGH", "500ms",
     "Arcane explosion — school-specific: fire=whoosh, ice=crack, earth=rumble",
     "6", "One per magic school: Fire, Ice, Nature, Shadow, Light, Storm"),

    ("Army/Combat", "army_navy_launch",
     "Naval unit deploys",
     "MEDIUM", "400ms",
     "Ship horn blast + water splash",
     "1", ""),

    # ── GOVERNORS / COMMANDERS ───────────────────────────────────────────────
    ("Governors/Commanders", "governor_hired",
     "Governor added to the player's roster",
     "HIGH", "400ms",
     "Parchment scroll + official wax seal stamp",
     "1", ""),

    ("Governors/Commanders", "governor_assigned",
     "Governor assigned to a tile (takes office)",
     "HIGH", "400ms",
     "Appointment chime + applause swell",
     "1", ""),

    ("Governors/Commanders", "governor_level_up",
     "Governor levels up (1→2→3)",
     "HIGH", "500ms",
     "Ascending 3-note chime + sparkle",
     "1", ""),

    ("Governors/Commanders", "governor_death",
     "A governor dies",
     "HIGH", "1s",
     "Single bell toll — mournful and resonant",
     "1", ""),

    ("Governors/Commanders", "commander_assigned",
     "Commander assigned to lead an army",
     "HIGH", "400ms",
     "Officer salute snap + short fanfare",
     "1", ""),

    ("Governors/Commanders", "commander_death",
     "A commander dies in battle",
     "HIGH", "1s",
     "Solemn trumpet + slow bell toll",
     "1", "More dramatic than governor death"),

    ("Governors/Commanders", "vp_succession",
     "Vice President succession triggers",
     "HIGH", "600ms",
     "Dramatic sting + gavel strike",
     "1", "Rare; should feel significant"),

    # ── TECH / LAWS / TRADITIONS ─────────────────────────────────────────────
    ("Tech/Laws/Traditions", "tech_unlock",
     "Technology researched and unlocked",
     "HIGH", "600ms",
     "Electric spark + rising sci-fi tone + completion chime",
     "1", ""),

    ("Tech/Laws/Traditions", "law_enacted",
     "Law passed and enacted",
     "HIGH", "500ms",
     "Gavel strike + brief applause from chamber",
     "1", ""),

    ("Tech/Laws/Traditions", "law_repealed",
     "Law repealed",
     "MEDIUM", "400ms",
     "Softer gavel + neutral crowd murmur",
     "1", ""),

    ("Tech/Laws/Traditions", "tradition_unlock",
     "Cultural tradition unlocked",
     "HIGH", "500ms",
     "Regional instrument sting appropriate to culture",
     "3", "Consider: Eastern, Native American, Colonial variants"),

    ("Tech/Laws/Traditions", "edict_issued",
     "Presidential edict issued",
     "HIGH", "500ms",
     "Short trumpet fanfare — regal and authoritative",
     "1", ""),

    ("Tech/Laws/Traditions", "crisis_declared",
     "National crisis declared",
     "HIGH", "600ms",
     "Alarm bell sequence + sustained low drone",
     "1", ""),

    # ── PROTECTORS / WIZARDS ─────────────────────────────────────────────────
    ("Protectors/Wizards", "protector_summon",
     "Protector summon event fires for the first time",
     "HIGH", "700ms",
     "Mystical horn + deep resonant rumble — something ancient awakens",
     "2", "Variants: benevolent-sounding vs threatening"),

    ("Protectors/Wizards", "protector_agree",
     "Player successfully binds a protector — they join as wizard",
     "HIGH", "1–1.5s",
     "Triumphant mystical fanfare + nature sound matching their school (e.g. Mothman = wings + wind)",
     "17", "Ideally one per protector; minimum one per magic school"),

    ("Protectors/Wizards", "protector_reject",
     "Player rejects a protector",
     "MEDIUM", "500ms",
     "Disappointed 'miss' horn + distant growl",
     "1", ""),

    ("Protectors/Wizards", "protector_tower_built",
     "Tower constructed at protector's home tile on agree",
     "HIGH", "600ms",
     "Stone rising + arcane seal being pressed",
     "1", ""),

    ("Protectors/Wizards", "protector_spell_grant",
     "Presidential Power spell granted to Ualani",
     "HIGH", "700ms",
     "Spellbook page burst open + arcane shimmer + Ualani's theme note",
     "1", ""),

    ("Protectors/Wizards", "wizard_spell_cast_fire",
     "Fire-school wizard casts a spell",
     "HIGH", "400ms",
     "Whoosh + crackle + heat distortion",
     "3", ""),

    ("Protectors/Wizards", "wizard_spell_cast_ice",
     "Ice/Storm-school wizard casts",
     "HIGH", "400ms",
     "Crystalline crack + freeze hiss",
     "3", ""),

    ("Protectors/Wizards", "wizard_spell_cast_nature",
     "Nature-school wizard casts",
     "HIGH", "400ms",
     "Rustling growth + animal call",
     "3", ""),

    ("Protectors/Wizards", "wizard_spell_cast_shadow",
     "Shadow-school wizard casts",
     "HIGH", "400ms",
     "Dark whoosh + whispers",
     "3", ""),

    ("Protectors/Wizards", "wizard_spell_cast_light",
     "Light-school wizard casts",
     "HIGH", "400ms",
     "Bell chime + radiant burst",
     "3", ""),

    ("Protectors/Wizards", "wizard_retaliation",
     "Enemy wizard retaliates / counterattacks",
     "HIGH", "500ms",
     "Dark arcane boom — heavier and more aggressive than normal cast",
     "2", ""),

    # ── FACTIONS / DIPLOMACY ──────────────────────────────────────────────────
    ("Factions/Diplomacy", "faction_discovered",
     "New faction encountered for the first time",
     "HIGH", "500ms",
     "Diplomatic fanfare + subtle cultural instrument",
     "3", "Different tone per faction type: military/civilian/religious"),

    ("Factions/Diplomacy", "faction_loyalty_gain",
     "Faction loyalty increases",
     "MEDIUM", "250ms",
     "Warm ascending tone + light applause",
     "2", ""),

    ("Factions/Diplomacy", "faction_loyalty_loss",
     "Faction loyalty decreases",
     "MEDIUM", "250ms",
     "Discordant descending tone",
     "2", ""),

    ("Factions/Diplomacy", "faction_alliance",
     "Alliance formed",
     "HIGH", "600ms",
     "Handshake impact + warm fanfare",
     "1", ""),

    ("Factions/Diplomacy", "faction_war_declared",
     "War declared against a faction",
     "HIGH", "600ms",
     "War horn blast + crowd noise surge",
     "1", ""),

    ("Factions/Diplomacy", "faction_defeated",
     "Faction eliminated from the game",
     "HIGH", "800ms",
     "Victory fanfare + crowd cheer (medium length)",
     "1", ""),

    # ── GAME STATE ────────────────────────────────────────────────────────────
    ("Game State", "game_start",
     "New game begins / world loads for first time",
     "HIGH", "3–5s",
     "Presidential inauguration fanfare — patriotic, grand, sets tone",
     "1", ""),

    ("Game State", "game_over_win",
     "Player achieves victory condition",
     "HIGH", "5–10s",
     "Full orchestral victory theme — uplifting and memorable",
     "1", "Should feel earned"),

    ("Game State", "game_over_loss",
     "Player loses the game",
     "HIGH", "5–10s",
     "Somber defeat theme — melancholy strings, fading away",
     "1", ""),

    ("Game State", "turn_start",
     "New turn begins after end-turn processing",
     "HIGH", "400ms",
     "Soft morning chime / new-day sound",
     "1", ""),

    ("Game State", "save_game",
     "Game saved",
     "MEDIUM", "300ms",
     "Paper stamp + confirmation click",
     "1", ""),

    ("Game State", "load_game",
     "Game loaded from save",
     "MEDIUM", "400ms",
     "Page flip + ambient fade-in",
     "1", ""),

    ("Game State", "notification",
     "Toast notification / alert popup appears",
     "MEDIUM", "200ms",
     "Soft notification ping",
     "2", ""),

    ("Game State", "approval_high",
     "Approval rating crosses into high zone",
     "MEDIUM", "400ms",
     "Small crowd cheer",
     "1", ""),

    ("Game State", "approval_low",
     "Approval rating falls to danger zone",
     "HIGH", "500ms",
     "Booing crowd + alarm",
     "1", ""),

    ("Game State", "ualani_ability",
     "Ualani uses a presidential power or unique ability",
     "HIGH", "600ms",
     "Unique Ualani signature sting — Hawaiian-inspired, powerful, memorable",
     "1", "Should be recognisable as distinctly hers"),
]


def priority_fill(p):
    if p == "HIGH":   return hex_fill(HIGH_BG)
    if p == "MEDIUM": return hex_fill(MED_BG)
    return hex_fill(LOW_BG)


def build_main_sheet(wb):
    ws = wb.active
    ws.title = "Sound Effects"

    hdr_font  = Font(bold=True, color=HDR_FG)
    hdr_fill  = hex_fill(HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    current_cat = None
    row = 2
    for entry in SOUNDS:
        cat, asset, when, prio, dur, char, variants, notes = entry

        if cat != current_cat:
            current_cat = cat
            for c in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row, column=c)
                cell.font      = Font(bold=True, color=CAT_FG)
                cell.fill      = hex_fill(CAT_BG)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = thin_border()
            ws.cell(row=row, column=1, value=f"── {cat.upper()} ──")
            ws.row_dimensions[row].height = 18
            row += 1

        alt  = (row % 2 == 0)
        base = hex_fill(ALT_BG) if alt else hex_fill(WHITE)
        vals = [cat, asset, when, prio, dur, char, variants, notes]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = thin_border()
            if c == 4:  # Priority
                cell.fill = priority_fill(prio)
                cell.font = Font(bold=True)
            else:
                cell.fill = base
        ws.row_dimensions[row].height = 42
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row - 1}"


def build_summary_sheet(wb):
    from collections import Counter
    ws = wb.create_sheet("Summary")

    cat_counts  = Counter(e[0] for e in SOUNDS)
    prio_counts = Counter(e[3] for e in SOUNDS)
    var_total   = sum(int(str(e[6]).split("–")[-1].strip()) for e in SOUNDS)

    hdr_font = Font(bold=True, color=HDR_FG)
    hdr_fill = hex_fill(HDR_BG)
    for col, w in [("A", 28), ("B", 14), ("D", 20), ("E", 14)]:
        ws.column_dimensions[col].width = w

    ws["A1"] = "Category"
    ws["B1"] = "Sound Count"
    for c in ["A1", "B1"]:
        ws[c].font = hdr_font
        ws[c].fill = hdr_fill

    for i, (cat, cnt) in enumerate(sorted(cat_counts.items()), 2):
        ws.cell(row=i, column=1, value=cat).fill  = hex_fill(ALT_BG)
        ws.cell(row=i, column=2, value=cnt).fill  = hex_fill(ALT_BG)

    tr = len(cat_counts) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font   = Font(bold=True)
    ws.cell(row=tr, column=2, value=len(SOUNDS)).font = Font(bold=True)
    ws.cell(row=tr + 1, column=1, value="Total incl. variants").font = Font(bold=True)
    ws.cell(row=tr + 1, column=2, value=f"~{var_total}").font        = Font(bold=True)

    ws["D1"] = "Priority"
    ws["E1"] = "Count"
    for c in ["D1", "E1"]:
        ws[c].font = hdr_font
        ws[c].fill = hdr_fill

    for i, (prio, cnt) in enumerate(sorted(prio_counts.items()), 2):
        ws.cell(row=i, column=4, value=prio).fill = priority_fill(prio)
        ws.cell(row=i, column=5, value=cnt).fill  = priority_fill(prio)

    ws.freeze_panes = "A2"


def main():
    wb = openpyxl.Workbook()
    build_main_sheet(wb)
    build_summary_sheet(wb)
    out = "sound_masterdoc.xlsx"
    wb.save(out)
    print(f"Saved {out}  ({len(SOUNDS)} unique sounds)")


if __name__ == "__main__":
    main()
