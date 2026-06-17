#!/usr/bin/env python3
"""
Generate flavordoc.xlsx — complete flavor & narrative tracker for USA and Canada.

Sheets:
  1. OVERVIEW      — status summary matrix across all flavor categories
  2. FACTIONS      — all USA/CA factions: leaders, tier rewards, laws granted
  3. LAWS          — all American and Canadian named laws
  4. DOCTRINES     — all belief doctrines (American + Canadian) with building effects
  5. ICONS         — all icon figures (American + Canadian) with effects and mil mods
  6. BELIEF MODS   — all icon/doctrine/axis mil mods with historical flavor
  7. GOVERNORS     — named governors with narrative hooks and faction affiliations
  8. VP ARC        — VP relationship arc events
  9. CA EVENTS     — Canadian events, alliance arc, and CA protectors

Status scale:
  FULL PASS   — fully written, implemented, and polished
  FIRST PASS  — functional and written; not fully polished
  FIRST DRAFT — skeleton exists; writing incomplete
  IDEA        — concept only; not yet written or implemented

Run from repo root:  python3 scripts/build_flavordoc.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "flavordoc.xlsx"

# ── PALETTE ──────────────────────────────────────────────────────────────────
HDR_BG, HDR_FG     = "1F2D3D", "FFFFFF"
USA_BG, USA_FG     = "1F3864", "FFFFFF"
CA_BG,  CA_FG      = "9B1C1C", "FFFFFF"
GEN_BG, GEN_FG     = "2E4057", "FFFFFF"
ALT_BG             = "EEF2F7"
WHITE              = "FFFFFF"

FULL_BG,   FULL_FG   = "00B050", "FFFFFF"
FPASS_BG,  FPASS_FG  = "92D050", "1A3A00"
FDRAFT_BG, FDRAFT_FG = "FFEB9C", "7A5A00"
IDEA_BG,   IDEA_FG   = "C9B1E8", "3A1A6A"

CAT_COLORS = {
    "USA":           "D6E4F7",
    "CA":            "FCE0D6",
    "GEN":           "EEF2F7",
    "Factions":      "D6EAF8",
    "Laws":          "EDE0F7",
    "Doctrines":     "F0E8C8",
    "Icons":         "FFE8D0",
    "Belief Mods":   "FDEBD0",
    "Governors":     "D6F5EA",
    "VP Arc":        "F5E6F0",
    "CA Events":     "D6EAF8",
}

STATUS_BG = {
    "FULL PASS":   FULL_BG,
    "FIRST PASS":  FPASS_BG,
    "FIRST DRAFT": FDRAFT_BG,
    "IDEA":        IDEA_BG,
}
STATUS_FG = {
    "FULL PASS":   FULL_FG,
    "FIRST PASS":  FPASS_FG,
    "FIRST DRAFT": FDRAFT_FG,
    "IDEA":        IDEA_FG,
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def _fill(h):
    return PatternFill("solid", fgColor=h)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, italic=False, color="000000", size=10):
    return Font(bold=bold, italic=italic, color=color, size=size, name="Calibri")

def _align(wrap=True, h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr(ws, row, cols, bg=HDR_BG, fg=HDR_FG, size=10):
    for col, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.fill = _fill(bg)
        c.font = _font(bold=True, color=fg, size=size)
        c.alignment = _align(wrap=False, h="center", v="center")
        c.border = _border()

def _row(ws, r, values, bg=WHITE, bold=False, italic=False, fg="1A1A1A", wrap=True):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.fill = _fill(bg)
        c.font = _font(bold=bold, italic=italic, color=fg)
        c.alignment = _align(wrap=wrap)
        c.border = _border()

def _status_cell(ws, r, col, status):
    c = ws.cell(row=r, column=col, value=status)
    bg = STATUS_BG.get(status, WHITE)
    fg = STATUS_FG.get(status, "1A1A1A")
    c.fill = _fill(bg)
    c.font = _font(bold=True, color=fg)
    c.alignment = _align(h="center")
    c.border = _border()

def _section_hdr(ws, r, num_cols, label, bg, fg):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws.cell(row=r, column=1, value=label)
    c.fill = _fill(bg)
    c.font = _font(bold=True, color=fg, size=11)
    c.alignment = _align(h="center", v="center")
    c.border = _border()

def _set_widths(ws, cols):
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════════════════════

# ── FACTIONS ─────────────────────────────────────────────────────────────────
# country | faction | leader | alignment | reward1 | reward2 | reward3 |
# laws_unlocked | notes | status

FACTIONS = [
    # ── USA ──────────────────────────────────────────────────────────────────
    ("USA", "Sons of Liberty", "Patrick Henry", "Freedom / Military",
     "Militia Muster", "Merchant Networks", "Letters of Marque",
     "Second Amendment, National Security Act, Merchant Marine Act",
     "Anti-Crown militia network. Henry is the loudest voice in the revolution and the most suspicious of the government it's building.",
     "FIRST PASS"),

    ("USA", "Continental Congress", "Abigail Adams", "Freedom / Order",
     "Articles of Confederation", "Foreign Diplomacy", "Constitutional Convention",
     "Municipal Reform Act, Voting Rights Act",
     "Constitutional democracy faction. Adams is the conscience of the movement — events about inclusion and what the revolution is willing to become.",
     "FIRST PASS"),

    ("USA", "Common Cause", "Francis Asbury", "Freedom / Equality",
     "Frontier Homesteads", "The People's Assembly", "Land Reform",
     "Second Amendment, Civil Rights Act",
     "Populist frontier movement. Asbury travels constantly — he knows what the ordinary people actually want from the revolution.",
     "FIRST DRAFT"),

    ("USA", "Abolitionist League", "Mercy Otis Warren", "Equality",
     "Freedom Papers", "Underground Railroad", "Universal Emancipation",
     "Civil Rights Act, Americans with Disabilities Act, Voting Rights Act",
     "Abolition-focused faction. Warren is the historian watching the revolution happen — she records everything and forgets nothing.",
     "FIRST PASS"),

    ("USA", "Free Workers Union", "Thomas Paine", "Equality",
     "Guild Charters", "General Strike", "Workers Commonwealth",
     "Americans with Disabilities Act, Civil Rights Act",
     "Labor rights faction. Paine trusts the people absolutely and is sometimes wrong. Events where principles collide with governance.",
     "FIRST PASS"),

    # ── CA ───────────────────────────────────────────────────────────────────
    ("CA", "French Habitants", "Marc Penoit / Pierre Renard (pool)", "Equality / Freedom",
     "Quebec Act Recognition", "Habitants Alliance", "Republic of Quebec",
     "Municipal Elections Act, Canadian Citizenship Act, Republic Elections Act",
     "Quebec sovereignty faction. Governor Carleton is the NPC counterpart. Events around language rights, dual governance, and Francophone identity.",
     "FIRST DRAFT"),

    ("CA", "Loyalist Settlers", "Benjamin Tallmadge (pool)", "Order",
     "Crown Defectors", "Pragmatic Compact", "New Republic Converts",
     "National Defence Act, Canada Shipping Act, Republic Elections Act",
     "Former Crown loyalists who've accepted the new reality. Provide military discipline and naval expertise. Uneasy but reliable.",
     "FIRST DRAFT"),

    ("CA", "Algonquin Nation", "Jessica Commanda Odjick", "Equality/Sovereignty",
     "Treaty of Friendship", "Haudenosaunee Alliance", "Sovereign Partnership",
     "Canadian Citizenship Act, Accessible Canada Act",
     "Indigenous sovereignty faction. The Confederacy has governed by consensus for centuries — the Republic would do well to learn from them. Key figure Joseph Brant occupies the NPC antagonist space.",
     "FIRST DRAFT"),

    ("CA", "Coureurs des Bois", "Louis Tremblant (pool)", "Freedom",
     "Trade Routes", "Frontier Network", "Continental Reach",
     "— (no laws currently; wilderness economy focus)",
     "Wilderness trade network. Know every river, every ridge, every tribe. Invaluable scouts. Culturally distinct from both Crown and Republic.",
     "IDEA"),

    ("CA", "Maritime Patriots", "— (uncast)", "Order / Freedom",
     "Port Alliance", "Atlantic Commerce", "Maritime Union",
     "Canada Shipping Act",
     "Atlantic coast faction. Halifax, Saint John, Prince Edward Island fishermen and dockhands who want the Crown out of their harbors.",
     "IDEA"),
]

# ── LAWS ─────────────────────────────────────────────────────────────────────
# country | law | quadrant | mechanical effect | real-world basis | description | icon_path | status

LAWS = [
    # ── USA ──────────────────────────────────────────────────────────────────
    ("USA", "Second Amendment", "Freedom",
     "+5 Weapons/Mansion lvl · +25 MP/Farm · −1 Mandate/Farm · +1 Mandate/Forge",
     "2nd Amendment (1791)", "The right of the people to keep and bear arms shall not be infringed. Every farm a muster point, every mansion an armoury — a Republic that arms its citizens does not ask permission to defend itself.",
     "", "FULL PASS"),

    ("USA", "Merchant Marine Act", "Order",
     "−2 Mandate/Workshop · +2 Gold/Workshop · +1 Gold/Farm · +1 Gold/Camp",
     "Merchant Marine Act (1920)", "American vessels carry American goods. The workshops that outfit the fleet answer to one flag, keep more of what they earn, and the river towns and frontier camps collect a share of every voyage.",
     "", "FULL PASS"),

    ("USA", "Municipal Reform Act", "Freedom",
     "−3 Mandate/Courthouse · +1 Happiness/Farm · +1 Happiness/Camp",
     "Municipal reform tradition (1800s–1900s)", "No appointed official may replace the voice of the people. Free elections in every township mean the courthouse runs on consent, not mandate — and the farms and camps that send their people there are better for it.",
     "", "FULL PASS"),

    ("USA", "Voting Rights Act", "Freedom",
     "−1 Mandate/Province · +1 Max Level Courthouse",
     "Voting Rights Act (1965)", "No citizen shall be denied the ballot on account of race or condition of previous servitude. The franchise is the foundation of the Republic — and the courthouse is where that foundation is enforced.",
     "", "FULL PASS"),

    ("USA", "Civil Rights Act", "Equality",
     "−1 Mandate/Population · +1 Happiness/Population",
     "Civil Rights Act (1964)", "Discrimination based on race, color, or creed is abolished. Every person the Republic governs is a person the Republic answers to — the mandate drops because the people it covers are finally counted.",
     "", "FULL PASS"),

    ("USA", "Americans with Disabilities Act", "Equality",
     "+1 Gold/Workshop · −10% Population Upgrade cost",
     "Americans with Disabilities Act (1990)", "No citizen is excluded from civic life on account of disability. The workshops that adapt their operations keep more of what they produce, and the cost of bringing more people into the Republic's economy falls.",
     "", "FULL PASS"),

    ("USA", "National Security Act", "Order",
     "−1 Mandate/Barracks · +5 Manpower/all buildings",
     "National Security Act (1947)", "One command, no gaps. The Republic draws its defence under the executive — the barracks run leaner, and every settlement from the frontier to the capital becomes a source of trained manpower.",
     "", "FULL PASS"),

    ("USA", "Homestead Act", "Freedom",
     "Farm: +1 Food · Camp: +1 Wood",
     "Homestead Act (1862)", "One hundred and sixty acres to any citizen willing to work it. The Republic built westward not by conquest alone but by offer — free land, free labour, and the promise that what you cultivate is yours. Every farm a frontier held; every camp a settlement begun.",
     "", "FULL PASS"),

    # ── CA ───────────────────────────────────────────────────────────────────
    ("CA", "Militia Act", "Freedom",
     "+5 Weapons/Mansion lvl · +25 MP/Farm · −1 Mandate/Farm · +1 Mandate/Forge",
     "Militia Act (1868)", "From Confederation onward, the Republic's defence rests on its people. Every farm a recruiting ground, every estate an armoury — the Militia Act turns the countryside into a reserve force that moves the moment the call goes out.",
     "", "FULL PASS"),

    ("CA", "Canada Shipping Act", "Order",
     "−1 Mandate/Dock · +2 Gold/Dock · +1 Happiness/Dock",
     "Canada Shipping Act (1936)", "Every vessel on Canadian waters submits to Republic safety standards — commercial freighters, fishing trawlers, and the pleasure yachts of whoever thought the St. Lawrence was a private amenity. The docks keep more of what they earn. In exchange, they fill out the forms.",
     "", "FULL PASS"),

    ("CA", "Municipal Elections Act", "Freedom",
     "−3 Mandate/Courthouse · +1 Happiness/Farm · +1 Happiness/Camp",
     "Municipal elections tradition (1849+)", "Responsible government runs all the way down. From Baldwin's day onward, every township elects its own council — the courthouse serves the community, not the administrator, and the people who work the land have a say in how it's governed.",
     "", "FULL PASS"),

    ("CA", "Republic Elections Act", "Freedom",
     "−1 Mandate/Province · +1 Max Level Courthouse",
     "Republic Elections Act (1920)", "One franchise, one standard, coast to coast. Ottawa's elections code unifies the ballot under federal law and elevates the courthouse as the seat of the democratic compact — no Province stands apart.",
     "", "FULL PASS"),

    ("CA", "Canadian Citizenship Act", "Equality",
     "−10% Mandate Cost (All Buildings)",
     "Canadian Citizenship Act (1947)", "For the first time, to live in the Republic is to belong to it. The British subject is gone; in its place, the Canadian citizen — French, Indigenous, newcomer alike, equal before every law that follows. The mandate slips because the Republic has grown wider than before. The harmony rises for exactly the same reason.",
     "", "FULL PASS"),

    ("CA", "Accessible Canada Act", "Equality",
     "+1 Gold/Courthouse · +1 Culture/Courthouse",
     "Accessible Canada Act (2019)", "No barrier — physical, digital, or architectural — shall prevent a Canadian from full participation in the life of the Republic. Workshops adapt and earn more; courthouses that serve everyone generate more revenue and more culture. The cost of bringing more citizens into the economy falls because the Republic has finally decided they belong there.",
     "", "FULL PASS"),

    ("CA", "National Defence Act", "Order",
     "−1 Mandate/Barracks · +5 Manpower/all buildings",
     "National Defence Act (1922)", "The Republic's defence is no longer a wartime improvisation — it is a permanent constitutional obligation. The National Defence Act unified command under federal authority, placing the armed forces on a standing footing that does not wait for the next crisis. Barracks run leaner; every settlement from the coasts to the interior contributes to the reserve.",
     "", "FULL PASS"),

    ("CA", "French Language Rights", "Equality",
     "+1 Culture/all Quebec buildings · +10% Manpower/Quebec Barracks · −2 Mandate/Quebec Courthouse",
     "French Language Rights (Emergency Recognition Act)", "The Prime Minister has answered the habitants with law. Every building in Quebec now operates in both official tongues of the Republic — the cultural output of the province rises, and the courthouses that administer federal authority in Quebec find their mandate stretched by the weight of recognition.",
     "", "FIRST PASS"),

    ("CA", "French Cultural Identity Enshrined", "Culture",
     "Quebec Resort: +1 Culture · +1 Gold · +100 Manpower · −1 Mandate/level · +50% resort dev speed (all CA)",
     "French Cultural Identity Enshrined (Cultural Institutes Act)", "The Republic funds what the Republic claims to value. Québécois cultural institutes receive direct patronage, turning resorts into cultural anchors — profitable, martial, and proud. Every leisure institution the Republic supports is, quietly, also a statement about who it has decided to be.",
     "", "FIRST PASS"),
]

# ── DOCTRINES ────────────────────────────────────────────────────────────────
# country | tier | name | real-world basis | year | building effects |
# mil mod granted | axis direction | status

DOCTRINES = [
    # ── USA Tier 1 ───────────────────────────────────────────────────────────
    ("USA", "Tier 1", "Nature Conservationists",      "Lacey Act / Conservation Movement", "1900",
     "Camp: +1 Culture · −0.1 Corruption/Farm & Camp level/turn", "—", "Providence", "FULL PASS"),
    ("USA", "Tier 1", "Civic Pride",                 "National Endowment for the Arts","1965",
     "Resort: +1 Culture · Monument: +0.5 Mandate/level/turn","—",       "Providence", "FULL PASS"),
    ("USA", "Tier 1", "Pioneer Heritage",            "Frontier Tradition",         "1800s",
     "Farm: +1 Food · −0.1 Corruption/Farm level/turn · Cost: 100 Culture",
     "—",       "Providence", "FULL PASS"),
    ("USA", "Tier 1", "Landmark Heritage",           "Antiquities Act",            "1906",
     "Monument: +1 Mandate",                                  "—",       "Providence", "FULL PASS"),
    ("USA", "Tier 1", "Sherman Antitrust Act",      "Sherman Antitrust Act",      "1890",
     "Market: +1 Gold · Mine: +1 Metal",                      "—",       "Providence", "FIRST DRAFT"),
    ("USA", "Tier 1", "Social Security Act",        "Social Security Act",        "1935",
     "Resort: +1 Happiness · Monument: +1 Culture",           "—",       "Providence", "FIRST DRAFT"),

    # ── USA Tier 2 ───────────────────────────────────────────────────────────
    ("USA", "Tier 2", "Inland Maritime Expertise",   "Inland Maritime Expertise",  "1964",
     "+1 Movement for units starting a turn on Major River or Major Lake tiles", "Woodsman","Providence", "FIRST DRAFT"),
    ("USA", "Tier 2", "First Amendment",            "1st Amendment (Bill of Rights)","1791",
     "Monument: +1 Mandate · Courthouse: +1 Mandate",         "—",       "Providence", "FIRST DRAFT"),
    ("USA", "Tier 2", "National Research Act",      "National Research Act",      "1974",
     "Library: +1 Science",                                   "—",       "Providence", "FIRST DRAFT"),
    ("USA", "Tier 2", "Height of Buildings Act",    "Height of Buildings Act",    "1910",
     "Monument: +1 Culture · +1 Mandate",                     "—",       "Providence", "FIRST DRAFT"),

    # ── USA Special ──────────────────────────────────────────────────────────
    ("USA", "Special", "Defense Production Act",   "Defense Production Act",     "1950",
     "Tower: +1 Magic · Library: +1 Science",                 "Vanguard","Providence", "FIRST DRAFT"),

    # ── CA Tier 1 ────────────────────────────────────────────────────────────
    ("CA", "Tier 1", "Nature Conservationists",           "Canada Wildlife Act / Conservation Movement", "1973",
     "Camp: +1 Culture · −0.1 Corruption/Farm & Camp level/turn", "—", "Providence", "FULL PASS"),
    ("CA", "Tier 1", "Canada Council for the Arts Act",  "Canada Council for the Arts",  "1957",
     "Theater/Faire: +1 Culture · Resort: +1 Happiness",      "—",       "Providence", "IDEA"),
    ("CA", "Tier 1", "Republic Lands Act",               "Republic Lands Act",            "1872",
     "Farm: +1 Food · Camp: +1 Wood",                         "—",       "Providence", "IDEA"),
    ("CA", "Tier 1", "Historic Sites and Monuments Act", "Historic Sites Act",            "1953",
     "Monument: +1 Mandate",                                  "—",       "Providence", "IDEA"),
    ("CA", "Tier 1", "Combines Investigation Act",       "Combines Investigation Act",    "1910",
     "Market: +1 Gold · Mine: +1 Metal",                      "—",       "Providence", "IDEA"),
    ("CA", "Tier 1", "Canada Health Act",                "Canada Health Act",             "1984",
     "Resort: +1 Happiness · Monument: +1 Culture",           "—",       "Providence", "IDEA"),

    # ── CA Tier 2 ────────────────────────────────────────────────────────────
    ("CA", "Tier 2", "National Parks Act",               "National Parks Act",            "1930",
     "Camp: +1 Wood · Barracks: +50 Manpower",                "Woodsman","Providence", "IDEA"),
    ("CA", "Tier 2", "Charter of Rights and Freedoms",   "Canadian Charter of R&F",      "1982",
     "Monument: +1 Mandate · Courthouse: +1 Mandate",         "—",       "Providence", "IDEA"),
    ("CA", "Tier 2", "Medical Research Council Act",     "Medical Research Council Act",  "1969",
     "Library: +1 Science",                                   "—",       "Providence", "IDEA"),
    ("CA", "Tier 2", "National Building Code of Canada", "Nat'l Building Code of Canada", "1941",
     "Monument: +1 Culture · +1 Mandate",                     "—",       "Providence", "IDEA"),

    # ── CA Special ───────────────────────────────────────────────────────────
    ("CA", "Special", "War Measures Act",            "War Measures Act",               "1914",
     "Tower: +1 Magic · Library: +1 Science",                 "Vanguard","Providence", "IDEA"),
]

# ── ICONS ────────────────────────────────────────────────────────────────────
# country | tier | name | era | building effects | mil mod | status | notes

ICONS = [
    # ── USA Tier 1 (Founding Era) ─────────────────────────────────────────────
    ("USA", "Tier 1 – Founding", "George Washington",  "1732–1799",
     "Barracks: +50 Manpower, +1 Weapons",
     "Crossing of the Delaware", "FIRST PASS",
     "Continental Army commander. Barracks/Fortress +3 Defense (mil mod)."),
    ("USA", "Tier 1 – Founding", "Benjamin Franklin",  "1706–1790",
     "Library: +1 Science · Workshop: +1 Gold",
     "—", "FIRST PASS",
     "Polymath inventor, diplomat, printer. Science and commercial output."),
    ("USA", "Tier 1 – Founding", "Abigail Adams",      "1744–1818",
     "Library: +1 Culture · Courthouse: +1 Mandate",
     "—", "FIRST PASS",
     "'Remember the Ladies.' Library culture and judicial mandate."),
    ("USA", "Tier 1 – Founding", "Alexander Hamilton", "1755–1804",
     "Market: +1 Gold · Workshop: +1 Gold",
     "—", "FIRST PASS",
     "First Treasury Secretary. Dual market/workshop gold bonus."),
    ("USA", "Tier 1 – Founding", "Phillis Wheatley",   "1753–1784",
     "Library: +1 Culture · Monument: +1 Culture",
     "—", "FIRST PASS",
     "First published African American poet. Dual culture bonus."),
    ("USA", "Tier 1 – Founding", "Thomas Jefferson",   "1743–1826",
     "Farm: +1 Food · Library: +1 Science",
     "—", "FIRST PASS",
     "Author of the Declaration, gentleman farmer, scholar. Farm and library."),

    # ── USA Tier 2 (1800s–Modern) ─────────────────────────────────────────────
    ("USA", "Tier 2 – 1800s/Modern", "Abraham Lincoln",       "1809–1865",
     "Barracks: +50 Manpower · Courthouse: +1 Mandate",
     "Emancipation Advance", "FIRST PASS",
     "16th President. All units +2 Atk +2 Def permanent (mil mod)."),
    ("USA", "Tier 2 – 1800s/Modern", "Harriet Tubman",        "1822–1913",
     "Barracks: +50 Manpower, +1 Weapons",
     "Combahee River Raid", "FIRST PASS",
     "Underground Railroad conductor, Union spy. +5 Atk first battle round (mil mod)."),
    ("USA", "Tier 2 – 1800s/Modern", "Frederick Douglass",    "1818–1895",
     "Library: +1 Culture · Courthouse: +1 Mandate",
     "North Star Address", "FIRST PASS",
     "Abolitionist orator. Morale loss reduced, rout threshold 15% (mil mod)."),
    ("USA", "Tier 2 – 1800s/Modern", "Sitting Bull",          "1831–1890",
     "Camp: +1 Wood, +1 Food",
     "Little Bighorn Ambush", "FIRST PASS",
     "Hunkpapa Lakota chief. +2 Atk +2 Def in Woods terrain (mil mod)."),
    ("USA", "Tier 2 – 1800s/Modern", "Sojourner Truth",       "1797–1883",
     "Farm: +1 Food · Monument: +1 Culture",
     "—", "FIRST PASS",
     "Abolitionist and women's rights activist. Farm and monument culture."),
    ("USA", "Tier 2 – 1800s/Modern", "Chief Joseph",          "1840–1904",
     "Courthouse: +1 Mandate",
     "—", "FIRST PASS",
     "Nez Perce leader. 'I will fight no more forever.' Courthouse justice."),
    ("USA", "Tier 2 – 1800s/Modern", "Theodore Roosevelt",    "1858–1919",
     "Mine: +1 Metal · Camp: +1 Wood · Barracks: +50 Manpower",
     "Rough Rider's Charge", "FIRST PASS",
     "Conservationist, Rough Rider. +4 Atk +4 Def in Woods/Wetlands (mil mod)."),
    ("USA", "Tier 2 – 1800s/Modern", "Susan B. Anthony",      "1820–1906",
     "Courthouse: +1 Mandate · Monument: +1 Culture",
     "—", "FIRST PASS",
     "Suffragist. Vote and civic culture. Courthouse and monument bonus."),
    ("USA", "Tier 2 – 1800s/Modern", "Ida B. Wells",          "1862–1931",
     "Library: +1 Culture",
     "—", "FIRST PASS",
     "Investigative journalist, anti-lynching crusader. Library culture."),
    ("USA", "Tier 2 – 1800s/Modern", "Eleanor Roosevelt",     "1884–1962",
     "Resort: +1 Happiness · Monument: +1 Culture",
     "—", "FIRST PASS",
     "UN diplomat, First Lady, human rights architect. Resort and monument."),
    ("USA", "Tier 2 – 1800s/Modern", "Martin Luther King Jr.","1929–1968",
     "Monument: +1 Culture, +1 Mandate · Courthouse: +1 Mandate",
     "—", "FULL PASS",
     "Baptist minister, Nobel laureate, apostle of nonviolent resistance. Monument yields culture AND mandate — his legacy reshaped civic identity; courthouse yields mandate because his movement forced the law to catch up."),
    ("USA", "Tier 2 – 1800s/Modern", "Cesar Chavez",          "1927–1993",
     "Farm: +1 Food · Barracks: +50 Manpower",
     "—", "FIRST PASS",
     "UFW co-founder. Farm labor and manpower bonus."),
    ("USA", "Tier 2 – 1800s/Modern", "Jimmy Carter",          "1924–2024",
     "Farm: +1 Food · Resort: +1 Happiness",
     "—", "FIRST PASS",
     "39th President, habitat builder, peacemaker. Farm and resort."),
    ("USA", "Tier 2 – 1800s/Modern", "Dolores Huerta",        "1930–",
     "Farm: +1 Food",
     "—", "FIRST PASS",
     "UFW co-founder. ¡Sí, se puede! Farm labor bonus."),

    # ── CA Tier 1 (Founding Era) ──────────────────────────────────────────────
    ("CA", "Tier 1 – Founding", "John A. Macdonald",          "1815–1891",
     "Courthouse: +1 Mandate · Monument: +1 Culture",
     "—", "IDEA",
     "First PM, Confederation architect. Nation-builder bonuses."),
    ("CA", "Tier 1 – Founding", "George-Étienne Cartier",     "1814–1873",
     "Library: +1 Culture · Courthouse: +1 Mandate",
     "—", "IDEA",
     "Father of Confederation (Quebec). Bilingual law and culture."),
    ("CA", "Tier 1 – Founding", "Wilfrid Laurier",            "1841–1919",
     "Market: +1 Gold · Monument: +1 Culture",
     "—", "IDEA",
     "First French-Canadian PM. Liberal economic prosperity and national pride."),
    ("CA", "Tier 1 – Founding", "Agnes Macphail",             "1890–1954",
     "Courthouse: +1 Mandate · Library: +1 Culture",
     "—", "IDEA",
     "First female MP (1921). Democracy and education."),
    ("CA", "Tier 1 – Founding", "Laura Secord",               "1775–1868",
     "Barracks: +50 Manpower, +1 Weapons",
     "Beaverdams Dispatch", "FIRST PASS",
     "Walked 20 miles to warn garrison at Beaverdams (1813). Barracks defense (mil mod)."),
    ("CA", "Tier 1 – Founding", "Louis-Hippolyte LaFontaine", "1807–1864",
     "Courthouse: +2 Mandate",
     "—", "IDEA",
     "Responsible government architect. Double courthouse mandate bonus."),

    # ── CA Tier 2 (Modern Era) ────────────────────────────────────────────────
    ("CA", "Tier 2 – Modern", "Tommy Douglas",          "1904–1986",
     "Resort: +1 Happiness · Monument: +1 Culture",
     "—", "IDEA",
     "Father of Canadian medicare. CCF leader. Resort and monument culture."),
    ("CA", "Tier 2 – Modern", "Viola Desmond",          "1882–1965",
     "Library: +1 Culture · Courthouse: +1 Mandate",
     "—", "IDEA",
     "Nova Scotia civil rights icon. Refused segregated seating (1946). Library and courthouse."),
    ("CA", "Tier 2 – Modern", "Lester B. Pearson",      "1897–1972",
     "Monument: +1 Culture, +1 Mandate",
     "—", "IDEA",
     "Nobel Peace Prize, peacekeeping founder, unified flag designer. Monument double."),
    ("CA", "Tier 2 – Modern", "Louis Riel",             "1844–1885",
     "Camp: +1 Wood, +1 Food",
     "Batoche's Stand", "FIRST PASS",
     "Métis leader, Father of Manitoba. Held Batoche with hunting rifles (1885). Woods terrain (mil mod)."),
    ("CA", "Tier 2 – Modern", "Emily Murphy",           "1868–1933",
     "Courthouse: +1 Mandate · Library: +1 Culture",
     "—", "IDEA",
     "Famous Five, 'Persons' case (1929). Law reform and civic culture."),
    ("CA", "Tier 2 – Modern", "Nellie McClung",         "1873–1951",
     "Farm: +1 Food · Monument: +1 Culture",
     "—", "IDEA",
     "Suffragist, Famous Five, temperance advocate. Prairie spirit, farm and monument."),
    ("CA", "Tier 2 – Modern", "Terry Fox",              "1958–1981",
     "Barracks: +50 Manpower · Resort: +1 Happiness",
     "—", "IDEA",
     "Marathon of Hope (1980). Physical endurance and national inspiration."),
    ("CA", "Tier 2 – Modern", "Chief Dan George",       "1899–1981",
     "Camp: +1 Wood · Library: +1 Culture",
     "—", "IDEA",
     "Tsleil-Waututh Chief, actor, Indigenous rights advocate. Camp and library wisdom."),
    ("CA", "Tier 2 – Modern", "Buffy Sainte-Marie",     "1941–",
     "Library: +1 Culture · Monument: +1 Culture",
     "—", "IDEA",
     "Cree artist, activist, educator. Dual culture bonus for arts and heritage."),
    ("CA", "Tier 2 – Modern", "David Suzuki",           "1936–",
     "Camp: +1 Wood · Farm: +1 Food",
     "—", "IDEA",
     "Environmentalist, The Nature of Things. Camp conservation and farm ecology."),
    ("CA", "Tier 2 – Modern", "Roméo Dallaire",         "1946–",
     "Barracks: +50 Manpower, +1 Weapons",
     "Peacekeeping Mandate", "FIRST PASS",
     "Lt. Gen., Rwanda genocide witness, peacekeeping advocate. Morale/rout reduction (mil mod)."),
    ("CA", "Tier 2 – Modern", "Thérèse Casgrain",       "1896–1981",
     "Courthouse: +1 Mandate · Library: +1 Culture",
     "—", "IDEA",
     "Quebec suffragist, CCF leader. Law reform and civic culture."),
    ("CA", "Tier 2 – Modern", "Mary Two-Axe Earley",    "1911–1996",
     "Courthouse: +1 Mandate · Farm: +1 Food",
     "—", "IDEA",
     "Mohawk activist. Fought to restore Indigenous women's status under Indian Act."),
    ("CA", "Tier 2 – Modern", "Pierre Elliott Trudeau", "1919–2000",
     "Monument: +1 Culture · Courthouse: +1 Mandate",
     "—", "IDEA",
     "15th PM, Charter of Rights architect, 'Just Society.' Monument and courthouse."),
]

# ── BELIEF MIL MODS ──────────────────────────────────────────────────────────
# mod name | source | figure/act | type | effect | historical flavor | status

BELIEF_MODS = [
    # ── American Icon Mods ───────────────────────────────────────────────────
    ("Crossing of the Delaware", "Icon — George Washington",
     "commanderMod", "+3 Defense in Barracks/Fortress tiles",
     "December 1776: Washington crossed the Delaware at midnight in sleet and darkness, surprising the Hessians at Trenton. The gambit saved the revolution.",
     "FIRST PASS"),

    ("Combahee River Raid", "Icon — Harriet Tubman",
     "infantryMod", "+5 Attack per Level in first battle round",
     "June 1863: Tubman led Col. James Montgomery's gunboats up the Combahee River, freeing 700+ enslaved people. No hesitation. No casualties on her side.",
     "FIRST PASS"),

    ("Emancipation Advance", "Icon — Abraham Lincoln",
     "commanderMod", "+2 Attack, +2 Defense per Level permanently",
     "January 1863: The Emancipation Proclamation transformed the war's moral stakes. Every soldier now fought for something larger than territory.",
     "FIRST PASS"),

    ("Rough Rider's Charge", "Icon — Theodore Roosevelt",
     "infantryMod (terrainMod: Woods)", "+4 Attack, +4 Defense in Woods or Wetlands terrain",
     "July 1898: Roosevelt led the Rough Riders charging up San Juan Hill under heavy fire. 'Bully!' he reportedly shouted. No one argued with him.",
     "FIRST PASS"),

    ("North Star Address", "Icon — Frederick Douglass",
     "commanderMod", "Morale loss reduced; rout threshold lowered to 15%",
     "Douglass's oratory — from escaped slave to the Republic's conscience. His newspaper The North Star (1847) gave voice to those the revolution was still failing.",
     "FIRST PASS"),

    ("Little Bighorn Ambush", "Icon — Sitting Bull",
     "infantryMod (terrainMod: Woods)", "+2 Attack, +2 Defense per Level in Woods terrain",
     "June 1876: Sitting Bull had a vision of soldiers falling like grasshoppers before the battle. The combined Lakota and Cheyenne forces destroyed Custer's 7th Cavalry.",
     "FIRST PASS"),

    # ── Canadian Icon Mods ───────────────────────────────────────────────────
    ("Beaverdams Dispatch", "Icon — Laura Secord",
     "commanderMod", "+3 Defense in Barracks/Fortress tiles",
     "June 1813: Secord walked 20 miles through American-occupied territory to warn Lt. FitzGibbon of an approaching attack. The garrison held. The Americans surrendered.",
     "FIRST PASS"),

    ("Batoche's Stand", "Icon — Louis Riel",
     "infantryMod (terrainMod: Woods)", "+2 Attack, +2 Defense per Level in Woods terrain",
     "May 1885: The Métis held Batoche for four days with hunting rifles, kitchen knives, and nails fired from improvised guns. Against a professional army. They lasted longer than anyone expected.",
     "FIRST PASS"),

    ("Peacekeeping Mandate", "Icon — Roméo Dallaire",
     "commanderMod", "Morale loss reduced; rout threshold lowered to 15%",
     "1994: Dallaire commanded UNAMIR in Rwanda with 2,500 troops and no mandate to stop a genocide. He stayed anyway. His testimony shaped every peacekeeping doctrine that followed.",
     "FIRST PASS"),

    # ── Doctrine Mods ────────────────────────────────────────────────────────
    ("Woodsman", "Doctrine — Inland Maritime Expertise / National Parks Act",
     "infantryMod (terrainMod: Woods)", "+2 Attack, +2 Defense per Level in Woods terrain",
     "The wilderness is not hostile to those who know it. Frontier rangers, trappers, and scouts treat the forest as cover, not obstacle.",
     "FIRST PASS"),

    ("Vanguard", "Doctrine — Defense Production Act / War Measures Act",
     "commanderMod", "+4 Attack per Level on first engagement in a fresh tile",
     "Industrial war footing means the army that strikes first has superior materiel. The first charge carries the advantage of preparation.",
     "FIRST PASS"),

    # ── Axis Mods (churchLevel ±3) ───────────────────────────────────────────
    ("Entrenched", "Axis — Providence III (churchLevel +3)",
     "commanderMod (entrenchMod)", "+5 Defense after 3 stationary turns; lost on movement",
     "A nation deeply committed to Providence digs in and holds. Faith becomes fortification. The Republic will not be moved from ground it has consecrated.",
     "FIRST PASS"),

    ("Sharpshooter", "Axis — Reason III (churchLevel −3)",
     "rangedMod", "Ranged attacks ignore 2 enemy Defense per Level",
     "An Enlightened Republic trains marksmen, not martyrs. Precision over prayer. The rifle doesn't need divine blessing — it needs a steady hand and a clear eye.",
     "FIRST PASS"),
]

# ── GOVERNORS ────────────────────────────────────────────────────────────────
# country | id | name | type | faction | alignment | narrative hook | status

GOVERNORS = [
    ("USA", "GOV_01", "Patrick Henry",       "Named Governor", "Sons of Liberty",       "Freedom/Military",
     "The revolution's loudest voice — and most suspicious of the government it's building. Events where ideological purity collides with pragmatic necessity.",
     "FIRST PASS"),
    ("USA", "GOV_02", "Abigail Adams",        "Named Governor", "Continental Congress",  "Freedom/Diplomat",
     "The voice that keeps the movement honest about who it's actually for. Events about inclusion, compromise, and what the revolution is willing to become.",
     "FIRST PASS"),
    ("USA", "GOV_03", "Thomas Paine",         "Named Governor", "Free Workers Union",    "Equality/Radical",
     "The man who lit the fire with a pamphlet. Trusts the people absolutely and is sometimes wrong about what they want.",
     "FIRST PASS"),
    ("USA", "GOV_04", "Mercy Otis Warren",    "Named Governor", "Abolitionist League",   "Equality/Patriot",
     "The historian watching the revolution happen. What's being recorded matters as much as what's being done. Abolitionist themes throughout.",
     "FIRST PASS"),
    ("USA", "GOV_05", "Daniel Shays",         "Named Governor", "Common Cause",          "Freedom/Populist",
     "Led the 1786 armed uprising against the Massachusetts government. Distrust of elites is his defining trait — including the elites now running the Republic.",
     "FIRST PASS"),
    ("USA", "GOV_06", "Ualani Carlisle",      "President (Player)", "Federal",            "Federal/Moderate",
     "The President. Hawaiian heritage; first woman in the role. The whole game runs through her — she IS the Republic's voice.",
     "FIRST PASS"),
    ("USA", "GOV_07", "Benjamin Tallmadge",   "Named Governor (pool)", "Loyalist Settlers","Order/Patriot",
     "Washington's spymaster, former Continental officer. Now in Canadian territory. Complex loyalty to both old and new orders.",
     "FIRST PASS"),
    ("USA", "GOV_08", "Phillis Wheatley",     "Named Governor (pool)", "Abolitionist League","Equality",
     "The first published African American poet. Awarded to player via Underground Railroad faction reward. Literary and moral gravitas.",
     "FIRST PASS"),
    ("USA", "GOV_09", "Francis Asbury",       "Named Governor", "Common Cause",          "Freedom/Populist",
     "Methodist circuit rider who covered 300,000 miles on horseback. He knows what ordinary people want from the revolution better than any congressman.",
     "FIRST PASS"),
    ("CA", "NPC_03", "Governor Carleton",     "NPC Faction Leader", "French Habitants",  "Order (Crown)",
     "British governor of Quebec who implemented the Quebec Act. Now a reluctant figure in a republic that's absorbed his colony. Complicated but not irredeemable.",
     "FIRST DRAFT"),
    ("CA", "NPC_08", "Jessica Commanda Odjick",      "Named Governor (PM)", "Algonquin Nation",  "Equality/Sovereignty",
     "Algonquin Prime Minister and diplomat. Carries the governance traditions of the Kitigan Zibi and Pikwakanagan into Continental politics. Named after William Commanda. Key voice for Indigenous sovereignty, treaty rights, and the future shape of the Canadian republic.",
     "FIRST DRAFT"),
    ("CA", "NPC_09", "Marc Penoit",           "Named Governor", "French Habitants",      "Equality/Freedom",
     "Québécois local leader navigating between French heritage, Crown legacy, and Continental alliance. Events around language rights, provincial autonomy, and what it means to be Canadien in a republic.",
     "FIRST DRAFT"),
]

# ── VP ARC ───────────────────────────────────────────────────────────────────
# id | headline | chain position | trigger | effect/buttons | status

VP_ARC = [
    ("VP_FIRST_MEETING", "THE VICE PRESIDENT REQUESTS A WORD: [COMMANDER_NAME] MAKES THEIR POSITION CLEAR",
     "Root (fires turn 5+, one-time)", "vp_met flag set; gates all other VP events",
     "BTN1: standard acknowledgment · BTN2: diplomatic response",
     "FIRST PASS"),
    ("VP_COUNSEL", "THE VICE PRESIDENT ARRIVES LATE WITH GOOD ADVICE",
     "Branch (req: vp_met, presidentialClaim < −2)", "Morale +10 or faction loyalty +10",
     "BTN1: Accept counsel (+10 morale) · BTN2: Defer (+10 VP faction loyalty)",
     "FIRST PASS"),
    ("VP_DOUBT", "THE VICE PRESIDENT AT 2 A.M.: [COMMANDER_NAME] IS QUESTIONING EVERYTHING",
     "Branch (req: vp_met)", "Relationship-building; no hard mechanical outcome",
     "BTN1: Reassure · BTN2: Share doubt (deepens relationship)",
     "FIRST PASS"),
    ("VP_LOYALTY_TEST", "[COMMANDER_NAME]'S COALITION WANTS ANSWERS — FACTION DEMANDS PUBLIC STATEMENT",
     "Branch (req: vp_met, VP faction loyalty < 20)", "VP faction loyalty +15 or +5",
     "BTN1: Full public statement (+15) · BTN2: Private assurance (+5)",
     "FIRST PASS"),
    ("VP_BATTLEFIELD", "THE VICE PRESIDENT PICKS UP A MUSKET: [COMMANDER_NAME] AT THE FRONT LINE",
     "Branch (req: vp_met)", "Military event; Crown advance repelled",
     "BTN1: Honor their service · BTN2: Order them to stand down",
     "FIRST PASS"),
    ("VP_PRE_ELECTION", "THE VICE PRESIDENT ISN'T SURE THEY WANT TO RUN",
     "Branch (req: vp_met, turns 88–92)", "Sets vp_declined_candidacy flag or morale +20",
     "BTN1: Accept withdrawal (sets flag; hides ELECTION_BTN1) · BTN2: Convince them to run (+20 morale)",
     "FIRST PASS"),
    ("VP_SACRIFICE", "THE VICE PRESIDENT OFFERS THEIR RESIGNATION",
     "Branch (req: vp_met, tile electionPressure < −20)", "VP faction +10 or VP resigned",
     "BTN1: Accept resignation (sets vp_resigned) · BTN2: Refuse (+10 VP faction loyalty)",
     "FIRST PASS"),
    ("VP_SOLIDARITY", "THE VICE PRESIDENT HAS BEEN PAYING ATTENTION — THEY KNOW ABOUT THE PROTECTORS",
     "Branch (req: vp_met)", "Relationship culmination; no hard outcome",
     "BTN1: Confirm everything · BTN2: Partial confirmation",
     "FIRST PASS"),
]

# ── CA PRESIDENT ARC (Marc Penoit as Deputy Governor for Canada playthrough) ──
CA_PM_ARC = [
    ("CA_PM_FIRST_MEETING", "MARC PENOIT REQUESTS A WAR COUNCIL",
     "Root (fires turn 5+, one-time)", "ca_pm_met flag set; gates all other CA_PM events",
     "BTN1: Standard acknowledgment · BTN2: Diplomatic response",
     "FIRST PASS"),
    ("CA_PM_COUNSEL", "PENOIT COUNSELS RESTRAINT — OR POSSIBLY AGGRESSION",
     "Branch (req: ca_pm_met, presidentialClaim < −2)", "Morale +10 or faction loyalty +10",
     "BTN1: Accept counsel · BTN2: Defer (+10 French Habitants loyalty)",
     "FIRST PASS"),
    ("CA_PM_DOUBT", "PENOIT QUESTIONS THE MILITIA'S RESOLVE",
     "Branch (req: ca_pm_met, tileMoralDecay > 30)", "Relationship-building event",
     "BTN1: Reassure · BTN2: Share doubt (deepens relationship)",
     "FIRST PASS"),
    ("CA_PM_LOYALTY_TEST", "THE FRENCH HABITANTS DEMAND RECOGNITION",
     "Branch (req: ca_pm_met, French Habitants loyalty < 20)", "Faction loyalty +15 or +5",
     "BTN1: Full public statement (+15) · BTN2: Private assurance (+5)",
     "FIRST PASS"),
    ("CA_PM_BATTLEFIELD", "JUST IN FROM THE FRONT — PRESIDENT PENOIT IS LEADING THE CHARGE PERSONALLY",
     "Branch (req: ca_pm_met, neighbor UK tile)", "BTN1: propaganda_spread all armies +2 attack −2 defence +30 mandate · BTN2: quebec_manpower_refill all QB armies full manpower",
     "BTN1: Spread the Images · BTN2: Limited Recruitment Drive",
     "FULL PASS"),
    ("CA_PM_PRE_ELECTION", "PENOIT ADDRESSES THE REPUBLIC: THE ELECTION IS COMING",
     "Branch (req: ca_pm_met, turns 88–92)", "Sets ca_pm_declined_candidacy or morale +20",
     "BTN1: Accept withdrawal · BTN2: Convince him to run (+20 morale)",
     "FIRST PASS"),
    ("CA_PM_SOLIDARITY", "PENOIT AND THE PROTECTORS — THE DEPUTY GOVERNOR ACKNOWLEDGES THE CREATURES",
     "Branch (req: ca_pm_met, 3+ CA protectors agreed)", "Relationship culmination; morale +15 or harmony +10",
     "BTN1: Forward to alliance (+15 morale) · BTN2: File with Records (+10 harmony)",
     "FULL PASS"),
    ("CA_PM_LEGACY", "MARC PENOIT'S FINAL ASSESSMENT — THE DEPUTY GOVERNOR WRITES HIS HISTORY",
     "Branch (req: ca_pm_met, turn 96+)", "Single dynamic letter (world.gd:_build_ca_pm_legacy_data). Paragraphs conditional on uk_ca_peace, all-Quebec ownership, named protectors, can_allied. Closes with Marc's independence decision: no referendum if Quebec freed + 2 Quebec protectors; referendum announced otherwise.",
     "BTN1: Acknowledge the Record (+1 claim)",
     "FULL PASS"),
]

# ── CA EVENTS ────────────────────────────────────────────────────────────────
# category | id | headline | chain_pos | trigger/flag | effect | status

CA_EVENTS = [
    # ── Canadian Alliance Arc ─────────────────────────────────────────────────
    ("Canadian Alliance", "UK_BUILDUP_01",     "UK FORCES MASS ON THE NORTHERN BORDER",
     "Root", "Fires when UK military presence grows; sets uk_buildup_known flag",
     "Sets up CAN_CALL_01; generates casus belli for Canadian contact",
     "FIRST PASS"),
    ("Canadian Alliance", "UK_DECLARATION_01", "CROWN DECLARES WAR ON THE CONTINENTAL REPUBLIC",
     "Root", "Fires on UK declaration",
     "Opens Canadian alliance tree; changes diplomatic options",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_CALL_01",       "THE REPUBLIC REACHES OUT TO CANADA",
     "Root", "Req: uk_buildup_known flag",
     "BTN1: Contact Canadian factions (→ CAN_PENOIT_01) · BTN2: Proceed without alliance",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_PENOIT_01",     "MARC PENOIT OF THE HABITANTS RECEIVES THE MESSAGE",
     "Branch", "Req: can_contact flag",
     "→ CAN_CLEARWATER_01; introduces Penoit as CA liaison",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_CLEARWATER_01", "JESSICA CLEAR-WATER REQUESTS A MEETING OF HER OWN",
     "Branch", "Req: CAN_PENOIT_01",
     "→ CAN_JOINT_OPS_01; introduces Haudenosaunee diplomatic track",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_JOINT_OPS_01",  "JOINT OPERATIONS: CONTINENTAL AND CANADIAN FORCES COORDINATE",
     "Branch", "Req: CAN_CLEARWATER_01",
     "→ CAN_SUMMIT_01; first military cooperation event",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_SUMMIT_01",     "THE CONTINENTAL SUMMIT — CANADA SENDS DELEGATES",
     "Branch", "Req: CAN_JOINT_OPS_01",
     "→ CAN_ALLIANCE_SIGNED; formal alliance negotiation",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_ALLIANCE_SIGNED", "CONTINENTAL-CANADIAN ALLIANCE FORMALIZED",
     "Branch", "Req: CAN_SUMMIT_01",
     "Sets can_allied flag; major diplomatic milestone",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_PEACE_01",      "THE ALLIANCE HOLDS — CANADA PROPOSES PEACE TERMS",
     "Followup", "Req: can_allied flag",
     "Diplomatic resolution options for joint prosecution of war",
     "FIRST PASS"),
    ("Canadian Alliance", "CAN_ELECTION_LUCK", "THE CANADIAN ELECTION: A CONTINENTAL ALLY WINS",
     "Followup", "Req: can_allied",
     "Bonus loyalty or resource gift; diplomatic goodwill event",
     "FIRST PASS"),
    ("Canadian Alliance", "PEACE_CA_AI_01",    "PEACE OF [TILE_NAME] — CANADA NEGOTIATES SEPARATE PEACE",
     "Branching outcome", "Fires if can_allied and CA player sues for peace",
     "Alliance fracture event; diplomatic fallout options",
     "FIRST DRAFT"),

    # ── CA Protectors ─────────────────────────────────────────────────────────
    ("CA Protector — Le Wendigo",          "CA_PROT_01_SUMMON", "LE WENDIGO WAKES IN THE LAURENTIAN FOREST",
     "Root",     "Fires when Laurentian region contested",
     "Summon/tame/agreement chain; grants Le Wendigo's Hunger mil mod",
     "FIRST PASS"),
    ("CA Protector — Le Wendigo",          "CA_PROT_01_TAME",   "THE WENDIGO RECEDES FROM THE HARVEST CORRIDOR",
     "Branch",   "Req: PROT_01_SUMMON",
     "→ CA_PROT_01_AGREE; negotiation event",
     "FIRST PASS"),
    ("CA Protector — Le Wendigo",          "CA_PROT_01_AGREE",  "LE WENDIGO STANDS DOWN: SAINT-GEORGES FOREST ACCORD",
     "Followup", "Req: CA_PROT_01_TAME",
     "Protector ally secured; mil mod applied",
     "FIRST PASS"),

    ("CA Protector — Le Loup-Garou",       "CA_PROT_02_SUMMON", "LE LOUP-GAROU RUNS THE RIVER ROAD — RIVIÈRE-DU-LOUP QUARANTINED",
     "Root",     "Fires near Rivière-du-Loup tile",
     "Summon/tame/agreement chain; grants Loup-Garou's Frenzy mil mod",
     "FIRST PASS"),
    ("CA Protector — Le Loup-Garou",       "CA_PROT_02_TAME",   "LE LOUP-GAROU ALLOWS PASSAGE — THE RIVER ROAD REOPENS",
     "Branch",   "Req: CA_PROT_02_SUMMON", "→ CA_PROT_02_AGREE", "FIRST PASS"),
    ("CA Protector — Le Loup-Garou",       "CA_PROT_02_AGREE",  "LE LOUP-GAROU OF RIVIÈRE-DU-LOUP: FORMALLY ACKNOWLEDGED BY THE REPUBLIC",
     "Followup", "Req: CA_PROT_02_TAME",  "Protector ally secured",  "FIRST PASS"),

    ("CA Protector — Les Feux Follets",    "CA_PROT_03_SUMMON", "LES FEUX FOLLETS RISE OVER THE SAINT JOHN MARSHES — NAVIGATION DISRUPTED",
     "Root",     "Fires near Saint John marshes", "Summon/tame chain; grants Feux Follets' Misdirection",  "FIRST PASS"),
    ("CA Protector — Les Feux Follets",    "CA_PROT_03_TAME",   "LES FEUX FOLLETS QUIET — MALISEET PROTOCOLS ESTABLISHED",
     "Branch",   "Req: CA_PROT_03_SUMMON", "→ CA_PROT_03_AGREE; cultural protocol required", "FIRST PASS"),
    ("CA Protector — Les Feux Follets",    "CA_PROT_03_AGREE",  "LES FEUX FOLLETS FORMALLY ACKNOWLEDGED: THE SAINT JOHN MARSH ACCORD",
     "Followup", "Req: CA_PROT_03_TAME",  "Protector ally secured",  "FIRST PASS"),

    ("CA Protector — Mishepeshu",          "CA_PROT_04_SUMMON", "MISHEPESHU STIRS BENEATH LAKE SIMCOE — CROSSINGS CONTESTED",
     "Root",     "Fires near Lake Simcoe", "Summon/tame chain; grants Mishepeshu's Depths", "FIRST PASS"),
    ("CA Protector — Mishepeshu",          "CA_PROT_04_TAME",   "MISHEPESHU WITHDRAWS FROM THE CROSSING LANES",
     "Branch",   "Req: CA_PROT_04_SUMMON", "→ CA_PROT_04_AGREE", "FIRST PASS"),
    ("CA Protector — Mishepeshu",          "CA_PROT_04_AGREE",  "MISHEPESHU PACT: THE GREAT LYNX HOLDS THE ONTARIO WATERWAYS",
     "Followup", "Req: CA_PROT_04_TAME",  "Protector ally secured",  "FIRST PASS"),

    ("CA Protector — La Corriveau",        "CA_PROT_05_SUMMON", "LA CORRIVEAU IS FREE — THE IRON CAGE HAS BEEN SWINGING IN TROIS-PISTOLES",
     "Root",     "Fires near Trois-Pistoles tile", "Summon/tame chain; grants La Corriveau's Cage", "FIRST PASS"),
    ("CA Protector — La Corriveau",        "CA_PROT_05_TAME",   "LA CORRIVEAU QUIETS — THE REPUBLIC ACKNOWLEDGES THE VERDICT WAS WRONG",
     "Branch",   "Req: CA_PROT_05_SUMMON", "→ CA_PROT_05_AGREE; justice acknowledgment required", "FIRST PASS"),
    ("CA Protector — La Corriveau",        "CA_PROT_05_AGREE",  "LA CORRIVEAU FORMALLY ALLIES WITH THE REPUBLIC",
     "Followup", "Req: CA_PROT_05_TAME",  "Protector ally secured",  "FIRST PASS"),

    ("CA Protector — Le Carcajou",         "CA_PROT_06_SUMMON", "LE CARCAJOU MOVES THROUGH THE MONCTON FOREST — NOTHING STOPS IT",
     "Root",     "Fires near Moncton tile", "Summon/tame chain; grants Le Carcajou's Tenacity", "FIRST PASS"),
    ("CA Protector — Le Carcajou",         "CA_PROT_06_TAME",   "LE CARCAJOU REDIRECTS — CONTINENTAL LINES EXEMPTED",
     "Branch",   "Req: CA_PROT_06_SUMMON", "→ CA_PROT_06_AGREE", "FIRST PASS"),
    ("CA Protector — Le Carcajou",         "CA_PROT_06_AGREE",  "LE CARCAJOU FORMALLY RECOGNIZED: THE MONCTON CORRIDOR ACCORD",
     "Followup", "Req: CA_PROT_06_TAME",  "Protector ally secured",  "FIRST PASS"),

    ("CA Protector — La Chasse-Galerie",   "CA_PROT_07_SUMMON", "THE FLYING CANOE RUNS THE OTTAWA RIVER — THREE SIGHTINGS, TWO MISSING PATROLS",
     "Root",     "Fires near Ottawa River corridor", "Summon/tame chain; grants La Chasse-Galerie mil mod", "FIRST PASS"),
    ("CA Protector — La Chasse-Galerie",   "CA_PROT_07_TAME",   "THE FLYING CANOE ACKNOWLEDGES THE REPUBLIC — OTTAWA CORRIDOR SECURED",
     "Branch",   "Req: CA_PROT_07_SUMMON", "→ CA_PROT_07_AGREE", "FIRST PASS"),
    ("CA Protector — La Chasse-Galerie",   "CA_PROT_07_AGREE",  "LA CHASSE-GALERIE: THE RIVER RUNS FOR THE REPUBLIC NOW",
     "Followup", "Req: CA_PROT_07_TAME",  "Protector ally secured",  "FIRST PASS"),

    ("CA Protector — Le Gougou",           "CA_PROT_08_SUMMON", "THE GOUGOU RISES IN THE CHALEUR BAY — CHAMPLAIN'S MONSTER IS REAL",
     "Root",     "Fires near Chaleur Bay tile", "Summon/tame chain; grants Le Gougou's Terror mil mod", "FIRST PASS"),
    ("CA Protector — Le Gougou",           "CA_PROT_08_TAME",   "THE GOUGOU WITHDRAWS FROM THE BAY LANES — RESTITUTION PROTOCOLS ACTIVE",
     "Branch",   "Req: CA_PROT_08_SUMMON", "→ CA_PROT_08_AGREE", "FIRST PASS"),
    ("CA Protector — Le Gougou",           "CA_PROT_08_AGREE",  "THE GOUGOU FORMALLY GUARDS THE CHALEUR BAY FOR THE CONTINENTAL ALLIANCE",
     "Followup", "Req: CA_PROT_08_TAME",  "Protector ally secured",  "FIRST PASS"),

    # ── CA Playthrough: Collapse & Loss Conditions ────────────────────────────
    ("CA Collapse", "CA_COLLAPSE_01",
     "OTTAWA HAS FALLEN", "Loss trigger (Ottawa tile lost to UK)",
     "Sets _ca_collapsed flag", "Triggers game-over scene; narrative resolution",
     "FIRST PASS"),
    ("CA Collapse", "CA_COLLAPSE_JESSICA",
     "JESSICA CLEAR-WATER HAS FALLEN", "Loss trigger (Jessica removed from unlockedGovernors)",
     "Sets _ca_collapsed flag", "Triggers game-over scene; Penoit takes command narrative",
     "FIRST PASS"),

    # ── CA Playthrough: USA Alliance Arc (from Canada's perspective) ──────────
    ("USA Alliance Arc", "USA_CALL_01",
     "A DISPATCH FROM WASHINGTON — THE AMERICANS WANT TO TALK",
     "Root (turn 8+, req: uk_buildup_known)", "Sets usa_contact flag; gates USA_SUMMIT_01",
     "BTN1: Hear them out (→ USA_SUMMIT_01) · BTN2: Decline (sets usa_rejected → CA_ALONE_01)",
     "FIRST PASS"),
    ("USA Alliance Arc", "USA_SUMMIT_01",
     "THE OTTAWA SUMMIT — JESSICA CLEAR-WATER MEETS THE AMERICAN DELEGATION",
     "Branch (req: usa_contact, turn 13+)", "Sets usa_summit_complete flag",
     "BTN1: Sign the accord (→ USA_ALLIANCE_SIGNED) · BTN2: Walk away (sets usa_rejected)",
     "FIRST PASS"),
    ("USA Alliance Arc", "USA_ALLIANCE_SIGNED",
     "THE CANADIAN-CONTINENTAL ACCORD IS SIGNED",
     "Branch (req: usa_summit_complete)", "Sets ca_allied flag; mirrors CAN_ALLIANCE_SIGNED from USA side",
     "Narrative resolution; alliance milestone event",
     "FIRST PASS"),
    ("USA Alliance Arc", "CA_ALONE_01",
     "CANADA STANDS ALONE — THE REPUBLIC DECLINES THE AMERICAN OFFER",
     "Branch (req: usa_rejected flag)", "Canada fights without USA alliance",
     "Narrative event; opens solo-play diplomatic tree",
     "FIRST PASS"),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_overview(wb):
    ws = wb.create_sheet("OVERVIEW")
    ws.sheet_properties.tabColor = "1F2D3D"

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "FARSWORD — FLAVOR DOCUMENT  ·  USA & CANADA"
    c.fill = _fill(HDR_BG)
    c.font = _font(bold=True, color=HDR_FG, size=14)
    c.alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 28

    headers = [("CATEGORY", 26), ("FULL PASS", 13), ("FIRST PASS", 13),
               ("FIRST DRAFT", 13), ("IDEA", 13), ("TOTAL", 10), ("NOTES", 40)]
    _hdr(ws, 2, headers)
    _set_widths(ws, headers)
    _freeze(ws, "B3")

    def count(data, col_idx):
        fp = sum(1 for r in data if r[col_idx] == "FULL PASS")
        p1 = sum(1 for r in data if r[col_idx] == "FIRST PASS")
        p2 = sum(1 for r in data if r[col_idx] == "FIRST DRAFT")
        id_ = sum(1 for r in data if r[col_idx] == "IDEA")
        return fp, p1, p2, id_

    rows = [
        ("Factions",      FACTIONS,    9,  "USA: 5 factions · CA: 5 factions"),
        ("Laws",          LAWS,        6,  "USA: 7 laws · CA: 7 laws"),
        ("Doctrines",     DOCTRINES,   8,  "USA: 11 doctrines · CA: 11 doctrines"),
        ("Icons",         ICONS,       7,  "USA: 20 icons · CA: 20 icons"),
        ("Belief Mods",   BELIEF_MODS, 5,  "6 icon mods · 2 doctrine mods · 2 axis mods · 1 Canadian"),
        ("Governors",     GOVERNORS,   7,  "9 USA named governors · 3 CA named governors"),
        ("VP Arc (USA)",  VP_ARC,      5,  "8 VP arc events — Ualani + elected VP"),
        ("PM Arc (CA)",   CA_PM_ARC,   5,  "9 CA President arc events — Jessica + Marc Penoit"),
        ("CA Events",     CA_EVENTS,   6,  "CA protectors (8×3) · alliance arc · collapse events · USA-CA alliance (CA side)"),
    ]

    alt = False
    for r_idx, (label, data, status_col, note) in enumerate(rows, 3):
        fp, p1, p2, id_ = count(data, status_col)
        total = fp + p1 + p2 + id_
        bg = ALT_BG if alt else WHITE
        vals = [label, fp or "", p1 or "", p2 or "", id_ or "", total, note]
        _row(ws, r_idx, vals, bg=bg)
        # color the count cells
        for col, (cnt, fbg, ffg) in enumerate([(fp, FULL_BG, FULL_FG), (p1, FPASS_BG, FPASS_FG),
                                                (p2, FDRAFT_BG, FDRAFT_FG), (id_, IDEA_BG, IDEA_FG)], 2):
            c = ws.cell(row=r_idx, column=col)
            if cnt:
                c.fill = _fill(fbg)
                c.font = _font(bold=True, color=ffg)
                c.alignment = _align(h="center")
        alt = not alt

    ws.row_dimensions[2].height = 20


def build_factions(wb):
    ws = wb.create_sheet("FACTIONS")
    ws.sheet_properties.tabColor = CAT_COLORS["Factions"]

    cols = [
        ("COUNTRY",        10), ("FACTION",          22), ("LEADER",          22),
        ("ALIGNMENT",      18), ("REWARD TIER 1",    22), ("REWARD TIER 2",   22),
        ("REWARD TIER 3",  22), ("LAWS UNLOCKED",    36), ("NOTES",           52), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "B2")

    _section_hdr(ws, 2, len(cols), "UNITED STATES OF AMERICA", USA_BG, USA_FG)
    r = 3
    for row in [f for f in FACTIONS if f[0] == "USA"]:
        bg = CAT_COLORS["USA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    _section_hdr(ws, r, len(cols), "CANADA", CA_BG, CA_FG)
    r += 1
    for row in [f for f in FACTIONS if f[0] == "CA"]:
        bg = CAT_COLORS["CA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    ws.row_dimensions[1].height = 20


def build_laws(wb):
    ws = wb.create_sheet("LAWS")
    ws.sheet_properties.tabColor = CAT_COLORS["Laws"]

    cols = [
        ("COUNTRY", 10), ("LAW NAME", 30), ("QUADRANT", 12),
        ("MECHANICAL EFFECT", 42), ("REAL-WORLD BASIS", 32),
        ("FLAVOR DESCRIPTION", 60), ("ICON PATH", 48), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "B2")

    _section_hdr(ws, 2, len(cols), "AMERICAN LAWS", USA_BG, USA_FG)
    r = 3
    for row in [l for l in LAWS if l[0] == "USA"]:
        bg = CAT_COLORS["USA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    _section_hdr(ws, r, len(cols), "CANADIAN LAWS", CA_BG, CA_FG)
    r += 1
    for row in [l for l in LAWS if l[0] == "CA"]:
        bg = CAT_COLORS["CA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    ws.row_dimensions[1].height = 20


def build_doctrines(wb):
    ws = wb.create_sheet("DOCTRINES")
    ws.sheet_properties.tabColor = CAT_COLORS["Doctrines"]

    cols = [
        ("COUNTRY", 10), ("TIER", 12), ("DOCTRINE NAME", 30), ("REAL-WORLD BASIS", 28),
        ("YEAR", 7), ("BUILDING EFFECTS", 40), ("MIL MOD", 18),
        ("AXIS DIR.", 12), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "C2")

    _section_hdr(ws, 2, len(cols), "AMERICAN DOCTRINES", USA_BG, USA_FG)
    r = 3
    for row in [d for d in DOCTRINES if d[0] == "USA"]:
        bg = CAT_COLORS["USA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    _section_hdr(ws, r, len(cols), "CANADIAN DOCTRINES", CA_BG, CA_FG)
    r += 1
    for row in [d for d in DOCTRINES if d[0] == "CA"]:
        bg = CAT_COLORS["CA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    ws.row_dimensions[1].height = 20


def build_icons(wb):
    ws = wb.create_sheet("ICONS")
    ws.sheet_properties.tabColor = CAT_COLORS["Icons"]

    cols = [
        ("COUNTRY", 10), ("TIER", 20), ("FIGURE", 24), ("DATES", 12),
        ("BUILDING EFFECTS", 38), ("MIL MOD GRANTED", 24),
        ("STATUS", 14), ("HISTORICAL NOTES", 54),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "C2")

    _section_hdr(ws, 2, len(cols), "AMERICAN ICONS", USA_BG, USA_FG)
    r = 3
    for row in [i for i in ICONS if i[0] == "USA"]:
        bg = CAT_COLORS["USA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, 7, row[6])
        ws.cell(row=r, column=8).value = row[7]
        r += 1

    _section_hdr(ws, r, len(cols), "CANADIAN ICONS", CA_BG, CA_FG)
    r += 1
    for row in [i for i in ICONS if i[0] == "CA"]:
        bg = CAT_COLORS["CA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, 7, row[6])
        ws.cell(row=r, column=8).value = row[7]
        r += 1

    ws.row_dimensions[1].height = 20


def build_belief_mods(wb):
    ws = wb.create_sheet("BELIEF MODS")
    ws.sheet_properties.tabColor = CAT_COLORS["Belief Mods"]

    cols = [
        ("MOD NAME", 28), ("SOURCE", 34), ("TYPE", 26),
        ("MECHANICAL EFFECT", 40), ("HISTORICAL FLAVOR", 66), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "B2")

    sections = [
        ("AMERICAN ICON MODS", USA_BG, USA_FG, lambda r: r[1].startswith("Icon — ") and any(
            n in r[1] for n in ["Washington","Tubman","Lincoln","Roosevelt","Douglass","Sitting Bull"])),
        ("CANADIAN ICON MODS", CA_BG, CA_FG, lambda r: r[1].startswith("Icon — ") and any(
            n in r[1] for n in ["Secord","Riel","Dallaire"])),
        ("DOCTRINE MODS", "2E4057", "FFFFFF", lambda r: r[1].startswith("Doctrine")),
        ("AXIS MODS — REASON ↔ PROVIDENCE", "4A2060", "FFFFFF", lambda r: r[1].startswith("Axis")),
    ]

    r = 2
    for sec_label, bg, fg, filt in sections:
        rows = [m for m in BELIEF_MODS if filt(m)]
        if not rows:
            continue
        _section_hdr(ws, r, len(cols), sec_label, bg, fg)
        r += 1
        for row in rows:
            rbg = ALT_BG if r % 2 == 0 else WHITE
            _row(ws, r, row[:-1], bg=rbg)
            _status_cell(ws, r, len(cols), row[-1])
            r += 1

    ws.row_dimensions[1].height = 20


def build_governors(wb):
    ws = wb.create_sheet("GOVERNORS")
    ws.sheet_properties.tabColor = CAT_COLORS["Governors"]

    cols = [
        ("COUNTRY", 10), ("ID", 10), ("NAME", 22), ("TYPE", 24),
        ("FACTION", 22), ("ALIGNMENT", 18), ("NARRATIVE HOOK", 66), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "C2")

    _section_hdr(ws, 2, len(cols), "USA NAMED GOVERNORS & NPCs", USA_BG, USA_FG)
    r = 3
    for row in [g for g in GOVERNORS if g[0] == "USA"]:
        bg = CAT_COLORS["USA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    _section_hdr(ws, r, len(cols), "CANADIAN GOVERNORS & NPCs", CA_BG, CA_FG)
    r += 1
    for row in [g for g in GOVERNORS if g[0] == "CA"]:
        bg = CAT_COLORS["CA"] if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    ws.row_dimensions[1].height = 20


def build_vp_arc(wb):
    ws = wb.create_sheet("VP ARC")
    ws.sheet_properties.tabColor = CAT_COLORS["VP Arc"]

    cols = [
        ("EVENT ID", 22), ("HEADLINE", 60), ("CHAIN POSITION", 34),
        ("TRIGGER / FLAG", 36), ("BUTTON EFFECTS", 40), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "B2")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    c = ws.cell(row=2, column=1, value="VICE PRESIDENT RELATIONSHIP ARC  ·  [COMMANDER_NAME] is the player-chosen VP")
    c.fill = _fill("F5E6F0")
    c.font = _font(bold=True, color="660044", size=11)
    c.alignment = _align(h="center")

    r = 3
    for row in VP_ARC:
        bg = ALT_BG if r % 2 == 0 else WHITE
        _row(ws, r, row[:-1], bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    ws.row_dimensions[1].height = 20


def build_ca_events(wb):
    ws = wb.create_sheet("CA EVENTS")
    ws.sheet_properties.tabColor = CAT_COLORS["CA Events"]

    cols = [
        ("CATEGORY", 28), ("EVENT ID", 22), ("HEADLINE", 52),
        ("CHAIN POSITION", 16), ("TRIGGER / FLAG", 38),
        ("EFFECT / BUTTONS", 42), ("STATUS", 14),
    ]
    _hdr(ws, 1, cols)
    _set_widths(ws, cols)
    _freeze(ws, "B2")

    # ── Alliance arc ──────────────────────────────────────────────────────────
    alliance = [e for e in CA_EVENTS if "Alliance" in e[0] or "UK" in e[1] or "PEACE" in e[1]]
    _section_hdr(ws, 2, len(cols), "CANADIAN ALLIANCE ARC", "1F3864", "FFFFFF")
    r = 3
    for row in alliance:
        bg = CAT_COLORS["CA Events"] if r % 2 == 0 else WHITE
        _row(ws, r, row, bg=bg)
        # last element is status
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    # ── CA Protectors (grouped by protector) ──────────────────────────────────
    prot_rows = [e for e in CA_EVENTS if "Protector" in e[0]]
    seen_prot = []
    _section_hdr(ws, r, len(cols), "CANADIAN PROTECTORS  (8 protectors × 3 events each)", CA_BG, CA_FG)
    r += 1
    for row in prot_rows:
        protector = row[0]
        if protector not in seen_prot:
            seen_prot.append(protector)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
            c = ws.cell(row=r, column=1, value=protector.upper())
            c.fill = _fill("EDD6F7")
            c.font = _font(bold=True, color="3A1A6A")
            c.alignment = _align(h="left")
            c.border = _border()
            r += 1
        bg = ALT_BG if r % 2 == 0 else WHITE
        _row(ws, r, row, bg=bg)
        _status_cell(ws, r, len(cols), row[-1])
        r += 1

    ws.row_dimensions[1].height = 20


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_overview(wb)
    build_factions(wb)
    build_laws(wb)
    build_doctrines(wb)
    build_icons(wb)
    build_belief_mods(wb)
    build_governors(wb)
    build_vp_arc(wb)
    build_ca_events(wb)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")

    # ── summary ───────────────────────────────────────────────────────────────
    def count_status(data, idx):
        from collections import Counter
        return Counter(r[idx] for r in data)

    total_entries = (len(FACTIONS) + len(LAWS) + len(DOCTRINES) +
                     len(ICONS) + len(BELIEF_MODS) + len(GOVERNORS) +
                     len(VP_ARC) + len(CA_PM_ARC) + len(CA_EVENTS))
    print(f"\n  Factions:    {len(FACTIONS):3d}  (USA: {sum(1 for f in FACTIONS if f[0]=='USA')}, CA: {sum(1 for f in FACTIONS if f[0]=='CA')})")
    print(f"  Laws:        {len(LAWS):3d}  (USA: {sum(1 for l in LAWS if l[0]=='USA')}, CA: {sum(1 for l in LAWS if l[0]=='CA')})")
    print(f"  Doctrines:   {len(DOCTRINES):3d}  (USA: {sum(1 for d in DOCTRINES if d[0]=='USA')}, CA: {sum(1 for d in DOCTRINES if d[0]=='CA')})")
    print(f"  Icons:       {len(ICONS):3d}  (USA: {sum(1 for i in ICONS if i[0]=='USA')}, CA: {sum(1 for i in ICONS if i[0]=='CA')})")
    print(f"  Belief Mods: {len(BELIEF_MODS):3d}")
    print(f"  Governors:   {len(GOVERNORS):3d}")
    print(f"  VP Arc:      {len(VP_ARC):3d}  (USA)")
    print(f"  PM Arc:      {len(CA_PM_ARC):3d}  (CA)")
    print(f"  CA Events:   {len(CA_EVENTS):3d}")
    print(f"  ─────────────────")
    print(f"  TOTAL:       {total_entries:3d} flavor entries")

if __name__ == "__main__":
    main()
