#!/usr/bin/env python3
"""
Generate preslib_masterdoc.xlsx — full content tracker for the Presidential Library.
Sheets: GALLERY · RECORDS · JOURNAL
Run from repo root: python3 scripts/build_preslib_masterdoc.py

HOW TO UPDATE:
  GALLERY → edit GALLERY_DATA list below
  RECORDS → edit RECORDS_DATA list below
  JOURNAL → edit JOURNAL_DATA list below
  Status values: "FULL PASS" | "FIRST PASS" | "FIRST DRAFT" | "IDEA" | "STUB"
  Art status:    "Not Started" | "In Progress" | "Done"
  Content flags: "sensual" | "explicit" | "kinky"
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "preslib_masterdoc.xlsx"

# ── palette ───────────────────────────────────────────────────────────────────
HDR_BG, HDR_FG   = "1F2D3D", "FFFFFF"
CAT_BG, CAT_FG   = "2E4057", "FFFFFF"
ALT_BG = "EEF2F7"
WHITE  = "FFFFFF"

SENSUAL_BG,  SENSUAL_FG  = "FFE0F0", "660033"
EXPLICIT_BG, EXPLICIT_FG = "FFB3D1", "660033"
KINKY_BG,    KINKY_FG    = "BB44AA", "FFFFFF"

ART_NONE_BG, ART_NONE_FG = "FFD7D7", "8B0000"
ART_WIP_BG,  ART_WIP_FG  = "FFEB9C", "7A5A00"
ART_DONE_BG, ART_DONE_FG = "C6EFCE", "1A4A1A"

FULL_BG,   FULL_FG   = "00B050", "FFFFFF"
FPASS_BG,  FPASS_FG  = "92D050", "1A3A00"
FDRAFT_BG, FDRAFT_FG = "BFEA7C", "2A4A00"
IDEA_BG,   IDEA_FG   = "FFEB9C", "7A5A00"
STUB_BG,   STUB_FG   = "FFD7D7", "8B0000"

EYES_BG, EYES_FG = "CC0000", "FFFFFF"
TOP_BG,  TOP_FG  = "FF4444", "FFFFFF"
SEC_BG,  SEC_FG  = "E06000", "FFFFFF"
CONF_BG, CONF_FG = "2255AA", "FFFFFF"
DECL_BG, DECL_FG = "226622", "FFFFFF"

GROUP_COLORS = {
    "White House Secrets":    "F5E6F0",
    "Ualani Personal Events": "D6F5EA",
    "Commander Thanks":       "D6E1F7",
    "Protector Arcs":         "EDD6F7",
}

CAT_COLORS = {
    "Game Systems":           "D6E4F7",
    "Terrain":                "D6F5D6",
    "Resources & Ores":       "FFF3CC",
    "Buildings":              "F7DDD6",
    "Governor Archetypes":    "D6EAF8",
    "Military Modifiers":     "FDEBD0",
    "Laws & Edicts":          "EDE0F7",
    "Technologies":           "D6F0D6",
    "Factions":               "FCE0D6",
    "Army Units":             "DCDCDC",
    "Magic Schools & Spells": "EDD6F7",
    "Protectors":             "B8D6F7",
    "Mythic Weapons":         "F5ECD5",
    "Lore & History":         "F5E6D6",
    "Faiths & Doctrines":     "F0E8C8",   # new — add to RecordsDatabase.gd CATEGORIES
    "American Icons":         "FFE8D0",   # new — historical hero patron figures
}


# ── helpers ───────────────────────────────────────────────────────────────────
def _fill(h):   return PatternFill("solid", fgColor=h)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _font(bold=False, italic=False, color="000000", size=10):
    return Font(bold=bold, italic=italic, color=color, name="Calibri", size=size)
def _wrap():    return Alignment(wrap_text=True, vertical="top")
def _center():  return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _hdr(ws, cols):
    for i, (label, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=label)
        c.fill = _fill(HDR_BG)
        c.font = _font(bold=True, color=HDR_FG, size=11)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 24

def _status_style(s):
    return {"FULL PASS":(FULL_BG,FULL_FG), "FIRST PASS":(FPASS_BG,FPASS_FG),
            "FIRST DRAFT":(FDRAFT_BG,FDRAFT_FG), "IDEA":(IDEA_BG,IDEA_FG),
            "STUB":(STUB_BG,STUB_FG)}.get(s, (WHITE,"000000"))

def _art_style(s):
    return {"Not Started":(ART_NONE_BG,ART_NONE_FG),
            "In Progress":(ART_WIP_BG,ART_WIP_FG),
            "Done":(ART_DONE_BG,ART_DONE_FG)}.get(s, (WHITE,"000000"))

def _cflag_style(f):
    return {"sensual":(SENSUAL_BG,SENSUAL_FG), "explicit":(EXPLICIT_BG,EXPLICIT_FG),
            "kinky":(KINKY_BG,KINKY_FG)}.get(f, (WHITE,"000000"))

def _classif_style(c):
    return {"EYES ONLY":(EYES_BG,EYES_FG), "TOP SECRET":(TOP_BG,TOP_FG),
            "SECRET":(SEC_BG,SEC_FG), "CONFIDENTIAL":(CONF_BG,CONF_FG),
            "DECLASSIFIED":(DECL_BG,DECL_FG)}.get(c, (WHITE,"000000"))

def _cat_banner(ws, row, label, ncols, bg=CAT_BG, fg=CAT_FG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(bg)
    c.font = _font(bold=True, color=fg, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border()
    ws.row_dimensions[row].height = 18
    return row + 1


# ═══════════════════════════════════════════════════════════════════════════════
# GALLERY DATA
# (group, event_id, title, content_flag, hint, flavor, writing_status, art_status, art_path, notes)
# ═══════════════════════════════════════════════════════════════════════════════
GALLERY_DATA = [
    # ── White House Secrets ───────────────────────────────────────────────────
    ("White House Secrets", "WH_SECRET_01", "The New Year, Privately",
     "sensual",
     "Ualani must be in Washington DC when the new year turns",
     "No ceremony. No camera. No country to perform for.",
     "FIRST PASS", "Not Started", "", ""),

    ("White House Secrets", "WH_SECRET_07", "Independence Day — Hers",
     "explicit",
     "Ualani must be in Washington DC on the Fourth of July",
     "The founding documents were written by people who did not mean her. She means them anyway.",
     "FIRST PASS", "Not Started", "", ""),

    ("White House Secrets", "WH_SECRET_10", "A Letter from Jessica",
     "sensual",
     "Reach the Alliance stage with Canada (Ualani in Washington DC)",
     "Not a diplomatic dispatch. A personal one.",
     "FIRST PASS", "Not Started", "", ""),

    ("White House Secrets", "WH_SECRET_12", "Christmas in Washington",
     "sensual",
     "Ualani must be in Washington DC in winter",
     "She is from Hawaii. This is not Christmas as she knows it.",
     "FIRST PASS", "Not Started", "", ""),

    # ── Ualani Personal Events ────────────────────────────────────────────────
    ("Ualani Personal Events", "UALANI_AMBUSH_01", "The Counteroffensive Briefing",
     "sensual",
     "Hold a tile against a Crown ambush with Ualani present",
     "Ualani was already there. The Crown did not know that.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_DIGNITARY_01", "The Full Presidential Reception",
     "sensual",
     "Host a diplomatic reception at a liberated courthouse tile",
     "The courthouse has never looked this good.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_MEMORIAL_01", "The Presidential Address",
     "sensual",
     "Visit a memorial tile with Ualani",
     "The ground has history. The President has remarks.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_WOUNDED_01", "The Field Hospital Visit",
     "sensual",
     "Ualani visits a field hospital after a battle",
     "She knew their names. That is not an expression.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_WINTER_01", "The Winter March",
     "sensual",
     "March an army through a cold terrain tile in winter",
     "Washington did it once. The precedent is established.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_FORGE_01", "The Surprise Inspection",
     "sensual",
     "Own a Level 2+ Forge with a garrison present",
     "Production numbers were good. The President made them better.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_CULPER_01", "The Culper Meeting",
     "sensual",
     "Activate the Culper Ring intelligence network",
     "The message arrived with no signature. She knew who sent it.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_CULPER_BRIEF_01", "Intelligence Confirmed",
     "sensual",
     "Receive a confirmed intelligence brief from the Culper Ring",
     "The asset delivered. The map is different now.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_ALLIANCE_01", "The Alliance Council",
     "sensual",
     "Meet with a Commander who holds two or more objectives",
     "An alliance was discussed. The discussion went well.",
     "FIRST PASS", "Not Started", "", ""),

    ("Ualani Personal Events", "UALANI_FRONTIER_01", "The Frontier Walk",
     "sensual",
     "Establish presence at a border tile adjacent to Canada",
     "The northern edge of the Republic. She walked it.",
     "FIRST PASS", "Not Started", "", ""),

    # ── Commander Thanks ──────────────────────────────────────────────────────
    ("Commander Thanks", "CMD_THANKS", "A Personal Presidential Visit",
     "explicit",
     "A Commander completes their full arc at maximum loyalty",
     "The President doesn't visit everyone. She visited them.",
     "FIRST PASS", "Not Started", "", ""),

    # ── Protector Arcs ────────────────────────────────────────────────────────
    ("Protector Arcs", "PROT_01_SUMMON", "Mothman Approaches",
     "kinky",
     "Begin the Mothman protector arc (Harper's Ferry region)",
     "A presence in the dark.",
     "FIRST PASS", "Not Started", "", ""),

    ("Protector Arcs", "PROT_01_AGREE", "The Harper's Ferry Accord",
     "kinky",
     "Complete the Mothman protector arc and accept the alliance",
     "The agreement required no words. The words came anyway.",
     "FIRST PASS", "Not Started", "", ""),

    ("Protector Arcs", "PROT_14_SUMMON", "Rushmore Awakens",
     "kinky",
     "Begin the Mount Rushmore protector arc (Gettysburg region)",
     "All four have arrived and they are already arguing.",
     "FIRST PASS", "Not Started", "", ""),

    ("Protector Arcs", "PROT_14_AGREE", "The Presidential Council",
     "kinky",
     "Complete the Mount Rushmore protector arc and accept the endorsement",
     "They came to a vote. The vote was unanimous. Lincoln broke the tie.",
     "FIRST PASS", "Not Started", "", ""),
]


# ═══════════════════════════════════════════════════════════════════════════════
# RECORDS DATA
# (category, id, name, entry_type, description, unlock_flag, has_icon, status, notes, references)
# entry_type: "Regular" | "Mystery"
# ═══════════════════════════════════════════════════════════════════════════════
RECORDS_DATA = [
    # ── GAME SYSTEMS ──────────────────────────────────────────────────────────
    ("Game Systems","sys_turns","Turns & Time","Regular",
     "Each turn represents a season. Four turns make a year. The game begins in 1782, Year One of the Carlisle Administration. Pressing End Turn advances time, triggers income, building output, army upkeep, event checks, and all AI actions.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_census","Tile Census","Regular",
     "At the end of each turn, every tile you own runs a census. Each building on the tile reports its resource output — Food, Dollars, Wood, Metal, Magic, Culture, Weapons, Science, Mandate, Happiness, Manpower, Influence, and Boats. Governor bonuses are applied first.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_governors","Governors","Regular",
     "Governors are administrators assigned to individual tiles. Each governor has an archetype that determines which buildings they enhance. Governors have three levels. A governor can only manage one tile at a time.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_corruption","Corruption","Regular",
     "Corruption erodes a tile's productivity and morale. It rises from enemy activity, neglect, and certain events. Above 25 it begins to affect output. Above 60 it becomes serious. At 100 the tile may revolt. Buildings that reduce corruption: Baths, Temples, Libraries.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_colonization","Colonization","Regular",
     "Unclaimed tiles can be colonized by accumulating Colonization Points on them. The required points vary by terrain and distance. Naval tiles require a Dock presence.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_factions","Factions","Regular",
     "Factions are political, military, or social groups with their own agendas. Each faction has a Loyalty score toward the player. High loyalty unlocks faction events and bonuses. Low loyalty causes crises, defections, and armed opposition.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_spells","Magic & Spells","Regular",
     "Spells are unlocked through the Spellbook panel. Each spell belongs to a Magic School (Fire, Ice, Nature, Shadow, Light, Storm). Towers produce magic income. Wizards assigned to tiles amplify magical output and may cast defensive or offensive spells each turn.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_autosave","Autosave","Regular",
     "The game saves automatically at the end of each turn to a single autosave slot. The main menu 'Continue' button always loads the most recent autosave. There is no manual save system.",
     "",False,"FIRST PASS","",""),
    ("Game Systems","sys_content","Content Flags","Regular",
     "The game contains optional adult content. Sensual, Explicit, and Kinky/Lewd flags can be individually enabled or disabled in Settings. Gallery entries are only unlocked when the corresponding event fires with the flag active.",
     "",False,"FIRST PASS","",""),

    # ── TERRAIN ───────────────────────────────────────────────────────────────
    ("Terrain","ter_jungle","Jungle","Regular",
     "Dense tropical growth. High food output from farms. Movement penalties for armies. Reduces corruption spread. Home to rare botanical ores.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_steppe","Steppe","Regular",
     "Open grassland ideal for cavalry movement. Moderate food and manpower output. Vulnerable to storm damage.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_bog","Bog","Regular",
     "Waterlogged ground. Slows all movement. Unique ore deposits. Penalties to building construction speed.",
     "",True,"FIRST PASS","","Mil Mod: Swamp Legs, Everglades Tracker, Bayou Warrior"),
    ("Terrain","ter_cold_coast","Cold Coast","Regular",
     "Northern shoreline battered by Atlantic wind. Strong fishing and Boat production from Docks.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_drylands","Drylands","Regular",
     "Arid scrubland with scarce water. Low food output. High weapon and metal production. Corruption spreads faster here.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_warm_coast","Warm Coast","Regular",
     "Southern shoreline with fertile soil and warm waters. High food output. Docks produce double the normal Boat yield. Vulnerable to storm events.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_floodplains","Floodplains","Regular",
     "River delta terrain. Exceptionally fertile — highest food output in the game. Vulnerable to seasonal flooding events.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_desert","Desert","Regular",
     "Barren and hostile. Minimal food. Armies suffer attrition here. Faith bonuses from Temples are doubled.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_meadow","Meadow","Regular",
     "Gentle grassland. Balanced output across most resource types. Easy to colonize. Preferred terrain for initial expansion.",
     "",True,"FIRST PASS","",""),
    ("Terrain","ter_mountaintop","Mountaintop","Regular",
     "High peaks. Strong mine output for metal and weapons. Armies move slowly. Difficult to colonize. Defensive bonus for stationed armies.",
     "",True,"FIRST PASS","","Mil Mod: Hill Runner | Resources: Metal"),
    ("Terrain","ter_mountaintop_cold","Frozen Mountaintop","Regular",
     "Snow-covered peaks. All mountaintop properties amplified. Higher defensive bonus. Army attrition risk in winter turns.",
     "",True,"FIRST PASS","","Mil Mod: Blizzard March, White Out Walker | Resources: Metal"),
    ("Terrain","ter_hills","Hills","Regular",
     "Rolling terrain with moderate defensive value. Good for mines and camps. Balanced movement penalty.",
     "",True,"FIRST PASS","","Mil Mod: Hill Runner, Frontier Marksman"),
    ("Terrain","ter_forest","Forest","Regular",
     "Dense woodland. High Wood output. Army movement penalty. Reduces enemy cavalry effectiveness. Source of herbal and magical ores.",
     "",True,"FIRST PASS","","Mil Mod: Woodsman, Guerrilla Tactics, Backcountry Rider | Resources: Wood"),
    ("Terrain","ter_taiga","Taiga","Regular",
     "Northern boreal forest. High Wood output. Harsh winter penalties apply here first. Unique ores unavailable elsewhere.",
     "",True,"FIRST PASS","",""),

    # ── RESOURCES & ORES ──────────────────────────────────────────────────────
    ("Resources & Ores","res_food","Food","Regular",
     "Feeds the population and sustains armies in the field. Produced by Farms, Granaries, and coastal fishing. Shortfalls cause happiness penalties and army attrition.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_dollars","Dollars","Regular",
     "The lifeblood of the republic. Pays for buildings, armies, and political favors. Produced by Markets, Faires, and trade routes. A negative balance triggers economic crisis events.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_wood","Wood","Regular",
     "Essential for construction and naval production. Produced by Camps, Docks, and forest terrain bonuses.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_metal","Metal","Regular",
     "Required for Weapons, advanced buildings, and certain tech upgrades. Produced primarily by Mines.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_weapons","Weapons","Regular",
     "Equips your armies and funds your war effort. Produced by Forges and Arsenals. Army upkeep consumes Weapons each turn.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_science","Science","Regular",
     "Powers the research of new Technologies. Produced by Libraries and Schools. Accumulated science is spent on techs in the Tech Tree panel.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_magic","Magic","Regular",
     "Fuel for spells and arcane infrastructure. Produced by Towers, Wizards, and certain protectors. Required to cast spells in battle and maintain magical buildings.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_faith","Faith","Regular",
     "Spiritual capital of the nation. Produced by Temples and Monasteries. Unlocks religious laws and certain diplomatic options with faith-based factions.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_culture","Culture","Regular",
     "Represents the artistic and intellectual vitality of the republic. Produced by Theaters, Baths, and cultural buildings. Required for Tradition unlocks.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_mandate","Mandate","Regular",
     "The political authority of the presidency. High Mandate unlocks stronger edicts and expands law options. Produced by government buildings and compliance events.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_happiness","Happiness","Regular",
     "National morale. Affects approval rating, faction loyalty, and productivity. Falls under corruption, war exhaustion, and resource shortfalls.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_manpower","Manpower","Regular",
     "The pool from which armies are drawn. Produced by Barracks and population-dense tiles. Consumed when armies are recruited or suffer heavy losses.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_influence","Influence","Regular",
     "Diplomatic capital. Used to sway factions, broker alliances, and suppress opposition. Produced by Forts, Embassies, and named governor bonuses.",
     "",True,"FIRST PASS","",""),
    ("Resources & Ores","res_boats","Boats","Regular",
     "Naval capacity. Produced by Docks. Required to move armies across water tiles and to maintain a navy.",
     "",False,"FIRST PASS","Icon pending — see MAN-001 editor task.",""),

    # ── BUILDINGS ─────────────────────────────────────────────────────────────
    ("Buildings","bld_farm","Farm","Regular",
     "The backbone of agricultural production. Generates Food and small amounts of Wood. FARMER governors dramatically increase output.",
     "",False,"FIRST PASS","","Governor: FARMER | Mil Mod: Farmhand, Seeder, Harvester | American Icons: Cesar Chavez, Dolores Huerta"),
    ("Buildings","bld_mine","Mine","Regular",
     "Extracts Metal from the earth. Output scales with terrain — Mountains and Hills provide the highest yields.",
     "",False,"FIRST PASS","",""),
    ("Buildings","bld_barracks","Barracks","Regular",
     "Trains and houses soldiers. Produces Manpower each turn. Stationed armies receive a combat bonus. Required for WARRIOR and SOLDIER governor assignments.",
     "",False,"FIRST PASS","","Governor: WARRIOR, SOLDIER | Mil Mod: Fortified Position, Steady Line, Iron Wall, Rampart"),
    ("Buildings","bld_market","Market","Regular",
     "Generates Dollars through trade. DIPLOMAT and ORATOR governors amplify output. Higher-level markets reduce corruption spread.",
     "",False,"FIRST PASS","",""),
    ("Buildings","bld_library","Library","Regular",
     "Produces Science and small amounts of Culture. Required for SCHOLAR governor assignment. Level 3 Libraries also produce Influence.",
     "",False,"FIRST PASS","","Governor: SCHOLAR | Mil Mod: Scholar, Translator | Laws: Navigation Acts"),
    ("Buildings","bld_tower","Tower","Regular",
     "Produces Magic each turn and provides a platform for Wizards. Protectors who join the republic are bound to a Tower at their home tile.",
     "",False,"FIRST PASS","","Governor: MAGE | Protectors: All (bound here) | Mil Mod: Druid"),
    ("Buildings","bld_granary","Granary","Regular",
     "Stores and distributes food surplus. Reduces the penalty of food shortfalls. With the Mandate from Granaries law active, also produces Mandate.",
     "",False,"FIRST PASS","",""),
    ("Buildings","bld_forge","Forge","Regular",
     "Produces Weapons from Metal. Required for WARRIOR governor assignment. Level 2+ Forges also produce a small amount of Science.",
     "",False,"FIRST PASS","","Governor: WARRIOR | Resources: Metal, Weapons | Mil Mod: Blacksmith"),
    ("Buildings","bld_temple","Temple","Regular",
     "Produces Faith and Happiness. Reduces corruption. HEALER governors amplify output. Required for religious law access.",
     "",False,"FIRST PASS","","Governor: HEALER, CIRCUIT PREACHER | Mil Mod: Healer | American Icons: Phillis Wheatley"),
    ("Buildings","bld_bath","Bath","Regular",
     "Generates Happiness and reduces corruption. Level 2+ Baths also produce Culture.",
     "",False,"FIRST PASS","",""),
    ("Buildings","bld_theater","Theater","Regular",
     "Cultural hub producing Culture and Happiness. ORATOR governors thrive here. Level 3 Theaters produce Influence.",
     "",False,"FIRST PASS","",""),
    ("Buildings","bld_dock","Dock","Regular",
     "Naval infrastructure on coastal tiles. Produces Boats and Wood. Required for colonizing across water. ADMIRAL governors dramatically amplify output.",
     "",False,"FIRST PASS","","Governor: ADMIRAL | Mil Mod: Marine, Coastal Watch, Chesapeake Sailor, Harbor Watch"),
    ("Buildings","bld_camp","Camp","Regular",
     "Frontier outpost. Produces Wood, Manpower, and Weapons. SCOUT governors thrive in Camps.",
     "",False,"FIRST PASS","","Governor: SCOUT | Mil Mod: Woodsman, Cartographer, Trapper"),
    ("Buildings","bld_monument","Monument","Regular",
     "A landmark of national identity. Produces Culture, Mandate, and Happiness. Required for high-level HERALD governor assignment.",
     "",False,"FIRST PASS","",""),

    # ── GOVERNOR ARCHETYPES ───────────────────────────────────────────────────
    ("Governor Archetypes","arc_farmer","FARMER","Regular",
     "Masters of the land. Dramatically boost Farm and Granary output. Level 3 unlocks crop rotation events and reduces food spoilage.",
     "",False,"FIRST PASS","","Buildings: Farm, Granary | Mil Mod: Farmhand, Seeder, Harvester"),
    ("Governor Archetypes","arc_scout","SCOUT","Regular",
     "Explorers and frontier agents. Boost Camp and Mine output. Level 2 reduces colonization costs on adjacent tiles. Level 3 expands tile visibility.",
     "",False,"FIRST PASS","","Buildings: Camp, Mine | Mil Mod: Woodsman, Swamp Legs, Hill Runner, Cartographer, Surveyor, Trapper"),
    ("Governor Archetypes","arc_warrior","WARRIOR","Regular",
     "Combat-focused administrators. Boost Barracks and Forge output and grant a tile defense bonus. Level 3 adds a standing combat modifier to stationed armies.",
     "",False,"FIRST PASS","","Buildings: Barracks, Forge | Mil Mod: Saber Drill, Iron Bayonet, Flanking Drill, Continental Line"),
    ("Governor Archetypes","arc_scholar","SCHOLAR","Regular",
     "Intellectuals and researchers. Boost Library output and generate bonus Science each turn. Level 3 occasionally unlocks free Technology events.",
     "",False,"FIRST PASS","","Buildings: Library | Mil Mod: Scholar, Translator"),
    ("Governor Archetypes","arc_engineer","ENGINEER","Regular",
     "Builders and infrastructure experts. Reduce building construction cost and boost Mine and Dock output. Level 3 allows one free building upgrade per year.",
     "",False,"FIRST PASS","","Buildings: Mine, Dock | Mil Mod: Engineer, Constructor, Architect"),
    ("Governor Archetypes","arc_diplomat","DIPLOMAT","Regular",
     "Political operators. Boost Market and Monument output. Generate Influence passively. Level 3 adds a loyalty buffer to the tile's dominant faction.",
     "",False,"FIRST PASS","","Buildings: Market, Monument | Mil Mod: Entertainer"),
    ("Governor Archetypes","arc_orator","ORATOR","Regular",
     "Speakers and agitators. Boost Theater and Market output. Generate Culture and Happiness. Level 3 triggers popular approval events.",
     "",False,"FIRST PASS","","Buildings: Theater, Market | Mil Mod: Entertainer, Preacher"),
    ("Governor Archetypes","arc_healer","HEALER","Regular",
     "Physicians and spiritual leaders. Boost Temple and Bath output. Reduce corruption. Level 3 provides army recovery bonuses when stationed on their tile.",
     "",False,"FIRST PASS","","Buildings: Temple, Bath | Mil Mod: Healer, Herbalist, Physician"),
    ("Governor Archetypes","arc_spy","SPYMASTER","Regular",
     "Intelligence operatives. Generate Influence and reduce enemy spy effectiveness on the tile. Level 3 can intercept enemy events.",
     "",False,"FIRST PASS","","Mil Mod: Visionary, Night Raider, Ghost March"),
    ("Governor Archetypes","arc_admiral","ADMIRAL","Regular",
     "Naval commanders in civilian governance. Dramatically boost Dock output. Level 3 doubles Boat production and reduces naval upkeep.",
     "",False,"FIRST PASS","","Buildings: Dock | Mil Mod: Marine, Coastal Watch, Naval Supremacy, Chesapeake Sailor, Harbor Watch"),
    ("Governor Archetypes","arc_mage","MAGE","Regular",
     "Arcane administrators. Boost Tower output and Magic generation. Level 3 learns a spell that fires once per year on their tile.",
     "",False,"FIRST PASS","","Buildings: Tower | Mil Mod: Druid"),
    ("Governor Archetypes","arc_soldier","SOLDIER","Regular",
     "Veteran fighters in peacetime roles. Boost Barracks and Camp output. Reduce army upkeep on the tile. Level 3 grants a morale bonus to stationed armies.",
     "",False,"FIRST PASS","","Buildings: Barracks, Camp | Mil Mod: Steady Line, Fortified Position, Iron Wall"),
    ("Governor Archetypes","arc_bureaucrat","BUREAUCRAT","Regular",
     "Administrative specialists. Boost Mandate and Influence generation. Reduce corruption. Level 3 reduces the cost of laws and edicts.",
     "",False,"FIRST PASS","","Buildings: Monument | Mil Mod: President, Election Season"),
    ("Governor Archetypes","arc_herald","HERALD","Regular",
     "Messengers and public figures. Boost Monument and Theater output. Generate Happiness and Culture across multiple tiles. Level 3 triggers national morale events.",
     "",False,"FIRST PASS","","Buildings: Monument, Theater"),
    ("Governor Archetypes","arc_preacher","CIRCUIT PREACHER","Regular",
     "Traveling ministers of faith. Boost Temple output significantly. Generate Faith across the region. Level 3 can suppress faction unrest through religious revival events.",
     "",False,"FIRST PASS","","Buildings: Temple | Mil Mod: Preacher | American Icons: All"),

    # ── MILITARY MODIFIERS ────────────────────────────────────────────────────
    # Country Modifiers
    ("Military Modifiers","milmod_berserkers","Berserkers","Regular",
     "Warriors are expected to kill or die trying. +3 Attack per level, -2 Harmony per level.",
     "",False,"FIRST PASS","Country milmod — national trait.",""),
    # Weapon Ore Mods
    ("Military Modifiers","milmod_club_bleed","ClubBleed","Regular",
     "Club weapon mod. +2 Attack, +1 Defense per level. -1 Weapons per level.",
     "",False,"FIRST PASS","","Weapon: Club | Resources: Weapons"),
    ("Military Modifiers","milmod_atlatl_pierce","AtlatlPierce","Regular",
     "Atlatl weapon mod. +1 Attack, +2 Defense per level. -1 Weapons per level.",
     "",False,"FIRST PASS","","Weapon: Atlatl | Resources: Weapons"),
    ("Military Modifiers","milmod_ore_wood","Wood Weapons","Regular",
     "Unit's weapons carved from wood. +1 Attack per level. -1 Wood per level.",
     "",False,"FIRST PASS","","Ore: Wood | Resources: Wood"),
    ("Military Modifiers","milmod_ore_copper","Copper Weapons","Regular",
     "Unit's weapons shaped by copper. +2 Attack per level. -1 Metal per level.",
     "",False,"FIRST PASS","","Ore: Copper | Resources: Metal"),
    ("Military Modifiers","milmod_ore_iron","Iron Weapons","Regular",
     "Unit's weapons forged from iron. +3 Attack per level. -3 Metal per level.",
     "",False,"FIRST PASS","","Ore: Iron | Resources: Metal"),
    ("Military Modifiers","milmod_ore_gold","Gold Weapons","Regular",
     "Unit's weapons built of gold. +2 Attack per level. -1 Metal, -3 Dollars per level.",
     "",False,"FIRST PASS","","Ore: Gold | Resources: Metal"),
    ("Military Modifiers","milmod_ore_floodstone","Floodstone Weapons","Regular",
     "Unit's weapons birthed from floodstone. +3 Attack per level. -2 Metal, -2 Magic per level.",
     "",False,"FIRST PASS","","Resources: Metal, Magic"),
    # Commander Mods
    ("Military Modifiers","milmod_visionary","Visionary","Regular",
     "Commander mod. This unit's commander is a genius. +3 Attack per level, +10% movement speed.",
     "",False,"FIRST PASS","","Governor: SPYMASTER"),
    ("Military Modifiers","milmod_champion_sun","Champion of the Sun","Regular",
     "Commander mod. While the sun is up, this unit will march. +1 Attack per level, +10% movement speed.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_healer","Healer","Regular",
     "Commander mod. This unit heals exceptionally fast. +10 reinforce rate.",
     "",False,"FIRST PASS","","Governor: HEALER | Buildings: Temple, Bath"),
    # Civilian Mods
    ("Military Modifiers","milmod_translator","Translator","Regular",
     "Civilian mod. Unit equipped with reference materials enabling translation of ancient writings. Generates Science.",
     "",False,"FIRST PASS","","Governor: SCHOLAR | Buildings: Library"),
    ("Military Modifiers","milmod_seeder","Seeder","Regular",
     "Civilian mod. Unit carries seeds enabling agricultural improvements on the tile. Generates Food.",
     "",False,"FIRST PASS","","Governor: FARMER | Buildings: Farm"),
    ("Military Modifiers","milmod_wooden_tools","Wooden Tools","Regular",
     "Civilian mod. Unit carries basic wooden tools.",
     "",False,"FIRST PASS","","Resources: Wood"),
    ("Military Modifiers","milmod_metal_tools","Metal Tools","Regular",
     "Civilian mod. Unit carries metal tools.",
     "",False,"FIRST PASS","","Resources: Metal"),
    ("Military Modifiers","milmod_steel_tools","Steel Tools","Regular",
     "Civilian mod. Unit carries advanced steel tools.",
     "",False,"FIRST PASS","","Resources: Metal | Technologies: Forging"),
    ("Military Modifiers","milmod_constructor","Constructor","Regular",
     "Civilian mod. Unit can build and upgrade structures faster.",
     "",False,"FIRST PASS","","Governor: ENGINEER | Resources: Wood"),
    ("Military Modifiers","milmod_adventurer","Adventurer","Regular",
     "Civilian mod. Unit trained in exploring ruins, caves, and frontier territories.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_scholar","Scholar","Regular",
     "Civilian mod. Unit spreads literacy and builds library infrastructure.",
     "",False,"FIRST PASS","","Governor: SCHOLAR | Buildings: Library"),
    ("Military Modifiers","milmod_entertainer","Entertainer","Regular",
     "Civilian mod. Unit can entertain, soothe, and manage morale in its care.",
     "",False,"FIRST PASS","","Governor: DIPLOMAT, ORATOR | Buildings: Theater"),
    ("Military Modifiers","milmod_harvester","Harvester","Regular",
     "Civilian mod. Unit manages crops, woodcutting, and rural land.",
     "",False,"FIRST PASS","","Governor: FARMER | Buildings: Farm, Camp"),
    ("Military Modifiers","milmod_prospector","Prospector","Regular",
     "Civilian mod. Unit trained in prospecting; can build mines and discover mineral deposits.",
     "",False,"FIRST PASS","","Governor: SCOUT | Buildings: Mine"),
    ("Military Modifiers","milmod_druid","Druid","Regular",
     "Civilian mod. Trained in druidic arts; can clear corruption and commune with nature.",
     "",False,"FIRST PASS","","Governor: MAGE | Buildings: Tower"),
    # Legacy Armor Mods
    ("Military Modifiers","milmod_chain","Chain Armor","Regular",
     "Unit wears chain armor. +5% Melee, +40% Ranged, +5% Spell damage block.",
     "",False,"FIRST PASS","","Ore: Iron"),
    ("Military Modifiers","milmod_shell","Shell Armor","Regular",
     "Unit wears a fully-enclosed shell. +50% Spell damage block.",
     "",False,"FIRST PASS","",""),
    # Tier 1 Military Mods
    ("Military Modifiers","milmod_woodsman","Woodsman","Regular",
     "Trained in forest fighting. +2 Attack, +2 Defense per level in Woods terrain.",
     "",False,"FIRST PASS","","Governor: SCOUT | Terrain: Forest, Taiga | Mil Mod: Guerrilla Tactics (upgrade)"),
    ("Military Modifiers","milmod_swamp_legs","Swamp Legs","Regular",
     "At home in the marshes. +2 Attack, +2 Defense per level in Wetlands terrain.",
     "",False,"FIRST PASS","","Terrain: Bog | Mil Mod: Everglades Tracker, Bayou Warrior (see also)"),
    ("Military Modifiers","milmod_hill_runner","Hill Runner","Regular",
     "Born on high ground. +2 Attack, +2 Defense per level in Foothills terrain.",
     "",False,"FIRST PASS","","Terrain: Hills, Mountaintop | Mil Mod: Frontier Marksman (see also)"),
    ("Military Modifiers","milmod_street_tough","Street Tough","Regular",
     "Raised fighting in alleyways. +2 Attack per level in Metro or Suburbs terrain.",
     "",False,"FIRST PASS","","Terrain: Metro"),
    ("Military Modifiers","milmod_farmhand","Farmhand","Regular",
     "Knows every row of every field. +1 Attack per level in Farmlands terrain.",
     "",False,"FIRST PASS","","Governor: FARMER | Terrain: Farmlands | American Icons: Cesar Chavez"),
    ("Military Modifiers","milmod_saber_drill","Saber Drill","Regular",
     "Relentless close-combat drilling. +3 Attack per level.",
     "",False,"FIRST PASS","","Governor: WARRIOR"),
    ("Military Modifiers","milmod_marksman","Marksman","Regular",
     "Trained to shoot straight and true. +2 Ranged Attack per level.",
     "",False,"FIRST PASS","","Governor: SCOUT | Mil Mod: Sharpshooter (upgrade)"),
    ("Military Modifiers","milmod_steady_line","Steady Line","Regular",
     "Hold the line at all costs. +3 Defense per level.",
     "",False,"FIRST PASS","","Governor: SOLDIER | Mil Mod: Continental Line (upgrade)"),
    ("Military Modifiers","milmod_quick_reload","Quick Reload","Regular",
     "Powder and ball faster than any rival. Reload reduced by 1 round.",
     "",False,"FIRST PASS","","Governor: SCOUT | Weapon: Musket (any)"),
    ("Military Modifiers","milmod_powder_shot","Powder & Shot","Regular",
     "The cannons never run dry. +3 Siege Attack per level.",
     "",False,"FIRST PASS","","Governor: ARTILLERIST | Weapon: Artillery | Mil Mod: Double Shot (upgrade)"),
    ("Military Modifiers","milmod_fortified_position","Fortified Position","Regular",
     "Commander mod. All units +3 Defense per level in tiles with Barracks or Fortress.",
     "",False,"FIRST PASS","","Governor: SOLDIER, COMMANDER | Buildings: Barracks"),
    ("Military Modifiers","milmod_coastal_watch","Coastal Watch","Regular",
     "Commander mod. All units +2 Defense per level when adjacent to naval tiles.",
     "",False,"FIRST PASS","","Governor: ADMIRAL | Buildings: Dock | Terrain: Cold Coast, Warm Coast"),
    # Tier 2 Military Mods
    ("Military Modifiers","milmod_marine","Marine","Regular",
     "Commander mod. Army may launch melee attacks into adjacent naval tile neighbors.",
     "",False,"FIRST PASS","","Governor: ADMIRAL | Buildings: Dock | Mil Mod: Naval Supremacy (upgrade)"),
    ("Military Modifiers","milmod_guerrilla","Guerrilla Tactics","Regular",
     "+4 Attack, +4 Defense per level in Woods or Wetlands terrain.",
     "",False,"FIRST PASS","","Governor: SCOUT | Terrain: Forest, Bog | Mil Mod: Woodsman (prerequisite)"),
    ("Military Modifiers","milmod_double_shot","Double Shot","Regular",
     "Siege fires twice per round — second shot at 50% power.",
     "",False,"FIRST PASS","","Weapon: Artillery | Mil Mod: Powder & Shot (prerequisite) | Upgrade: Double Cannonade"),
    ("Military Modifiers","milmod_iron_bayonet","Iron Bayonet","Regular",
     "+5 Attack per level in first battle round.",
     "",False,"FIRST PASS","","Governor: WARRIOR | Weapon: Musket (any)"),
    ("Military Modifiers","milmod_sharpshooter","Sharpshooter","Regular",
     "Ranged attacks ignore 2 enemy Defense per level.",
     "",False,"FIRST PASS","","Governor: SCOUT | Mil Mod: Marksman (prerequisite)"),
    ("Military Modifiers","milmod_corrupted_ground","Corrupted Ground","Regular",
     "Commander mod. Army presence reduces tile corruption by 1 per turn.",
     "",False,"FIRST PASS","","Governor: DIPLOMAT | Buildings: Bath, Temple"),
    ("Military Modifiers","milmod_rallying_voice","Rallying Voice","Regular",
     "Commander mod. Morale loss reduced; rout threshold lowered to 15%.",
     "",False,"FIRST PASS","","Governor: DIPLOMAT, STRATEGIST"),
    ("Military Modifiers","milmod_night_raider","Night Raider","Regular",
     "Commander mod. Army may move and attack in the same turn without penalty.",
     "",False,"FIRST PASS","","Governor: SPYMASTER"),
    ("Military Modifiers","milmod_flanking_drill","Flanking Drill","Regular",
     "+3 Attack per level when fighting in a contested tile.",
     "",False,"FIRST PASS","","Governor: WARRIOR"),
    ("Military Modifiers","milmod_vanguard","Vanguard","Regular",
     "Commander mod. All units +4 Attack per level on first engagement in a fresh tile.",
     "",False,"FIRST PASS","","Governor: GENERAL"),
    ("Military Modifiers","milmod_siege_line","Siege Line","Regular",
     "Siege attacks against fortified tiles suffer no defensive penalty.",
     "",False,"FIRST PASS","","Weapon: Artillery | Buildings: Barracks"),
    ("Military Modifiers","milmod_cleaner","Cleaner","Regular",
     "Commander mod. Army presence reduces tile moral decay by 1 per turn.",
     "",False,"FIRST PASS","","Governor: DIPLOMAT | Buildings: Bath"),
    # Tier 3 Military Mods
    ("Military Modifiers","milmod_entrenched","Entrenched","Regular",
     "Commander mod. After 3 stationary turns, all units gain +5 Defense per level. Lost on movement.",
     "",False,"FIRST PASS","","Governor: COMMANDER | Buildings: Barracks"),
    ("Military Modifiers","milmod_continental_line","Continental Line","Regular",
     "Commander mod. All units +2 Attack, +2 Defense per level permanently.",
     "",False,"FIRST PASS","","Governor: GENERAL, STRATEGIST | American Icons: George Washington | Lore: First Reconquest War"),
    ("Military Modifiers","milmod_last_stand","Last Stand","Regular",
     "Units below 25% manpower gain +6 Attack per level.",
     "",False,"FIRST PASS","","Governor: MARSHAL"),
    ("Military Modifiers","milmod_terror","Terror","Regular",
     "Commander mod. Enemy loses 10 Morale at start of each battle round.",
     "",False,"FIRST PASS","","Governor: GENERAL, MARSHAL"),
    ("Military Modifiers","milmod_iron_wall","Iron Wall","Regular",
     "Commander mod. +8 Defense per level when defending the commander's home tile.",
     "",False,"FIRST PASS","","Governor: COMMANDER, SOLDIER"),
    ("Military Modifiers","milmod_rampart","Rampart","Regular",
     "Commander mod. +5 Defense per level in any Fortress tile.",
     "",False,"FIRST PASS","","Governor: COMMANDER | Buildings: Barracks"),
    ("Military Modifiers","milmod_naval_supremacy","Naval Supremacy","Regular",
     "Commander mod. Marine melee attacks deal +5 additional damage per level.",
     "",False,"FIRST PASS","","Governor: ADMIRAL | Mil Mod: Marine (prerequisite)"),
    ("Military Modifiers","milmod_ghost_march","Ghost March","Regular",
     "Commander mod. Army ignores enemy zone of control.",
     "",False,"FIRST PASS","","Governor: SPYMASTER, ADMIRAL | American Icons: Harriet Tubman"),
    ("Military Modifiers","milmod_undaunted","Undaunted","Regular",
     "Commander mod. Ignore the first retreat check each battle.",
     "",False,"FIRST PASS","","Governor: GENERAL, MARSHAL"),
    ("Military Modifiers","milmod_double_cannonade","Double Cannonade","Regular",
     "Siege fires twice per round AND +3 Attack on all shots.",
     "",False,"FIRST PASS","","Weapon: Artillery | Mil Mod: Double Shot (prerequisite)"),
    ("Military Modifiers","milmod_liberators_will","Liberator's Will","Regular",
     "Commander mod. After liberating a tile, +15% manpower and commander gains +2 Loyalty.",
     "",False,"FIRST PASS","","Governor: DIPLOMAT | American Icons: Abraham Lincoln, Harriet Tubman"),
    ("Military Modifiers","milmod_long_march","The Long March","Regular",
     "Commander mod. +2 Movement Points; full movement may be used before attacking.",
     "",False,"FIRST PASS","","Governor: MARSHAL | American Icons: George Washington"),
    # Mythic Weapon Mods
    ("Military Modifiers","milmod_bat_sweep","BatSweep","Regular",
     "Baseball Bat mythic mod. +10% attack; first hit ignores shields.",
     "",False,"FIRST PASS","","Mythic Weapon: Baseball Bat"),
    ("Military Modifiers","milmod_trident_pierce","TridentPierce","Regular",
     "Trident mythic mod. Pierces shields; +3 attack per level near naval tiles.",
     "",False,"FIRST PASS","","Mythic Weapon: Trident | Buildings: Dock"),
    ("Military Modifiers","milmod_mythic_atlatl","MythicAtlatl","Regular",
     "Mythic Atlatl mod. +2 ranged per level; critical hits stun target 1 turn.",
     "",False,"FIRST PASS","","Mythic Weapon: Mythic Atlatl"),
    ("Military Modifiers","milmod_sharp_shot","SharpShot","Regular",
     "Sharps Carbine mythic mod. Ignores 4 enemy defense per level; terrain cover ignored.",
     "",False,"FIRST PASS","","Mythic Weapon: Sharps Carbine"),
    ("Military Modifiers","milmod_pirate_volley","PirateVolley","Regular",
     "Blackbeard's Pistols mythic mod. Fires twice per ranged round; +5 enemy morale terror.",
     "",False,"FIRST PASS","","Mythic Weapon: Blackbeard's Pistols"),
    ("Military Modifiers","milmod_cylinder_fire","CylinderFire","Regular",
     "Colt Revolver mythic mod. No reload for first 3 shots; 2-turn reload thereafter.",
     "",False,"FIRST PASS","","Mythic Weapon: Colt Revolver"),
    ("Military Modifiers","milmod_rocket_barrage","RocketBarrage","Regular",
     "Rocket Artillery mythic mod. Massive area damage; leaves fire modifier on tile for 2 turns.",
     "",False,"FIRST PASS","","Mythic Weapon: Rocket Artillery"),
    ("Military Modifiers","milmod_trebuchet_launch","TrebuchetLaunch","Regular",
     "Trebuchet mythic mod. +50% siege damage vs Fortress tiles; stuns defenders 1 round.",
     "",False,"FIRST PASS","","Mythic Weapon: Trebuchet"),
    ("Military Modifiers","milmod_aerial_bombing","AerialBombing","Regular",
     "Wright Flyer mythic mod. Ignores all ground defensive bonuses; terrifies enemy.",
     "",False,"FIRST PASS","","Mythic Weapon: Wright Flyer"),
    # Uniform Mods
    ("Military Modifiers","milmod_quick_draw","QuickDraw","Regular",
     "Tombstone Cap uniform mod. First ranged attack each battle deals +5 bonus damage.",
     "",False,"FIRST PASS","","Uniform: Tombstone Cap"),
    ("Military Modifiers","milmod_hardee_disc","HardeeDisc","Regular",
     "Hardee Hat uniform mod. +3 defense per level when an adjacent friendly unit is present.",
     "",False,"FIRST PASS","","Uniform: Hardee Hat"),
    # Storm Counter Mods
    ("Military Modifiers","milmod_fog_born","Fog-Born","Regular",
     "Commander mod. +5 attack per level in Fog storm tiles.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_storm_rider","Storm Rider","Regular",
     "Commander mod. Movement not reduced by any active storm.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_thunder_proof","Thunder Proof","Regular",
     "Commander mod. Immune to Thunderstorm morale penalty.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_blizzard_march","Blizzard March","Regular",
     "Commander mod. No movement or supply penalty in Blizzard tiles. Valley Forge was just training.",
     "",False,"FIRST PASS","","Terrain: Frozen Mountaintop | American Icons: George Washington | Magic Schools: Storm"),
    ("Military Modifiers","milmod_hurricane_eyes","Hurricane Eyes","Regular",
     "Commander mod. +5 attack per level in Hurricane storm tiles.",
     "",False,"FIRST PASS","","Magic Schools: Storm | Terrain: Warm Coast"),
    ("Military Modifiers","milmod_tornado_dancer","Tornado Dancer","Regular",
     "Commander mod. Army ignores Tornado scatter and manpower drain effects.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_norester_veteran","Nor'easter Veteran","Regular",
     "Commander mod. +3 defense per level in Nor'easter tiles.",
     "",False,"FIRST PASS","","Terrain: Cold Coast | Magic Schools: Storm"),
    ("Military Modifiers","milmod_rain_reader","Rain Reader","Regular",
     "Ranged mod. +3 ranged attack per level during Thunderstorm.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_white_out_walker","White Out Walker","Regular",
     "Infantry mod. +3 attack per level during Blizzard.",
     "",False,"FIRST PASS","","Terrain: Frozen Mountaintop | Magic Schools: Storm"),
    ("Military Modifiers","milmod_storm_chaser","Storm Chaser","Regular",
     "Commander mod. +1 movement point in any active storm tile.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_lightning_rod","Lightning Rod","Regular",
     "Siege mod. Artillery units ignore storm ranged accuracy penalty.",
     "",False,"FIRST PASS","","Weapon: Artillery | Magic Schools: Storm | Technologies: Lenscraft"),
    ("Military Modifiers","milmod_eye_of_storm","Eye of the Storm","Regular",
     "Commander mod. +4 attack, +4 defense per level while any storm is active in tile.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    # Cultural / State Mods
    ("Military Modifiers","milmod_country_musician","Country Musician","Regular",
     "Tennessee cultural mod. +3 morale; +2 attack per level in Farmlands.",
     "",False,"FIRST PASS","","State: TN | Terrain: Farmlands"),
    ("Military Modifiers","milmod_virginia_gentry","Virginia Gentry","Regular",
     "Virginia cultural mod. +3 ranged defense per level.",
     "",False,"FIRST PASS","","State: VA"),
    ("Military Modifiers","milmod_minutemans_pride","Minuteman's Pride","Regular",
     "Massachusetts cultural mod. +5 attack per level in the first 3 battle rounds.",
     "",False,"FIRST PASS","","State: MA | Lore: First Reconquest War"),
    ("Military Modifiers","milmod_quaker_steel","Quaker Steel","Regular",
     "Pennsylvania cultural mod. +2 defense per level, -1 attack per level.",
     "",False,"FIRST PASS","","State: PA | Faiths & Doctrines: Quaker traditions"),
    ("Military Modifiers","milmod_georgia_peach","Georgia Peach","Regular",
     "Georgia cultural mod. +3 food efficiency; +1 ranged per level.",
     "",False,"FIRST PASS","","State: GA | Buildings: Farm"),
    ("Military Modifiers","milmod_backcountry_rider","Backcountry Rider","Regular",
     "South Carolina cultural mod. +4 attack per level in Woods terrain.",
     "",False,"FIRST PASS","","State: SC | Terrain: Forest"),
    ("Military Modifiers","milmod_harbor_watch","Harbor Watch","Regular",
     "New York cultural/marine mod. +3 attack per level near naval tiles.",
     "",False,"FIRST PASS","","State: NY | Buildings: Dock"),
    ("Military Modifiers","milmod_chesapeake_sailor","Chesapeake Sailor","Regular",
     "Maryland cultural/marine mod. +2 melee per level; Marine melee attack enabled.",
     "",False,"FIRST PASS","","State: MD | Buildings: Dock"),
    ("Military Modifiers","milmod_frontier_marksman","Frontier Marksman","Regular",
     "Kentucky cultural mod. +4 ranged per level in Foothills terrain.",
     "",False,"FIRST PASS","","State: KY | Terrain: Hills"),
    ("Military Modifiers","milmod_river_runner","River Runner","Regular",
     "Ohio cultural mod. +2 movement points; +2 attack per level near water tiles.",
     "",False,"FIRST PASS","","State: OH | Terrain: Floodplains"),
    ("Military Modifiers","milmod_everglades_tracker","Everglades Tracker","Regular",
     "Florida cultural mod. +4 attack and defense per level in Wetlands.",
     "",False,"FIRST PASS","","State: FL | Terrain: Bog"),
    ("Military Modifiers","milmod_bayou_warrior","Bayou Warrior","Regular",
     "Louisiana cultural mod. +5 attack per level in Wetlands terrain.",
     "",False,"FIRST PASS","","State: LA | Terrain: Bog"),
    # Tool Mods (Civilian)
    ("Military Modifiers","milmod_cartographer","Cartographer","Regular",
     "Civilian tool mod. Maps explored tiles, revealing terrain bonuses and hidden resources.",
     "",False,"FIRST PASS","","Governor: SCOUT"),
    ("Military Modifiers","milmod_herbalist","Herbalist","Regular",
     "Civilian tool mod. Gathers medicinal herbs, healing +5 manpower per turn in the tile.",
     "",False,"FIRST PASS","","Governor: HEALER | Buildings: Temple"),
    ("Military Modifiers","milmod_engineer_tool","Engineer","Regular",
     "Civilian tool mod. Builds roads and improves structures faster than standard workers.",
     "",False,"FIRST PASS","","Governor: ENGINEER | Buildings: Camp"),
    ("Military Modifiers","milmod_blacksmith","Blacksmith","Regular",
     "Civilian tool mod. Reduces weapon upkeep costs by 1 per level per turn for stationed armies.",
     "",False,"FIRST PASS","","Governor: ENGINEER | Buildings: Forge | Resources: Weapons"),
    ("Military Modifiers","milmod_physician","Physician","Regular",
     "Civilian tool mod. Provides advanced medical care, restoring +10 manpower per turn.",
     "",False,"FIRST PASS","","Governor: HEALER | Buildings: Bath"),
    ("Military Modifiers","milmod_merchant","Merchant","Regular",
     "Civilian tool mod. Conducts trade, generating +3 Dollars per turn.",
     "",False,"FIRST PASS","","Governor: DIPLOMAT | Buildings: Market"),
    ("Military Modifiers","milmod_preacher","Preacher","Regular",
     "Civilian tool mod. Delivers sermons raising tile governor loyalty by +5 per turn.",
     "",False,"FIRST PASS","","Governor: CIRCUIT PREACHER | Buildings: Temple | American Icons: All"),
    ("Military Modifiers","milmod_architect","Architect","Regular",
     "Civilian tool mod. Designs buildings more efficiently, reducing construction costs by 15%.",
     "",False,"FIRST PASS","","Governor: ENGINEER"),
    ("Military Modifiers","milmod_hunter","Hunter","Regular",
     "Civilian tool mod. Hunts game in surrounding wilderness, generating +5 Food per turn.",
     "",False,"FIRST PASS","","Governor: SCOUT | Terrain: Forest"),
    ("Military Modifiers","milmod_fisherman","Fisherman","Regular",
     "Civilian tool mod. Fishes from rivers or coastlines, generating +3 Food per turn near water tiles.",
     "",False,"FIRST PASS","","Governor: ADMIRAL | Terrain: Cold Coast, Warm Coast"),
    ("Military Modifiers","milmod_surveyor","Surveyor","Regular",
     "Civilian tool mod. Surveys the land, revealing terrain bonuses of all adjacent tiles.",
     "",False,"FIRST PASS","","Governor: SCOUT, ENGINEER"),
    ("Military Modifiers","milmod_trapper","Trapper","Regular",
     "Civilian tool mod. Sets trap lines through wilderness, generating +2 Food and +1 trade per turn.",
     "",False,"FIRST PASS","","Governor: SCOUT | Terrain: Forest, Taiga"),
    ("Military Modifiers","milmod_park_ranger","Park Ranger","Regular",
     "Civilian tool mod. Army is immune to corruption-based disease.",
     "",False,"FIRST PASS","","Governor: SCOUT | Terrain: Forest | Laws: Environmental Protection"),
    # Presidential Mods
    ("Military Modifiers","milmod_president","President","Regular",
     "Commander mod. The Republic moves when she moves. +3 movement points per turn.",
     "",False,"FIRST PASS","","American Icons: Ualani Carlisle | Lore: President Ualani Carlisle"),
    ("Military Modifiers","milmod_election_season","Election Season","Regular",
     "Commander mod. History is watching. Keep up. +3 movement points per turn, for the rest of the game.",
     "",False,"FIRST PASS","","American Icons: Ualani Carlisle | Laws: Democratic Mandate"),
    # State Guard Mods
    ("Military Modifiers","milmod_pa_guard","Pennsylvania Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Pennsylvania tiles.",
     "",False,"FIRST PASS","","State: PA"),
    ("Military Modifiers","milmod_va_guard","Virginia Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Virginia tiles.",
     "",False,"FIRST PASS","","State: VA"),
    ("Military Modifiers","milmod_ny_guard","New York Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in New York tiles.",
     "",False,"FIRST PASS","","State: NY"),
    ("Military Modifiers","milmod_ma_guard","Massachusetts Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Massachusetts tiles.",
     "",False,"FIRST PASS","","State: MA"),
    ("Military Modifiers","milmod_md_guard","Maryland Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Maryland tiles.",
     "",False,"FIRST PASS","","State: MD"),
    ("Military Modifiers","milmod_nc_guard","North Carolina Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in North Carolina tiles.",
     "",False,"FIRST PASS","","State: NC"),
    ("Military Modifiers","milmod_sc_guard","South Carolina Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in South Carolina tiles.",
     "",False,"FIRST PASS","","State: SC"),
    ("Military Modifiers","milmod_ga_guard","Georgia Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Georgia tiles.",
     "",False,"FIRST PASS","","State: GA"),
    ("Military Modifiers","milmod_ct_guard","Connecticut Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Connecticut tiles.",
     "",False,"FIRST PASS","","State: CT"),
    ("Military Modifiers","milmod_nj_guard","New Jersey Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in New Jersey tiles.",
     "",False,"FIRST PASS","","State: NJ"),
    ("Military Modifiers","milmod_de_guard","Delaware Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Delaware tiles.",
     "",False,"FIRST PASS","","State: DE"),
    ("Military Modifiers","milmod_nh_guard","New Hampshire Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in New Hampshire tiles.",
     "",False,"FIRST PASS","","State: NH"),
    ("Military Modifiers","milmod_ri_guard","Rhode Island Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Rhode Island tiles.",
     "",False,"FIRST PASS","","State: RI"),
    ("Military Modifiers","milmod_vt_guard","Vermont Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Vermont tiles.",
     "",False,"FIRST PASS","","State: VT"),
    ("Military Modifiers","milmod_me_guard","Maine Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Maine tiles.",
     "",False,"FIRST PASS","","State: ME"),
    ("Military Modifiers","milmod_tn_guard","Tennessee Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Tennessee tiles.",
     "",False,"FIRST PASS","","State: TN"),
    ("Military Modifiers","milmod_al_guard","Alabama Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Alabama tiles.",
     "",False,"FIRST PASS","","State: AL"),
    ("Military Modifiers","milmod_fl_guard","Florida Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Florida tiles.",
     "",False,"FIRST PASS","","State: FL"),
    ("Military Modifiers","milmod_wv_guard","West Virginia Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in West Virginia tiles.",
     "",False,"FIRST PASS","","State: WV"),
    ("Military Modifiers","milmod_dc_guard","DC Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Washington DC tiles.",
     "",False,"FIRST PASS","","State: DC"),
    ("Military Modifiers","milmod_ca_qb_guard","Quebec Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Quebec tiles.",
     "",False,"FIRST PASS","","State: CA-QB"),
    ("Military Modifiers","milmod_ca_ot_guard","Ontario Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Ontario tiles.",
     "",False,"FIRST PASS","","State: CA-OT"),
    ("Military Modifiers","milmod_ca_ns_guard","Nova Scotia Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Nova Scotia tiles.",
     "",False,"FIRST PASS","","State: CA-NS"),
    ("Military Modifiers","milmod_ca_nb_guard","New Brunswick Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in New Brunswick tiles.",
     "",False,"FIRST PASS","","State: CA-NB"),
    ("Military Modifiers","milmod_ca_pei_guard","Prince Edward Island Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Prince Edward Island tiles.",
     "",False,"FIRST PASS","","State: CA-PEI"),
    ("Military Modifiers","milmod_ba_guard","Bahamas Guard","Regular",
     "Home-soil bonus. +2 Attack, +2 Defence per level when fighting in Bahamas tiles.",
     "",False,"FIRST PASS","","State: BA"),
    # Protector Buffs
    ("Military Modifiers","milmod_mothman","Mothman Presence","Regular",
     "Mothman Presence: +20 Ranged Attack, +15 Ranged Defence.",
     "",False,"FIRST PASS","","Protectors: Mothman (PROT_01)"),
    ("Military Modifiers","milmod_jersey_devil","Jersey Devil's Fury","Regular",
     "Jersey Devil's Fury: +25 Attack, +10 Ranged, +10 Block.",
     "",False,"FIRST PASS","","Protectors: Jersey Devil (PROT_02)"),
    ("Military Modifiers","milmod_bigfoot","Bigfoot's Solidarity","Regular",
     "Bigfoot's Solidarity: +30 Block, +15 Attack.",
     "",False,"FIRST PASS","","Protectors: Bigfoot (PROT_03)"),
    ("Military Modifiers","milmod_thunderbird","Thunderbird's Sovereignty","Regular",
     "Thunderbird's Sovereignty: +25 Ranged Attack, +10 Melee Attack.",
     "",False,"FIRST PASS","","Protectors: Thunderbird (PROT_04)"),
    ("Military Modifiers","milmod_headless","Headless Terror","Regular",
     "Headless Terror: +20 Attack, +10 Block; attackers become Terrified.",
     "",False,"FIRST PASS","","Protectors: Headless Horseman (PROT_05)"),
    ("Military Modifiers","milmod_chessie","Chessie's Blessing","Regular",
     "Chessie's Blessing: +20 Block, +15 Ranged Defence.",
     "",False,"FIRST PASS","","Protectors: Chessie (PROT_06)"),
    ("Military Modifiers","milmod_bell_witch","Bell Witch's Harassment","Regular",
     "Bell Witch's Harassment: +15 Attack, +20 Defence; attackers become Demoralized.",
     "",False,"FIRST PASS","","Protectors: Bell Witch (PROT_07)"),
    ("Military Modifiers","milmod_old_ironsides","Old Ironsides' Hull","Regular",
     "Old Ironsides' Hull: +30 Shield, +20 Block.",
     "",False,"FIRST PASS","","Protectors: Old Ironsides (PROT_08)"),
    ("Military Modifiers","milmod_valley_forge","Valley Forge's Will","Regular",
     "Valley Forge's Will: +10 Attack, +25 Block, +20 Defence.",
     "",False,"FIRST PASS","","Protectors: Valley Forge (PROT_09)"),
    ("Military Modifiers","milmod_snallygaster","Snallygaster's Claim","Regular",
     "Snallygaster's Claim: +20 Attack, +10 Ranged, +10 Block.",
     "",False,"FIRST PASS","","Protectors: Snallygaster (PROT_10)"),
    ("Military Modifiers","milmod_paul_revere","Paul Revere's Ride","Regular",
     "Paul Revere's Ride: +15 Ranged, +10 Attack, +3 Movement.",
     "",False,"FIRST PASS","","Protectors: Paul Revere (PROT_11)"),
    ("Military Modifiers","milmod_liberty_bell","Liberty Bell's Resonance","Regular",
     "Liberty Bell's Resonance: +25 Block, +15 Ranged Defence.",
     "",False,"FIRST PASS","","Protectors: Liberty Bell (PROT_12)"),
    ("Military Modifiers","milmod_green_mountain","Green Mountain Haunting","Regular",
     "Green Mountain Haunting: +20 Block, +15 Ranged Defence.",
     "",False,"FIRST PASS","","Protectors: Green Mountain (PROT_13)"),
    ("Military Modifiers","milmod_presidential_decree","Presidential Decree","Regular",
     "Presidential Decree: +20 Attack, +20 Block, +15 Ranged, +15 Defence.",
     "",False,"FIRST PASS","","Protectors: Presidential Decree (PROT_14)"),
    ("Military Modifiers","milmod_skunk_ape","Skunk Ape's Domain","Regular",
     "Skunk Ape's Domain: +20 Attack, +15 Block.",
     "",False,"FIRST PASS","","Protectors: Skunk Ape (PROT_15)"),
    ("Military Modifiers","milmod_eternal_vigilance","Eternal Vigilance","Regular",
     "Eternal Vigilance: +25 Block, +10 Attack.",
     "",False,"FIRST PASS","","Protectors: Eternal Vigilance (PROT_16)"),
    ("Military Modifiers","milmod_lincolns_mandate","Lincoln's Mandate","Regular",
     "Lincoln's Mandate: +15 Attack, +15 Block, +15 Ranged, +10 Defence.",
     "",False,"FIRST PASS","","Protectors: Lincoln's Ghost (PROT_17)"),
    ("Military Modifiers","milmod_wendigo","Le Wendigo's Hunger","Regular",
     "Le Wendigo's Hunger: +30 Melee Attack.",
     "",False,"FIRST PASS","","Protectors: Wendigo (CA_PROT_03)"),
    ("Military Modifiers","milmod_loup_garou","Loup-Garou's Frenzy","Regular",
     "Loup-Garou's Frenzy: +25 Attack, +15 Block, +10 Defence.",
     "",False,"FIRST PASS","","Protectors: Loup-Garou (CA_PROT_01)"),
    ("Military Modifiers","milmod_feux_follets","Feux Follets' Misdirection","Regular",
     "Feux Follets' Misdirection: +25 Ranged Defence, +15 Block.",
     "",False,"FIRST PASS","","Protectors: Feux Follets (CA_PROT_07)"),
    ("Military Modifiers","milmod_mishepeshu","Mishepeshu's Depths","Regular",
     "Mishepeshu's Depths: +20 Block, +20 Ranged Defence.",
     "",False,"FIRST PASS","","Protectors: Mishepeshu (CA_PROT_04)"),
    ("Military Modifiers","milmod_la_corriveau","La Corriveau's Cage","Regular",
     "La Corriveau's Cage: +20 Ranged, +15 Attack.",
     "",False,"FIRST PASS","","Protectors: La Corriveau (CA_PROT_05)"),
    ("Military Modifiers","milmod_le_carcajou","Le Carcajou's Tenacity","Regular",
     "Le Carcajou's Tenacity: +20 Attack, +15 Block, +10 Defence.",
     "",False,"FIRST PASS","","Protectors: Le Carcajou (CA_PROT_06)"),
    ("Military Modifiers","milmod_la_chasse_galerie","La Chasse-Galerie","Regular",
     "La Chasse-Galerie: +15 Attack, +15 Ranged, +4 Movement.",
     "",False,"FIRST PASS","","Protectors: La Chasse-Galerie (CA_PROT_07)"),
    ("Military Modifiers","milmod_le_gougou","Le Gougou's Terror","Regular",
     "Le Gougou's Terror: +15 Attack, +20 Defence; attackers become Terrified for 3 turns.",
     "",False,"FIRST PASS","","Protectors: Le Gougou (CA_PROT_08)"),
    # Negative Status Effects
    ("Military Modifiers","milmod_stunned","Stunned","Regular",
     "Dazed and unable to act: Cannot make melee attacks this turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_suppressed","Suppressed","Regular",
     "Pinned under fire: Cannot fire ranged attacks this turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_shaken","Shaken","Regular",
     "Formation broken: Melee block halved.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_terrified","Terrified","Regular",
     "Frozen in fear: Melee attack -50%, melee block -30%.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_routed","Routed","Regular",
     "Fleeing the field: Cannot attack; melee attack zeroed; ranged defence halved.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_burning","Burning","Regular",
     "The ranks are on fire: Loses manpower each turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_blinded","Blinded","Regular",
     "Can't see a thing: Ranged attack and ranged defence halved.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_hexed","Hexed","Regular",
     "Cursed by dark magic: Magic defence eliminated.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_diseased","Diseased","Regular",
     "Plague has struck the ranks: Loses manpower each turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_waterlogged","Waterlogged","Regular",
     "Soaked to the bone: Melee and ranged attack -20%.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_frostbitten","Frostbitten","Regular",
     "Fingers too cold to hold a weapon: Melee and ranged attack -30%.",
     "",False,"FIRST PASS","","Magic Schools: Storm"),
    ("Military Modifiers","milmod_demoralized","Demoralized","Regular",
     "Spirit broken: Melee attack -20%, block -20%, all mil mods disabled.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_exhausted","Exhausted","Regular",
     "Worn out from the march: Movement reduced to 1 this turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_bogged_down","Bogged Down","Regular",
     "Stuck fast: Cannot move this turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_pacified","Pacified","Regular",
     "Ordered to stand down: Cannot initiate attacks this turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_supply_cut","Supply Cut","Regular",
     "Supplies intercepted: Cannot reinforce or resupply.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_quarantined","Quarantined","Regular",
     "Plague quarantine in effect: No reinforcement; manpower drains each turn.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_seduced","Seduced","Regular",
     "The commander has found better things to do: Melee zeroed; cannot attack.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_starstruck","Starstruck","Regular",
     "Encountered a celebrity. Requested autograph. Please advise: All stats -30%.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_hangover","Hangover","Regular",
     "Someone opened the wrong cask: All stats -50%.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_love_struck","Love-Struck","Regular",
     "Cupid's arrow strikes without warning: Melee attack -70%, block -70%; cannot attack.",
     "",False,"FIRST PASS","",""),
    ("Military Modifiers","milmod_mutinous","Mutinous","Regular",
     "The ranks are restless: Melee attack -40%; 50% chance army refuses orders each turn.",
     "",False,"FIRST PASS","",""),

    # ── LAWS & EDICTS ─────────────────────────────────────────────────────────
    # 7 existing laws (text updated from DODK 'Demon King' to colonial themes)
    ("Laws & Edicts","law_armed_peasantry","Armed Peasantry","Regular",
     "Citizen militias will defend their homes and their republic with their lives. Every federal armory maintains stocks for civilian use. The Crown disarmed its subjects. The republic does the opposite.\n\nQuadrant: Freedom | Cost: 50 Mandate\n+5 Weapons per Gov Mansion level · +25 Manpower per Farm · −1 Mandate per Farm · +1 Mandate per Forge",
     "",False,"FIRST PASS","Existing law — description updated from DODK 'Demon King' reference.",""),
    ("Laws & Edicts","law_navigation_acts","Navigation Acts","Regular",
     "A corps of federal customs agents regulates trade routes and enforces export priorities. American goods travel on American ships to American advantage. We have replaced the Crown's Navigation Acts with our own — the difference is that ours benefit us.\n\nQuadrant: Order | Cost: 25 Mandate\n−2 Mandate +2 Dollars per Workshop · +1 Dollar per Farm · +1 Dollar per Camp",
     "",False,"FIRST PASS","Existing law — historically authentic name. Description updated.",""),
    ("Laws & Edicts","law_local_elections","Local Elections","Regular",
     "Each province holds elections where every citizen of legal age may cast a ballot for local administration. Democratic accountability begins at home. The courthouse answers to the people who built it.\n\nQuadrant: Freedom | Cost: 10 Mandate\n−3 Mandate +1 Harmony per Courthouse · +1 Harmony per Farm · +1 Harmony per Camp",
     "",False,"FIRST PASS","Existing law — description updated.",""),
    ("Laws & Edicts","law_democratic_mandate","Democratic Mandate","Regular",
     "It is the will of the people which directs the republic. Monarchism died at Yorktown. Let its grave be well-marked and well-tended, as a reminder of what we replaced it with.\n\nQuadrant: Freedom | Cost: 30 Mandate\n−1 Mandate per Province · +1 Max Level Courthouse",
     "",False,"FIRST PASS","Existing law — description updated from DODK 'Demon King' reference.",""),
    ("Laws & Edicts","law_universal_citizenship","Universal Citizenship","Regular",
     "No person shall be turned away from the republic for reasons of birth, origin, or personal circumstance. All who commit to republican principles are welcome here. The republic is not an ethnicity — it is a proposition.\n\nQuadrant: Equality | Cost: 15 Mandate\n−1 Mandate per Population · +1 Harmony per Population",
     "",False,"FIRST PASS","Existing law — description updated from DODK reference.",""),
    ("Laws & Edicts","law_disability_care","Disability Care","Regular",
     "The toll of war has left many citizens physically unable to care for themselves. The republic cares for those who cannot. A nation that abandons its wounded has already lost something more important than a battle.\n\nQuadrant: Equality | Cost: 145 Mandate\n+1 Dollar per Workshop · −10% cost for Population Upgrade",
     "",False,"FIRST PASS","Existing law — description updated.",""),
    ("Laws & Edicts","law_homeland_defense","Homeland Defense","Regular",
     "Every citizen is expected to contribute to national defense in case of invasion. Freedom is not inherited — it is defended, repeatedly, by people willing to stand in the road.\n\nQuadrant: Order | Cost: 15 Mandate\n−1 Mandate per Barracks · +5 Manpower from all buildings",
     "",False,"FIRST PASS","Existing law — description updated from DODK reference.",""),
    # 8 proposed laws (not yet in code)
    ("Laws & Edicts","law_press_freedom","Freedom of the Press","Regular",
     "No law shall abridge the freedom of speech or of the press. The printing presses of Philadelphia won this revolution as surely as the riflemen at Lexington. We protect them accordingly — and not merely when we agree with what they print.\n\nQuadrant: Freedom | Cost: TBD\nProposed bonuses: Culture output, Faction loyalty, Influence generation",
     "",False,"FIRST PASS","PROPOSED — not yet implemented in code.",""),
    ("Laws & Edicts","law_abolition","Abolition Decree","Regular",
     "Slavery is incompatible with the founding principles of this republic. No territory under federal jurisdiction shall permit the ownership of persons. The hypocrisy of the founding documents is noted. This law is the correction.\n\nQuadrant: Equality | Cost: TBD (high)\nProposed bonuses: Large Happiness gain, major Faction loyalty shifts, Mandate cost",
     "",False,"FIRST PASS","PROPOSED — major political event, major faction consequences. Not yet in code.",""),
    ("Laws & Edicts","law_continental_army","Continental Army","Regular",
     "A standing professional army under federal command, separate from state militias. Washington understood this distinction. So did every foreign power that encountered American troops still in the field in December.\n\nQuadrant: Order | Cost: TBD\nProposed bonuses: Manpower efficiency, army upkeep reduction, combat bonuses",
     "",False,"FIRST PASS","PROPOSED — not yet in code.",""),
    ("Laws & Edicts","law_frontier_land","Frontier Land Act","Regular",
     "Western territories are surveyed, parceled, and made available to citizens in good standing. The republic expands by law, not by chaos — which is the polite way of saying 'also by chaos, but with paperwork.'\n\nQuadrant: Freedom | Cost: TBD\nProposed bonuses: Colonization cost reduction, Manpower boost in frontier tiles",
     "",False,"FIRST PASS","PROPOSED — not yet in code.",""),
    ("Laws & Edicts","law_religious_toleration","Religious Toleration","Regular",
     "No official religion. No state tithe. No persecution for peaceful practice of any faith. Washington wrote to the Jewish congregation in Newport that this republic gives 'to bigotry no sanction, to persecution no assistance.' We intend to honor that.\n\nQuadrant: Freedom | Cost: TBD\nProposed bonuses: Happiness boost, Faction loyalty across religious factions",
     "",False,"FIRST PASS","PROPOSED — pairs with Faiths & Doctrines system. Not yet in code.",""),
    ("Laws & Edicts","law_bill_of_rights","Bill of Rights","Regular",
     "The rights of citizens are not granted by the government — they are protected by it. The Bill of Rights defines what the state may not do, which is more important than what it may. Madison wrote most of it. It was an argument he won.\n\nQuadrant: Freedom | Cost: TBD (requires Democratic Mandate)\nProposed bonuses: Comprehensive Happiness and Mandate generation",
     "",False,"FIRST PASS","PROPOSED — capstone Freedom law. Not yet in code.",""),
    ("Laws & Edicts","law_tariff_protection","Tariff Protection","Regular",
     "American manufacturing must be shielded from British underselling until it is strong enough to compete. Protectionism is not failure — it is agriculture: you protect the seedling until it can stand on its own.\n\nQuadrant: Order | Cost: TBD\nProposed bonuses: Dollar and Weapons output from domestic production",
     "",False,"FIRST PASS","PROPOSED — not yet in code.",""),
    ("Laws & Edicts","law_habeas_corpus","Habeas Corpus","Regular",
     "No person shall be held without charge, evidence, or trial. The writ of habeas corpus is the bedrock of civil liberty — the thing that distinguishes a republic from a dungeon that happens to hold elections.\n\nQuadrant: Freedom | Cost: TBD\nProposed bonuses: Mandate generation, Happiness in high-population tiles",
     "",False,"FIRST PASS","PROPOSED — not yet in code.",""),

    # ── 19th Century Expansion ────────────────────────────────────────────────
    ("Laws & Edicts","law_homestead_act","Homestead Act","Regular",
     "One hundred sixty acres of federal land to any citizen willing to work it for five years. The frontier is opened not by conquest alone but by a paperwork process that rewards the persistent and the willing. The land is there. The republic is making it available.\n\nEra: 1862 | Quadrant: Freedom\nProposed bonuses: Colonization cost −50% · Farm output on new tiles · Manpower from frontier",
     "",False,"FIRST PASS","PROPOSED — 1862 historical law.",""),
    ("Laws & Edicts","law_land_grant_education","Land-Grant Education","Regular",
     "Federal land allocated for the establishment of colleges of agriculture and the mechanical arts. The Morrill Act's conviction: the republic needs educated engineers and farmers, not merely educated gentlemen. An educated citizenry governs better.\n\nEra: 1862 | Quadrant: Equality\nProposed bonuses: Science +30% · Library output increased · Technology research costs reduced",
     "",False,"FIRST PASS","PROPOSED — Morrill Act 1862.",""),
    ("Laws & Edicts","law_transcontinental_railroad","Transcontinental Railroad","Regular",
     "Federal land grants and subsidies to connect the continent by rail. The United States of America becomes economically real when you can ship goods from California to New York in a week. The workers who built it are not in the history books as often as they should be.\n\nEra: 1860s | Quadrant: Order\nProposed bonuses: Logistics bonus nationwide · Trade income +25% · Colonization speed increased",
     "",False,"FIRST PASS","PROPOSED — 1862-1869 historical construction.",""),
    ("Laws & Edicts","law_emancipation","Emancipation Proclamation","Regular",
     "As a war measure and a moral declaration: all persons held as slaves in states in rebellion against the United States are henceforth and forever free. The republic is correcting its founding contradiction. Late. Imperfectly. But definitively.\n\nEra: 1863 | Quadrant: Equality\nProposed bonuses: Major Happiness · Manpower from freed population · Faction loyalty shifts · Enables Reconstruction laws",
     "",False,"FIRST PASS","PROPOSED — requires active war state. Major event trigger.",""),
    ("Laws & Edicts","law_reconstruction_amendments","Reconstruction Amendments","Regular",
     "The Thirteenth, Fourteenth, and Fifteenth Amendments to the Constitution: the abolition of slavery, equal citizenship regardless of race, and the right to vote regardless of race. The republic's second founding. The attempt to mean what it said in 1776 begins now.\n\nEra: 1865–1870 | Quadrant: Equality\nProposed bonuses: Mandate generation · Universal Citizenship bonuses doubled · Voting faction loyalty",
     "",False,"FIRST PASS","PROPOSED — requires Emancipation Proclamation.",""),
    ("Laws & Edicts","law_sherman_antitrust","Sherman Antitrust Act","Regular",
     "No combination of companies shall exist in restraint of trade or to monopolize markets. The republic is large enough for many competitors and too large to be divided as an inheritance among a few wealthy families. The government will enforce this.\n\nEra: 1890 | Quadrant: Equality\nProposed bonuses: Dollar income more evenly distributed · Monopoly events prevented · Faction loyalty with labor factions",
     "",False,"FIRST PASS","PROPOSED — 1890 historical law.",""),
    ("Laws & Edicts","law_civil_service_reform","Civil Service Reform","Regular",
     "Federal employment based on merit, examination, and demonstrated competence rather than political connection or patronage. The Spoils System is formally over. The informal version persists, but we are now required to be embarrassed about it.\n\nEra: 1883 | Quadrant: Order\nProposed bonuses: Corruption −20% nationwide · Government efficiency bonus · Mandate from administration",
     "",False,"FIRST PASS","PROPOSED — Pendleton Act 1883.",""),
    ("Laws & Edicts","law_eight_hour_day","Eight-Hour Day","Regular",
     "No worker shall labor more than eight hours in a day without additional compensation. The republic belongs to those who build it, not merely to those who own the tools they build it with. Time is the one thing that cannot be manufactured.\n\nEra: 1866 onward | Quadrant: Equality\nProposed bonuses: Manpower productivity · Worker happiness · Reduced unrest from labor faction",
     "",False,"FIRST PASS","PROPOSED — multiple historical labor laws coalesced.",""),

    # ── Progressive Era ───────────────────────────────────────────────────────
    ("Laws & Edicts","law_womens_suffrage","Women's Suffrage","Regular",
     "The right of citizens to vote shall not be denied or abridged on account of sex. Seneca Falls was in 1848. The Nineteenth Amendment passed in 1920. The argument was won by the people who refused to stop making it — which is the method for most things worth winning.\n\nEra: 1920 | Quadrant: Freedom\nProposed bonuses: Mandate +30% · Faction loyalty from women's rights legislation · Election outcomes more favorable",
     "",False,"FIRST PASS","PROPOSED — 19th Amendment.",""),
    ("Laws & Edicts","law_federal_reserve","Federal Reserve Act","Regular",
     "A central bank of the United States to stabilize currency, regulate credit, and prevent the periodic banking collapses that have disrupted the economy every decade since the republic's founding. Hamilton was right about this in 1791. The republic is catching up.\n\nEra: 1913 | Quadrant: Order\nProposed bonuses: Dollar income stabilized · Banking crises prevented · Inflation events reduced",
     "",False,"FIRST PASS","PROPOSED — 1913 Federal Reserve Act.",""),
    ("Laws & Edicts","law_income_tax","Graduated Income Tax","Regular",
     "Those who earn more shall contribute more. The republic requires revenue; it will not collect that revenue solely from those who can least afford to provide it. The proportion is adjustable. The principle is not.\n\nEra: 1913 | Quadrant: Equality\nProposed bonuses: Steady Mandate income · Reduces wealth-based faction tension · Enables New Deal laws",
     "",False,"FIRST PASS","PROPOSED — 16th Amendment 1913.",""),
    ("Laws & Edicts","law_direct_senate","Direct Senate Elections","Regular",
     "Senators of the United States are henceforth elected directly by the citizens of their states, not appointed by state legislatures. The people choose their representatives. This is the direction the republic continues to move: more people, more direct.\n\nEra: 1913 | Quadrant: Freedom\nProposed bonuses: Mandate from popular elections · Reduces elite faction dominance",
     "",False,"FIRST PASS","PROPOSED — 17th Amendment 1913.",""),
    ("Laws & Edicts","law_pure_food","Pure Food and Drug Act","Regular",
     "What citizens eat and what medicines they take shall be subject to federal inspection and truthful labeling. The republic protects its people from profitable deceptions. Upton Sinclair wrote The Jungle to inspire labor reform; it inspired food safety law instead. The meatpackers were furious.\n\nEra: 1906 | Quadrant: Equality\nProposed bonuses: Happiness +15% · Population health bonus · Reduces disease event frequency",
     "",False,"FIRST PASS","PROPOSED — 1906 Pure Food and Drug Act.",""),
    ("Laws & Edicts","law_conservation","Conservation and National Parks","Regular",
     "The great natural landscapes of this continent — the forests, the canyons, the geysers, the coastlines — are held in federal trust for all citizens and all future generations. Some things cannot be owned. Some things must be protected from the people who would like to own them.\n\nEra: 1906 | Quadrant: Freedom\nProposed bonuses: Wood output sustained indefinitely · Nature magic bonus · Happiness from wilderness tiles",
     "",False,"FIRST PASS","PROPOSED — Antiquities Act and National Park Service.",""),
    ("Laws & Edicts","law_prohibition","Temperance Decree","Regular",
     "The manufacture, sale, and transportation of intoxicating liquors is hereby prohibited. This law was enacted in 1920 and repealed in 1933. Both the enacting and the repealing are instructive. The Constitution can prohibit behavior; it cannot make people comply. This is a lesson the republic keeps learning.\n\nEra: 1920–1933 | Quadrant: Order\nProposed bonuses/penalties: Happiness −20% · Corruption +10% · Black market faction events · Repeal restores status quo",
     "",False,"FIRST PASS","PROPOSED — 18th Amendment (1920) and repeal via 21st (1933).",""),

    # ── New Deal Era ──────────────────────────────────────────────────────────
    ("Laws & Edicts","law_social_safety_net","Social Safety Net","Regular",
     "Unemployment insurance, aid to dependent families, and federal welfare programs. The republic does not abandon its citizens when the economy fails them. The failure, when it occurs, is structural. The response is collective. The republic takes responsibility.\n\nEra: 1930s | Quadrant: Equality\nProposed bonuses: Happiness +20% · Stability during economic events · Faction loyalty from workers",
     "",False,"FIRST PASS","PROPOSED — New Deal programs 1933–1938.",""),
    ("Laws & Edicts","law_social_security","Social Security Act","Regular",
     "A federal insurance program for old age and disability. Citizens contribute throughout their working lives; they are entitled to support in retirement and in crisis. This is what collective social obligation looks like when a republic decides to formalize it.\n\nEra: 1935 | Quadrant: Equality\nProposed bonuses: Happiness +25% · Population stability · Reduces crisis mortality events",
     "",False,"FIRST PASS","PROPOSED — Social Security Act 1935.",""),
    ("Laws & Edicts","law_labor_compact","Labor Compact","Regular",
     "Workers have the right to organize, to bargain collectively, and to strike. The Wagner Act made explicit what the Declaration of Independence implied: the governed must consent, including those whose labor makes the machines run. The bosses objected. The Supreme Court disagreed.\n\nEra: 1935 | Quadrant: Equality\nProposed bonuses: Manpower +30% from Barracks and Forges · Worker happiness · Reduces strikes and labor unrest events",
     "",False,"FIRST PASS","PROPOSED — National Labor Relations Act (Wagner Act) 1935.",""),
    ("Laws & Edicts","law_rural_electrification","Rural Electrification","Regular",
     "Federal support for electrical infrastructure in areas the private market deemed unprofitable to serve. The New Deal's conviction: the government's job is to make modernity available to everyone, not just the accessible markets. The lights come on in places that had never seen them.\n\nEra: 1936 | Quadrant: Equality\nProposed bonuses: Magic/Power output in frontier tiles · Building productivity in rural regions",
     "",False,"FIRST PASS","PROPOSED — Rural Electrification Act 1936.",""),
    ("Laws & Edicts","law_gi_bill","GI Bill","Regular",
     "Veterans of military service receive education benefits, housing assistance, and employment support upon their return. The republic acknowledges a debt to those who served it and offers repayment in the form of opportunity — not ceremony, not a parade, but a loan for a house and tuition for a college.\n\nEra: 1944 | Quadrant: Equality\nProposed bonuses: Manpower recovery after battles · Science bonus from returning veterans · Happiness after war ends",
     "",False,"FIRST PASS","PROPOSED — Servicemen's Readjustment Act 1944.",""),

    # ── Civil Rights Era ──────────────────────────────────────────────────────
    ("Laws & Edicts","law_civil_rights_act","Civil Rights Act","Regular",
     "Discrimination based on race, color, religion, sex, or national origin is prohibited in employment, public accommodations, and federal programs. The republic is being made to demonstrate that its founding documents mean what they say. The demonstration takes longer than it should.\n\nEra: 1964 | Quadrant: Equality\nProposed bonuses: Happiness +30% · Major faction loyalty shifts · Enables Voting Rights Act",
     "",False,"FIRST PASS","PROPOSED — Civil Rights Act 1964.",""),
    ("Laws & Edicts","law_voting_rights","Voting Rights Act","Regular",
     "Federal protection of the right to vote, with oversight mechanisms for jurisdictions with histories of systematic disenfranchisement. The Fifteenth Amendment meant this in 1870. The republic is now enforcing it.\n\nEra: 1965 | Quadrant: Freedom\nProposed bonuses: Mandate from elections · Faction representation bonus · Democratic Mandate law amplified",
     "",False,"FIRST PASS","PROPOSED — Voting Rights Act 1965.",""),
    ("Laws & Edicts","law_medicare_medicaid","Medicare and Medicaid","Regular",
     "Federal healthcare coverage for citizens over 65 and for low-income families. A republic that allows its elderly and its poor to die of preventable illness for lack of money is administering a sorting process, not a government. The republic declines to sort.\n\nEra: 1965 | Quadrant: Equality\nProposed bonuses: Happiness +25% · Population health bonus · Disability Care law amplified",
     "",False,"FIRST PASS","PROPOSED — Medicare and Medicaid Act 1965.",""),
    ("Laws & Edicts","law_immigration_reform","Immigration Reform","Regular",
     "The national origins quota system, which favored Northern Europeans and excluded most of humanity, is abolished. Families may reunite; refugees may apply; talent and labor from every nation may contribute. The republic is not an ethnicity. It is a proposition, and the proposition is open.\n\nEra: 1965 | Quadrant: Equality\nProposed bonuses: Population growth · Manpower · Happiness from diversity events",
     "",False,"FIRST PASS","PROPOSED — Hart-Celler Act 1965.",""),
    ("Laws & Edicts","law_fair_housing","Fair Housing Act","Regular",
     "Racial discrimination in the sale, rental, and financing of housing is prohibited. Where a person can live shall not be determined by what they look like. The republic's geography has been shaped by deliberate exclusion. This law begins the reversal.\n\nEra: 1968 | Quadrant: Equality\nProposed bonuses: Happiness +20% · Corruption reduced in urban tiles · Faction loyalty",
     "",False,"FIRST PASS","PROPOSED — Fair Housing Act 1968.",""),

    # ── Modern Era ────────────────────────────────────────────────────────────
    ("Laws & Edicts","law_environmental_protection","Environmental Protection","Regular",
     "Federal standards for clean air, clean water, and the management of toxic materials. The republic holds the natural environment in trust; it does not have the right to permanently degrade it for temporary profit. The rivers should not be flammable. The air should be breathable. These are not radical positions.\n\nEra: 1970 | Quadrant: Freedom\nProposed bonuses: Wood and Food output sustained · Magic bonus · Nature magic school amplified",
     "",False,"FIRST PASS","PROPOSED — Clean Air Act, Clean Water Act, EPA founding 1970.",""),
    ("Laws & Edicts","law_ada","Americans with Disabilities Act","Regular",
     "Public accommodations, employment, and transportation shall be fully accessible to citizens with disabilities. Full participation in the republic is not conditional on able-bodiedness. This law states what the republic has always implied and frequently failed to build.\n\nEra: 1990 | Quadrant: Equality\nProposed bonuses: Happiness +20% · Disability Care law amplified · Mandate from accessibility events",
     "",False,"FIRST PASS","PROPOSED — ADA 1990. Extends Disability Care law.",""),
    ("Laws & Edicts","law_family_leave","Family and Medical Leave","Regular",
     "Workers may take unpaid leave for family medical emergencies without losing their employment. The republic acknowledges that workers are people with lives, families, and crises — not production units with personal inconveniences.\n\nEra: 1993 | Quadrant: Equality\nProposed bonuses: Worker happiness · Manpower retention · Reduced population attrition events",
     "",False,"FIRST PASS","PROPOSED — Family and Medical Leave Act 1993.",""),
    ("Laws & Edicts","law_healthcare_reform","Healthcare Reform","Regular",
     "Expanded access to health insurance through market regulation and government subsidy. The republic's citizens should not face bankruptcy for the cost of survival. This is not a radical position; every other developed republic reached it earlier.\n\nEra: 2010 | Quadrant: Equality\nProposed bonuses: Happiness +25% · Population growth · Medicare and Medicaid amplified",
     "",False,"FIRST PASS","PROPOSED — Affordable Care Act 2010.",""),
    ("Laws & Edicts","law_marriage_equality","Marriage Equality","Regular",
     "The right to marry is not contingent on the sex of the parties involved. All citizens stand equal before the law in the formation of families, the recognition of partnership, and the inheritance of property. The republic extends its foundational promise another step.\n\nEra: 2015 | Quadrant: Freedom\nProposed bonuses: Happiness +20% · Faction loyalty from progressive factions · Freedom of Conscience doctrine amplified",
     "",False,"FIRST PASS","PROPOSED — Obergefell v. Hodges 2015.",""),
    ("Laws & Edicts","law_climate_accord","Climate Accord","Regular",
     "The republic commits to reducing carbon emissions and participating in international frameworks for environmental protection. The climate is a public good. Its degradation is a public harm. Future generations are stakeholders in decisions made today. They do not have votes, but they have inheritances.\n\nEra: Modern | Quadrant: Freedom\nProposed bonuses: Environmental Protection amplified · International influence bonus · Long-term resource sustainability",
     "",False,"FIRST PASS","PROPOSED — Paris Agreement and domestic legislation.",""),
    ("Laws & Edicts","law_criminal_justice","Criminal Justice Reform","Regular",
     "Revision of mandatory minimum sentencing, investment in rehabilitation over incarceration, and accountability mechanisms for law enforcement. A justice system that punishes poverty more reliably than crime is not administering justice. It is administering something else.\n\nEra: Modern | Quadrant: Equality\nProposed bonuses: Happiness +20% · Corruption reduced · Faction loyalty from oppressed factions",
     "",False,"FIRST PASS","PROPOSED — ongoing legislative effort.",""),

    # ── TECHNOLOGIES ──────────────────────────────────────────────────────────
    # All 24 techs reframed for 1782 colonial America.
    # Tech IDs match in-game techID values from tech_unlock_button.gd.
    # Display names in parentheses where they differ from techID.
    ("Technologies","tech_language","Language","Regular",
     "The printed word as political weapon. In Philadelphia, a pamphlet costs a shilling and can topple a government. Every broadside posted on a tavern door is a political act. The republic runs on language — and the presses that multiply it.\n\nUnlocks: no building reward. Enables all that follows.",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_writing","Writing","Regular",
     "Standardized script enables military orders, property records, tax filings, and the first presidential correspondence. A nation that cannot put words on paper and have them read correctly on the other end cannot govern itself.\n\nUnlocks: Library Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_alphabet","Alphabet","Regular",
     "A shared alphabet across a continent of dialects. Without common script, the republic writes thirteen different letters and receives thirteen different replies, none of which agree.\n\nUnlocks: Library Upgrade (institution tech — requires 3 prior unlocks)",
     "",False,"FIRST PASS","Institution tech — requires 3 prior technologies.",""),
    ("Technologies","tech_mathematics","Mathematics","Regular",
     "Surveying land boundaries. Calculating cannon trajectories. Balancing the republic's accounts. Without mathematics, we cannot know what we own, cannot aim what we carry, and cannot balance what we owe. Hamilton did not invent this, but he applied it more aggressively than anyone else.\n\nUnlocks: Library Upgrade (institution tech)",
     "",False,"FIRST PASS","Institution tech.",""),
    ("Technologies","tech_agriculture","Agriculture","Regular",
     "Scientific farming: crop rotation, soil analysis, drainage, and improved seed selection. A well-fed army marches. A well-fed population governs. The republic's foundation is dirt, and proud of it.\n\nUnlocks: Farm Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_calendar","Calendar","Regular",
     "The republican calendar coordinates planting seasons, election cycles, campaign windows, and the administration's schedule. An organized republic runs on shared time. One that cannot count its seasons cannot count its votes.\n\nUnlocks: Tower Upgrade · Granary Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_irrigation","Irrigation","Regular",
     "Water directed is food secured. Canals, drainage channels, and river works are the republic's commitment to permanent infrastructure over seasonal luck. Virginia's tobacco and Pennsylvania's grain both depend on it.\n\nUnlocks: Farm Upgrade · Bath Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_engineering","Engineering","Regular",
     "Applied mathematics becomes physical reality: bridges that cross the Delaware, fortifications that held Bunker Hill, roads that connect Philadelphia to the frontier. ENGINEER governors were made for this age. Valley Forge was a failure of construction — it will not happen again.\n\nUnlocks: Farm Upgrade · Camp Upgrade · Faire Upgrade · Bath Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_craftmanship","Craftmanship (Copper Working)","Regular",
     "The artisan tradition: copper, tin, and bronze worked into tools, instruments, and household goods. American craftmanship is not yet industrial, but it is skilled, proud, and consistently underselling British imports in the domestic market.\n\nUnlocks: Camp Upgrade · Mine Upgrade",
     "",False,"FIRST PASS","In-game techID is 'Copper Working'; displayed as 'Craftmanship'.",""),
    ("Technologies","tech_metal_casting","Metal Casting (Bronze Working)","Regular",
     "The foundry arts. Cannons are cast. Bells are rung. Mill mechanisms are poured into molds by men who have learned to read the color of molten iron. The republic's industrial voice is cast here, and it is loud.\n\nUnlocks: Forge Upgrade · Camp Upgrade",
     "",False,"FIRST PASS","In-game techID is 'Bronze Working'; displayed as 'Metal Casting'.",""),
    ("Technologies","tech_forging","Forging (Iron Working)","Regular",
     "Iron shaped under the hammer into the republic's spine. Better muskets, stronger bridges, sharper plows. The difference between a colony and a sovereign nation is often a matter of what it can smelt and what it can make of what it smelts.\n\nUnlocks: Mine Upgrade · Forge Upgrade",
     "",False,"FIRST PASS","In-game techID is 'Iron Working'; displayed as 'Forging'.",""),
    ("Technologies","tech_tempuring","Tempering (Tempuring)","Regular",
     "Controlled heating and cooling hardens steel beyond what simple forging achieves. American tempered steel holds its edge through a winter campaign where British iron cracks in the cold. The original records spell this 'Tempuring.' The steel does not care.\n\nUnlocks: Workshop Upgrade · Forge Upgrade",
     "",False,"FIRST PASS","In-game techID is 'Tempuring' (typo for Tempering).",""),
    ("Technologies","tech_artistry","Artistry","Regular",
     "The fine arts document the revolution and build the republic's cultural identity. Painters, sculptors, silversmiths, and poets are arguing with Europe simultaneously, which Europe finds irritating. That is the intended effect.\n\nUnlocks: Temple Upgrade · Workshop Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_masonry","Masonry","Regular",
     "Stone and mortar laid to last centuries. Courthouses, capitols, and customs houses built to say the republic is not going anywhere — regardless of what Parliament thinks. Good masonry outlasts good intentions.\n\nUnlocks: Bath Upgrade · Temple Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_architecture","Architecture","Regular",
     "Neoclassical design borrowed from Rome and Athens, adapted for a continent still deciding what it wants to look like. Every public building is a political argument. Most of them are columns and optimism, which is the right combination.\n\nUnlocks: Workshop Upgrade · Temple Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_banking","Banking","Regular",
     "The Bank of the Republic extends credit, stabilizes the currency, and enables the government to finance wars without immediate bankruptcy. Hamilton won this argument. Jefferson lost. The republic is still having the argument, but the banks are cashing the drafts either way.\n\nUnlocks: Banking Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_sailing","Sailing","Regular",
     "Wind, current, and tide mastered in the service of trade, strategy, and national reach. The republic's commercial prosperity and diplomatic access depend entirely on sailors who know the Atlantic — which is less romantic but more accurate than any poem about it.\n\nUnlocks: Dock Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_statecraft","Statecraft","Regular",
     "The art of republican governance: constitutional procedure, diplomatic protocol, and the ancient practice of getting politicians to agree on anything at all. The founders were good at this when they agreed on the goal. Less so afterward.\n\nUnlocks: Courthouse Upgrade · Faire Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_shipbuilding","Shipbuilding","Regular",
     "American white oak produces frigates that have surprised the Royal Navy on multiple occasions and will surprise them again. The republic's navy grows from converted merchant vessels to purpose-built warships, one hull at a time.\n\nUnlocks: Dock Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_lenscraft","Lenscraft","Regular",
     "The grinding of lenses for telescopes, sextants, microscopes, and surveying instruments. Science advances through careful observation. Navigation advances through accurate instruments. The difference between 'approximately' and 'precisely' matters more at sea than it does anywhere on land.\n\nUnlocks: Tower Upgrade · Dock Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_organization","Organization","Regular",
     "Chain of command, census records, legislative procedure, and the administrative infrastructure of a functioning state. The Continental Army's greatest weakness was never courage or arms — it was paperwork and command structure. We fixed the paperwork. The courage was always there.\n\nUnlocks: Barracks Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_logistics","Logistics","Regular",
     "Supply chains, quartermaster records, and the unglamorous administrative work that actually determines which side wins campaigns. Valley Forge is the republic's permanent monument to the cost of ignoring this. We did not put it on the currency, but we remember.\n\nUnlocks: Barracks Upgrade · Dock Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_tactics","Tactics","Regular",
     "American irregular warfare: riflemen in the tree line, night river crossings, hit-and-run engagements that drained a professional army over eight years. We won by refusing to fight the way they expected. We continue to develop this philosophy.\n\nUnlocks: Barracks Upgrade",
     "",False,"FIRST PASS","",""),
    ("Technologies","tech_authority","Authority","Regular",
     "Legitimate governmental power derived from the consent of the governed. The founding documents wrote it down. The muskets at Concord enforced it. The elections every four years renew it. Constitutional authority is the republic's most important technology — and the hardest to maintain.\n\nUnlocks: Barracks Upgrade · Courthouse Upgrade",
     "",False,"FIRST PASS","",""),

    # ── FACTIONS ──────────────────────────────────────────────────────────────
    ("Factions","fac_sons_of_liberty","Sons of Liberty","Regular",
     "Led by Patrick Henry, the Sons of Liberty are the revolutionary vanguard — the men who lit the fire before anyone else was ready. They believe the republic is won at the barrel of a musket and maintained by the vigilance of armed citizens.\n\nT1 Militia Muster (loyalty 30): Armed Peasantry + Homeland Defense laws; Barracks +50 manpower.\nT2 Merchant Networks (loyalty 60): Navigation Acts law; Markets and Workshops +1 gold.\nT3 Letters of Marque (loyalty 90): Calico Jack joins your governors; Docks produce weapons.",
     "",False,"FIRST PASS","","law_armed_peasantry|law_homeland_defense|law_navigation_acts"),

    ("Factions","fac_continental_congress","Continental Congress","Regular",
     "Led by Abigail Adams, the Continental Congress believes the republic must be built on institutions — law, consent, legitimacy. They are the architects.\n\nT1 Articles of Confederation (loyalty 30): Local Elections + Democratic Mandate laws; Courthouses +1 mandate.\nT2 Foreign Diplomacy (loyalty 60): Monuments +1 influence.\nT3 Constitutional Convention (loyalty 90): Democratic Mandate law; Libraries +1 science.",
     "",False,"FIRST PASS","","law_local_elections|law_democratic_mandate"),

    ("Factions","fac_common_cause","Common Cause","Regular",
     "Led by Daniel Shays, Common Cause is the faction of the ordinary settler — the farmer, the homesteader, the person who fought a war and came home to find a debt collector.\n\nT1 Frontier Homesteads (loyalty 30): Armed Peasantry law; Farms +1 food.\nT2 The People's Assembly (loyalty 60): Universal Citizenship law; Baths and Temples +1 happiness.\nT3 Land Reform (loyalty 90): Farms +1 additional food.",
     "",False,"FIRST PASS","","law_armed_peasantry|law_universal_citizenship"),

    ("Factions","fac_abolitionist_league","Abolitionist League","Regular",
     "Led by Mercy Otis Warren, the Abolitionist League holds that no republic worthy of the name can tolerate slavery.\n\nT1 Freedom Papers (loyalty 30): Universal Citizenship + Disability Care laws.\nT2 Underground Railroad (loyalty 60): Phillis Wheatley joins your governors; Temples +1 culture.\nT3 Universal Emancipation (loyalty 90): Democratic Mandate law; Temples +1 happiness, Courthouses +1 mandate.",
     "",False,"FIRST PASS","","law_universal_citizenship|law_disability_care|law_democratic_mandate"),

    ("Factions","fac_free_workers_union","Free Workers Union","Regular",
     "Led by Thomas Paine, the Free Workers Union represents artisans, dockworkers, and the laboring poor.\n\nT1 Guild Charters (loyalty 30): Forges +1 weapons, Markets and Workshops +1 gold.\nT2 General Strike (loyalty 60): Disability Care law; Camps +1 wood, Mines +1 metal.\nT3 Workers Commonwealth (loyalty 90): Universal Citizenship law; Farms +1 food, Libraries +1 science.",
     "",False,"FIRST PASS","","law_disability_care|law_universal_citizenship"),

    ("Factions","fac_french_habitants","French Habitants","Regular",
     "The French-speaking farming communities of Quebec are cautious allies. They want their language, their faith, and their legal traditions respected.\n\nT1 Quebec Act Recognition (loyalty 30): Local Elections law; Temples +1 culture.\nT2 Habitants Alliance (loyalty 60): Pierre Renard joins your governors; Farms +1 food.\nT3 Republic of Quebec (loyalty 90): Universal Citizenship + Democratic Mandate laws; Farms +1 additional food, Temples +1 happiness.",
     "",False,"FIRST PASS","","law_local_elections|law_universal_citizenship|law_democratic_mandate"),

    ("Factions","fac_loyalist_settlers","Loyalist Settlers","Regular",
     "Former Crown loyalists who decided the republic is the safer bet. They bring British administrative competence and, occasionally, British guilt.\n\nT1 Crown Defectors (loyalty 30): Homeland Defense law; Courthouses +1 mandate.\nT2 Pragmatic Compact (loyalty 60): Benjamin Tallmadge joins at level 2; Barracks +50 manpower.\nT3 New Republic Converts (loyalty 90): Navigation Acts + Democratic Mandate laws; Libraries +1 science, Markets and Workshops +1 gold.",
     "",False,"FIRST PASS","","law_homeland_defense|law_navigation_acts|law_democratic_mandate"),

    ("Factions","fac_haudenosaunee_confederacy","Haudenosaunee Confederacy","Regular",
     "The Six Nations have governed this land by confederation for centuries. They want sovereignty, respect, and a treaty that is actually honored.\n\nT1 Treaty of Friendship (loyalty 30): Universal Citizenship law; Camps +1 wood.\nT2 Haudenosaunee Alliance (loyalty 60): Camps +1 weapons.\nT3 Sovereign Partnership (loyalty 90): Universal Citizenship + Disability Care laws; Barracks +50 manpower.",
     "",False,"FIRST PASS","","law_universal_citizenship|law_disability_care"),

    ("Factions","fac_coureurs_des_bois","Coureurs des Bois","Regular",
     "The woodsmen who mapped the continent before any government claimed it. Their loyalty is earned slowly and kept carefully. When they commit to a cause, the forest moves with them.\n\nT1 Trade Routes (loyalty 30): Camps +1 wood.\nT2 Frontier Network (loyalty 60): Louis Tremblant joins your governors; Camps +2 wood total, Mines +1 metal.\nT3 Continental Reach (loyalty 90): Camps +1 weapons.",
     "",False,"FIRST PASS","",""),

    ("Factions","fac_maritime_patriots","Maritime Patriots","Regular",
     "The fishing and trading communities of the Atlantic coast — Nova Scotia, New Brunswick, Prince Edward Island. The republic offers them something the Crown never did: a charter that treats them as citizens instead of subjects.\n\nT1 Port Alliance (loyalty 30): Docks +1 boats.\nT2 Atlantic Commerce (loyalty 60): Navigation Acts law; Markets, Workshops, and Docks +1 gold.\nT3 Maritime Union (loyalty 90): Docks +1 additional boats.",
     "",False,"FIRST PASS","","law_navigation_acts"),

    # ── ARMY UNITS (STUB) ─────────────────────────────────────────────────────
    ("Army Units","_stub_army","— stub —","Regular",
     "Army Units entries not yet written. Add them to RecordsDatabase.gd.",
     "",False,"STUB","Needs entries written.",""),

    # ── MAGIC SCHOOLS & SPELLS ────────────────────────────────────────────────
    # Six schools referenced in sys_spells. Themed for 1782 colonial-supernatural.
    ("Magic Schools & Spells","magic_fire","Fire","Regular",
     "The oldest magical tradition — command of flame, heat, and the fundamental energy of transformation. Fire mages are deployed offensively in the republic's service, which is to say they are deployed first and asked questions later. The Towers that produce magical output generate Fire-school energy first.\n\nPrimary use: offensive combat spells · bonus Weapons output from Forge tiles",
     "",False,"FIRST PASS","",""),
    ("Magic Schools & Spells","magic_ice","Ice","Regular",
     "Cold, stillness, and preservation. Ice magic slows what should be stopped and preserves what would otherwise be lost — food stores, documents, armies positioned for a long winter campaign. Valley Forge would have gone differently.\n\nPrimary use: defensive spells, army preservation · bonus Food output from cold terrain",
     "",False,"FIRST PASS","",""),
    ("Magic Schools & Spells","magic_nature","Nature","Regular",
     "Growth, healing, and the power of living systems. Nature magic is practiced in the forests and bogs of the American frontier, where it is largely indistinguishable from advanced herbalism by everyone except the plants.\n\nPrimary use: healing, crop enhancement · bonus Food and Wood output",
     "",False,"FIRST PASS","",""),
    ("Magic Schools & Spells","magic_shadow","Shadow","Regular",
     "Concealment, illusion, and the manipulation of perception. The republic has the Culper Ring for conventional intelligence operations. Shadow practitioners operate on stranger material — and are not officially acknowledged.\n\nPrimary use: intelligence operations, counter-espionage · bonus Influence generation",
     "",False,"FIRST PASS","",""),
    ("Magic Schools & Spells","magic_light","Light","Regular",
     "Revelation, clarity, and the exposure of hidden things. Light magic has applications in healing, morale, and the detection of deception. It is the magic of courthouses and public squares and uncomfortable truths spoken at the wrong moment.\n\nPrimary use: healing, morale · bonus Happiness and Mandate output",
     "",False,"FIRST PASS","",""),
    ("Magic Schools & Spells","magic_storm","Storm","Regular",
     "The power of weather, tide, and the great atmospheric systems that shaped this continent long before anyone tried to govern it. Storm practitioners are rare, strong-willed, and frequently useful at inconvenient moments. The founding documentation on this school is sparse.\n\nPrimary use: weather manipulation, naval advantage · bonus Boats and Magic output",
     "",False,"FIRST PASS","",""),

    # ── PROTECTORS (all mystery until agree) ──────────────────────────────────
    ("Protectors","PROT_01","???","Mystery",
     "Something moves in the mountain passes of West Virginia at night. Old miners refuse to speak its name.",
     "PROT_01",False,"FIRST PASS","Mothman. Revealed entry not yet written.",""),
    ("Protectors","PROT_02","???","Mystery",
     "Hunters along the Pine Barrens have reported a winged figure that leaves no tracks and makes no sound.",
     "PROT_02",False,"FIRST PASS","Jersey Devil. Revealed entry not yet written.",""),
    ("Protectors","PROT_03","???","Mystery",
     "The Blue Ridge holds something ancient. Larger than a man. Older than the republic.",
     "PROT_03",False,"FIRST PASS","Bigfoot. Revealed entry not yet written.",""),
    ("Protectors","PROT_04","???","Mystery",
     "Storm riders speak of a great shape seen above the clouds near the Great Lakes. The thunder that follows it is not natural.",
     "PROT_04",False,"FIRST PASS","Thunderbird. Revealed entry not yet written.",""),
    ("Protectors","PROT_05","???","Mystery",
     "A horseman without a head has been reported along the Hudson. It rides hard and it rides at night.",
     "PROT_05",False,"FIRST PASS","Headless Horseman. Revealed entry not yet written.",""),
    ("Protectors","PROT_06","???","Mystery",
     "Chesapeake fishermen have stopped working the deep water. Something beneath the surface watches back.",
     "PROT_06",False,"FIRST PASS","Chessie/sea serpent. Revealed entry not yet written.",""),
    ("Protectors","PROT_07","???","Mystery",
     "In the hills of Tennessee, a farmhouse was visited nightly by something that could not be touched. It knew names. It remembered.",
     "PROT_07",False,"FIRST PASS","Bell Witch. Revealed entry not yet written.",""),
    ("Protectors","PROT_08","???","Mystery",
     "A ship that cannot be sunk has been sighted in Boston Harbor. Its crew does not age. Its guns never run dry.",
     "PROT_08",False,"FIRST PASS","Ghost ship / Constitution spirit. Revealed entry not yet written.",""),
    ("Protectors","PROT_09","???","Mystery",
     "Near Valley Forge, sentries report a figure walking the old encampment grounds in the fog. It wears Continental blue.",
     "PROT_09",False,"FIRST PASS","Washington's Ghost. Revealed entry not yet written.",""),
    ("Protectors","PROT_10","???","Mystery",
     "Something nests in the Catoctin Mountains. Travelers between the capital and the north have gone missing.",
     "PROT_10",False,"FIRST PASS","Snallygaster. Revealed entry not yet written.",""),
    ("Protectors","PROT_11","???","Mystery",
     "A rider was seen near Lexington moving faster than any horse alive. The message he carries has not yet been delivered.",
     "PROT_11",False,"FIRST PASS","Paul Revere ghost. Revealed entry not yet written.",""),
    ("Protectors","PROT_12","???","Mystery",
     "In Philadelphia, on quiet nights, a ringing is heard with no source. It comes from Independence Hall. Nothing is there.",
     "PROT_12",False,"FIRST PASS","Liberty Bell spirit. Revealed entry not yet written.",""),
    ("Protectors","PROT_13","???","Mystery",
     "The mountains of Vermont are haunted by something patriotic and enormous. It has opinions about taxation.",
     "PROT_13",False,"FIRST PASS","Green Mountain giant (Ethan Allen). Revealed entry not yet written.",""),
    ("Protectors","PROT_14","???","Mystery",
     "The faces in the rock at Gettysburg open their eyes sometimes. Only sometimes. But when they do, they are looking south.",
     "PROT_14",False,"FIRST PASS","Mount Rushmore (anachronistic). Revealed entry not yet written.",""),
    ("Protectors","PROT_15","???","Mystery",
     "The Everglades hold something enormous and foul-smelling that the local Seminole call very old. They do not explain further.",
     "PROT_15",False,"FIRST PASS","Skunk Ape. Revealed entry not yet written.",""),
    ("Protectors","PROT_16","???","Mystery",
     "There is a musket that fires without being loaded, held by someone who cannot be seen, near the Connecticut River valley.",
     "PROT_16",False,"FIRST PASS","Invisible musketeer. Revealed entry not yet written.",""),
    ("Protectors","PROT_17","???","Mystery",
     "The White House has had a permanent guest since 1862. He walks the halls at night. He is tall. He is patient.",
     "PROT_17",False,"FIRST PASS","Lincoln's Ghost. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_01","???","Mystery",
     "Something pulls the ice floes apart along the St. Lawrence. The voyageurs call it a bad crossing year. It is not a crossing year.",
     "CA_PROT_01",False,"FIRST PASS","Canadian protector. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_02","???","Mystery",
     "A shape has been seen beneath Lake Ontario for three hundred years. It has not moved. It is waiting.",
     "CA_PROT_02",False,"FIRST PASS","Lake monster (Manipogo). Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_03","???","Mystery",
     "The Wendigo of the northern forests is not a story parents tell children. It is a warning parents tell each other.",
     "CA_PROT_03",False,"FIRST PASS","Wendigo. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_04","???","Mystery",
     "The Ojibwe speak of a great lynx that controls the deep water. It has not been seen since the last winter that killed everyone who saw it.",
     "CA_PROT_04",False,"FIRST PASS","Mishibizhiw. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_05","???","Mystery",
     "In the villages near Québec City, they remember a woman who was executed. They do not say she stayed dead.",
     "CA_PROT_05",False,"FIRST PASS","La Corriveau. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_06","???","Mystery",
     "The wolverine of Moncton is not an animal. The trappers learned this. The trappers are gone now.",
     "CA_PROT_06",False,"FIRST PASS","Acadian monster. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_07","???","Mystery",
     "The flying canoe that travels the river at night does not appear on maps. The men inside it have been paddling for a very long time.",
     "CA_PROT_07",False,"FIRST PASS","Chasse-galerie. Revealed entry not yet written.",""),
    ("Protectors","CA_PROT_08","???","Mystery",
     "In the bay, where the Chaleur meets the ocean, fishermen sometimes see a light beneath the water. They go home. They do not explain why.",
     "CA_PROT_08",False,"FIRST PASS","Phantom ship of Chaleur Bay. Revealed entry not yet written.",""),

    # ── MYTHIC WEAPONS ────────────────────────────────────────────────────────
    ("Mythic Weapons","_stub_mythic","— stub —","Regular",
     "Mythic Weapons entries not yet written. Weapons show as mystery entries until first found. Add them to RecordsDatabase.gd.",
     "",False,"STUB","Needs entries written. Each weapon needs a mystery + revealed version.",""),

    # ── LORE & HISTORY ────────────────────────────────────────────────────────
    ("Lore & History","lore_first_war","The First British Reconquest War","Regular",
     "In 1782, following the unexpected death of King George III, the British Parliament authorized a full military reconquest of the former colonies. What followed was the First Reconquest War — a brutal, satirical, and deeply inconvenient reminder that revolution is easier the second time.\n\nPresident Ualani Carlisle faced the full weight of the British Empire with a standing army, a treasury of questionable depth, and the most politically functional cabinet in American history.",
     "",False,"FIRST PASS","",""),
    ("Lore & History","lore_ualani","President Ualani Carlisle","Regular",
     "Hawaii's first President of the United States. Former senator. Former general. Current problem for the British Empire.\n\nUalani Carlisle was elected on a platform of infrastructure, diplomacy, and what her opponents called 'an alarming amount of common sense.' She is known for her directness and her habit of personally responding to threatening letters from foreign heads of state.",
     "",False,"FIRST PASS","",""),

    # ── FAITHS & DOCTRINES ────────────────────────────────────────────────────
    # NEW CATEGORY — add "Faiths & Doctrines" to RecordsDatabase.gd CATEGORIES list.
    # Replaces the DODK fantasy religion system (Benaxtara, Tyla-Dyn, etc.)
    # with colonial-American themed doctrines and faith traditions.
    #
    # Doctrines (faithBelief=false in belief.gd): political/philosophical stances
    # Faith Traditions (faithBelief=true): actual religious denominations
    # The -3 to +3 churchLevel scale = Secular Republic ↔ Providence Republic

    # ── Doctrines (Tier 1) ────────────────────────────────────────────────────
    ("Faiths & Doctrines","doc_natural_rights","Natural Rights","Regular",
     "Jefferson's premise, Locke's theory: rights exist before governments do. Life, liberty, and the pursuit of happiness are not gifts from the state — they are the reason the state exists at all. Any government that forgets this will hear about it eventually.\n\nType: Doctrine (Tier 1) | Effect: Mandate and Happiness generation | Replaces: 'Sacred Groves'",
     "",False,"FIRST PASS","Replaces DODK 'Sacred Groves'. Needs implementation in belief.gd and religion_data.gd.",""),
    ("Faiths & Doctrines","doc_freedom_conscience","Freedom of Conscience","Regular",
     "The right to hold private beliefs without state interference or coercion. What a person believes between themselves and God — and how they practice that belief privately — is no business of any government, however republican.\n\nType: Doctrine (Tier 1) | Effect: Happiness and Faith generation | Replaces: 'Midsummer Celebrations'",
     "",False,"FIRST PASS","Replaces DODK 'Midsummer Celebrations'. Has mild sensual undertone — private life is private.",""),
    ("Faiths & Doctrines","doc_republican_virtue","Republican Virtue","Regular",
     "A republic requires citizens of virtue: willing to set personal interest aside for the public good, capable of self-governance, and informed enough to choose wisely. Washington demonstrated this. Franklin also demonstrated this, in a different and equally impressive sense.\n\nType: Doctrine (Tier 1) | Effect: Mandate generation, reduces Corruption | Replaces: 'Tree of Life'",
     "",False,"FIRST PASS","Replaces DODK 'Tree of Life'. Franklin reference is deliberate.",""),
    ("Faiths & Doctrines","doc_common_sense","Common Sense","Regular",
     "Thomas Paine's pamphlet outsold the Bible in the year of its publication. Common Sense is the doctrine that ordinary people are capable of self-governance — that kings are not chosen by God, they are merely born with better tailors.\n\nType: Doctrine (Tier 1) | Effect: Culture generation, Faction loyalty boost | Replaces: 'Standing Stones'",
     "",False,"FIRST PASS","Replaces DODK 'Standing Stones'.",""),
    ("Faiths & Doctrines","doc_providence","Providence & Republic","Regular",
     "Many founders believed the republic was not merely a political experiment but a providential one — that God's hand was in the revolution, that the new nation had a divine mandate to demonstrate republican governance to the world. God's involvement is debated. His apparent preference for constitutions over hereditary monarchy is less so.\n\nType: Doctrine (Tier 1) | Effect: Faith and Mandate generation | Replaces: 'Valued Idolatry'",
     "",False,"FIRST PASS","Replaces DODK 'Valued Idolatry'.",""),
    ("Faiths & Doctrines","doc_enlightenment","Enlightenment Reason","Regular",
     "The application of scientific method to political philosophy. Reason — not tradition, bloodline, or divine right — as the foundation for law and governance. Franklin, Jefferson, and Madison were largely convinced this was sufficient. They were largely right, which is different from being entirely right.\n\nType: Doctrine (Tier 1) | Effect: Science and Culture generation | Replaces: 'Healing Waters'",
     "",False,"FIRST PASS","Replaces DODK 'Healing Waters'.",""),

    # ── Doctrines (Tier 2) ────────────────────────────────────────────────────
    ("Faiths & Doctrines","doc_federal_theology","Federal Theology","Regular",
     "The Puritan covenant tradition: a binding agreement between God and the people, adapted for federal governance. New England's political culture is still recognizably covenantal — organized, literate, deeply convinced of its own righteousness, and extremely good at keeping records.\n\nType: Doctrine (Tier 2) | Effect: Faith + Mandate generation | Replaces: 'Nature Sanctuaries'",
     "",False,"FIRST PASS","Replaces DODK 'Nature Sanctuaries'.",""),
    ("Faiths & Doctrines","doc_classical_republicanism","Classical Republicanism","Regular",
     "Governance lessons drawn from Rome and Athens: civic virtue, balance of powers, rotation of office, and the constant danger of demagogues who appeal to the crowd. The founders read the classics obsessively. Cicero appears in the Federalist Papers more often than is comfortable.\n\nType: Doctrine (Tier 2) | Effect: Mandate generation, increased Courthouse output | Replaces: 'Conservative Orthodoxy'",
     "",False,"FIRST PASS","Replaces DODK 'Conservative Orthodoxy'.",""),
    ("Faiths & Doctrines","doc_manifest_destiny","Manifest Destiny","Regular",
     "The belief that the republic is providentially ordained to expand across the continent — that the land to the west belongs to those with the will and the means to claim it. It is a doctrine of enormous confidence. The nations already present on that land have opinions about it which are not represented in this pamphlet.\n\nType: Doctrine (Tier 2) | Effect: Colonization bonuses, Frontier expansion | Replaces: 'Sanctioned Cadaver Research'",
     "",False,"FIRST PASS","Replaces DODK 'Sanctioned Cadaver Research'. Historical complexity noted in description.",""),
    ("Faiths & Doctrines","doc_civic_nationalism","Civic Nationalism","Regular",
     "National identity defined by shared ideals rather than ethnicity, bloodline, or region of origin. You are American because you hold these truths — not because of where you were born. This is an unusual claim for 1782. The republic is attempting it anyway.\n\nType: Doctrine (Tier 2) | Effect: Population growth, Happiness, reduces faction penalties | Replaces: 'Temple Height Restrictions'",
     "",False,"FIRST PASS","Replaces DODK 'Temple Height Restrictions'.",""),

    # ── Faith Traditions (Tier 1) ─────────────────────────────────────────────
    ("Faiths & Doctrines","faith_congregationalist","Congregationalist","Regular",
     "The Standing Order of New England: town meetings, covenant theology, and a deep suspicion of bishops. The Puritan tradition shaped colonial political culture more profoundly than it is generally credited for, which may explain why Massachusetts is the way it is.\n\nType: Faith (Tier 1) | Effect: Faith + Mandate generation, Courthouse output bonus | Replaces: 'Benaxtara'",
     "",False,"FIRST PASS","Replaces DODK 'Benaxtara'. New England denomination.",""),
    ("Faiths & Doctrines","faith_quaker","Quaker Meeting","Regular",
     "The Society of Friends opposes war, slavery, and the swearing of oaths — and built Pennsylvania into one of the wealthiest and most stable colonies. Their influence on the republic's founding is considerable and consistently underacknowledged by people who like wars.\n\nType: Faith (Tier 1) | Effect: Happiness + Influence generation | Replaces: 'Tyla-Dyn'",
     "",False,"FIRST PASS","Replaces DODK 'Tyla-Dyn'. Pennsylvania Quaker tradition.",""),
    ("Faiths & Doctrines","faith_presbyterian","Presbyterian Assembly","Regular",
     "The Scots-Irish Presbyterian tradition fueled revolutionary sentiment throughout the Appalachian backcountry. They arrived in the colonies already furious at various kings and simply redirected their considerable energy accordingly.\n\nType: Faith (Tier 1) | Effect: Manpower + Faith generation, army morale bonus | Replaces: 'Fa Enepo'",
     "",False,"FIRST PASS","Replaces DODK 'Fa Enepo'. Scots-Irish frontier tradition.",""),
    ("Faiths & Doctrines","faith_baptist","Baptist Congregation","Regular",
     "The Baptist movement spreads rapidly through the frontier: believer's baptism, congregational authority, and a firm conviction that the government should leave them alone. They are not wrong about that last part. Roger Williams invented religious liberty in this country. He was a Baptist.\n\nType: Faith (Tier 1) | Effect: Faith + Happiness generation | Replaces: 'Bibwey'",
     "",False,"FIRST PASS","Replaces DODK 'Bibwey'. Frontier/Southern denomination.",""),
    ("Faiths & Doctrines","faith_methodist","Methodist Circuit","Regular",
     "John Wesley's traveling ministers reach communities no permanent church can serve. The circuit rider is the fastest information network in the frontier — they arrive before the roads do. The Methodist movement spreads through the republic faster than any official institution can organize it, which is either the Holy Spirit or excellent logistics.\n\nType: Faith (Tier 1) | Effect: Faith generation across frontier tiles | Replaces: 'Dilnith-Amen'",
     "",False,"FIRST PASS","Replaces DODK 'Dilnith-Amen'. Wesleyan tradition.",""),
    ("Faiths & Doctrines","faith_anglican","Anglican Remnant","Regular",
     "The Church of England in America. Many Anglican clergy departed with the Tories. Those who remained tend toward moderation — or, if they remained in Virginia, toward loudly and historically claiming they were never really that committed to the Anglican position on anything in particular.\n\nType: Faith (Tier 1) | Effect: Influence + Culture generation | Replaces: 'Ornil-Ra'",
     "",False,"FIRST PASS","Replaces DODK 'Ornil-Ra'. Complicated loyalist undertones.",""),

    # ── Faith Traditions (Tier 2) ─────────────────────────────────────────────
    ("Faiths & Doctrines","faith_catholic","Catholic Parish","Regular",
     "The Catholic communities of Maryland, French Canada, and Spanish Florida represent a substantial and frequently overlooked part of the republic's religious landscape. Washington's letter to Bishop Carroll confirms the republic intends to honor its promise of religious liberty. The Catholics are watching to see if it does.\n\nType: Faith (Tier 2) | Effect: Faith + Culture generation, Canada diplomatic bonus | Replaces: 'Vibian Karik'",
     "",False,"FIRST PASS","Replaces DODK 'Vibian Karik'. Maryland/Canadian Catholic tradition.",""),
    ("Faiths & Doctrines","faith_deist","Deist Philosophy","Regular",
     "God wound the universe and stepped back. The founders who held this view include several of the ones whose faces appear on the currency, and others whose private correspondence is kept in sealed archives for the benefit of their reputations. It is a faith tradition that does not require a building, which saves on construction costs.\n\nType: Faith (Tier 2) | Effect: Science + Mandate generation | Replaces: 'Venodam'",
     "",False,"FIRST PASS","Replaces DODK 'Venodam'. Franklin, Jefferson, Washington adjacent.",""),
    ("Faiths & Doctrines","faith_jewish","Jewish Congregation","Regular",
     "Newport's Touro Synagogue is the oldest Jewish house of worship in the country. Washington wrote personally to the congregation to assure them that the republic 'gives to bigotry no sanction, to persecution no assistance.' He meant it. The republic is still deciding whether it means it.\n\nType: Faith (Tier 2) | Effect: Dollars + Influence generation | Replaces: 'Jerriwix'",
     "",False,"FIRST PASS","Replaces DODK 'Jerriwix'. Newport/Philadelphia communities.",""),
    ("Faiths & Doctrines","faith_native_traditions","Native Traditions","Regular",
     "The spiritual traditions of nations who were present on this continent before the republic existed, who will be present after its various experiments have been tested, and who have their own relationship with the land and the supernatural that predates every European document by several thousand years. Their relationship with the new government is complicated, which is a significant understatement.\n\nType: Faith (Tier 2) | Effect: Magic + Nature bonuses, protector arc connections | Replaces: 'Qalin Ling & Tyrus'",
     "",False,"FIRST PASS","Replaces DODK 'Qalin Ling & Tyrus'. Historical complexity intentional.",""),

    # ── Expanded Universal Spiritual Traditions ───────────────────────────────
    ("Faiths & Doctrines","faith_islam","Islam in America","Regular",
     "Islam arrived in America with enslaved Africans before the republic existed. It continued through immigration, through conversion, through the Nation of Islam, through Malcolm X, through Dearborn. It is the most diverse religious community on the continent, and it has been American longer than most Americans know.\n\nType: Spiritual Tradition | Effect: Faith + Influence generation · Community solidarity bonus",
     "",False,"FIRST PASS","Deeply American tradition. Pre-dates the republic in the form of enslaved Muslims.",""),
    ("Faiths & Doctrines","faith_buddhism","Buddhism in America","Regular",
     "Chinese laborers brought Buddhism to California in the 1840s. Japanese American Buddhists carried it through internment camps and out the other side. The countercultural movement of the 1960s adopted it from a different direction. It arrived on three separate tracks and stayed.\n\nType: Spiritual Tradition | Effect: Happiness + wisdom bonus · Reduces internal conflict",
     "",False,"FIRST PASS","Chinese railroad workers, Japanese internment community, Beat Generation adoption.",""),
    ("Faiths & Doctrines","faith_hinduism","Hinduism and Dharmic Traditions","Regular",
     "Arriving primarily through 20th century immigration, the Hindu, Sikh, Jain, and related Dharmic communities built temples in Texas, software companies in California, and medical practices in New York. The republic is large enough for all of it, which is why they came.\n\nType: Spiritual Tradition | Effect: Science + Culture bonus · Merchant productivity",
     "",False,"FIRST PASS","Indian subcontinent diaspora. Hart-Celler 1965 opened immigration.",""),
    ("Faiths & Doctrines","faith_vodou","Vodou and African Diaspora Traditions","Regular",
     "The spiritual traditions carried by enslaved Africans survived forced conversion by adapting, syncreting, and persisting in ways no colonial power could completely suppress. Louisiana Vodou, Candomblé, Santería, Hoodoo — spiritual systems of resistance, healing, and memory that predate many of the denominations that tried to replace them.\n\nType: Spiritual Tradition | Effect: Shadow magic amplified · Healing bonuses · Manpower recovery",
     "",False,"FIRST PASS","African diaspora traditions. Major presence in Louisiana, New York, the Caribbean rim.",""),
    ("Faiths & Doctrines","faith_sikhism","Sikhism","Regular",
     "Punjabi Sikh farmers came to California's Central Valley in the early 1900s. They built gurdwaras, farmed in a valley that everyone said was too arid, and stayed. In 2012, a mass shooting at a gurdwara in Wisconsin prompted many Americans to ask what Sikhism was. The Sikhs had been here for a century. The answer was available.\n\nType: Spiritual Tradition | Effect: Food + Manpower from Farm tiles · Community resilience bonus",
     "",False,"FIRST PASS","Central Valley farming community. One of the oldest South Asian communities in the US.",""),
    ("Faiths & Doctrines","faith_ame","African Methodist Episcopal","Regular",
     "The AME Church was founded in 1816 by Richard Allen, who walked out of a Philadelphia church when an usher told him to move to the segregated gallery. He founded a denomination instead. The Black church became the organizational backbone of civil rights, mutual aid, education, and resistance for two hundred years.\n\nType: Spiritual Tradition | Effect: Mandate + Faith generation · Civil rights law costs reduced · Faction loyalty",
     "",False,"FIRST PASS","Richard Allen, 1816. First major Black institution in the United States.",""),
    ("Faiths & Doctrines","faith_secular_humanism","Secular Humanism","Regular",
     "No gods. No divine authority. Ethics derived from reason, human dignity, and the observable consequences of our choices. The republic's founding documents are partly a product of this tradition — they derive authority from consent, not from heaven. The rest is application.\n\nType: Secular Tradition | Effect: Science + Mandate · Reason-based law costs reduced · Enlightenment doctrine amplified",
     "",False,"FIRST PASS","Major American tradition. Strongly represented in academic and scientific communities.",""),
    ("Faiths & Doctrines","faith_uu","Unitarian Universalism","Regular",
     "A living tradition that draws from all the world's wisdom — Christian, Jewish, Buddhist, Pagan, Humanist, Indigenous, and more. The only faith tradition in America whose founding text includes the possibility that its founding text is not definitive. John Adams and John Quincy Adams were Unitarians.\n\nType: Spiritual Tradition | Effect: All Doctrine bonuses +10% · Interfaith relations bonus",
     "",False,"FIRST PASS","John and John Quincy Adams were Unitarians. Uniquely inclusive tradition.",""),
    ("Faiths & Doctrines","faith_spiritualism","American Spiritualism","Regular",
     "The belief that the dead can communicate with the living through mediums, séances, and spiritual practice. It emerged in upstate New York in 1848 and spread nationally within a decade. Mary Todd Lincoln held séances in the White House. The republic has always maintained a back channel to whatever comes next.\n\nType: Spiritual Tradition | Effect: Protector arc bonds stronger · Faith + Magic generation",
     "",False,"FIRST PASS","American-invented tradition, 1848. Lincoln White House connection.",""),
    ("Faiths & Doctrines","faith_lds","Latter-day Saints","Regular",
     "Founded in 1830 in upstate New York by Joseph Smith, the Church of Jesus Christ of Latter-day Saints is one of the few world religions invented entirely on American soil. Its history includes persecution, migration, the founding of Salt Lake City, and a current membership larger than Judaism. It is uniquely, unavoidably American.\n\nType: Spiritual Tradition | Effect: Colonization bonus · Community building output · Frontier expansion amplified",
     "",False,"FIRST PASS","Founded 1830, Western migration, Utah territorial founding.",""),

    # ── Movement Doctrines (political traditions that function as civic faiths) ─
    ("Faiths & Doctrines","doc_labor_rights","Labor Rights Doctrine","Regular",
     "The political faith that workers are the true producers of wealth and have the right to organize, bargain, and share in the prosperity they create. From the AFL to the CIO to the United Farm Workers to today, it is the tradition that built the weekend, the minimum wage, and the eight-hour day — and continues to argue that these are not the ceiling.\n\nType: Movement Doctrine | Effect: Manpower + Forge and Farm output · Happiness from worker laws",
     "",False,"FIRST PASS","Movement tradition from 1860s to present.",""),
    ("Faiths & Doctrines","doc_civil_rights_covenant","Civil Rights Covenant","Regular",
     "The tradition of nonviolent struggle for the full inclusion of all citizens in the promise of the republic's founding documents. It holds that the republic has written checks it has not cashed — and presents them, persistently, until they clear. Martin Luther King Jr. was its most eloquent articulation. He was not its origin, and he was not its end.\n\nType: Movement Doctrine | Effect: Happiness + Mandate from equality laws · Coalition faction bonuses",
     "",False,"FIRST PASS","Ongoing tradition from Reconstruction through present.",""),
    ("Faiths & Doctrines","doc_environmentalism","Environmental Stewardship","Regular",
     "The republic holds the natural world in trust for all future generations. The land, the water, the air, and the living systems that make them habitable are not resources to be liquidated — they are the commons that make the republic possible. John Muir walked alone into the wilderness and came back to argue for everyone.\n\nType: Movement Doctrine | Effect: Wood and Food output sustained long-term · Conservation laws cost less",
     "",False,"FIRST PASS","Thoreau → Muir → Sierra Club → EPA → modern environmentalism.",""),
    ("Faiths & Doctrines","doc_womens_liberation","Women's Liberation","Regular",
     "From Abigail Adams to Seneca Falls to the suffragettes to Betty Friedan to the present, the tradition that the republic's promise of equality is not complete until it applies to more than half the population. 'Remember the ladies,' Adams wrote. The republic is still in the process of remembering.\n\nType: Movement Doctrine | Effect: Culture + Mandate · Women's Suffrage law costs reduced · Election outcomes improved",
     "",False,"FIRST PASS","Continuous tradition 1770s to present.",""),
    ("Faiths & Doctrines","doc_nonviolent_resistance","Nonviolent Resistance","Regular",
     "The political method developed by Thoreau, refined by Gandhi, and practiced by the American civil rights movement: the deliberate, public, nonviolent violation of unjust laws in order to expose them, oppose them, and change them. It requires extraordinary courage. It works.\n\nType: Movement Doctrine | Effect: Mandate from protest events · Faction loyalty from injustice events · Reduces military costs in civil conflicts",
     "",False,"FIRST PASS","Thoreau 1849 → Gandhi → King → present.",""),

    # ═══════════════════════════════════════════════════════════════════════════
    # AMERICAN ICONS
    # NEW CATEGORY — add "American Icons" to RecordsDatabase.gd CATEGORIES list.
    # Replaces the DODK fantasy gods with historical American patron figures.
    # Players choose one Icon as their inspiration; personality determines bonuses.
    # faithBelief=true in belief.gd for all entries.
    # ═══════════════════════════════════════════════════════════════════════════

    # ── Founding Era ──────────────────────────────────────────────────────────
    ("American Icons","icon_washington","George Washington","Regular",
     "Commander. President. The man who refused to be king, which at the time was the most radical thing an American had ever done. His farewell address warned against factionalism and entangling foreign alliances. Neither warning was fully heeded. His armies did not quit, which mattered more than any single battle.\n\nBonuses: Manpower +25% from Barracks · Army morale holds in winter · Defensive combat modifier for fortified tiles · No retreat penalties | Replaces: 'Benaxtara'",
     "",False,"FIRST PASS","Founding Era. Replaces DODK 'Benaxtara'.",""),
    ("American Icons","icon_franklin","Benjamin Franklin","Regular",
     "Printer. Philosopher. Diplomat. Inventor. The man who charmed the French into financing the American Revolution, which is arguably the most important thing anyone did in 1778. His experiment with the kite and key was either science or folly depending on how closely you were watching. Both things are possible simultaneously.\n\nBonuses: Science doubled from Libraries · Influence +50% · Banking research halved · Diplomatic options unlock earlier | Replaces: 'Tyla-Dyn'",
     "",False,"FIRST PASS","Founding Era. Replaces DODK 'Tyla-Dyn'.",""),
    ("American Icons","icon_abigail_adams","Abigail Adams","Regular",
     "'Remember the ladies.' The most important political correspondent of the founding era wrote letters that shaped a presidency. Her advice was given and frequently ignored, which tells you something about the founding period and explains why she is worth remembering now more than many of the people who ignored her.\n\nBonuses: Culture +40% · Mandate from Courthouses doubled · Library output increased · Faction loyalty from women's rights legislation | Replaces: 'Fa Enepo'",
     "",False,"FIRST PASS","Founding Era. Replaces DODK 'Fa Enepo'.",""),
    ("American Icons","icon_hamilton","Alexander Hamilton","Regular",
     "Secretary of the Treasury. Author of the Federalist Papers. Father of the American financial system. Shot by a sitting Vice President in a duel over personal reputation, which was how business was sometimes conducted. His vision of an industrial republic survived him, which was the point.\n\nBonuses: Dollars +30% national income · Forge and Market output increased · Banking upgrades cost 50% less · Tariff Protection law amplified | Replaces: 'Bibwey'",
     "",False,"FIRST PASS","Founding Era. Replaces DODK 'Bibwey'.",""),
    ("American Icons","icon_phillis_wheatley","Phillis Wheatley","Regular",
     "Born in Senegal, enslaved as a child, became the first African American published poet at age nineteen. Her work argued in classical verse that a people capable of poetry are capable of liberty. The founders who owned slaves found this inconvenient. She published it anyway.\n\nBonuses: Culture +60% · Faith generation across all tiles · Abolition law costs halved · Faction loyalty from oppressed communities | Replaces: 'Dilnith-Amen'",
     "",False,"FIRST PASS","Founding Era. Replaces DODK 'Dilnith-Amen'.",""),
    ("American Icons","icon_jefferson","Thomas Jefferson","Regular",
     "Author of the Declaration of Independence. Architect. Inventor. President. The man who wrote 'all men are created equal' while owning several hundred people across his lifetime. His vision of an educated agrarian republic of free citizens was always more aspirational than descriptive. History is still working out what to do with the distance between the vision and the record.\n\nBonuses: Science +40% · Colonization costs −30% · Culture from Monuments and Libraries · Land-Grant Education law amplified | Replaces: 'Ornil-Ra'",
     "",False,"FIRST PASS","Founding Era. Replaces DODK 'Ornil-Ra'. Contradiction acknowledged.",""),

    # ── 19th Century ──────────────────────────────────────────────────────────
    ("American Icons","icon_lincoln","Abraham Lincoln","Regular",
     "Sixteenth President. Rail-splitter. Self-taught lawyer. The man who held the republic together by force of will, careful argument, and a war that cost 620,000 lives. He did not want the war. He did not end it until it accomplished what was necessary. The Gettysburg Address contains two minutes of prose that the republic has not yet lived up to.\n\nBonuses: Mandate +30% in wartime · Corruption reduced across all tiles · War exhaustion reduced · Morale holds through significant losses | Replaces: 'Vibian Karik'",
     "",False,"FIRST PASS","1800s. Replaces DODK 'Vibian Karik'.",""),
    ("American Icons","icon_tubman","Harriet Tubman","Regular",
     "She called herself Moses. She operated the Underground Railroad. She scouted for the Union Army. She lobbied for women's suffrage until her death. She never lost a passenger on the Railroad. When asked if she was afraid, she said she would have gone back regardless. She went back thirteen times.\n\nBonuses: Manpower does not decay in enemy territory · Shadow magic doubled · Colonization of contested tiles −50% cost · Emancipation and Reconstruction laws cost less Mandate | Replaces: 'Venodam'",
     "",False,"FIRST PASS","1800s. Replaces DODK 'Venodam'.",""),
    ("American Icons","icon_douglass","Frederick Douglass","Regular",
     "Born into slavery. Escaped. Wrote three autobiographies. Advised Lincoln. Argued that the Constitution — properly read — obligated the republic to end slavery, at a time when most abolitionists called the Constitution irredeemably corrupt. The most powerful orator in nineteenth-century America argued for a republic that did not yet exist. He lived to see it partially arrive.\n\nBonuses: Culture +50% · Press and journalism events more frequent · Mandate from public approval events · Civil Rights laws amplified | Replaces: 'Jerriwix'",
     "",False,"FIRST PASS","1800s. Replaces DODK 'Jerriwix'.",""),
    ("American Icons","icon_sitting_bull","Sitting Bull","Regular",
     "Hunkpapa Lakota leader and holy man. Refused to sell Lakota territory. Directed the victory at Little Bighorn. Fled to Canada. Returned. Was killed by Indian Agency police in 1890 during an attempted arrest. He spent his life defending a way of life the republic was determined to eliminate. His resistance outlasted his enemies' confidence that it was finished.\n\nBonuses: Faith +50% · Nature magic amplified · Manpower from frontier tiles · All protector agreements cost 25% less Influence | Replaces: 'Qalin Ling & Tyrus'",
     "",False,"FIRST PASS","1800s. Replaces DODK 'Qalin Ling & Tyrus'.",""),
    ("American Icons","icon_sojourner_truth","Sojourner Truth","Regular",
     "Born into slavery in New York. Freed herself when the law moved too slowly. Walked to freedom and then walked to every lecture hall that would have her. Argued for abolition and women's rights simultaneously when the movements preferred to work separately. 'Ain't I a Woman?' is a question that still requires an answer.\n\nBonuses: Happiness +30% from equality laws · Faction loyalty from oppressed factions · Coalition events more frequent | New entry",
     "",False,"FIRST PASS","1800s. New entry — not a DODK replacement.",""),
    ("American Icons","icon_chief_joseph","Chief Joseph","Regular",
     "Nez Perce leader who led his people on a 1,400-mile retreat rather than accept relocation, outmaneuvering four armies before being stopped forty miles from the Canadian border. 'From where the sun now stands, I will fight no more forever.' The dignity of that statement says more about the republic's failure than any argument could.\n\nBonuses: Tactical retreat bonus — armies escape encirclement · Influence with frontier factions · Morale does not collapse in defeat | New entry",
     "",False,"FIRST PASS","1800s. New entry — not a DODK replacement.",""),

    # ── 20th Century ──────────────────────────────────────────────────────────
    ("American Icons","icon_teddy_roosevelt","Theodore Roosevelt","Regular",
     "Cowboy. Soldier. Author. Trust-buster. Conservationist. The twenty-sixth president put two million acres of wilderness into protected federal status and personally led a charge up a hill in Cuba. Not all of his enthusiasms aged well. The national parks did.\n\nBonuses: Wood output +40% · Manpower from wilderness tiles · Antitrust and Conservation laws amplified · Army morale bonus | New entry",
     "",False,"FIRST PASS","Progressive Era. New entry.",""),
    ("American Icons","icon_susan_b_anthony","Susan B. Anthony","Regular",
     "She dedicated fifty years to women's suffrage and died fourteen years before it passed. Her face appeared on a dollar coin in 1979. The movement she built outlasted every setback it encountered, which is what successful movements do. She understood that organizing is not an event — it is a practice.\n\nBonuses: Mandate from election events · Women's Suffrage law costs halved · Faction loyalty from organizing · Democratic Mandate amplified | New entry",
     "",False,"FIRST PASS","Progressive Era. New entry.",""),
    ("American Icons","icon_ida_wells","Ida B. Wells","Regular",
     "Journalist. Anti-lynching crusader. Co-founder of the NAACP. She investigated racial terror in the South by traveling there and writing about it with precision and fury, which was both dangerous and necessary and effective. Her reporting changed what the country was willing to pretend it did not know.\n\nBonuses: Culture +40% · Press events more frequent · Corruption and injustice events trigger Mandate gains · Civil rights laws cost less | New entry",
     "",False,"FIRST PASS","Progressive Era. New entry.",""),
    ("American Icons","icon_eleanor_roosevelt","Eleanor Roosevelt","Regular",
     "First Lady, diplomat, UN delegate, chair of the committee that wrote the Universal Declaration of Human Rights in 1948. She had her own schedule, her own column, her own politics, and her own conception of what a public figure owed the public. 'No one can make you feel inferior without your consent.' She did not consent.\n\nBonuses: Influence +50% · Happiness from welfare and rights laws · International diplomacy events more favorable · All social laws cost less | New entry",
     "",False,"FIRST PASS","Mid-20th Century. New entry.",""),
    ("American Icons","icon_mlk","Martin Luther King Jr.","Regular",
     "The most consequential American of the twentieth century argued that the republic's founding documents obligated it to racial equality — that the check had been written in 1776, marked 'insufficient funds' for a hundred and eighty years, and he was presenting it for payment. The argument was won. The payment is ongoing.\n\nBonuses: Happiness +40% from civil rights legislation · Faction loyalty across all factions · Mandate holds through crises · Nonviolent Resistance doctrine amplified | New entry",
     "",False,"FIRST PASS","Civil Rights Era. New entry.",""),
    ("American Icons","icon_cesar_chavez","Cesar Chavez","Regular",
     "Farm worker, organizer, co-founder of the United Farm Workers with Dolores Huerta. Led boycotts of grapes and lettuce that changed American agricultural labor law. 'Sí, se puede.' He understood that power is not given — it is built, meeting by meeting, contract by contract, until the other side decides it is cheaper to negotiate.\n\nBonuses: Manpower +30% from Farm tiles · Worker happiness · Food production increased from labor laws · Labor Compact amplified | New entry",
     "",False,"FIRST PASS","Civil Rights / Labor Era. New entry.",""),
    ("American Icons","icon_jimmy_carter","Jimmy Carter","Regular",
     "Thirty-ninth President. Naval officer. Peanut farmer. He brokered the Camp David Accords between Egypt and Israel in 1978. When his term ended, he spent the next four decades building houses with Habitat for Humanity. He may be the best person to have served as president. Not the most powerful. Possibly the most decent.\n\nBonuses: Influence in peace negotiations · Faith and Happiness bonuses · Diplomatic costs reduced · Post-conflict tiles stabilize faster | New entry",
     "",False,"FIRST PASS","Late 20th Century. New entry.",""),
    ("American Icons","icon_dolores_huerta","Dolores Huerta","Regular",
     "Labor organizer. Civil rights activist. Co-founder of the United Farm Workers alongside Cesar Chavez. She coined 'Sí, se puede' — yes, we can. She organized for six decades without stopping. The labor movement, she said, is about hope. It is also about showing up to the same meeting seventeen times until something changes. She showed up.\n\nBonuses: Manpower and Happiness from Farm tiles · Coalition faction events more frequent · Labor and women's rights laws amplified | New entry",
     "",False,"FIRST PASS","Civil Rights / Labor Era. New entry.",""),
]


# ═══════════════════════════════════════════════════════════════════════════════
# JOURNAL DATA
# (event_type, classification, description, status, notes)
# ═══════════════════════════════════════════════════════════════════════════════
JOURNAL_DATA = [
    ("ualani_event",        "EYES ONLY",    "Ualani personal arc events — intimate moments away from official duties",                             "FIRST PASS",""),
    ("white_house_secret",  "EYES ONLY",    "White House holiday intimacy events — seasonal moments in Washington DC",                             "FIRST PASS",""),
    ("vp_event",            "EYES ONLY",    "Vice President relationship arc — private events involving the VP",                                   "FIRST PASS",""),
    ("protector_agree",     "TOP SECRET",   "USA protector alliance agreements — fires when a protector joins the republic",                       "FIRST PASS",""),
    ("ca_protector_agree",  "TOP SECRET",   "Canadian protector alliance agreements",                                                              "FIRST PASS",""),
    ("war_declaration",     "SECRET",       "UK declares war — major historical inflection point",                                                 "FIRST PASS",""),
    ("war_buildup",         "SECRET",       "Pre-war intelligence events — buildup and warning signs before conflict",                             "FIRST PASS",""),
    ("peace",               "SECRET",       "Peace treaty events — resolution of conflicts",                                                       "FIRST PASS",""),
    ("election_season",     "CONFIDENTIAL", "Election campaign events — political maneuvering and platform speeches",                              "FIRST PASS",""),
    ("election_night_win",  "CONFIDENTIAL", "Ualani wins re-election — victory speech and aftermath",                                              "FIRST PASS",""),
    ("election_night_lose", "CONFIDENTIAL", "Ualani loses the election — concession and transition of power",                                      "FIRST PASS",""),
    ("city_liberated",      "CONFIDENTIAL", "Major city liberated from Crown control",                                                             "FIRST PASS",""),
    ("city_lost",           "CONFIDENTIAL", "Major city lost to enemy forces",                                                                     "FIRST PASS",""),
    ("state_liberated",     "CONFIDENTIAL", "Full state liberated — major territorial milestone",                                                  "FIRST PASS",""),
    ("collapse",            "SECRET",       "Republic collapse event — catastrophic failure condition",                                            "FIRST PASS",""),
    ("secession",           "SECRET",       "State secession — a state breaks from the republic",                                                  "FIRST PASS",""),
    ("reintegration",       "CONFIDENTIAL", "State reintegration — a seceded state rejoins the republic",                                         "FIRST PASS",""),
    ("commander_complete",  "DECLASSIFIED", "Commander arc completion — a commander finishes their full loyalty arc",                              "FIRST PASS",""),
]


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_gallery_sheet(wb):
    ws = wb.create_sheet("GALLERY")
    ws.freeze_panes = "A2"
    COLS = [
        ("Group",           20),
        ("Event ID",        24),
        ("Title",           30),
        ("Content Flag",    14),
        ("Unlock Hint",     44),
        ("Flavor Text",     48),
        ("Writing Status",  16),
        ("Art Status",      14),
        ("Art Path",        28),
        ("Notes",           32),
    ]
    _hdr(ws, COLS)

    row = 2
    current_group = None
    alt = False

    for entry in GALLERY_DATA:
        grp, eid, title, cflag, hint, flavor, wstatus, astatus, apath, notes = entry

        if grp != current_group:
            current_group = grp
            row = _cat_banner(ws, row, f"  ● {grp}", len(COLS),
                              bg=GROUP_COLORS.get(grp, CAT_BG), fg="1A2A3A")
            alt = False

        bg = ALT_BG if alt else WHITE
        alt = not alt
        cflag_bg, cflag_fg = _cflag_style(cflag)
        wst_bg,  wst_fg   = _status_style(wstatus)
        art_bg,  art_fg   = _art_style(astatus)

        vals = [grp, eid, title, cflag, hint, flavor, wstatus, astatus, apath, notes]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _border()
            c.alignment = _wrap()
            c.fill = _fill(bg)
            c.font = _font()
            if ci == 4:
                c.fill = _fill(cflag_bg); c.font = _font(bold=True, color=cflag_fg)
                c.alignment = _center()
            elif ci == 7:
                c.fill = _fill(wst_bg);  c.font = _font(bold=True, color=wst_fg)
                c.alignment = _center()
            elif ci == 8:
                c.fill = _fill(art_bg);  c.font = _font(bold=True, color=art_fg)
                c.alignment = _center()

        ws.row_dimensions[row].height = 50
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def build_records_sheet(wb):
    ws = wb.create_sheet("RECORDS")
    ws.freeze_panes = "A2"
    COLS = [
        ("Category",            22),
        ("ID",                  22),
        ("Name",                26),
        ("Type",                12),
        ("Description / Hint",  60),
        ("Unlock Flag",         14),
        ("Has Icon",            10),
        ("Status",              14),
        ("Notes",               32),
        ("References",          36),
    ]
    _hdr(ws, COLS)

    row = 2
    current_cat = None
    alt = False

    for entry in RECORDS_DATA:
        cat, eid, name, etype, desc, unlock_flag, has_icon, status, notes, refs = entry

        if cat != current_cat:
            current_cat = cat
            row = _cat_banner(ws, row, f"  ● {cat}", len(COLS),
                              bg=CAT_COLORS.get(cat, CAT_BG), fg="1A2A3A")
            alt = False

        is_stub = eid.startswith("_stub_")
        bg = STUB_BG if is_stub else (ALT_BG if alt else WHITE)
        alt = not alt

        st_bg, st_fg = _status_style(status)
        type_bg = "BB44AA" if etype == "Mystery" else WHITE
        type_fg = "FFFFFF" if etype == "Mystery" else "000000"

        vals = [cat, eid, name, etype, desc, unlock_flag,
                "Yes" if has_icon else "No", status, notes, refs]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _border()
            c.alignment = _wrap()
            c.fill = _fill(bg)
            c.font = _font(italic=is_stub)
            if ci == 4:
                c.fill = _fill(type_bg); c.font = _font(bold=True, color=type_fg)
                c.alignment = _center()
            elif ci == 8:
                c.fill = _fill(st_bg); c.font = _font(bold=True, color=st_fg)
                c.alignment = _center()
            elif ci == 10 and refs:
                c.fill = _fill("D6EAD0"); c.font = _font(color="1A3A0A")

        ws.row_dimensions[row].height = 55
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def build_journal_sheet(wb):
    ws = wb.create_sheet("JOURNAL")
    ws.freeze_panes = "A2"
    COLS = [
        ("Event Type",     28),
        ("Classification", 18),
        ("Description",    55),
        ("Status",         14),
        ("Notes",          32),
    ]
    _hdr(ws, COLS)

    for i, entry in enumerate(JOURNAL_DATA):
        etype, classif, desc, status, notes = entry
        bg = ALT_BG if i % 2 else WHITE
        cl_bg, cl_fg = _classif_style(classif)
        st_bg, st_fg = _status_style(status)

        row = i + 2
        vals = [etype, classif, desc, status, notes]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _border()
            c.alignment = _wrap()
            c.fill = _fill(bg)
            c.font = _font()
            if ci == 2:
                c.fill = _fill(cl_bg); c.font = _font(bold=True, color=cl_fg)
                c.alignment = _center()
            elif ci == 4:
                c.fill = _fill(st_bg); c.font = _font(bold=True, color=st_fg)
                c.alignment = _center()

        ws.row_dimensions[row].height = 40

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_gallery_sheet(wb)
    build_records_sheet(wb)
    build_journal_sheet(wb)

    wb.save(OUT_PATH)

    n_gal   = len(GALLERY_DATA)
    n_sens  = sum(1 for e in GALLERY_DATA if e[3] == "sensual")
    n_expl  = sum(1 for e in GALLERY_DATA if e[3] == "explicit")
    n_kink  = sum(1 for e in GALLERY_DATA if e[3] == "kinky")
    art_done= sum(1 for e in GALLERY_DATA if e[7] == "Done")

    real_rec = [e for e in RECORDS_DATA if not e[1].startswith("_stub_")]
    stub_rec = [e for e in RECORDS_DATA if e[1].startswith("_stub_")]
    n_myst  = sum(1 for e in real_rec if e[3] == "Mystery")

    n_jour  = len(JOURNAL_DATA)

    print(f"Saved: {OUT_PATH}")
    print(f"GALLERY : {n_gal} entries  ({n_sens} sensual, {n_expl} explicit, {n_kink} kinky)"
          f"  |  art done: {art_done}/{n_gal}")
    print(f"RECORDS : {len(real_rec)} entries ({n_myst} mystery)  +  {len(stub_rec)} stub categories")
    print(f"JOURNAL : {n_jour} event types tracked")


if __name__ == "__main__":
    main()
