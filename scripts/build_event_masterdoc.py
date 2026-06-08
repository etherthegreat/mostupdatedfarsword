"""
Generates event_masterdoc.xlsx — a complete tracker for all Farsword events.
Run from the repo root: python3 scripts/build_event_masterdoc.py
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUT_PATH = "event_masterdoc.xlsx"

# ── PALETTE ──────────────────────────────────────────────────────────────────
CAT_COLORS = {
    "Date":           "D6E4F7",
    "City Liberated": "D6F0D6",
    "City Lost":      "FCE0D6",
    "State":          "FFF3CC",
    "Protector":      "EDD6F7",
    "Commander":      "D6E1F7",
    "Fort Chain":     "F7DDD6",
    "Collapse":       "DCDCDC",
    "Tile Crisis":    "FDEBD0",
    "Ualani":         "D6F5EA",
    "VP Arc":         "F5E6D6",
    "Canada":         "D6EAF8",
    "CA Protector":   "B8D6F7",
    "Idea":           "EDE0F7",
}

STATUS_COLORS = {
    "FULL COMPLETION": "00B050",
    "FIRST PASS":      "92D050",
    "IDEA":            "FFEB9C",
}
STATUS_TEXT = {
    "FULL COMPLETION": "FFFFFF",
    "FIRST PASS":      "1A3A00",
    "IDEA":            "7A5A00",
}

ART_COLORS = {
    "Not Started": "FFD7D7",
    "In Progress": "FFEB9C",
    "Integrated":  "C6EFCE",
}
ART_TEXT = {
    "Not Started": "8B0000",
    "In Progress": "7A5A00",
    "Integrated":  "1A4A1A",
}

HEADER_BG   = "1F3864"
HEADER_FG   = "FFFFFF"
EXPLICIT_BG = "FF6B35"
EXPLICIT_FG = "FFFFFF"
NORMAL_FG   = "1A1A1A"

# ── THIN BORDER ──────────────────────────────────────────────────────────────
thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── COLUMN DEFINITIONS ───────────────────────────────────────────────────────
#  (header, width)
COLUMNS = [
    ("CATEGORY",          14),
    ("CHAIN / GROUP",     18),
    ("EVENT ID",          24),
    ("HEADLINE",          42),
    ("CHAIN POSITION",    13),
    ("PARENT EVENT(S)",   24),
    ("CHILD EVENT(S)",    24),
    ("STATUS",            16),
    ("ART TAG",           22),
    ("ART STATUS",        13),
    ("EXPLICIT ART?",     13),
    ("CONTENT FLAG",      13),
    ("COOLDOWN",          10),
    ("NOTES / TRIGGERS",  46),
]

# ── ALL EVENT DATA ────────────────────────────────────────────────────────────
# Tuple order matches COLUMNS:
#  category, chain_group, event_id, headline,
#  chain_pos, parent, children,
#  status, art_tag, art_status, explicit_art, content_flag, cooldown, notes
EVENTS = [

    # ── DATE ─────────────────────────────────────────────────────────────────
    ("Date","Annual Calendar","DE_001",
     "A NEW WAR FOR AN OLD REPUBLIC",
     "Root","—","—",
     "FIRST PASS","president_address","Not Started","NO","","0",
     "Opening address; non-repeatable; fires turn 1"),

    ("Date","Annual Calendar","DE_002",
     "HAPPY NEW YEAR: Redcoats Ring in Revolution",
     "Root","—","—",
     "FIRST PASS","fireworks_muskets","Not Started","NO","","300",
     "Repeatable; month 1; +food 10"),

    ("Date","Annual Calendar","DE_003",
     "LOVE IN A TIME OF MUSKETS",
     "Root","—","—",
     "FIRST PASS","heart_cannon","Not Started","NO","","300",
     "Repeatable; month 2; +gold 15 or +loyalty Sons of Liberty"),

    ("Date","Annual Calendar","DE_004",
     "INDEPENDENCE DAY",
     "Root","—","—",
     "FIRST PASS","fourth_of_july","Not Started","NO","","300",
     "Repeatable; month 7; +morale"),

    ("Date","Annual Calendar","DE_005",
     "HALLOWS' EVE: Mothman Spotted Over West Virginia",
     "Root","—","Sets mothman_available flag",
     "FIRST PASS","mothman_silhouette","Not Started","NO","","300",
     "Repeatable; month 10; sets flag for PROT_01 unlock"),

    # ── CITY LIBERATED ───────────────────────────────────────────────────────
    ("City Liberated","Nassau","CL_182",
     "NASSAU FALLS: PIRATES STAND WITH THE REPUBLIC",
     "Root","—","—",
     "FIRST PASS","nassau_alliance","Not Started","NO","","0",
     "Tile 182; non-repeatable; 3 alliance buttons"),

    ("City Liberated","New York City","CL_001",
     "NYC FALLS TO THE REPUBLIC",
     "Root","—","Triggers PROT_09_SUMMON via BTN3",
     "FIRST PASS","nyc_liberation","Not Started","NO","","0",
     "Tile 14; non-repeatable; BTN3 summons Valley Forge Guardian"),

    ("City Liberated","Gettysburg","CL_006",
     "GETTYSBURG SECURED",
     "Root","—","—",
     "FIRST PASS","gettysburg_memorial","Not Started","NO","","0",
     "Tile 11; tile_feature=Gettysburg Memorial; non-repeatable"),

    ("City Liberated","Kennett Square","CL_007",
     "KENNETT SQUARE FREE",
     "Root","—","—",
     "FIRST PASS","mushroom_victory","Not Started","NO","","0",
     "Tile 10; tile_feature=Mushroom Capital; non-repeatable"),

    ("City Liberated","Brattleboro","CL_009",
     "BRATTLEBORO LIBERATED",
     "Root","—","—",
     "FIRST PASS","brattleboro_liberation","Not Started","NO","","0",
     "Tile 76; tile_feature=Green Mountains; non-repeatable"),

    # ── CITY LOST ────────────────────────────────────────────────────────────
    ("City Lost","Washington DC","CX_001",
     "DC FALLS: Government in Temporary Relocation",
     "Root","—","—",
     "FIRST PASS","dc_falls","Not Started","NO","","50",
     "Tile 188; repeatable cooldown 50"),

    ("City Lost","Gettysburg","CX_005",
     "GETTYSBURG TAKEN",
     "Root","—","—",
     "FIRST PASS","gettysburg_lost","Not Started","NO","","50",
     "Tile 11; tile_feature=Gettysburg Memorial; repeatable"),

    # ── STATE ────────────────────────────────────────────────────────────────
    ("State","Pennsylvania","SL_001",
     "PENNSYLVANIA FULLY LIBERATED",
     "Root","—","—",
     "FIRST PASS","pa_liberated","Not Started","NO","","0",
     "state_liberated type; non-repeatable"),

    ("State","Vermont","SL_004",
     "VERMONT LIBERATED",
     "Root","—","—",
     "FIRST PASS","vermont_liberated","Not Started","NO","","0",
     "state_liberated type; non-repeatable"),

    ("State","Secession Chain","STATE_REBEL_01",
     "STATE DECLARES INDEPENDENCE: [TILE_NAME] RAISES A NEW FLAG",
     "Root","—","→ STATE_REINTEGRATED_01 (after reconquest)",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires when presidentialClaim very low per-state; secession type"),

    ("State","Secession Chain","STATE_REINTEGRATED_01",
     "[TILE_NAME] RETURNS TO THE REPUBLIC",
     "Followup","STATE_REBEL_01","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires after reconquest; pardon vs replace governors choice"),

    # ── PROTECTOR ────────────────────────────────────────────────────────────
    # Mothman
    ("Protector","Mothman","PROT_01_SUMMON",
     "MOTHMAN APPROACHES",
     "Root","—","→ PROT_01_TAME",
     "FIRST PASS","mothman_approach","Not Started","YES","kinky_lewd","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 10+; kinky_lewd gated"),
    ("Protector","Mothman","PROT_01_TAME",
     "MOTHMAN DEESCALATED: THE RIDGE MEETING HOLDS",
     "Branch","PROT_01_SUMMON","→ PROT_01_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_01_tame flag; immediately chains to AGREE"),
    ("Protector","Mothman","PROT_01_AGREE",
     "MOTHMAN FORMALLY ALLIES WITH THE CONTINENTAL REPUBLIC",
     "Followup","PROT_01_TAME","—",
     "FIRST PASS","","Not Started","YES","kinky_lewd","0",
     "Sets prot_01_agreed; BTN2 kinky_lewd gated"),

    # Jersey Devil
    ("Protector","Jersey Devil","PROT_02_SUMMON",
     "JERSEY DEVIL EMERGES FROM THE PINE BARRENS",
     "Root","—","→ PROT_02_TAME",
     "FIRST PASS","jersey_devil","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 15+"),
    ("Protector","Jersey Devil","PROT_02_TAME",
     "JERSEY DEVIL CEASES SUPPLY ROUTE DISRUPTIONS — PINE BARRENS ACCORD",
     "Branch","PROT_02_SUMMON","→ PROT_02_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_02_tame flag; immediately chains to AGREE"),
    ("Protector","Jersey Devil","PROT_02_AGREE",
     "JERSEY DEVIL FORMALLY ENDORSES THE CONTINENTAL CAUSE",
     "Followup","PROT_02_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_02_agreed; BTN2 explicit gated"),

    # Bigfoot
    ("Protector","Bigfoot","PROT_03_SUMMON",
     "BIGFOOT SIDES WITH THE REVOLUTION",
     "Root","—","→ PROT_03_TAME",
     "FIRST PASS","bigfoot","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 20+"),
    ("Protector","Bigfoot","PROT_03_TAME",
     "APPALACHIAN PASSES REOPEN: BIGFOOT ACKNOWLEDGES THE REPUBLIC",
     "Branch","PROT_03_SUMMON","→ PROT_03_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_03_tame flag; immediately chains to AGREE"),
    ("Protector","Bigfoot","PROT_03_AGREE",
     "BIGFOOT FORMALLY GUIDES THE CONTINENTAL CAUSE THROUGH THE MOUNTAINS",
     "Followup","PROT_03_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_03_agreed; BTN2 explicit gated"),

    # Thunderbird
    ("Protector","Thunderbird","PROT_04_SUMMON",
     "THUNDERBIRD ANSWERS THE CALL",
     "Root","—","→ PROT_04_TAME",
     "FIRST PASS","thunderbird","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 25+"),
    ("Protector","Thunderbird","PROT_04_TAME",
     "THUNDERBIRD STANDS DOWN: STORM ACTIVITY RETURNS TO PREDICTABLE PATTERNS",
     "Branch","PROT_04_SUMMON","→ PROT_04_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_04_tame flag; immediately chains to AGREE"),
    ("Protector","Thunderbird","PROT_04_AGREE",
     "THUNDERBIRD FORMALLY ALLIES WITH THE CONTINENTAL REPUBLIC",
     "Followup","PROT_04_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_04_agreed; BTN2 explicit gated"),

    # Headless Horseman
    ("Protector","Headless Horseman","PROT_05_SUMMON",
     "THE HORSEMAN RIDES AGAIN",
     "Root","—","→ PROT_05_TAME",
     "FIRST PASS","headless_horseman","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 30+"),
    ("Protector","Headless Horseman","PROT_05_TAME",
     "THE HORSEMAN ACCEPTS IDENTIFICATION PROTOCOLS — CONTINENTAL COURIERS SAFE",
     "Branch","PROT_05_SUMMON","→ PROT_05_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_05_tame flag; immediately chains to AGREE"),
    ("Protector","Headless Horseman","PROT_05_AGREE",
     "THE HORSEMAN RIDES FOR THE REPUBLIC — FORMALLY",
     "Followup","PROT_05_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_05_agreed; BTN2 explicit gated"),

    # Chessie
    ("Protector","Chessie","PROT_06_SUMMON",
     "CHESSIE RISES FROM THE BAY",
     "Root","—","→ PROT_06_TAME",
     "FIRST PASS","chessie","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 35+"),
    ("Protector","Chessie","PROT_06_TAME",
     "CHESSIE REDIRECTS MARITIME OPERATIONS: CHESAPEAKE FISHING RESTORED",
     "Branch","PROT_06_SUMMON","→ PROT_06_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_06_tame flag; immediately chains to AGREE"),
    ("Protector","Chessie","PROT_06_AGREE",
     "CHESSIE FORMALLY PATROLS FOR THE CONTINENTAL NAVY",
     "Followup","PROT_06_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_06_agreed; BTN2 explicit gated"),

    # Bell Witch
    ("Protector","Bell Witch","PROT_07_SUMMON",
     "THE BELL WITCH MANIFESTS IN TENNESSEE",
     "Root","—","→ PROT_07_TAME",
     "FIRST PASS","bell_witch","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 40+"),
    ("Protector","Bell Witch","PROT_07_TAME",
     "BELL WITCH REDIRECTS HAUNTING OPERATIONS TO CROWN-OCCUPIED TERRITORY",
     "Branch","PROT_07_SUMMON","→ PROT_07_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_07_tame flag; immediately chains to AGREE"),
    ("Protector","Bell Witch","PROT_07_AGREE",
     "THE BELL WITCH FORMALLY COMMITS TO THE CONTINENTAL CAUSE",
     "Followup","PROT_07_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_07_agreed; BTN2 explicit gated"),

    # Old Ironsides
    ("Protector","Old Ironsides","PROT_08_SUMMON",
     "OLD IRONSIDES RETURNS TO SERVICE",
     "Root","—","→ PROT_08_TAME",
     "FIRST PASS","old_ironsides","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 45+"),
    ("Protector","Old Ironsides","PROT_08_TAME",
     "OLD IRONSIDES ACCEPTS CONTINENTAL NAVAL COMMAND — BOSTON HARBOR STABLE",
     "Branch","PROT_08_SUMMON","→ PROT_08_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_08_tame flag; immediately chains to AGREE"),
    ("Protector","Old Ironsides","PROT_08_AGREE",
     "OLD IRONSIDES FORMALLY JOINS THE CONTINENTAL NAVY",
     "Followup","PROT_08_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_08_agreed; BTN2 explicit gated"),

    # Valley Forge Guardian
    ("Protector","Valley Forge Guardian","PROT_09_SUMMON",
     "THE GUARDIAN OF VALLEY FORGE STIRS",
     "Followup","CL_001 BTN3 (trigger_event)","→ PROT_09_TAME",
     "FIRST PASS","valley_forge_guardian","Not Started","NO","","0",
     "Only fires via NYC liberation button; not staggered-turn triggered"),
    ("Protector","Valley Forge Guardian","PROT_09_TAME",
     "VALLEY FORGE SPIRITS SETTLE — CARLISLE'S ARMY WELCOMED",
     "Branch","PROT_09_SUMMON","→ PROT_09_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_09_tame flag; immediately chains to AGREE"),
    ("Protector","Valley Forge Guardian","PROT_09_AGREE",
     "VALLEY FORGE GUARDIAN FORMALLY BLESSES THE CONTINENTAL ARMY",
     "Followup","PROT_09_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_09_agreed; BTN2 explicit gated"),

    # Snallygaster
    ("Protector","Snallygaster","PROT_10_SUMMON",
     "THE SNALLYGASTER DESCENDS ON MARYLAND",
     "Root","—","→ PROT_10_TAME",
     "FIRST PASS","snallygaster","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 55+"),
    ("Protector","Snallygaster","PROT_10_TAME",
     "SNALLYGASTER REDIRECTS — CONTINENTAL SUPPLY ROUTES NOW CLEAR",
     "Branch","PROT_10_SUMMON","→ PROT_10_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_10_tame flag; immediately chains to AGREE"),
    ("Protector","Snallygaster","PROT_10_AGREE",
     "SNALLYGASTER FORMALLY GUARDS THE CONTINENTAL SUPPLY NETWORK",
     "Followup","PROT_10_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_10_agreed; BTN2 explicit gated"),

    # Paul Revere
    ("Protector","Paul Revere","PROT_11_SUMMON",
     "PAUL REVERE RIDES AGAIN",
     "Root","—","→ PROT_11_TAME",
     "FIRST PASS","paul_revere","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 60+"),
    ("Protector","Paul Revere","PROT_11_TAME",
     "PAUL REVERE INTEGRATES WITH CONTINENTAL COMMAND — MIDNIGHT RIDES COORDINATED",
     "Branch","PROT_11_SUMMON","→ PROT_11_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_11_tame flag; immediately chains to AGREE"),
    ("Protector","Paul Revere","PROT_11_AGREE",
     "PAUL REVERE FORMALLY INTEGRATES WITH CONTINENTAL INTELLIGENCE",
     "Followup","PROT_11_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_11_agreed; BTN2 explicit gated"),

    # Liberty Bell
    ("Protector","Liberty Bell","PROT_12_SUMMON",
     "THE LIBERTY BELL RINGS",
     "Root","—","→ PROT_12_TAME",
     "FIRST PASS","liberty_bell_ring","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 65+"),
    ("Protector","Liberty Bell","PROT_12_TAME",
     "LIBERTY BELL RINGS ON SCHEDULE — CIVILIAN PANIC CEASES",
     "Branch","PROT_12_SUMMON","→ PROT_12_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_12_tame flag; immediately chains to AGREE"),
    ("Protector","Liberty Bell","PROT_12_AGREE",
     "THE LIBERTY BELL FORMALLY RINGS FOR THE REPUBLIC",
     "Followup","PROT_12_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_12_agreed; BTN2 explicit gated"),

    # Green Mountain Ghost
    ("Protector","Green Mountain Ghost","PROT_13_SUMMON",
     "THE GREEN MOUNTAIN GHOST RIDES",
     "Root","—","→ PROT_13_TAME",
     "FIRST PASS","green_mountain_ghost","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 70+"),
    ("Protector","Green Mountain Ghost","PROT_13_TAME",
     "GREEN MOUNTAIN GHOST IDENTIFIES VERMONT MILITIA AS NON-HOSTILE",
     "Branch","PROT_13_SUMMON","→ PROT_13_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_13_tame flag; immediately chains to AGREE"),
    ("Protector","Green Mountain Ghost","PROT_13_AGREE",
     "GREEN MOUNTAIN GHOST FORMALLY PROTECTS VERMONT FOR THE REPUBLIC",
     "Followup","PROT_13_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_13_agreed; BTN2 explicit gated"),

    # Mount Rushmore
    ("Protector","Mount Rushmore","PROT_14_SUMMON",
     "MOUNT RUSHMORE AWAKENS",
     "Root","—","→ PROT_14_TAME",
     "FIRST PASS","rushmore_awakens","Not Started","YES","kinky_lewd","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 75+; kinky_lewd gated"),
    ("Protector","Mount Rushmore","PROT_14_TAME",
     "MOUNT RUSHMORE COUNCIL REACHES INTERNAL CONSENSUS — DEBATES FORMALIZED",
     "Branch","PROT_14_SUMMON","→ PROT_14_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_14_tame flag; immediately chains to AGREE"),
    ("Protector","Mount Rushmore","PROT_14_AGREE",
     "THE RUSHMORE COUNCIL FORMALLY ENDORSES PRESIDENT CARLISLE",
     "Followup","PROT_14_TAME","—",
     "FIRST PASS","","Not Started","YES","kinky_lewd","0",
     "Sets prot_14_agreed; BTN2 kinky_lewd gated"),

    # Skunk Ape
    ("Protector","Skunk Ape","PROT_15_SUMMON",
     "THE SKUNK APE EMERGES FROM FLORIDA",
     "Root","—","→ PROT_15_TAME",
     "FIRST PASS","skunk_ape","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 80+"),
    ("Protector","Skunk Ape","PROT_15_TAME",
     "SKUNK APE WITHDRAWS FROM ACTIVE DISRUPTION ZONES — SOUTHERN OPERATIONS STABLE",
     "Branch","PROT_15_SUMMON","→ PROT_15_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_15_tame flag; immediately chains to AGREE"),
    ("Protector","Skunk Ape","PROT_15_AGREE",
     "SKUNK APE FORMALLY PATROLS FLORIDA FOR THE CONTINENTAL REPUBLIC",
     "Followup","PROT_15_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_15_agreed; BTN2 explicit gated"),

    # Minuteman
    ("Protector","Minuteman","PROT_16_SUMMON",
     "THE MINUTEMAN IS READY",
     "Root","—","→ PROT_16_TAME",
     "FIRST PASS","eternal_minuteman","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 85+"),
    ("Protector","Minuteman","PROT_16_TAME",
     "THE MINUTEMAN ACKNOWLEDGES CONTINENTAL FORCES — LEXINGTON GREEN ACCESS RESTORED",
     "Branch","PROT_16_SUMMON","→ PROT_16_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_16_tame flag; immediately chains to AGREE"),
    ("Protector","Minuteman","PROT_16_AGREE",
     "THE MINUTEMAN FORMALLY DECLARES FOR THE CONTINENTAL REPUBLIC",
     "Followup","PROT_16_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_16_agreed; BTN2 explicit gated"),

    # Lincoln's Ghost
    ("Protector","Lincoln's Ghost","PROT_17_SUMMON",
     "LINCOLN'S GHOST APPEARS IN WASHINGTON",
     "Root","—","→ PROT_17_TAME",
     "FIRST PASS","lincolns_ghost","Not Started","NO","","0",
     "Wild: +1 corruption/turn 25% chance; fires turn 90+"),
    ("Protector","Lincoln's Ghost","PROT_17_TAME",
     "LINCOLN'S GHOST ESTABLISHES OFFICE HOURS — PRESIDENTIAL SCHEDULE RESTORED",
     "Branch","PROT_17_SUMMON","→ PROT_17_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets prot_17_tame flag; immediately chains to AGREE"),
    ("Protector","Lincoln's Ghost","PROT_17_AGREE",
     "LINCOLN FORMALLY ENDORSES PRESIDENT CARLISLE'S ADMINISTRATION",
     "Followup","PROT_17_TAME","—",
     "FIRST PASS","","Not Started","YES","","0",
     "Sets prot_17_agreed; BTN2 explicit gated"),

    # ── COMMANDER ────────────────────────────────────────────────────────────
    ("Commander","Commander Arc","CMD_MERIT",
     "FIELD PROMOTION: [COMMANDER_NAME] DISTINGUISHED IN EARLY OPERATIONS",
     "Root","—","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires at governorLevel==1 after 5 turns stationed; promote_commander BTN1; morale_boost BTN2"),

    ("Commander","Commander Arc","CMD_RECOGNITION",
     "CITATION FOR EXCEPTIONAL SERVICE: [COMMANDER_NAME] AT [TILE_NAME]",
     "Branch","CMD_MERIT","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires at governorLevel==2 after 20 turns stationed; promote_commander BTN1; claim_change BTN2"),

    ("Commander","Commander Arc","CMD_THANKS",
     "PRESIDENT CARLISLE DELIVERS PERSONAL THANKS TO [COMMANDER_NAME] AT [TILE_NAME]",
     "Followup","CMD_RECOGNITION","—",
     "FIRST PASS","","Not Started","YES","explicit","0",
     "Fires at governorLevel==3 after 50 turns stationed; BTN2 explicit morale_boost +40; non-repeatable"),

    # ── FORT CHAIN ───────────────────────────────────────────────────────────
    ("Fort Chain","Fort Disrepair","FORT_001",
     "FORT [TILE_NAME]: STRUCTURAL CONCERNS AND A PAMPHLET PROBLEM",
     "Root","—","→ FORT_003 (BTN1) | → FORT_004 (BTN2)",
     "FIRST PASS","fort_pamphlets","Not Started","NO","","0",
     "Fires at Fortress tiles; requires governor + no fortDisrepair; army station required"),

    ("Fort Chain","Fort Disrepair","FORT_003",
     "PRESIDENT CARLISLE FIRES [COMMANDER_NAME] IN FRONT OF GOD AND THE GARRISON",
     "Branch","FORT_001 BTN1","—",
     "FIRST PASS","fort_confrontation","Not Started","NO","","0",
     "Public confrontation path; removes governor; generates replacement"),

    ("Fort Chain","Fort Disrepair","FORT_004",
     "FORT [TILE_NAME] REPAIR ORDER SIGNED — GRIEVANCES RESOLVED",
     "Branch","FORT_001 BTN2","→ FORT_005",
     "FIRST PASS","fort_handshake","Not Started","NO","","0",
     "Private resolution path; repair_fort outcome; chains to FORT_005"),

    ("Fort Chain","Fort Disrepair","FORT_005",
     "CLASSIFIED DIPLOMATIC PROCEEDINGS AT [TILE_NAME]",
     "Followup","FORT_004","—",
     "FIRST PASS","fort_diplomatic","Not Started","YES","explicit","0",
     "Explicit reward event; BTN1 explicit morale+30 / BTN2 standard morale+10"),

    # ── COLLAPSE ─────────────────────────────────────────────────────────────
    ("Collapse","Republic Falls","COLLAPSE_01",
     "THE REPUBLIC HAS FALLEN",
     "Root","—","→ COLLAPSE_02",
     "FIRST PASS","collapse_republic","Not Started","NO","","0",
     "trigger_collapse outcome; fires when all metros fall"),

    ("Collapse","Republic Falls","COLLAPSE_02",
     "WASHINGTON STANDS ALONE",
     "Followup","COLLAPSE_01","—",
     "FIRST PASS","ualani_washington","Not Started","YES","explicit option","0",
     "BTN1 explicit 'The Long Night' morale+30 / BTN2 standard 'Hold the Line' morale+10"),

    # ── TILE CRISIS ──────────────────────────────────────────────────────────
    ("Tile Crisis","Harvest Crisis","HARVEST_001",
     "THE ARMY ATE THE SEED CORN",
     "Root","—","→ HARVEST_VISIT_01 (BTN1 mission)",
     "FIRST PASS","harvest_crisis","Not Started","NO","","0",
     "Fires when food < 50; Farm tile required; set_mission_flag BTN1"),

    ("Tile Crisis","Harvest Crisis","HARVEST_VISIT_01",
     "PRESIDENT CARLISLE SAVES THE HARVEST",
     "Followup","HARVEST_001 BTN1 (army mission)","—",
     "FIRST PASS","harvest_complete","Not Started","YES","explicit option","0",
     "BTN2 explicit 'Presidential Persuasion' +food 80 vs BTN1 standard +food 60"),

    ("Tile Crisis","Harbor Threat","HARBOR_001",
     "CROWN VESSELS SPOTTED: DOCK AT [TILE_NAME] AT RISK",
     "Root","—","→ HARBOR_ESCORT_01 (BTN1) | → HARBOR_BURN_01 (BTN2)",
     "FIRST PASS","harbor_threat","Not Started","NO","","0",
     "Dock tile required + UK neighbor; army station required"),

    ("Tile Crisis","Harbor Threat","HARBOR_ESCORT_01",
     "HARBOR SECURED: CROWN FRIGATE RETREATS",
     "Branch","HARBOR_001 BTN1 (army mission)","—",
     "FIRST PASS","harbor_secure","Not Started","NO","","0",
     "+food 30; garrison path"),

    ("Tile Crisis","Harbor Threat","HARBOR_BURN_01",
     "DOCK FACILITIES DESTROYED: ACCESS DENIED",
     "Branch","HARBOR_001 BTN2 (army mission)","—",
     "FIRST PASS","harbor_burned","Not Started","NO","","0",
     "disable_building Dock 8 turns; destruction path"),

    ("Tile Crisis","Forge Threat","FORGE_001",
     "INTELLIGENCE ALERT: CROWN FORCES LOCATE THE FORGE AT [TILE_NAME]",
     "Root","—","→ FORGE_REINFORCE_01 (BTN1) | → FORGE_EVACUATE_01 (BTN2)",
     "FIRST PASS","forge_threat","Not Started","NO","","0",
     "Forge tile required + UK neighbor; army station required"),

    ("Tile Crisis","Forge Threat","FORGE_REINFORCE_01",
     "FORGE SECURED: CROWN ASSAULT PLAN ABANDONED",
     "Branch","FORGE_001 BTN1 (army mission)","—",
     "FIRST PASS","forge_secure","Not Started","NO","","0",
     "+weapons 30; defense path"),

    ("Tile Crisis","Forge Threat","FORGE_EVACUATE_01",
     "EMERGENCY EXTRACTION: WEAPONS CACHE SECURED",
     "Branch","FORGE_001 BTN2 (army mission)","—",
     "FIRST PASS","forge_evac","Not Started","NO","","0",
     "+weapons 25; Forge temporarily offline"),

    ("Tile Crisis","Corruption Crisis","CORRUPT_001",
     "MAGISTRATE [COMMANDER_NAME] UNDER INVESTIGATION",
     "Root","—","→ CORRUPT_TRIAL_01 (BTN1) | → CORRUPT_DEAL_01 (BTN2)",
     "FIRST PASS","corruption_crisis","Not Started","NO","","0",
     "Fires when tileMoralDecay >= 50 on Courthouse tile; army station required"),

    ("Tile Crisis","Corruption Crisis","CORRUPT_TRIAL_01",
     "MAGISTRATE REMOVED IN DISGRACE",
     "Branch","CORRUPT_001 BTN1 (army mission)","→ CORRUPT_REFORM_01",
     "FIRST PASS","corrupt_trial","Not Started","NO","","0",
     "remove_governor; chains to CORRUPT_REFORM_01"),

    ("Tile Crisis","Corruption Crisis","CORRUPT_REFORM_01",
     "COURTHOUSE REFORMS UNDERWAY AT [TILE_NAME]",
     "Followup","CORRUPT_TRIAL_01","—",
     "FIRST PASS","courthouse_reform","Not Started","NO","","0",
     "tile_moral_decay_change -50"),

    ("Tile Crisis","Corruption Crisis","CORRUPT_DEAL_01",
     "ARRANGEMENT REACHED AT [TILE_NAME]",
     "Branch","CORRUPT_001 BTN2 (army mission)","—",
     "FIRST PASS","corrupt_deal","Not Started","NO","","0",
     "+gold 40; private deal path; magistrate stays"),

    ("Tile Crisis","Border Dispute","BORDER_001",
     "CANADIAN BORDER INCIDENT AT [TILE_NAME]: THE LINE IS IN QUESTION",
     "Root","—","→ BORDER_ASSERT_01 (BTN1) | → BORDER_SURVEY_01 (BTN3)",
     "FIRST PASS","border_dispute","Not Started","NO","","0",
     "Woods/Foothills tile + CA neighbor; BTN2 claim_change -1 (diplomacy)"),

    ("Tile Crisis","Border Dispute","BORDER_ASSERT_01",
     "FEDERAL BOUNDARY MARKER PLANTED AT [TILE_NAME]",
     "Branch","BORDER_001 BTN1 (army mission)","—",
     "FIRST PASS","border_assert","Not Started","NO","","0",
     "+claim 2; garrison asserts sovereignty"),

    ("Tile Crisis","Border Dispute","BORDER_SURVEY_01",
     "SURVEY COMPLETE AT [TILE_NAME]: THE LAND IS OURS",
     "Branch","BORDER_001 BTN3 (army mission)","—",
     "FIRST PASS","border_survey","Not Started","NO","","0",
     "+gold 20; diplomatic survey path"),

    ("Tile Crisis","Garrison Hunger","GARRISON_001",
     "FIELD REPORT: [TILE_NAME] GARRISON REPORTING CRITICAL SUPPLY SHORTFALL",
     "Root","—","—",
     "FIRST PASS","garrison_starving","Not Started","NO","","0",
     "Fires when any USA army < 50 manpower; BTN1 spend -30 gold / BTN2 retreat"),

    ("Tile Crisis","Legitimacy Crisis","CRISIS_CLAIM_001",
     "LEGITIMACY CRISIS: THE REPUBLIC'S CLAIM ON ITSELF IS IN QUESTION",
     "Root","—","—",
     "FIRST PASS","legitimacy_crisis","Not Started","NO","","0",
     "Fires when presidentialClaim <= -5; 3 claim_change response buttons"),

    ("Tile Crisis","Turncoat General","TURNCOAT_001",
     "INTELLIGENCE REPORT: COMMANDER [COMMANDER_NAME] IN CONTACT WITH CROWN",
     "Root","—","—",
     "FIRST PASS","turncoat_report","Not Started","NO","","0",
     "Fires when any governor loyalty <= -8; remove / double-agent / watch"),

    ("Tile Crisis","Memorial Lost","MEMORIAL_001",
     "[TILE_NAME] FALLS TO THE CROWN: HISTORICAL SITE UNDER OCCUPATION",
     "Root","—","→ MEMORIAL_RESTORE_01 (BTN1 own-mission)",
     "FIRST PASS","memorial_lost","Not Started","NO","","0",
     "Fires when UK captures tile with tileSpecialFeatures; set_mission_flag_own"),

    ("Tile Crisis","Memorial Lost","MEMORIAL_RESTORE_01",
     "[TILE_NAME] LIBERATED: THE MEMORIAL IS OURS AGAIN",
     "Followup","MEMORIAL_001 BTN1 (own-mission)","—",
     "FIRST PASS","memorial_restored","Not Started","NO","","0",
     "Fires when USA retakes the special-feature tile; +morale 20"),

    # ── UALANI ───────────────────────────────────────────────────────────────
    ("Ualani","Ambush","UALANI_AMBUSH_01",
     "CROWN AMBUSH AVERTED: PRESIDENT CARLISLE HOLDS THE WETLANDS",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","15",
     "Ualani in Wetlands + UK neighbor; BTN3 explicit +weapons 40"),

    ("Ualani","Dignitary Reception","UALANI_DIGNITARY_01",
     "STATE RECEPTION AT [TILE_NAME]",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","18",
     "Ualani in Metro/Suburbs + Courthouse + tileMoralDecay < 30; BTN3 explicit -20 moralDecay"),

    ("Ualani","Memorial Address","UALANI_MEMORIAL_01",
     "PRESIDENT CARLISLE VISITS [TILE_NAME]: PRESIDENTIAL ADDRESS AT THE MEMORIAL",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","20",
     "Ualani at tile with tileSpecialFeatures; BTN3 explicit +morale 50"),

    ("Ualani","Field Hospital","UALANI_WOUNDED_01",
     "PRESIDENT CARLISLE VISITS FIELD HOSPITAL AT [TILE_NAME]",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","12",
     "Ualani army < 80% manpower; BTN2 explicit +manpower 40"),

    ("Ualani","Winter March","UALANI_WINTER_01",
     "PRESIDENT CARLISLE MARCHES THROUGH [TILE_NAME] IN THE DEAD OF WINTER",
     "Root","—","—",
     "FIRST PASS","","Not Started","NO","","20",
     "Ualani in Foothills/Woods; winter months [11,12,1,2]; winterScore > 0; no explicit"),

    ("Ualani","Forge Inspection","UALANI_FORGE_01",
     "SURPRISE INSPECTION OF FORGE AT [TILE_NAME]",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","15",
     "Ualani at tile with enabled Forge; BTN3 explicit +weapons 60"),

    ("Ualani","Culper Ring","UALANI_CULPER_01",
     "CULPER RING CONTACT AT [TILE_NAME]",
     "Root","—","→ UALANI_CULPER_BRIEF_01 (BTN1/BTN3 next_event_id)",
     "FIRST PASS","","Not Started","YES","explicit option","18",
     "Ualani neighbors tile with espionageActive; BTN3 explicit +gold 40 then brief"),

    ("Ualani","Culper Ring","UALANI_CULPER_BRIEF_01",
     "CULPER RING INTELLIGENCE RECEIVED: CROWN PLANS CONFIRMED",
     "Followup","UALANI_CULPER_01 BTN1 / BTN3","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets culper_active flag; single button; no explicit"),

    ("Ualani","Alliance Council","UALANI_ALLIANCE_01",
     "PRESIDENT CARLISLE MEETS WITH [COMMANDER_NAME] AT [TILE_NAME]",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","18",
     "Fires at NEIGHBOR tile (named hist. governor); BTN3 explicit governor_loyalty +12"),

    ("Ualani","Frontier","UALANI_FRONTIER_01",
     "PRESIDENT CARLISLE ESTABLISHES FRONTIER PRESENCE AT [TILE_NAME]",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","20",
     "Ualani in Woods/Foothills bordering CA - * tile; BTN3 explicit +morale 30"),

    # ── ELECTION ─────────────────────────────────────────────────────────────
    ("Tile Crisis","Election Arc","ELECTION_SEASON",
     "UALANI'S TERM IS ENDING — THE REPUBLIC MUST CHOOSE A SUCCESSOR",
     "Root","—","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires once turns 93-95; presents 5 faction endorsement buttons; ELECTION_BTN1 gated by !vp_declined_candidacy"),

    ("Tile Crisis","Election Arc","STUMP_SPEECH_01",
     "PRESIDENT CARLISLE TAKES THE FLOOR AT [TILE_NAME] — THE CAMPAIGN IS HERE",
     "Root","—","—",
     "FIRST PASS","","Not Started","YES","explicit option","2",
     "Ualani at Courthouse turns 90-100; BTN1 +15 pressure / BTN2 explicit +25 pressure"),

    ("Tile Crisis","Election Arc","ELECTION_NIGHT_WIN",
     "THE REPUBLIC ENDURES: UALANI'S COALITION CARRIES THE ELECTION",
     "Followup","turn 100 (pressure > 0)","—",
     "FIRST PASS","","Not Started","NO","","0",
     "End-game WIN; fires when sum(electionPressure) > 0 at turn 100"),

    ("Tile Crisis","Election Arc","ELECTION_NIGHT_LOSE",
     "WELL. THIS IS SOMETHING: KING GEORGE III WINS THE AMERICAN POPULAR VOTE",
     "Followup","turn 100 (pressure ≤ 0)","—",
     "FIRST PASS","","Not Started","NO","","0",
     "End-game LOSE; fires when sum(electionPressure) ≤ 0 at turn 100"),

    # ── VICE PRESIDENT ───────────────────────────────────────────────────────
    ("VP Arc","VP Arc","VP_FIRST_MEETING",
     "THE VICE PRESIDENT REQUESTS A WORD: [COMMANDER_NAME] MAKES THEIR POSITION CLEAR",
     "Root","—","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires turn 5+, one-time; sets vp_met flag; gates all other VP events"),

    ("VP Arc","VP Arc","VP_COUNSEL",
     "THE VICE PRESIDENT ARRIVES LATE WITH GOOD ADVICE: [COMMANDER_NAME] READS THE SITUATION",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires when presidentialClaim < -2; BTN1 morale +10 / BTN2 loyalty vp_faction +10"),

    ("VP Arc","VP Arc","VP_DOUBT",
     "THE VICE PRESIDENT AT 2 A.M.: [COMMANDER_NAME] IS QUESTIONING EVERYTHING",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","NO","sensual option","0",
     "Fires when VP tile moralDecay >= 30; BTN2 sensual morale +25"),

    ("VP Arc","VP Arc","VP_LOYALTY_TEST",
     "[COMMANDER_NAME]'S COALITION WANTS ANSWERS — THE FACTION DEMANDS A PUBLIC STATEMENT",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires when VP faction loyalty < 20; BTN1 vp_faction +15 / BTN2 vp_faction +5"),

    ("VP Arc","VP Arc","VP_BATTLEFIELD",
     "THE VICE PRESIDENT PICKS UP A MUSKET: [COMMANDER_NAME] AT THE FRONT LINE",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Fires when VP tile has UK neighbor + stationed army; BTN2 explicit morale +30"),

    ("VP Arc","VP Arc","VP_PRE_ELECTION",
     "THE VICE PRESIDENT CALLS AN UNSCHEDULED MEETING — [COMMANDER_NAME] ISN'T SURE THEY WANT TO RUN",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Fires turns 88-92; BTN1 sets vp_declined_candidacy (hides ELECTION_BTN1); BTN2 explicit morale +20"),

    ("VP Arc","VP Arc","VP_SACRIFICE",
     "THE VICE PRESIDENT OFFERS THEIR RESIGNATION — [COMMANDER_NAME] WANTS TO PROTECT THE REPUBLIC",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires when VP tile electionPressure < -20; BTN1 sets vp_resigned / BTN2 vp_faction +10"),

    ("VP Arc","VP Arc","VP_SOLIDARITY",
     "THE VICE PRESIDENT HAS BEEN PAYING ATTENTION — [COMMANDER_NAME] MAKES A SPECIFIC OBSERVATION",
     "Branch","VP_FIRST_MEETING (vp_met flag)","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires when agreed_protectors >= 3; BTN1 morale +15 / BTN2 vp_faction +10"),

    ("Canada","War Arc","UK_BUILDUP_01",
     "INTELLIGENCE REPORT: CROWN FORCES AMASSING ON ALL BORDERS",
     "Root","—","→ CAN_CALL_01 (uk_buildup_known flag)",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires turn 7, one-time; BTN1 alert / BTN2 contact Ottawa; both set uk_buildup_known"),

    ("Canada","War Arc","UK_DECLARATION_01",
     "KING GEORGE III DECLARES WAR: THE CROWN HAS MOVED ON BOTH FRONTS",
     "Branch","UK_BUILDUP_01","—",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires turns 10-16 (30%/turn) once uk_buildup_known; sets uk_declared_war"),

    ("Canada","Canadian Alliance","CAN_CALL_01",
     "THE OTTAWA CALL: PRESIDENT PENOIT AND PM CLEAR-WATER ARE ON THE LINE",
     "Root","UK_BUILDUP_01 (flag)","→ CAN_PENOIT_01 (BTN1) or END (BTN2)",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires turn 8+ after uk_buildup_known; BTN1 sets can_contact / BTN2 sets can_rejected (ends arc)"),

    ("Canada","Canadian Alliance","CAN_PENOIT_01",
     "PRESIDENT PENOIT'S CONDITIONS: MARK PENOIT PUTS IT IN WRITING",
     "Branch","CAN_CALL_01 (can_contact)","→ CAN_CLEARWATER_01",
     "FIRST PASS","","Not Started","NO","","0",
     "Fires turn 12+ after can_contact; BTN1 can_penoit_agreed / BTN2 can_penoit_negotiating"),

    ("Canada","Canadian Alliance","CAN_CLEARWATER_01",
     "PM CLEAR-WATER'S PRIVATE CHANNEL: JESSICA WRITES OUTSIDE OFFICIAL CHANNELS",
     "Branch","CAN_PENOIT_01","→ CAN_JOINT_OPS_01",
     "FIRST PASS","","Not Started","NO","sensual option","0",
     "After penoit path + can_contact; BTN1 can_clearwater_warm / BTN2 sensual can_clearwater_close"),

    ("Canada","Canadian Alliance","CAN_JOINT_OPS_01",
     "JOINT INTELLIGENCE SUMMIT: THREE GOVERNMENTS SHARE ONE MAP",
     "Branch","CAN_CLEARWATER_01","→ CAN_SUMMIT_01",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "After penoit+clearwater paths; BTN1 +weapons 20 / BTN2 explicit +morale 25"),

    ("Canada","Canadian Alliance","CAN_SUMMIT_01",
     "THE OTTAWA SUMMIT: ALL THREE IN THE SAME ROOM FOR THE FIRST TIME",
     "Branch","CAN_JOINT_OPS_01","→ CAN_ALLIANCE_SIGNED",
     "FIRST PASS","","Not Started","YES","explicit","0",
     "EXPLICIT #1 (all three); turn 13+, penoit_agreed+clearwater; BTN1/BTN2 set can_summit_complete"),

    ("Canada","Canadian Alliance","CAN_ALLIANCE_SIGNED",
     "THE CONTINENTAL DEFENSE ALLIANCE IS SIGNED: THREE NATIONS — ONE FRONT",
     "Branch","CAN_SUMMIT_01","→ CAN_PEACE_01",
     "FIRST PASS","","Not Started","YES","explicit","0",
     "EXPLICIT #2 (all three); fires after can_summit_complete; form_alliance CA outcome"),

    ("Canada","Canadian Alliance","CAN_PEACE_01",
     "THE TIDE HAS TURNED: PENOIT AND CLEAR-WATER COME TO WASHINGTON",
     "Followup","CAN_ALLIANCE_SIGNED (can_allied flag)","—",
     "FIRST PASS","","Not Started","YES","explicit","0",
     "EXPLICIT #3 (all three); fires when can_allied + UK tiles ≤ 20 + turn ≥ 60; BTN2 +morale 40"),

    ("Canada","Canadian Alliance","CAN_ELECTION_LUCK",
     "PM CLEAR-WATER'S CALL: JESSICA HAS SOMETHING TO SAY BEFORE THE ELECTION",
     "Followup","CAN_PEACE_01 (can_allied)","—",
     "FIRST PASS","","Not Started","NO","sensual option","0",
     "End-game; fires turn 93+ if can_allied; BTN1 +pressure 10 / BTN2 sensual +pressure 15"),

    # ── CANADIAN PROTECTORS ──────────────────────────────────────────────────
    # 8 creatures from Algonquin, Mi'kmaq, and French-Canadian folklore.
    # Fire as dispatches from Jessica Clear-Water to Ualani Carlisle.
    # Turns 8–43, staggered at 5-turn intervals.

    # Le Wendigo
    ("CA Protector","Le Wendigo","CA_PROT_01_SUMMON",
     "LE WENDIGO WAKES IN THE LAURENTIAN FOREST",
     "Root","—","→ CA_PROT_01_TAME",
     "FIRST PASS","wendigo_approach","Not Started","NO","","0",
     "Wild: +1 corruption/turn (25% chance) to CA tiles; fires turn 8+; Saint-Georges tile 99"),
    ("CA Protector","Le Wendigo","CA_PROT_01_TAME",
     "THE WENDIGO RECEDES FROM THE HARVEST CORRIDOR",
     "Branch","CA_PROT_01_SUMMON","→ CA_PROT_01_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_01_tame flag; immediately chains to AGREE"),
    ("CA Protector","Le Wendigo","CA_PROT_01_AGREE",
     "LE WENDIGO STANDS DOWN: SAINT-GEORGES FOREST ACCORD",
     "Followup","CA_PROT_01_TAME","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Sets ca_prot_01_agreed; BTN2 explicit — private dispatch from Jessica"),

    # Le Loup-Garou
    ("CA Protector","Le Loup-Garou","CA_PROT_02_SUMMON",
     "LE LOUP-GAROU RUNS THE RIVER ROAD — RIVIÈRE-DU-LOUP QUARANTINED",
     "Root","—","→ CA_PROT_02_TAME",
     "FIRST PASS","loup_garou","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 13+; Rivière-du-Loup tile 105"),
    ("CA Protector","Le Loup-Garou","CA_PROT_02_TAME",
     "LE LOUP-GAROU ALLOWS PASSAGE — THE RIVER ROAD REOPENS",
     "Branch","CA_PROT_02_SUMMON","→ CA_PROT_02_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_02_tame flag; immediately chains to AGREE"),
    ("CA Protector","Le Loup-Garou","CA_PROT_02_AGREE",
     "LE LOUP-GAROU OF RIVIÈRE-DU-LOUP: FORMALLY ACKNOWLEDGED BY THE REPUBLIC",
     "Followup","CA_PROT_02_TAME","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Sets ca_prot_02_agreed; BTN2 explicit — the wolf stops running"),

    # Les Feux Follets
    ("CA Protector","Les Feux Follets","CA_PROT_03_SUMMON",
     "LES FEUX FOLLETS RISE OVER THE SAINT JOHN MARSHES — NAVIGATION DISRUPTED",
     "Root","—","→ CA_PROT_03_TAME",
     "FIRST PASS","feux_follets","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 18+; Saint John tile 194"),
    ("CA Protector","Les Feux Follets","CA_PROT_03_TAME",
     "LES FEUX FOLLETS QUIET — MALISEET PROTOCOLS ESTABLISHED",
     "Branch","CA_PROT_03_SUMMON","→ CA_PROT_03_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_03_tame flag; immediately chains to AGREE"),
    ("CA Protector","Les Feux Follets","CA_PROT_03_AGREE",
     "LES FEUX FOLLETS FORMALLY ACKNOWLEDGED: THE SAINT JOHN MARSH ACCORD",
     "Followup","CA_PROT_03_TAME","—",
     "FIRST PASS","","Not Started","NO","sensual option","0",
     "Sets ca_prot_03_agreed; BTN2 sensual — Jessica walks you through the marsh"),

    # Mishepeshu
    ("CA Protector","Mishepeshu","CA_PROT_04_SUMMON",
     "MISHEPESHU STIRS BENEATH LAKE SIMCOE — CROSSINGS CONTESTED",
     "Root","—","→ CA_PROT_04_TAME",
     "FIRST PASS","mishepeshu","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 23+; Barrie tile 131"),
    ("CA Protector","Mishepeshu","CA_PROT_04_TAME",
     "MISHEPESHU WITHDRAWS FROM THE CROSSING LANES — LAKE SIMCOE NAVIGABLE",
     "Branch","CA_PROT_04_SUMMON","→ CA_PROT_04_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_04_tame flag; immediately chains to AGREE"),
    ("CA Protector","Mishepeshu","CA_PROT_04_AGREE",
     "MISHEPESHU PACT FORMALIZED: THE GREAT LYNX HOLDS THE ONTARIO WATERWAYS",
     "Followup","CA_PROT_04_TAME","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Sets ca_prot_04_agreed; BTN2 explicit — why the Republic had to apologize first"),

    # La Corriveau
    ("CA Protector","La Corriveau","CA_PROT_05_SUMMON",
     "LA CORRIVEAU IS FREE — THE IRON CAGE HAS BEEN SWINGING IN TROIS-PISTOLES",
     "Root","—","→ CA_PROT_05_TAME",
     "FIRST PASS","la_corriveau","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 28+; Trois-Pistoles tile 106"),
    ("CA Protector","La Corriveau","CA_PROT_05_TAME",
     "LA CORRIVEAU QUIETS — THE REPUBLIC ACKNOWLEDGES THE VERDICT WAS WRONG",
     "Branch","CA_PROT_05_SUMMON","→ CA_PROT_05_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_05_tame flag; immediately chains to AGREE"),
    ("CA Protector","La Corriveau","CA_PROT_05_AGREE",
     "LA CORRIVEAU FORMALLY ALLIES WITH THE REPUBLIC — THREE GARRISONS EVACUATED THIS WEEK",
     "Followup","CA_PROT_05_TAME","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Sets ca_prot_05_agreed; BTN2 explicit — what Corriveau actually wanted to say"),

    # Le Carcajou
    ("CA Protector","Le Carcajou","CA_PROT_06_SUMMON",
     "LE CARCAJOU MOVES THROUGH THE MONCTON FOREST — NOTHING STOPS IT",
     "Root","—","→ CA_PROT_06_TAME",
     "FIRST PASS","carcajou","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 33+; Moncton tile 193"),
    ("CA Protector","Le Carcajou","CA_PROT_06_TAME",
     "LE CARCAJOU REDIRECTS — CONTINENTAL LINES EXEMPTED FROM DESTRUCTION",
     "Branch","CA_PROT_06_SUMMON","→ CA_PROT_06_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_06_tame flag; immediately chains to AGREE"),
    ("CA Protector","Le Carcajou","CA_PROT_06_AGREE",
     "LE CARCAJOU FORMALLY RECOGNIZED: THE MONCTON CORRIDOR ACCORD",
     "Followup","CA_PROT_06_TAME","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Sets ca_prot_06_agreed; BTN2 explicit — why the Carcajou accepted the accord"),

    # La Chasse-Galerie
    ("CA Protector","La Chasse-Galerie","CA_PROT_07_SUMMON",
     "THE FLYING CANOE RUNS THE OTTAWA RIVER — THREE SIGHTINGS, TWO MISSING PATROLS",
     "Root","—","→ CA_PROT_07_TAME",
     "FIRST PASS","chasse_galerie","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 38+; Deep River tile 128"),
    ("CA Protector","La Chasse-Galerie","CA_PROT_07_TAME",
     "THE FLYING CANOE ACKNOWLEDGES THE REPUBLIC — THE OTTAWA RIVER CORRIDOR SECURED",
     "Branch","CA_PROT_07_SUMMON","→ CA_PROT_07_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_07_tame flag; immediately chains to AGREE"),
    ("CA Protector","La Chasse-Galerie","CA_PROT_07_AGREE",
     "LA CHASSE-GALERIE: THE RIVER RUNS FOR THE REPUBLIC NOW",
     "Followup","CA_PROT_07_TAME","—",
     "FIRST PASS","","Not Started","NO","sensual option","0",
     "Sets ca_prot_07_agreed; BTN2 sensual — Mark Penoit on the river"),

    # Le Gougou
    ("CA Protector","Le Gougou","CA_PROT_08_SUMMON",
     "THE GOUGOU RISES IN THE CHALEUR BAY — CHAMPLAIN'S MONSTER IS REAL",
     "Root","—","→ CA_PROT_08_TAME",
     "FIRST PASS","le_gougou","Not Started","NO","","0",
     "Wild: +1 corruption/turn to CA tiles; fires turn 43+; Bathurst tile 109"),
    ("CA Protector","Le Gougou","CA_PROT_08_TAME",
     "THE GOUGOU WITHDRAWS FROM THE BAY LANES — RESTITUTION PROTOCOLS ACTIVE",
     "Branch","CA_PROT_08_SUMMON","→ CA_PROT_08_AGREE (next_event_id)",
     "FIRST PASS","","Not Started","NO","","0",
     "Sets ca_prot_08_tame flag; immediately chains to AGREE"),
    ("CA Protector","Le Gougou","CA_PROT_08_AGREE",
     "THE GOUGOU FORMALLY GUARDS THE CHALEUR BAY FOR THE CONTINENTAL ALLIANCE",
     "Followup","CA_PROT_08_TAME","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Sets ca_prot_08_agreed; BTN2 explicit — the Gougou's real name"),

    ("VP Arc","VP Arc","VP_LEGACY",
     "THE VICE PRESIDENT LEAVES A NOTE — [COMMANDER_NAME] HAS SOMETHING TO SAY",
     "Followup","VP_FIRST_MEETING (turn 96+)","—",
     "FIRST PASS","","Not Started","YES","explicit option","0",
     "Fires turn 96+; BTN1 claim +1 / BTN2 explicit morale +30"),

    # ── IDEAS ────────────────────────────────────────────────────────────────
    ("Idea","Ualani Expansion","UALANI_DOCK_01",
     "PRESIDENT CARLISLE SURVEYS THE HARBOR AT [TILE_NAME]  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","YES","explicit option","15",
     "Ualani at tile with enabled Dock; follows UALANI_FORGE_01 pattern"),

    ("Idea","Ualani Expansion","UALANI_FARM_01",
     "PRESIDENT CARLISLE VISITS THE FARMLANDS AT [TILE_NAME]  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","YES","explicit option","15",
     "Ualani at Farm tile; variant of HARVEST_VISIT_01 but Ualani-led"),

    ("Idea","Ualani Expansion","UALANI_BARRACKS_01",
     "PRESIDENT CARLISLE CONDUCTS REGIMENTAL REVIEW AT [TILE_NAME]  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","YES","explicit option","15",
     "Ualani at tile with Barracks; army strength / loyalty focus"),

    ("Idea","Ualani Expansion","UALANI_COURTHOUSE_01",
     "PRESIDENT CARLISLE HOLDS TOWN HALL AT [TILE_NAME]  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","YES","explicit option","15",
     "Ualani at Courthouse tile; complement to UALANI_DIGNITARY but different trigger (no moralDecay gate)"),

    ("Idea","More City Events","CL_WASHINGTON_01",
     "WASHINGTON DC LIBERATED  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","NO","","0",
     "Would fire when tile 188 is recaptured; natural counterpart to CX_001"),

    ("Idea","More City Events","CL_VALLEY_FORGE_01",
     "VALLEY FORGE SECURED  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","NO","","0",
     "Tile 1; historically significant; would fire on liberation"),

    ("Idea","More City Events","CX_NYC_01",
     "NEW YORK CITY FALLS AGAIN  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","NO","","0",
     "Would fire if NYC is re-taken after CL_001; repeatable pair"),

    ("Idea","Espionage Events","ESPIONAGE_DISCOVERY_01",
     "SPY NETWORK DISCOVERED AT [TILE_NAME]  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","NO","","0",
     "Would fire when espionageActive + enemy army arrives; uses existing espionage vars"),

    ("Idea","Commander Expansion","CMD_OBJECTIVE_2",
     "A COMMANDER'S SECOND TASK  ← IDEA",
     "Followup","CMD_OBJECTIVE_1","→ CMD_OBJECTIVE_3?",
     "IDEA","","Not Started","NO","","0",
     "Extend the commander arc beyond first objective"),

    ("Idea","Seasonal Events","WINTER_SIEGE_01",
     "VALLEY FORGE CONDITIONS: ARMY SUFFERING IN THE FIELD  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","NO","","0",
     "Generic winter hardship event; non-Ualani; uses winterScore + month check"),

    ("Idea","Seasonal Events","SPRING_OFFENSIVE_01",
     "THE THAW: CROWN FORCES RESUME OFFENSIVE OPERATIONS  ← IDEA",
     "Root","—","—",
     "IDEA","","Not Started","NO","","0",
     "Month 3-4; UK neighbor + player army present; tactical pressure event"),
]


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_font(bold=False, color=NORMAL_FG, size=10) -> Font:
    return Font(bold=bold, color=color, name="Calibri", size=size)


def apply_header(ws):
    ws.row_dimensions[1].height = 32
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill    = make_fill(HEADER_BG)
        cell.font    = make_font(bold=True, color=HEADER_FG, size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center",
            wrap_text=True
        )
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_row(ws, row_num: int, event: tuple):
    (cat, chain, eid, headline,
     chain_pos, parent, children,
     status, art_tag, art_status, explicit_art, content_flag, cooldown, notes) = event

    values = [
        cat, chain, eid, headline,
        chain_pos, parent, children,
        status, art_tag, art_status, explicit_art, content_flag, cooldown, notes
    ]

    cat_color    = CAT_COLORS.get(cat, "FFFFFF")
    status_bg    = STATUS_COLORS.get(status, "FFFFFF")
    status_fg    = STATUS_TEXT.get(status, NORMAL_FG)
    art_bg       = ART_COLORS.get(art_status, "FFFFFF")
    art_fg       = ART_TEXT.get(art_status, NORMAL_FG)
    is_idea_row  = (cat == "Idea" or status == "IDEA")

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.border = BORDER

        # Default styling for the row
        base_color = cat_color if not is_idea_row else CAT_COLORS["Idea"]
        cell.fill  = make_fill(base_color)
        cell.font  = make_font(color=NORMAL_FG, size=9)
        cell.alignment = Alignment(
            horizontal="left", vertical="center",
            wrap_text=(col_idx in (4, 6, 7, 14))  # headline, parent, children, notes
        )

        # Column-specific overrides
        col_header = COLUMNS[col_idx - 1][0]

        if col_header == "STATUS":
            cell.fill = make_fill(status_bg)
            cell.font = make_font(bold=True, color=status_fg, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col_header == "ART STATUS":
            cell.fill = make_fill(art_bg)
            cell.font = make_font(bold=False, color=art_fg, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col_header == "EXPLICIT ART?":
            if val == "YES":
                cell.fill = make_fill(EXPLICIT_BG)
                cell.font = make_font(bold=True, color=EXPLICIT_FG, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col_header in ("CATEGORY", "CHAIN / GROUP"):
            cell.font = make_font(bold=True, color=NORMAL_FG, size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        elif col_header == "CHAIN POSITION":
            pos_colors = {
                "Root":    ("D6EAF8", "154360"),
                "Branch":  ("D5F5E3", "145A32"),
                "Followup":("FDEBD0", "784212"),
            }
            if val in pos_colors:
                bg, fg = pos_colors[val]
                cell.fill = make_fill(bg)
                cell.font = make_font(bold=True, color=fg, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        elif col_header == "COOLDOWN":
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row_num].height = 30


def add_canadian_events_sheet(wb):
    """Dedicated tab showing all Canada + CA Protector category events."""
    ws = wb.create_sheet("Canadian Events")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    apply_header(ws)

    ca_events = [e for e in EVENTS if e[0] in ("Canada", "CA Protector")]

    prev_cat = None
    row_num = 2
    for event in ca_events:
        cat = event[0]
        if prev_cat is not None and cat != prev_cat:
            ws.row_dimensions[row_num].height = 8
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = make_fill("E8E8E8")
            row_num += 1
        write_row(ws, row_num, event)
        prev_cat = cat
        row_num += 1

    # Summary box
    row_num += 1
    summary_data = [
        ("Canadian events total",              len(ca_events)),
        ("Canada (Alliance arc)",              sum(1 for e in ca_events if e[0] == "Canada")),
        ("CA Protector (Jessica's creatures)", sum(1 for e in ca_events if e[0] == "CA Protector")),
        ("Need explicit art (YES)",            sum(1 for e in ca_events if e[10] == "YES")),
    ]
    for label, count in summary_data:
        lc = ws.cell(row=row_num, column=1, value=label)
        cc = ws.cell(row=row_num, column=2, value=count)
        lc.fill = make_fill(HEADER_BG)
        cc.fill = make_fill(HEADER_BG)
        lc.font = make_font(bold=True, color=HEADER_FG, size=10)
        cc.font = make_font(bold=True, color=HEADER_FG, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        cc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 20
        row_num += 1


def add_legend_sheet(wb):
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18

    def legend_header(r, text):
        cell = ws.cell(row=r, column=1, value=text)
        cell.fill = make_fill(HEADER_BG)
        cell.font = make_font(bold=True, color=HEADER_FG, size=11)
        ws.merge_cells(f"A{r}:C{r}")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22

    def legend_row(r, label, desc, color, fg=NORMAL_FG):
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=desc)
        c = ws.cell(row=r, column=3)
        for cell in (a, b, c):
            cell.fill   = make_fill(color)
            cell.font   = make_font(color=fg, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        a.font = make_font(bold=True, color=fg, size=10)
        ws.row_dimensions[r].height = 22

    r = 1
    legend_header(r, "STATUS LEGEND"); r += 1
    legend_row(r, "FULL COMPLETION", "Final pass done — art integrated, text finalized, fully tested", STATUS_COLORS["FULL COMPLETION"], STATUS_TEXT["FULL COMPLETION"]); r += 1
    legend_row(r, "FIRST PASS",      "Event is coded, in CSV, all buttons wired; no final polish yet", STATUS_COLORS["FIRST PASS"],      STATUS_TEXT["FIRST PASS"]);      r += 1
    legend_row(r, "IDEA",            "Concept stage: not in events.csv; some referenced in button CSV", STATUS_COLORS["IDEA"],            STATUS_TEXT["IDEA"]);            r += 1

    r += 1
    legend_header(r, "ART STATUS LEGEND"); r += 1
    legend_row(r, "Not Started", "No art asset exists yet",                         ART_COLORS["Not Started"], ART_TEXT["Not Started"]); r += 1
    legend_row(r, "In Progress", "Art is being created but not in engine",          ART_COLORS["In Progress"], ART_TEXT["In Progress"]); r += 1
    legend_row(r, "Integrated",  "Art is in the project and referenced correctly",  ART_COLORS["Integrated"],  ART_TEXT["Integrated"]);  r += 1

    r += 1
    legend_header(r, "EXPLICIT ART LEGEND"); r += 1
    legend_row(r, "YES — Explicit Art Needed",
               "Event has explicit or kinky_lewd content flag OR has at least one explicit button option that will need unique explicit artwork",
               EXPLICIT_BG, EXPLICIT_FG); r += 1
    legend_row(r, "NO", "Standard event — no adult artwork required", "F5F5F5"); r += 1

    r += 1
    legend_header(r, "CHAIN POSITION LEGEND"); r += 1
    pos_colors = {"Root": ("D6EAF8","154360"), "Branch": ("D5F5E3","145A32"), "Followup": ("FDEBD0","784212")}
    descs = {
        "Root":    "Standalone event or first event in a chain; fires from world.gd check function or EventDatabase date trigger",
        "Branch":  "Fires as outcome of a choice button on a Root/prior event; represents one path in a fork",
        "Followup":"Fires automatically after a mission completes (via next_event_id or mission flag return)",
    }
    for pos, (bg, fg) in pos_colors.items():
        legend_row(r, pos, descs[pos], bg, fg); r += 1

    r += 1
    legend_header(r, "CATEGORY COLOR LEGEND"); r += 1
    cat_descs = {
        "Date":           "Calendar-triggered events; month or turn-based",
        "City Liberated": "Fires when a specific tile is liberated (city_liberated type)",
        "City Lost":      "Fires when a specific tile falls to the enemy (city_lost type)",
        "State":          "State-level liberation, secession, and reintegration",
        "Protector":      "Mythological entity summon events (protector_summon type)",
        "Commander":      "Named commander arc events (commander_* types)",
        "Fort Chain":     "Fort disrepair investigation chain (fort_* types)",
        "Collapse":       "Republic collapse sequence",
        "Tile Crisis":    "Dynamic tile-based crisis chains; triggered by world.gd check functions",
        "Ualani":         "Ualani Carlisle army presence events; personal storyline",
        "VP Arc":         "Vice President relationship arc; 9 events tied to the VP governor",
        "Canada":         "Canadian Alliance diplomatic arc (war declaration + 8 CAN_ events)",
        "CA Protector":   "Jessica Clear-Water's 8 creature arcs; Algonquin, Mi'kmaq, French-Canadian folklore",
        "Idea":           "Not yet implemented; concept stage or referenced-but-missing events",
    }
    for cat, color in CAT_COLORS.items():
        legend_row(r, cat, cat_descs.get(cat, ""), color); r += 1


def build_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Events"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    apply_header(ws)

    # Group events by category for visual separation
    prev_cat = None
    row_num = 2
    for event in EVENTS:
        cat = event[0]
        if prev_cat is not None and cat != prev_cat:
            # Blank separator row between categories
            ws.row_dimensions[row_num].height = 8
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = make_fill("E8E8E8")
            row_num += 1
        write_row(ws, row_num, event)
        prev_cat = cat
        row_num += 1

    # Summary counts in a box below the data
    row_num += 1
    summary_data = [
        ("Total events in spreadsheet", len(EVENTS)),
        ("FIRST PASS (coded & wired)", sum(1 for e in EVENTS if e[7] == "FIRST PASS")),
        ("IDEA (not yet implemented)",  sum(1 for e in EVENTS if e[7] == "IDEA")),
        ("Need explicit art (YES)",      sum(1 for e in EVENTS if e[10] == "YES")),
    ]
    for label, count in summary_data:
        lc = ws.cell(row=row_num, column=1, value=label)
        cc = ws.cell(row=row_num, column=2, value=count)
        lc.fill = make_fill(HEADER_BG)
        cc.fill = make_fill(HEADER_BG)
        lc.font = make_font(bold=True, color=HEADER_FG, size=10)
        cc.font = make_font(bold=True, color=HEADER_FG, size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        cc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    add_canadian_events_sheet(wb)
    add_legend_sheet(wb)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Total events: {len(EVENTS)}")
    print(f"  FIRST PASS: {sum(1 for e in EVENTS if e[7] == 'FIRST PASS')}")
    print(f"  IDEA:       {sum(1 for e in EVENTS if e[7] == 'IDEA')}")
    print(f"  Explicit art needed: {sum(1 for e in EVENTS if e[10] == 'YES')}")


if __name__ == "__main__":
    build_workbook()
