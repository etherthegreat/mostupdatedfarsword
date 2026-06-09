"""
Generates army_masterdoc.xlsx — complete reference for all Farsword army/combat systems.
Run from repo root: python3 scripts/build_army_masterdoc.py

Sheets:
  1. Systems Overview    — every army subsystem with FULL PASS / FIRST PASS / FIRST DRAFT / IDEA label
  2. Mil Mods            — all 60+ mil mods (weapon-embedded, armor-embedded, commander, legacy)
  3. Weapons             — all weapon types by class with stats and embedded mods
  4. Armor & Uniforms    — all armor/uniform types with stats and embedded mods
  5. Archetypes          — all 25 procedural archetypes with assigned mil mod tiers
  6. Battle Mechanics    — combat subsystem breakdown (damage formula, shields, morale, retreat)
  7. Known Bugs          — confirmed code-level bugs flagged for fixing
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "army_masterdoc.xlsx"

# ── PALETTE ──────────────────────────────────────────────────────────────────
HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
NORMAL_FG = "1A1A1A"
BUG_BG    = "FF4444"
BUG_FG    = "FFFFFF"

STATUS_COLORS = {
    "FULL PASS":   "00B050",
    "FIRST PASS":  "92D050",
    "FIRST DRAFT": "FFEB9C",
    "IDEA":        "C9B1E8",
}
STATUS_FG = {
    "FULL PASS":   "FFFFFF",
    "FIRST PASS":  "1A3A00",
    "FIRST DRAFT": "7A5A00",
    "IDEA":        "3A1A6A",
}

CAT_COLORS = {
    "Army Core":       "D6E4F7",
    "Unit Core":       "D6F0F7",
    "Combat":          "FCE0D6",
    "Weapons":         "FFF3CC",
    "Armor":           "FDEBD0",
    "Mil Mods":        "EDD6F7",
    "Commander":       "D6E1F7",
    "Movement":        "D6F0D6",
    "Weather":         "DCDCDC",
    "Spawn":           "D6F5EA",
    "Special":         "F5E6D6",
    "Status Effects":  "FFEBEE",
}

TIER_COLORS = {
    "T1 (123)":    "D6E4F7",
    "T2 (23)":     "FFF3CC",
    "T3 (3)":      "FCE0D6",
    "Weapon":      "D6F0D6",
    "Armor":       "FDEBD0",
    "Legacy":      "DCDCDC",
    "Civilian":    "EDD6F7",
    "Resource":    "F5E6D6",
    "Storm":       "D6E8F7",
    "Cultural":    "F7EDD6",
    "Tool":        "D6F7E8",
    "Special":     "F0E6F7",
    "State Guard": "E8F5E9",
    "Protector":   "FFF8E1",
    "Negative":    "FFEBEE",
}

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def mf(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def mfont(bold=False, color=NORMAL_FG, size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def hcell(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = mf(HEADER_BG)
    c.font = mfont(bold=True, color=HEADER_FG, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def pcell(ws, row, col, value, bg=None, fg=None, bold=False, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    c.font = mfont(bold=bold, color=fg or NORMAL_FG)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = BORDER
    if bg:
        c.fill = mf(bg)
    return c

def divider(ws, row, ncols, text, bg="2F4F6F"):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row, column=ci, value=text if ci == 1 else "")
        c.fill = mf(bg)
        c.font = mfont(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 16


# ── SHEET 1: SYSTEMS OVERVIEW ─────────────────────────────────────────────────
# status, category, system, key_file(s), description, notes/gaps

SYSTEMS = [
    # ── ARMY CORE ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Army Core", "Army Class & Properties",
     "army.gd",
     "Army node with armyPunch/Block/Launch/Defence/MagicDefense/Shield; "
     "maxMovementPoints (3); armyTags[]; commander; commanderModifiers1/2/3; "
     "unitsList; parentCountry; inTile; sabotaged/reconDebuffed flags",
     "All core vars declared and used. armyCurse/armyCharm vars declared but inactive."),

    ("FULL PASS", "Army Core", "Army Survey (stat recalc)",
     "army.gd surveySelf()",
     "Rebuilds armyPunch/Block/Launch/Defence from all units; applies morale "
     "multiplier from commander (1.0–1.25); calculates resource costs; disables mods "
     "when resources run out. Called by updateArmyUI().",
     "calculateMilMods() has inverted disabled check — see Known Bugs sheet."),

    ("FULL PASS", "Army Core", "Army UI Update",
     "army.gd updateArmyUI()",
     "Chain: getUnitAttributes → commanderCheck → surveySelf → updateUnitUIs → "
     "updateCommanderUI → updateFinalTotals. Called on turn end, after battle, "
     "after promotion.",
     "promote_commander now calls updateArmyUI() on stationedArmy."),

    ("FULL PASS", "Army Core", "Turn-End Reset",
     "army.gd onTurnEnd()",
     "Restores currentMovementPoints = maxMovementPoints; ticks sabotage/recon debuff "
     "timers; calls refillManpower(armyReinforceRate) per unit; calls updateArmyUI().",
     ""),

    ("FULL PASS", "Army Core", "Resource Cost Aggregation",
     "army.gd surveySelf()",
     "Sums food/wood/gold/metal/manpower/weapons/magic/science/culture/influence/"
     "harmony/faith costs across all units. Mercantilism law adds +2 harmonyCost/unit-level.",
     "Weapon costs use old DODK weapon names (Spear/Club/Atlatl etc), not new Saber/Musket names."),

    ("FULL PASS", "Army Core", "Spy Effects (Sabotage / Recon)",
     "army.gd",
     "sabotaged bool + sabotageTimer: blocks movement/action for N turns. "
     "reconDebuffed bool + reconDebuffTimer: reduces armyDefence in battle. "
     "Both tick down on onTurnEnd().",
     "Wiring from espionage events to set these flags is partial."),

    ("FULL PASS", "Army Core", "Army Tags",
     "army.gd, army_database.gd",
     "armyTags[] populated from army_templates CSV armyMods column (pipe-delimited). "
     "Tags: Cold Weather, Naval Power, Redcoats, Pirates. "
     "Cold Weather tag exempts army from winter movement and supply drain penalties.",
     ""),

    # ── UNIT CORE ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Unit Core", "Unit Class & Stats",
     "unit.gd",
     "unitType, unitLevel (1+), unitOffensiveScore, unitDefensiveScore, "
     "unitRangedOffence, unitRangedDefence, unitMagicDefence, unitMaxShield/unitShield. "
     "unitMaxManpower = 100 × level; unitMaxWeapons = 100 × level.",
     ""),

    ("FULL PASS", "Unit Core", "Unit Attribute Calculation",
     "unit.gd getUnitAttributes() + calculateGrossValues()",
     "effectMultiplier = (manPower% + weapons%) / 2; "
     "artillery uses weapons% only. "
     "meleePenalty = 0.5 for muskets (half melee), 1.0 for sabers. "
     "offScore = level × weaponOffence × meleePenalty × effectMultiplier.",
     ""),

    ("FULL PASS", "Unit Core", "Weapon / Ore / Armor Integration",
     "unit.gd calculateWeaponsOresArmor()",
     "Pulls stats from Weapon, Ore, Armor nodes; copies their embedded milMods into "
     "unit.militaryModifierList. Shield from ore.oreMaxShield × unitLevel.",
     ""),

    ("FULL PASS", "Unit Core", "Damage Reception (takeLosses)",
     "unit.gd takeLosses(type, amount)",
     "Melee: blocked = amount × unitDefensiveScore; shield absorbs first; rest hits manpower. "
     "Ranged: same with unitRangedDefence. "
     "Magic: blocked = amount × unitMagicDefence; no shield absorption.",
     ""),

    ("FULL PASS", "Unit Core", "Manpower Refill",
     "unit.gd refillManpower(RR)",
     "Adds armyReinforceRate to currentManpower each turn, capped at max. "
     "Requires parentCountry.TotalManpower > 0 (checked in army.onTurnEnd).",
     ""),

    ("FULL PASS", "Unit Core", "Combat Effectiveness",
     "unit.gd get_combat_effectiveness()",
     "Returns 0.0–1.0. Sabers: manpower% only. Others: (manpower% + weapons%) / 2.",
     ""),

    # ── COMBAT ─────────────────────────────────────────────────────────────────
    ("FULL PASS", "Combat", "Battle Class",
     "battle.gd",
     "Instantiated by army.calculateBattle(). Snapshots both armies' manpower/shield. "
     "Calculates projected damage immediately for display. Player clicks Attack to apply.",
     ""),

    ("FULL PASS", "Combat", "Melee Damage Calculation",
     "battle.gd _calculate_melee_damage()",
     "raw_attack = attacker.armyPunch. "
     "block_ratio = defender.armyBlock / unit_count (capped 0–0.9). "
     "net_to_defender = raw_attack × (1 − block_ratio). "
     "Shield absorbs first, remainder hits manpower. "
     "Defender counter-attacks simultaneously with same formula.",
     "Saber charge manpower cost applied to attacker on top of counter damage."),

    ("FULL PASS", "Combat", "Ranged Damage Calculation",
     "battle.gd _calculate_ranged_damage()",
     "effective_launch = sum of get_effective_ranged_offence() (0 if reloading). "
     "ranged_block_ratio = defender.armyDefence / unit_count (capped 0–0.9). "
     "Defender can counter-fire if they have ready ranged units. "
     "Pure artillery attackers have no defensive counter.",
     ""),

    ("FULL PASS", "Combat", "Shield System",
     "unit.gd + battle.gd",
     "unitMaxShield = unitLevel × ore.oreMaxShield. "
     "Shield absorbs damage before manpower in both melee and ranged. "
     "restore_shield() resets shield to max at start of new engagement.",
     "Magic damage bypasses shield (no shield absorption in takeLosses magic branch)."),

    ("FULL PASS", "Combat", "Saber Charge Cost",
     "battle.gd _apply_charge_costs(), unit.gd apply_charge_cost()",
     "Each saber unit pays chargeManpowerCost × currentManpower on every melee attack. "
     "Cutlass/Cavalry Saber/Light Saber/Heavy Saber: 10% manpower cost per charge.",
     ""),

    ("FULL PASS", "Combat", "Civ-Style Army Death",
     "army.gd armyDestroyed signal, country.gd _on_army_destroyed()",
     "When manpowerInArmy ≤ 0, army emits signal armyDestroyed(self). "
     "country.gd connects handler in addArmy(): erases army from countryArmyList, "
     "nulls tile.stationedArmy, calls army.queue_free(). "
     "No retreat system — armies simply die at 0 manpower.",
     "inRetreat and retreatTarget vars still declared but no longer emitted. "
     "May clean up later."),

    ("FULL PASS", "Combat", "Morale Multiplier",
     "army.gd surveySelf()",
     "mm = 1.0 + (commander.morale / 100) × 0.25. "
     "At morale 100: ×1.25 to armyPunch and armyDefence. "
     "At morale 0: ×1.0 (no penalty). Commander morale range: 0–100.",
     "Negative morale not modeled (army just gets no bonus). "
     "governor.morale range actually −20 to +10 via update_loyalty()… "
     "the 0–100 scale in surveySelf appears to use loyalty, not a separate morale stat."),

    ("FIRST PASS", "Combat", "Weapon Mods — Embedded Effects (Stubs)",
     "weapon.gd + unit.gd calculateMilMods()",
     "VolleyFire (Brown Bess), RapidFire (Lever Repeater), AreaDamage (Field Gun+), "
     "CavalryMorale (Cavalry Saber), HeavyCharge (Heavy Saber), "
     "Siege/Mortar (Mortar): all declared, added to weaponMilMods[], "
     "but calculateMilMods() has no case for any of them. Pass stubs only.",
     "Effects all need implementation in calculateMilMods or battle.gd."),

    ("FIRST PASS", "Combat", "Armor Mods — Formation Effects (Stubs)",
     "armor.gd + unit.gd calculateMilMods()",
     "DrillFormation (Continental), LineFormation (Redcoat), "
     "ShockTroop (Heavy Infantry), Skirmish (Light Infantry), "
     "MinutemanSpirit (Minuteman): all declared as 'handled at army level' "
     "but no army-level check exists.",
     "All formation bonuses need army-level logic in battle.gd or army.gd."),

    ("FIRST PASS", "Combat", "Siege Score",
     "army.gd surveySelf(), tile.gd get_siege_difficulty()",
     "armySiegeScore = unitCount × 0.1 × tile.get_siege_difficulty(). "
     "tile.get_siege_difficulty() returns value based on tile buildings/fortifications.",
     "Siege score calculated but never read in battle.gd. "
     "No siege-phase battle type implemented."),

    ("FIRST PASS", "Combat", "Terrain Combat Bonuses (36 commander mods)",
     "mil_mod.gd, unit.gd calculateMilMods(), army.gd surveySelf()",
     "All 36 commander mil mods have cases in calculateMilMods(). "
     "surveySelf() sets unit.currentTerrain = inTile.terrain and unit.currentStorm before calling calculateMilMods(). "
     "Woodsman, Swamp Legs, Hill Runner, Street Tough, Farmhand, Guerrilla Tactics, "
     "Backcountry Rider, Frontier Marksman, Everglades Tracker, Bayou Warrior all apply terrain bonuses. "
     "Continental Line, Last Stand, Saber Drill, Marksman, Steady Line apply directly. "
     "Marine, Entrenched, Night Raider, Iron Wall etc. still flagged IDEA pending subsystems.",
     "Marine: navalTileNeighbors melee logic still needed. Entrenched: stationary turn counter still needed."),

    ("FIRST PASS", "Combat", "Marine Mechanic",
     "mil_mod.gd (marineMod), world.gd meleePressed(), path_point_button.gd navalPathPoints",
     "Army with Marine mod may melee-attack into adjacent navalPathPoints (occupied PPBs). "
     "_army_has_active_marine() checks units for enabled Marine mod. "
     "meleePressed() loops all inTile.navalPathPoints and builds melee battles on occupied ones. "
     "path_point_button.buildSelf() populates navalPathPoints from navalPathPointsEXP array. "
     "Naval Supremacy mod adds +5×unitLevel to raw_attack in _calculate_melee_damage().",
     "PPBs must be pre-wired in editor navalPathPointsEXP. No separate naval-target UI; "
     "naval attacks triggered from same melee button as land attacks."),

    ("FIRST DRAFT", "Combat", "Entrenched Mechanic",
     "mil_mod.gd (entrenchMod flag)",
     "entrenchMod bool added to MilMod class. Design: after 3 turns stationary, "
     "all units get +5 def/level; bonus lost on movement.",
     "No turn counter tracks stationary status per army. "
     "No flag is set/cleared on movement. Nothing reads entrenchMod."),

    ("IDEA", "Combat", "Zone of Control",
     "(Ghost March mod hints)",
     "Ghost March commander mod description says 'ignore enemy zone of control'. "
     "No ZoC system exists in tile.gd or world.gd.",
     "Would require: ZoC radius per army; movement penalty when entering ZoC tile."),

    ("IDEA", "Combat", "Naval Combat",
     "(Marine mod, navalTileNeighbors tile property)",
     "Marine and Naval Supremacy mods imply a naval battle system. "
     "tile.navalTileNeighbors exists. No Naval Army type or naval battle logic defined.",
     ""),

    ("IDEA", "Combat", "Formation System (Army-Level)",
     "(DrillFormation, LineFormation, VolleyFire, ShockTroop mods)",
     "Several weapon and armor mods reference formation bonuses "
     "(bonus when adjacent to same-type units, bonus to ranged in line, etc.). "
     "Completely unimplemented.",
     ""),

    ("IDEA", "Combat", "Unit Promotion / Experience",
     "army.gd (averageExperience var)",
     "averageExperience and armyLevel vars declared on Army. "
     "armyAbility (recruit=100 to veteran=115) commented with 15% veteran bonus. "
     "No experience gain logic exists anywhere.",
     ""),

    ("IDEA", "Combat", "Army Spell System",
     "army.gd (armyCurse, armyCharm vars, commented-out spell section)",
     "armyCurse/armyCharm vars declared. Commented-out code shows "
     "apply_spell()/tick_spell()/has_active_spell() design. Never implemented.",
     ""),

    # ── WEAPONS ────────────────────────────────────────────────────────────────
    ("FULL PASS", "Weapons", "Saber Class (4 types)",
     "weapon.gd",
     "Cutlass (L1, +3/+2), Cavalry Saber (L2, +5/+3), "
     "Light Saber (L3, +7/+4), Heavy Saber (L4, +9/+5). "
     "reloadTurns = 0; no ammo; chargeManpowerCost = 10% per charge. "
     "All have SaberCharge mod; Cavalry Saber adds CavalryMorale; Heavy Saber adds HeavyCharge.",
     ""),

    ("FULL PASS", "Weapons", "Musket Class (4 types)",
     "weapon.gd",
     "Flintlock (L1, ranged +4, reload 2), Brown Bess (L2, ranged +6, reload 2), "
     "Percussion Cap (L3, ranged +8, reload 1), Lever Repeater (L4, ranged +10, no reload). "
     "All have Bayonet (limited melee); Lever burns 2 ammo/level.",
     ""),

    ("FULL PASS", "Weapons", "Artillery Class (4 types)",
     "weapon.gd",
     "Falconet (L1, ranged +12, reload 3, ammo 3/level), "
     "Field Gun (L2, ranged +18, reload 3), Howitzer (L3, ranged +25, reload 2), "
     "Mortar (L4, ranged +35, reload 2, ammo 5/level + Siege mod). "
     "No melee at all. CannonBlast on all; AreaDamage on L2+.",
     ""),

    ("FULL PASS", "Weapons", "Reload System",
     "unit.gd + battle.gd",
     "reloadCounter counts down to 0 (ready). start_reload() sets counter = reloadTurns. "
     "tick_reload() decrements counter each round. "
     "GunCrewEfficiency armor mod reduces reloadTurns by 1 (min 0) at start_reload(). "
     "get_effective_ranged_offence() returns 0 if reloading.",
     "Reload only checked in ranged combat. Melee rounds don't tick reload."),

    ("FIRST PASS", "Weapons", "Legacy Weapon Class (12 types)",
     "weapon.gd",
     "Atlatl, Club, Double Axe, Flail, Longsword, Mace, Machete, Macuahuitl, "
     "Pike, Shortsword, Tomahawk, Spear. Kept for DODK compatibility. "
     "No weaponClass specialization (all 'Legacy'). AtlatlPierce and ClubBleed mods applied.",
     "Cost calculation in surveySelf uses legacy names only. New Saber/Musket/Artillery names "
     "not in the cost match block → army shows 0 weapons cost for all new weapons."),

    # ── ARMOR ──────────────────────────────────────────────────────────────────
    ("FULL PASS", "Armor", "Uniform Class (8 types)",
     "armor.gd",
     "Continental (M15/R20/S10 + DrillFormation), Redcoat (M20/R15/S10 + LineFormation), "
     "Militia (M10/R10/S5), Light Infantry (M10/R30/S10 + Skirmish), "
     "Heavy Infantry (M35/R10/S10 + ShockTroop), Cavalry (M25/R5/S15 + MountedCharge + CavalryMorale), "
     "Artillery Corps (M5/R5/S20 + GunCrewEfficiency), Minuteman (M12/R18/S8 + MinutemanSpirit). "
     "M/R/S = melee/ranged/spell block %.",
     ""),

    ("FIRST PASS", "Armor", "Legacy Armor Class (11 types)",
     "armor.gd",
     "Canine, Cast, Chain, Archer, Plate, Padded, Scout, Scale, Shell, Point, Feline, Otter. "
     "Kept for DODK compatibility. Chain and Shell legacy armors have embedded mods.",
     ""),

    # ── MIL MODS ───────────────────────────────────────────────────────────────
    ("FULL PASS", "Mil Mods", "Mil Mod Class & Pipeline",
     "mil_mod.gd, governor.gd, army.gd, unit.gd",
     "MilMod node with milModType, milModResource, disabled bool, "
     "infantryMod/rangedMod/siegeMod/commanderMod/civilianMod/resourceMod/marineMod/entrenchMod/terrainMod bools. "
     "Flow: governor.addMilMod() → govMilModsLvl1/2/3 → army.commanderCheck() → "
     "unit.addMilMod() → unit.militaryModifierList → calculateMilMods().",
     "addMilMod in governor.gd does NOT call buildSelf() — only milModType is set, "
     "all boolean flags remain false on programmatic mods. "
     "calculateMilMods has inverted disabled check (see Known Bugs)."),

    ("FULL PASS", "Mil Mods", "Resource Disable System",
     "army.gd surveySelf(), mil_mod.gd disableMilModType()",
     "surveySelf() calls enableMilModType('All') on all units first (reset), "
     "then disableMilModType(resource) for each depleted resource. "
     "Any mod with matching milModResource gets disabled = true.",
     "Due to inverted check bug, disabled mods actually DO apply and enabled mods DON'T."),

    ("FULL PASS", "Mil Mods", "Commander Mod Tiers",
     "governor.gd addMilMod(), army.gd commanderCheck()",
     "Levels encoded: 123 = all 3 tiers, 23 = tiers 2+3, 3 = tier 3 only. "
     "commanderCheck() selects commanderModifiers1/2/3 array based on governorLevel. "
     "VP commanders (isVicePresident = true) get all mods doubled.",
     ""),

    ("FULL PASS", "Mil Mods", "Archetype Mod Assignment",
     "world.gd _apply_archetype_mods()",
     "Each of 25 archetypes receives 4 mods: 2 at T1(123), 1 at T2(23), 1 at T3(3). "
     "Called immediately after governorLevel = 1 in generateBarracksCommanders().",
     ""),

    # ── COMMANDER ──────────────────────────────────────────────────────────────
    ("FULL PASS", "Commander", "Governor Class",
     "governor.gd",
     "governorType (name), governorLevel 1–3, governorArchetypeId (ARC_01–25 or CA_ARC_01+), "
     "loyalty −20 to +10, morale 0–100, isVicePresident bool, questComplete bool. "
     "govMilModsLvl1/2/3 arrays hold tier-tiered mods.",
     ""),

    ("FULL PASS", "Commander", "Named Historical Governors",
     "governor.gd buildSelf()",
     "Ualani Carlisle (Visionary×123, Champion of the Sun×23, Healer×3), "
     "Benjamin Tallmadge (Translator×123). "
     "Patrick Henry, Abigail Adams, Thomas Paine, Joseph Warren, Daniel Shays, "
     "Phillis Wheatley, Francis Asbury also defined.",
     ""),

    ("FULL PASS", "Commander", "Procedural Commander Generation",
     "world.gd generateBarracksCommanders()",
     "Runs on game start. For each player barracks tile: picks archetype by terrain match, "
     "picks name pool by cultural affinity, generates name, "
     "calls _apply_archetype_mods(), registers War Room arc, auto-assigns to tile army.",
     ""),

    ("FULL PASS", "Commander", "Commander Arc Objectives",
     "WarRoomPanel.gd registerCommanderArc(), _get_commander_objective()",
     "Each commander gets 3 objectives tied to archetype. "
     "15 condition types: kill_count, liberate_tiles, tile_corruption_below, "
     "tile_buildings_count, build_in_home_tile, home_tile_corruption_below, "
     "home_tile_moral_decay_below, resource_threshold, liberate_tiles_in_state, "
     "home_tile_has_building, liberate_any_with_building, liberate_state_count, + terrain types.",
     ""),

    ("FULL PASS", "Commander", "Commander Turn-Count Progression",
     "world.gd _commander_turns dict",
     "Tracks turns_served per governor. CMD_MERIT fires at 5 turns (level 1), "
     "CMD_RECOGNITION at 20 turns (level 2), CMD_THANKS at 50 turns (level 3). "
     "promote_commander outcome bumps governorLevel and calls stationedArmy.updateArmyUI().",
     ""),

    ("FULL PASS", "Commander", "Commander Narrative Arc Events",
     "data/events.csv, data/event_triggers.csv",
     "ARC_01_FIRST through ARC_25_FIRST: fire when objective 1 is met. "
     "ARC_01_DONE through ARC_25_DONE: fire when objective 3 is met. "
     "50 events total; 48 triggers (TRIG_036–083) wired in event_triggers.csv.",
     ""),

    # ── MOVEMENT ───────────────────────────────────────────────────────────────
    ("FULL PASS", "Movement", "Movement Point System",
     "army.gd, tile.gd get_move_cost()",
     "maxMovementPoints = 3 per army per turn. "
     "Terrain base costs: Metro/Fort/Suburbs/Farmlands = 1; "
     "Woods/Foothills/Wetlands = 2. "
     "Road level reduces cost (max(1, base − road_level)). "
     "Metro acts as road_level 2. Points restored to max each turn.",
     ""),

    ("FULL PASS", "Movement", "Winter Movement Penalty",
     "tile.gd get_move_cost(), army.armyTags",
     "Winter amplification: factor = 1.0 + (1.0 − get_winter_army_modifier()) × 1.0. "
     "Applied when winterScore > 0 and army lacks 'Cold Weather' tag. "
     "Cold/Harsh/Blizzard zones increase tile entry cost significantly.",
     ""),

    ("FIRST DRAFT", "Movement", "Pathfinding",
     "army.gd pathLine, targetTile",
     "pathLine and targetTile vars declared on Army. "
     "Description says 'builds a path from inTile to targetTile'. "
     "No pathfinding algorithm is visible in army.gd.",
     "Actual pathfinding logic likely in world.gd or path_control.gd."),

    ("IDEA", "Movement", "Night Raider / Ghost March (movement mods)",
     "mil_mod.gd Night Raider, Ghost March",
     "Night Raider: 'move and attack same turn without penalty'. "
     "Ghost March: 'ignore enemy zone of control'. "
     "Neither movement penalty nor ZoC system exists to modify.",
     ""),

    ("IDEA", "Movement", "The Long March (+2 MP mod)",
     "mil_mod.gd The Long March",
     "'+ 2 Movement Points; full movement may be used before attacking'. "
     "No code reads this mod and adds to maxMovementPoints.",
     ""),

    # ── WEATHER ────────────────────────────────────────────────────────────────
    ("FULL PASS", "Weather", "Winter Score System",
     "tile.gd winterScore",
     "winterScore int per tile. Range: −80 (extreme hurricane) to 80+ (blizzard). "
     "Categories: Extreme Hurricane ≤−80, Hurricane ≤−60, Storm Prone ≤−20, "
     "Mild −20 to 20, Cold 20–60, Harsh 60–80, Blizzard ≥80.",
     ""),

    ("FULL PASS", "Weather", "Winter Army Supply Drain",
     "world.gd _apply_winter_army_drain()",
     "Runs months 11–2. modifier 1.0 = 0 food drain, 0.85 (cold) = 3 food, "
     "0.65 (harsh) = 7 food. Tropical tiles (winterScore ≤ 0) exempt. "
     "Cold Weather armyTag exempts army.",
     ""),

    ("FULL PASS", "Weather", "Seasonal Eco Modifiers",
     "tile.gd _apply_winter_eco_modifier()",
     "Applies tileEcoModifier entries based on winterScore category. "
     "Summer/Winter season labels applied to tile.season var.",
     ""),

    ("FIRST PASS", "Weather", "Storm Event System",
     "tile.gd (storm vars), world.gd (_tick_storms, _spawn_storm, _spread_storm, _determine_storm_type)",
     "Storms spawn each turn at 3% chance on a random tile. "
     "Storm type determined by tile.winterScore + terrain (Tornado in Farmlands/Foothills, "
     "Hurricane/Thunderstorm in tropics, Blizzard/Nor'easter in cold zones, Fog/Thunderstorm/Nor'easter in temperate). "
     "Storm spreads to all TileNeighbors; duration 2–6 turns, intensity 1–3. "
     "_tick_storms() decrements duration each turn and clears expired storms. "
     "tile.stormActive, stormType, stormDuration, stormIntensity, stormOriginId properties on Tile.",
     "Storm combat effects (morale, movement penalties, supply) are NOT yet applied. "
     "Only the 12 storm counter mods in calculateMilMods read currentStorm for bonuses."),

    ("FIRST PASS", "Weather", "Storm Counter Mods (12 mods)",
     "mil_mod.gd, unit.gd calculateMilMods()",
     "12 storm counter mil mods: Fog-Born, Storm Rider, Thunder Proof, Blizzard March, "
     "Hurricane Eyes, Tornado Dancer, Nor'easter Veteran, Rain Reader, White Out Walker, "
     "Storm Chaser, Lightning Rod, Eye of the Storm. "
     "All have stormMod=true and stormType set. Direct stat bonuses read currentStorm in calculateMilMods(). "
     "Movement/immunity effects flagged IDEA pending army-level storm checks.",
     ""),

    # ── SPAWN ──────────────────────────────────────────────────────────────────
    ("FULL PASS", "Spawn", "Army Spawn from Barracks",
     "world.gd generateBarracksCommanders()",
     "On game start, scans all player barracks tiles. "
     "For each: picks archetype, generates commander, auto-assigns to tile and tile army.",
     ""),

    ("FULL PASS", "Spawn", "Unit Template Spawn",
     "unit.gd buildSelf()",
     "buildSelf(playerCountry, Type, Level, WeaponType, OreType, ArmorType, CurMan, CurWeapons). "
     "Instantiates Weapon/Ore/Armor sub-nodes; calls getUnitAttributes().",
     ""),

    ("FULL PASS", "Spawn", "Canadian Army Generation",
     "world.gd (generateBarracksCommanders companion section)",
     "Parallel archetype table for Canadian tiles: CA_ARC_01–CA_ARC_10 with "
     "French-Canadian name pools (NP_06) and terrain affinities.",
     ""),

    ("FIRST PASS", "Spawn", "Army Template from CSV",
     "army_database.gd",
     "army_templates CSV supplies ArmyName, country, tile, armyMods (pipe-delimited tags). "
     "armyTags populated from armyMods column.",
     ""),

    # ── SPECIAL ────────────────────────────────────────────────────────────────
    ("FULL PASS", "Special", "Vice President Army Bonus",
     "army.gd addUnitCommander()",
     "isVicePresident governors (Ualani Carlisle) get ALL their mods doubled — "
     "govMilModsLvl1/2/3 appended twice into commanderModifiers arrays.",
     ""),

    ("FIRST PASS", "Special", "Barracks Max Unit Level",
     "army.gd calculateMaxUnitLevel()",
     "maxUnitLevel = barracks building level in tile.tileBuildingsList. "
     "Higher barracks level allows higher-tier units.",
     "maxUnitLevel calculated but nothing enforces it during unit recruitment."),

    ("FIRST PASS", "Special", "Double Shot / Double Cannonade",
     "battle.gd _calculate_ranged_damage()",
     "Double Shot (T2): second volley at 50% power for all artillery units. "
     "Double Cannonade (T3): second volley at 100% power + 3 bonus per artillery unit. "
     "Detected via _army_has_siege_mod(). Second shot calculates against remaining shield "
     "after first shot, then adds to defenderManpowerLoss.",
     "Second shot only uses artillery units' ranged offence (not muskets/sabers)."),

    ("FIRST DRAFT", "Special", "Terrain Defensive Bonuses (new mods)",
     "mil_mod.gd (terrainMod, terrainType flags)",
     "terrainMod bool + terrainType String added to MilMod. "
     "Woodsman/Swamp Legs/Hill Runner/Street Tough/Farmhand all set these flags. "
     "No code in calculateMilMods() or battle.gd reads terrainType or applies bonus.",
     "Needs: army.inTile.terrain passed to calculateMilMods(); "
     "match block entries for each terrain mod type."),

    ("FIRST PASS", "Special", "Mythic Weapons (9 Easter Egg)",
     "weapon.gd",
     "9 Mythic-class weapons unlockable only via events: Baseball Bat, Trident, Mythic Atlatl, "
     "Sharps Carbine, Blackbeard's Pistols, Colt Revolver, Rocket Artillery, Trebuchet, Wright Flyer. "
     "All have Mythic-class embedded mods defined in mil_mod.gd. is_mythic() helper added to Weapon class. "
     "can_melee() excludes Rocket Artillery and Wright Flyer.",
     "Mythic mod effects partially stubbed (PirateVolley, CylinderFire, TrebuchetLaunch, etc. "
     "need battle.gd cases). RocketBarrage/AerialBombing tile effects not yet implemented."),

    ("FIRST PASS", "Special", "Cultural / State-Specific Mods (12 mods)",
     "mil_mod.gd, governor.gd",
     "12 state-exclusive cultural mods: Country Musician (TN), Virginia Gentry (VA), "
     "Minuteman's Pride (MA), Quaker Steel (PA), Georgia Peach (GA), Backcountry Rider (SC), "
     "Harbor Watch (NY), Chesapeake Sailor (MD), Frontier Marksman (KY), River Runner (OH), "
     "Everglades Tracker (FL), Bayou Warrior (LA). "
     "All have culturalMod=true and culturalState set. "
     "Stat bonuses apply via calculateMilMods(). Gate enforcement (only correct-state governors can have it) "
     "is handled at assignment time (event or archetype assignment — not auto-enforced in code).",
     "No code auto-prevents wrong-state governors from having cultural mods. "
     "Must be assigned correctly at event/archetype assignment time."),

    ("FIRST PASS", "Special", "Tool Mods (12 expanded civilian mods)",
     "mil_mod.gd",
     "12 new civilian tool mods: Cartographer, Herbalist, Engineer, Blacksmith, Physician, "
     "Merchant, Preacher, Architect, Hunter, Fisherman, Surveyor, Trapper. "
     "All have toolMod=true flag. Effects described in milModDescription. "
     "Added alongside existing civilian mods (Wooden Tools, Metal Tools, etc.).",
     "All tool mod gameplay effects are IDEA tier — no per-turn hooks wired to any of them yet."),

    ("FIRST PASS", "Special", "New Uniforms: Tombstone Cap + Hardee Hat",
     "armor.gd",
     "Tombstone Cap: Uniform class, M10/R20/S5, QuickDraw mod (+5 first ranged attack/battle). "
     "Hardee Hat: Uniform class, M15/R25/S10, HardeeDisc mod (+3 DEF/level adjacent friendly). "
     "Both mods defined in mil_mod.gd.",
     "QuickDraw and HardeeDisc effects are FIRST DRAFT — mods defined but "
     "first-attack tracking and adjacent-unit checks not yet in battle.gd/army.gd."),

    ("IDEA", "Special", "Supply Lines",
     "(army cost vars hint at it)",
     "armyFoodCost/armyWeaponsCost accumulated in surveySelf(). "
     "Country resources depleted by these costs is not fully automated per turn.",
     ""),

    # ── STATUS EFFECTS ─────────────────────────────────────────────────────────
    ("FULL PASS", "Mil Mods", "Army Status Effect System",
     "army.gd armyStatusEffects, apply_status(), _tick_status_effects()",
     "armyStatusEffects: Array of dicts {type, turnsLeft, magicCostPerTurn}. "
     "apply_status(type, duration, magic_cost=0): refresh/replace existing. "
     "_tick_status_effects() on onTurnEnd: drains magic for protector buffs (self-removes at 0); "
     "DoT for Burning (5×unitCount/turn) and Diseased/Quarantined (3×/1× unitCount/turn); "
     "ticks turnsLeft and removes expired. "
     "_apply_status_effects_to_stats(): called from surveySelf(); applies all 22 debuff "
     "and 25 protector buff stat modifiers to army scores. "
     "_apply_status_flags(): sets attackBlocked, reinforcementBlocked, movement from statuses.",
     "No UI widget to display active status effects on army panel."),

    ("FULL PASS", "Combat", "Combat Status Effects",
     "battle.gd _apply_combat_status_effects()",
     "Called from applyBattleResults() before deleteBattles signal. "
     "Rout: >40% manpower lost in one battle → Routed 2 turns. "
     "Weapon-based: Baseball Bat→Shaken 1t, Trident→Terrified 2t, Mythic Atlatl→Stunned 1t, "
     "Sharps Carbine→Blinded 1t, Blackbeard's Pistols→Terrified 2t, Colt Revolver→Shaken 1t, "
     "Rocket Artillery→Burning 2t+Suppressed 1t, Trebuchet→Stunned 1t (ranged only), "
     "Wright Flyer→Suppressed 1t. "
     "Protector retaliations: Bell Witch's Harassment→Demoralize attacker 2t, "
     "Headless Terror→Terrify attacker 2t, Le Gougou's Terror→Terrify attacker 3t.",
     "Weapon statuses apply on every attack (no crit condition)."),

    ("FULL PASS", "Mil Mods", "State Guard Mod System",
     "mil_mod.gd, unit.gd calculateMilMods(), army.gd surveySelf()",
     "26 state/provincial guard mods (Pennsylvania Guard → Bahamas Guard). "
     "Each has culturalMod=true and culturalState=state code (PA, VA, NY, etc.). "
     "Generic post-match handler in calculateMilMods(): "
     "if MilMod.culturalMod and MilMod.culturalState != '' and currentState == MilMod.culturalState: "
     "+2 attack +2 defence per level in any tile of that state. "
     "surveySelf() sets unit.currentState = inTile.tileContinent before calculateMilMods().",
     "tileContinent value must match culturalState codes exactly. "
     "No gate prevents wrong-state governors from having guard mods."),

    ("FIRST PASS", "Mil Mods", "Protector Buff Casting",
     "world.gd executeOutcome(), army.gd apply_status(), _tick_status_effects()",
     "25 protector buff statuses (Mothman Presence, Bell Witch's Harassment, etc.) "
     "applied as army status effects via magic upkeep. "
     "executeOutcome() 'cast_protector_buff' case: tile.stationedArmy.apply_status(name, 9999, cost). "
     "turnsLeft=9999 (infinite) but drains parentCountry.TotalMagic each tick. "
     "Self-removes when magic < magicCostPerTurn. "
     "_apply_status_effects_to_stats() applies buff bonuses to armyPunch/Launch/Block/Defence/Shield.",
     "Wiring 'cast_protector_buff' to specific event outcomes still needed per protector. "
     "No UI shows active protector buffs or magic upkeep cost."),

    ("FULL PASS", "Special", "Corruption Disease Check",
     "army.gd onTurnEnd()",
     "On each turn end: if inTile.tileCorrution > 0, army has tileCorrution% chance to gain "
     "Diseased status for 2 turns. Park Ranger mod (civilianMod, toolMod) grants immunity. "
     "Disease check: randf() < float(inTile.tileCorrution) / 100.0.",
     "Spelling 'tileCorrution' matches the in-game variable name."),

    ("FULL PASS", "Special", "President + Election Season Movement",
     "governor.gd, army.gd _commander_movement_bonus(), world.gd _grant_election_season_mods()",
     "Ualani Carlisle has President mod (commanderMod, +3 movement). "
     "_commander_movement_bonus() sums +3 for President and +3 for Election Season "
     "from govMilModsLvl1/2/3 arrays. Applied in onTurnEnd(): "
     "currentMovementPoints = maxMovementPoints + bonus (before _apply_status_flags() clamping). "
     "_check_election_season() calls _grant_election_season_mods() → grants Election Season "
     "to Ualani and any governor where isVicePresident = true.",
     "Election Season is permanent once granted. Commander must be in an army for bonus to apply."),

    ("FIRST PASS", "Weather", "Storm Status Debuffs",
     "world.gd _apply_storm_debuffs(), _apply_storm_status_to_army()",
     "Applied at turn start via _on_next_turn_pressed() after _tick_storms(). "
     "Thunderstorm/Hurricane → Waterlogged 2 turns (−20% melee+ranged). "
     "Blizzard/Nor'easter → Frostbitten 2 turns (−30% melee+ranged). "
     "Fog → Blinded 1 turn (ranged halved). "
     "Tornado → Bogged Down 1t (no move) + Shaken 1t (block halved).",
     "Armies moving INTO storm tiles mid-turn don't get debuff until next turn start. "
     "Storm debuffs use same apply_status() refresh/replace logic."),
]


def build_systems_sheet(wb):
    ws = wb.create_sheet("Systems Overview")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("STATUS",       13),
        ("CATEGORY",     14),
        ("SYSTEM",       30),
        ("KEY FILE(S)",  30),
        ("DESCRIPTION",  62),
        ("GAPS / NOTES", 44),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    current_cat = None
    row = 2
    for (status, cat, name, files, desc, notes) in SYSTEMS:
        if cat != current_cat:
            divider(ws, row, len(COLS), cat)
            row += 1
            current_cat = cat

        sbg = STATUS_COLORS.get(status, "FFFFFF")
        sfg = STATUS_FG.get(status, NORMAL_FG)
        cbg = CAT_COLORS.get(cat, "FFFFFF")

        pcell(ws, row, 1, status, bg=sbg,  fg=sfg,    bold=True, align="center")
        pcell(ws, row, 2, cat,    bg=cbg)
        pcell(ws, row, 3, name,   bg=cbg,  bold=True)
        pcell(ws, row, 4, files,  bg=cbg)
        pcell(ws, row, 5, desc,   bg="FAFAFA")
        pcell(ws, row, 6, notes,  bg="FFFBE6")
        ws.row_dimensions[row].height = 44
        row += 1

    ws.freeze_panes = "A2"


# ── SHEET 2: MIL MODS ─────────────────────────────────────────────────────────
# tier_label, category, mod_type, type_flags, description, resource, notes

MIL_MODS = [
    # ── TIER 1 COMMANDER MODS ──────────────────────────────────────────────────
    ("T1 (123)", "Commander/Infantry", "Woodsman",
     "infantryMod, terrainMod", "Farmlands", "+2 Attack +2 Defense per level in Woods terrain",
     "FIRST DRAFT — terrainType set; effect not in calculateMilMods"),
    ("T1 (123)", "Commander/Infantry", "Swamp Legs",
     "infantryMod, terrainMod", "Wetlands", "+2 Attack +2 Defense per level in Wetlands terrain",
     "FIRST DRAFT — terrainType set; effect not in calculateMilMods"),
    ("T1 (123)", "Commander/Infantry", "Hill Runner",
     "infantryMod, terrainMod", "Foothills", "+2 Attack +2 Defense per level in Foothills terrain",
     "FIRST DRAFT — terrainType set; effect not in calculateMilMods"),
    ("T1 (123)", "Commander/Infantry", "Street Tough",
     "infantryMod, terrainMod", "Metro/Suburbs", "+2 Attack per level in Metro or Suburbs terrain",
     "FIRST DRAFT — terrainType set; effect not in calculateMilMods"),
    ("T1 (123)", "Commander/Infantry", "Farmhand",
     "infantryMod, terrainMod", "Farmlands", "+1 Attack per level in Farmlands terrain",
     "FIRST DRAFT — terrainType set; effect not in calculateMilMods"),
    ("T1 (123)", "Commander/Infantry", "Saber Drill",
     "infantryMod", "—", "+3 Attack per level",
     "FIRST DRAFT — defined; no case in calculateMilMods"),
    ("T1 (123)", "Commander/Ranged", "Marksman",
     "rangedMod", "—", "+2 Ranged Attack per level",
     "FIRST DRAFT — defined; no case in calculateMilMods"),
    ("T1 (123)", "Commander/Infantry", "Steady Line",
     "infantryMod", "—", "+3 Defense per level",
     "FIRST DRAFT — defined; no case in calculateMilMods"),
    ("T1 (123)", "Commander/Ranged", "Quick Reload",
     "rangedMod", "—", "Reload reduced by 1 round (min 0)",
     "FIRST DRAFT — defined; no case in calculateMilMods or start_reload"),
    ("T1 (123)", "Commander/Siege", "Powder & Shot",
     "siegeMod", "—", "+3 Siege Attack per level",
     "FIRST DRAFT — defined; no case in calculateMilMods"),
    ("T1 (123)", "Commander/All", "Fortified Position",
     "commanderMod", "—", "All units +3 Defense per level in tiles with Barracks or Fortress",
     "FIRST DRAFT — defined; no terrain/building check in calculateMilMods"),
    ("T1 (123)", "Commander/All", "Coastal Watch",
     "commanderMod", "—", "All units +2 Defense per level when adjacent to naval tiles",
     "FIRST DRAFT — defined; no naval neighbor check in calculateMilMods"),

    # ── TIER 2 COMMANDER MODS ──────────────────────────────────────────────────
    ("T2 (23)", "Commander/Special", "Marine",
     "commanderMod, marineMod", "—", "Army may launch melee attacks into adjacent navalPathPoints (occupied PPBs)",
     "FIRST PASS — _army_has_active_marine() wired; meleePressed() loops navalPathPoints for targets"),
    ("T2 (23)", "Commander/Infantry", "Guerrilla Tactics",
     "infantryMod, terrainMod", "Woods/Wetlands", "+4 Attack +4 Defense per level in Woods or Wetlands",
     "FIRST DRAFT — terrainType set; effect not in calculateMilMods"),
    ("T2 (23)", "Commander/Siege", "Double Shot",
     "siegeMod", "—", "Siege fires twice per round — second shot at 50% power",
     "FIRST DRAFT — defined; no second-shot loop in battle.gd"),
    ("T2 (23)", "Commander/Infantry", "Iron Bayonet",
     "infantryMod", "—", "+5 Attack per level in first battle round",
     "FIRST DRAFT — defined; no round-counter in battle.gd"),
    ("T2 (23)", "Commander/Ranged", "Sharpshooter",
     "rangedMod", "—", "Ranged attacks ignore 2 enemy Defense per level",
     "FIRST DRAFT — defined; no armor-pierce logic in _calculate_ranged_damage"),
    ("T2 (23)", "Commander/Tile", "Corrupted Ground",
     "commanderMod", "—", "Army presence reduces tile corruption by 1 per turn",
     "FIRST DRAFT — defined; no per-turn tile.corruption modifier hooked to army"),
    ("T2 (23)", "Commander/All", "Rallying Voice",
     "commanderMod", "—", "Morale loss reduced; rout threshold lowered to 15%",
     "FIRST DRAFT — defined; retreat threshold hardcoded at 25% in army.gd"),
    ("T2 (23)", "Commander/All", "Night Raider",
     "commanderMod", "—", "Army may move and attack same turn without penalty",
     "IDEA — no day/night cycle; movement-attack penalty doesn't exist yet"),
    ("T2 (23)", "Commander/Infantry", "Flanking Drill",
     "infantryMod", "—", "+3 Attack per level when fighting in a contested tile",
     "FIRST DRAFT — defined; no contested-tile check in battle.gd"),
    ("T2 (23)", "Commander/All", "Vanguard",
     "commanderMod", "—", "All units +4 Attack per level on first engagement in a fresh tile",
     "FIRST DRAFT — defined; no first-engagement tracker per tile"),
    ("T2 (23)", "Commander/Siege", "Siege Line",
     "siegeMod", "—", "Siege attacks against fortified tiles suffer no defensive penalty",
     "FIRST DRAFT — defined; no fortification check in _calculate_ranged_damage"),
    ("T2 (23)", "Commander/Tile", "Cleaner",
     "commanderMod", "—", "Army presence reduces tile moral decay by 1 per turn",
     "FIRST DRAFT — defined; no per-turn tileMoralDecay modifier hooked to army"),

    # ── TIER 3 COMMANDER MODS ──────────────────────────────────────────────────
    ("T3 (3)", "Commander/Special", "Entrenched",
     "commanderMod, entrenchMod", "—", "After 3 stationary turns, all units +5 Defense/level — lost on movement",
     "FIRST DRAFT — entrenchMod flag set; no stationary turn counter anywhere"),
    ("T3 (3)", "Commander/All", "Continental Line",
     "commanderMod", "—", "All units +2 Attack +2 Defense per level permanently",
     "FIRST DRAFT — defined; no case in calculateMilMods"),
    ("T3 (3)", "Commander/Infantry", "Last Stand",
     "infantryMod", "—", "Units below 25% manpower gain +6 Attack per level",
     "FIRST DRAFT — defined; no manpower% check in calculateMilMods"),
    ("T3 (3)", "Commander/All", "Terror",
     "commanderMod", "—", "Enemy loses 10 Morale at start of each battle round",
     "IDEA — no round-start morale tick in battle.gd"),
    ("T3 (3)", "Commander/All", "Iron Wall",
     "commanderMod", "—", "+8 Defense per level when defending the commander's home tile",
     "FIRST DRAFT — defined; no home-tile check in battle.gd"),
    ("T3 (3)", "Commander/All", "Rampart",
     "commanderMod", "—", "+5 Defense per level in any Fortress tile",
     "FIRST DRAFT — defined; no fortress-tile check in calculateMilMods"),
    ("T3 (3)", "Commander/Special", "Naval Supremacy",
     "commanderMod, marineMod", "—", "Marine melee attacks deal +5 additional damage per unit level",
     "FIRST PASS — _army_naval_supremacy_bonus() in battle.gd adds +5×unitLevel to raw_attack in melee"),
    ("T3 (3)", "Commander/All", "Ghost March",
     "commanderMod", "—", "Army ignores enemy zone of control",
     "IDEA — no ZoC system exists"),
    ("T3 (3)", "Commander/All", "Undaunted",
     "commanderMod", "—", "Ignore the first retreat check each battle",
     "FIRST DRAFT — retreat check hardcoded in army.gd; no first-check skip logic"),
    ("T3 (3)", "Commander/Siege", "Double Cannonade",
     "siegeMod", "—", "Siege fires twice per round AND +3 Attack on all shots",
     "FIRST DRAFT — no second-shot loop in battle.gd"),
    ("T3 (3)", "Commander/All", "Liberator's Will",
     "commanderMod", "—", "After liberating a tile: +15% manpower, commander +2 loyalty",
     "FIRST DRAFT — defined; no post-liberation hook reads this mod"),
    ("T3 (3)", "Commander/All", "The Long March",
     "commanderMod", "—", "+2 Movement Points; full movement usable before attacking",
     "IDEA — no code reads this mod and adds to maxMovementPoints"),

    # ── LEGACY RESOURCE MODS ───────────────────────────────────────────────────
    ("Resource", "Resource/Infantry", "ClubBleed",
     "infantryMod", "Weapons", "CLUB: +2 Attack +1 Defense per level; costs Weapons",
     "FIRST PASS — in calculateMilMods() but inverted disabled check prevents it firing"),
    ("Resource", "Resource/Ranged", "AtlatlPierce",
     "rangedMod", "Weapons", "ATLATL: +1 Attack +2 Defense per level; costs Weapons",
     "FIRST PASS — in calculateMilMods() but inverted disabled check prevents it firing"),
    ("Resource", "Resource", "Wood",
     "resourceMod", "Wood", "+1 Attack per level; costs Wood",
     "FIRST PASS — in buildSelf but no case in calculateMilMods"),
    ("Resource", "Resource", "Copper",
     "resourceMod", "Metal", "+2 Attack per level (+magic defense); costs Metal",
     "FIRST PASS — Copper adds magic defense in calculateMilMods (inverted check bug)"),
    ("Resource", "Resource", "Iron",
     "resourceMod", "Metal", "+3 Attack per level; costs Metal ×3",
     "FIRST PASS — defined in buildSelf, no case in calculateMilMods"),
    ("Resource", "Resource", "Gold",
     "resourceMod", "Metal+Gold", "+2 Attack per level (+magic defense); costs Metal+Gold",
     "FIRST PASS — Gold adds magic defense in calculateMilMods (inverted check bug)"),
    ("Resource", "Resource", "Floodstone",
     "resourceMod", "Metal+Magic", "+3 Attack per level; costs Metal+Magic",
     "FIRST PASS — defined in buildSelf, no case in calculateMilMods"),
    ("Resource", "Country/Infantry", "Berserkers",
     "infantryMod", "Harmony", "+3 Attack per level, −2 Harmony per level",
     "FIRST PASS — defined in buildSelf, no case in calculateMilMods"),

    # ── ORIGINAL COMMANDER MODS ────────────────────────────────────────────────
    ("Legacy", "Commander", "Visionary",
     "commanderMod", "None", "+3 Attack per level, +10% speed",
     "FIRST PASS — defined, no case in calculateMilMods; Ualani Carlisle has this"),
    ("Legacy", "Commander", "Champion of the Sun",
     "commanderMod", "None", "+1 Attack per level, +10% speed (daytime)",
     "FIRST PASS — defined, no case in calculateMilMods; daytime check unimplemented"),
    ("Legacy", "Commander", "Healer",
     "commanderMod", "None", "+10 reinforce rate",
     "FIRST PASS — defined, no case in calculateMilMods; no reinforce rate modifier hooked"),
    ("Legacy", "Commander", "Chain (armor)",
     "commanderMod", "Weapons", "+5% Melee, +40% Ranged, +5% Spell block",
     "FIRST PASS — Chain adds unitRangedOffence in calculateMilMods (inverted check bug)"),
    ("Legacy", "Commander", "Shell (armor)",
     "commanderMod", "Weapons", "+50% Spell block",
     "FIRST PASS — defined, no direct effect in calculateMilMods"),

    # ── WEAPON-EMBEDDED MODS ───────────────────────────────────────────────────
    ("Weapon", "Saber", "SaberCharge",
     "(weapon mod)", "—", "Enables saber charge: attacker pays 10% manpower to strike",
     "FULL PASS — charge cost applied in battle.gd _apply_charge_costs()"),
    ("Weapon", "Saber", "CavalryMorale",
     "(weapon mod)", "—", "Bonus vs infantry morale on charge",
     "IDEA — declared, never read anywhere"),
    ("Weapon", "Saber", "HeavyCharge",
     "(weapon mod)", "—", "Extra punch damage on Heavy Saber charge",
     "IDEA — declared, never read anywhere"),
    ("Weapon", "Musket", "Bayonet",
     "(weapon mod)", "—", "Musket unit can melee (limited effectiveness, 50% penalty)",
     "FULL PASS — can_charge_melee() returns true for Musket; 0.5 penalty in calculateGrossValues"),
    ("Weapon", "Musket", "VolleyFire",
     "(weapon mod)", "—", "Bonus when multiple musket units fire simultaneously",
     "IDEA — declared on Brown Bess, never read in battle.gd"),
    ("Weapon", "Musket", "RapidFire",
     "(weapon mod)", "—", "Lever Repeater fires every round without reload",
     "FIRST PASS — already handled by reloadTurns = 0 on weapon; mod redundant"),
    ("Weapon", "Artillery", "CannonBlast",
     "(weapon mod)", "—", "Artillery ranged attack",
     "FIRST PASS — all artillery can fire ranged via can_fire_ranged(); mod declared 'handled in battle.gd' but no special branch"),
    ("Weapon", "Artillery", "AreaDamage",
     "(weapon mod)", "—", "Hits multiple units (Field Gun, Howitzer, Mortar)",
     "IDEA — declared, never read in battle.gd; all ranged attacks currently single-target"),
    ("Weapon", "Artillery", "Siege (Mortar)",
     "(weapon mod)", "—", "Bonus vs fortified tiles",
     "IDEA — declared on Mortar, never read in battle.gd"),

    # ── STORM COUNTER MODS ─────────────────────────────────────────────────────
    ("Storm", "Storm/Commander", "Fog-Born",
     "commanderMod, stormMod", "Fog", "+5 attack per level in Fog storm tiles",
     "FIRST DRAFT — stormMod flag + stormType set; calculateMilMods reads currentStorm"),
    ("Storm", "Storm/Commander", "Storm Rider",
     "commanderMod, stormMod", "Any", "Movement not reduced by any active storm",
     "IDEA — movement exemption needs army-level storm check"),
    ("Storm", "Storm/Commander", "Thunder Proof",
     "commanderMod, stormMod", "Thunderstorm", "Immune to Thunderstorm morale penalty",
     "IDEA — morale immunity needs army-level check"),
    ("Storm", "Storm/Commander", "Blizzard March",
     "commanderMod, stormMod", "Blizzard", "No movement or supply penalty in Blizzard tiles",
     "IDEA — supply/movement exemption needs army-level check"),
    ("Storm", "Storm/Commander", "Hurricane Eyes",
     "commanderMod, stormMod", "Hurricane", "+5 attack per level in Hurricane storm tiles",
     "FIRST DRAFT — reads currentStorm in calculateMilMods"),
    ("Storm", "Storm/Commander", "Tornado Dancer",
     "commanderMod, stormMod", "Tornado", "Army ignores Tornado scatter and manpower drain",
     "IDEA — Tornado scatter effect not yet implemented"),
    ("Storm", "Storm/Commander", "Nor'easter Veteran",
     "commanderMod, stormMod", "Nor'easter", "+3 defense per level in Nor'easter tiles",
     "FIRST DRAFT — reads currentStorm in calculateMilMods"),
    ("Storm", "Storm/Ranged", "Rain Reader",
     "rangedMod, stormMod", "Thunderstorm", "+3 ranged attack per level during Thunderstorm",
     "FIRST DRAFT — reads currentStorm in calculateMilMods"),
    ("Storm", "Storm/Infantry", "White Out Walker",
     "infantryMod, stormMod", "Blizzard", "+3 attack per level during Blizzard",
     "FIRST DRAFT — reads currentStorm in calculateMilMods"),
    ("Storm", "Storm/Commander", "Storm Chaser",
     "commanderMod, stormMod", "Any", "+1 movement point in any active storm tile",
     "IDEA — movement bonus needs army-level check"),
    ("Storm", "Storm/Siege", "Lightning Rod",
     "siegeMod, stormMod", "Thunderstorm", "Artillery units ignore storm ranged accuracy penalty",
     "IDEA — storm accuracy penalty not yet implemented"),
    ("Storm", "Storm/Commander", "Eye of the Storm",
     "commanderMod, stormMod", "Any", "+4 attack +4 defense per level while storm is active",
     "FIRST DRAFT — reads currentStorm != '' in calculateMilMods"),

    # ── CULTURAL / STATE-SPECIFIC MODS ─────────────────────────────────────────
    ("Cultural", "Cultural/TN", "Country Musician",
     "commanderMod, culturalMod", "TN only", "+3 morale; +2 ATK/level in Farmlands. Tennessee governors only.",
     "FIRST DRAFT — culturalState='TN'; terrain check in calculateMilMods"),
    ("Cultural", "Cultural/VA", "Virginia Gentry",
     "commanderMod, culturalMod", "VA only", "+3 ranged defense per level. Virginia governors only.",
     "FIRST DRAFT — culturalState='VA'; calculateMilMods applies ranged def bonus"),
    ("Cultural", "Cultural/MA", "Minuteman's Pride",
     "commanderMod, culturalMod", "MA only", "+5 ATK/level first 3 battle rounds. Massachusetts governors only.",
     "IDEA — first-round tracking not yet in battle.gd"),
    ("Cultural", "Cultural/PA", "Quaker Steel",
     "commanderMod, culturalMod", "PA only", "+2 DEF −1 ATK per level. Pennsylvania governors only.",
     "FIRST DRAFT — calculateMilMods applies both bonuses"),
    ("Cultural", "Cultural/GA", "Georgia Peach",
     "infantryMod, culturalMod", "GA only", "+3 food efficiency; +1 ranged/level. Georgia governors only.",
     "FIRST DRAFT — ranged bonus in calculateMilMods; food handled separately"),
    ("Cultural", "Cultural/SC", "Backcountry Rider",
     "infantryMod, terrainMod, culturalMod", "SC only + Woods", "+4 ATK/level in Woods. Carolina governors only.",
     "FIRST DRAFT — terrain + cultural flags; calculateMilMods checks currentTerrain"),
    ("Cultural", "Cultural/NY", "Harbor Watch",
     "commanderMod, marineMod, culturalMod", "NY only", "+3 ATK/level near naval tiles. New York governors only.",
     "IDEA — near-naval check not in calculateMilMods yet"),
    ("Cultural", "Cultural/MD", "Chesapeake Sailor",
     "commanderMod, marineMod, culturalMod", "MD only", "+2 melee/level + Marine ability. Maryland governors only.",
     "FIRST DRAFT — melee bonus in calculateMilMods; Marine handled at army level"),
    ("Cultural", "Cultural/KY", "Frontier Marksman",
     "rangedMod, terrainMod, culturalMod", "KY only + Foothills", "+4 ranged/level in Foothills. Kentucky governors only.",
     "FIRST DRAFT — terrain + cultural flags; calculateMilMods checks currentTerrain"),
    ("Cultural", "Cultural/OH", "River Runner",
     "commanderMod, culturalMod", "OH only", "+2 movement; +2 ATK near water. Ohio governors only.",
     "IDEA — movement bonus and near-water check not yet wired"),
    ("Cultural", "Cultural/FL", "Everglades Tracker",
     "infantryMod, terrainMod, culturalMod", "FL only + Wetlands", "+4 ATK+DEF/level in Wetlands. Florida governors only.",
     "FIRST DRAFT — terrain check in calculateMilMods"),
    ("Cultural", "Cultural/LA", "Bayou Warrior",
     "infantryMod, terrainMod, culturalMod", "LA only + Wetlands", "+5 ATK/level in Wetlands. Louisiana governors only.",
     "FIRST DRAFT — terrain check in calculateMilMods"),

    # ── TOOL MODS (expanded civilian list) ─────────────────────────────────────
    ("Tool", "Civilian/Tool", "Cartographer",
     "civilianMod, toolMod", "Science", "Maps explored tiles, revealing terrain bonuses and hidden resources.",
     "IDEA — reveal logic not yet wired"),
    ("Tool", "Civilian/Tool", "Herbalist",
     "civilianMod, toolMod", "Food", "+5 manpower per turn in tile.",
     "IDEA — per-turn heal not yet wired"),
    ("Tool", "Civilian/Tool", "Engineer",
     "civilianMod, toolMod", "Wood", "Builds roads and improves structures faster.",
     "IDEA — build speed modifier not yet wired"),
    ("Tool", "Civilian/Tool", "Blacksmith",
     "civilianMod, toolMod", "Metal", "Reduces weapon upkeep by 1/level/turn for stationed armies.",
     "IDEA — weapon cost reduction not yet wired"),
    ("Tool", "Civilian/Tool", "Physician",
     "civilianMod, toolMod", "Food", "+10 manpower per turn to armies in tile.",
     "IDEA — army heal not yet wired"),
    ("Tool", "Civilian/Tool", "Merchant",
     "civilianMod, toolMod", "Gold", "+3 gold per turn.",
     "IDEA — gold generation not yet wired"),
    ("Tool", "Civilian/Tool", "Preacher",
     "civilianMod, toolMod", "Influence", "+5 loyalty/turn for tile governor.",
     "IDEA — loyalty tick not yet wired"),
    ("Tool", "Civilian/Tool", "Architect",
     "civilianMod, toolMod", "Wood", "Reduces building costs by 15%.",
     "IDEA — cost reduction not yet wired"),
    ("Tool", "Civilian/Tool", "Hunter",
     "civilianMod, toolMod", "Food", "+5 food per turn.",
     "IDEA — food generation not yet wired"),
    ("Tool", "Civilian/Tool", "Fisherman",
     "civilianMod, toolMod", "Food", "+3 food per turn near water tiles.",
     "IDEA — water-adjacency food not yet wired"),
    ("Tool", "Civilian/Tool", "Surveyor",
     "civilianMod, toolMod", "Science", "Reveals terrain bonuses of adjacent tiles.",
     "IDEA — reveal logic not yet wired"),
    ("Tool", "Civilian/Tool", "Trapper",
     "civilianMod, toolMod", "Food", "+2 food +1 trade per turn.",
     "IDEA — trade/food generation not yet wired"),

    # ── ARMOR-EMBEDDED MODS ────────────────────────────────────────────────────
    ("Armor", "Uniform", "DrillFormation",
     "(armor mod)", "—", "Continental: bonus when adjacent to other Continental units",
     "IDEA — declared on Continental armor, 'handled at army level', no army-level check"),
    ("Armor", "Uniform", "LineFormation",
     "(armor mod)", "—", "Redcoat: bonus to ranged when in line formation",
     "IDEA — declared on Redcoat armor, 'handled at army level', no army-level check"),
    ("Armor", "Uniform", "Skirmish",
     "(armor mod)", "—", "Light Infantry: can move and fire same turn",
     "IDEA — declared on Light Infantry, 'handled at army level', no movement-fire logic"),
    ("Armor", "Uniform", "ShockTroop",
     "(armor mod)", "—", "Heavy Infantry: bonus damage on first melee attack",
     "FIRST PASS — declared 'handled in battle.gd first strike bonus'; no first-round check in battle.gd"),
    ("Armor", "Uniform", "MountedCharge",
     "(armor mod)", "—", "Cavalry: amplifies saber charge damage by 20%",
     "FIRST PASS — case in calculateMilMods() adds 20% to offScore IF saber; inverted disabled check prevents firing"),
    ("Armor", "Uniform", "GunCrewEfficiency",
     "(armor mod)", "—", "Artillery Corps: reduces reload by 1 turn",
     "FULL PASS — checked in unit.gd start_reload() directly by milModType string; bypasses disabled check bug"),
    ("Armor", "Uniform", "MinutemanSpirit",
     "(armor mod)", "—", "Minuteman: bonus morale in home territory",
     "IDEA — declared on Minuteman, never read anywhere"),
    ("Armor", "Civilian", "Translator",
     "civilianMod", "Science", "Enable reading of ancient writings",
     "CIVILIAN — not a combat mod"),
    ("Armor", "Civilian", "Seeder",
     "civilianMod", "Food", "Change agricultural output of tile",
     "CIVILIAN — not a combat mod"),

    # ── SPECIAL COMMANDER / CIVILIAN MODS ──────────────────────────────────────
    ("Special", "Civilian/Tool", "Park Ranger",
     "civilianMod, toolMod", "None", "Immunity to corruption-based disease on turn end.",
     "FULL PASS — corruption disease check in army.gd onTurnEnd() skips armies whose commander has Park Ranger"),
    ("Special", "Commander/Movement", "President",
     "commanderMod", "None", "+3 movement points per turn. Ualani Carlisle only.",
     "FULL PASS — _commander_movement_bonus() sums +3; applied before status flag clamping in onTurnEnd()"),
    ("Special", "Commander/Movement", "Election Season",
     "commanderMod", "None", "+3 movement points per turn; granted to Ualani and active VP during election season.",
     "FULL PASS — _grant_election_season_mods() called by _check_election_season(); permanent once granted"),

    # ── STATE / PROVINCIAL GUARD MODS ──────────────────────────────────────────
    ("State Guard", "State/PA", "Pennsylvania Guard",
     "commanderMod, culturalMod", "PA tiles", "+2 Attack +2 Defence per level in Pennsylvania tiles",
     "FULL PASS — generic culturalMod handler in calculateMilMods(); culturalState='PA'"),
    ("State Guard", "State/VA", "Virginia Guard",
     "commanderMod, culturalMod", "VA tiles", "+2 Attack +2 Defence per level in Virginia tiles",
     "FULL PASS — culturalState='VA'"),
    ("State Guard", "State/NY", "New York Guard",
     "commanderMod, culturalMod", "NY tiles", "+2 Attack +2 Defence per level in New York tiles",
     "FULL PASS — culturalState='NY'"),
    ("State Guard", "State/MA", "Massachusetts Guard",
     "commanderMod, culturalMod", "MA tiles", "+2 Attack +2 Defence per level in Massachusetts tiles",
     "FULL PASS — culturalState='MA'"),
    ("State Guard", "State/MD", "Maryland Guard",
     "commanderMod, culturalMod", "MD tiles", "+2 Attack +2 Defence per level in Maryland tiles",
     "FULL PASS — culturalState='MD'"),
    ("State Guard", "State/NC", "North Carolina Guard",
     "commanderMod, culturalMod", "NC tiles", "+2 Attack +2 Defence per level in North Carolina tiles",
     "FULL PASS — culturalState='NC'"),
    ("State Guard", "State/SC", "South Carolina Guard",
     "commanderMod, culturalMod", "SC tiles", "+2 Attack +2 Defence per level in South Carolina tiles",
     "FULL PASS — culturalState='SC'"),
    ("State Guard", "State/GA", "Georgia Guard",
     "commanderMod, culturalMod", "GA tiles", "+2 Attack +2 Defence per level in Georgia tiles",
     "FULL PASS — culturalState='GA'"),
    ("State Guard", "State/CT", "Connecticut Guard",
     "commanderMod, culturalMod", "CT tiles", "+2 Attack +2 Defence per level in Connecticut tiles",
     "FULL PASS — culturalState='CT'"),
    ("State Guard", "State/NJ", "New Jersey Guard",
     "commanderMod, culturalMod", "NJ tiles", "+2 Attack +2 Defence per level in New Jersey tiles",
     "FULL PASS — culturalState='NJ'"),
    ("State Guard", "State/DE", "Delaware Guard",
     "commanderMod, culturalMod", "DE tiles", "+2 Attack +2 Defence per level in Delaware tiles",
     "FULL PASS — culturalState='DE'"),
    ("State Guard", "State/NH", "New Hampshire Guard",
     "commanderMod, culturalMod", "NH tiles", "+2 Attack +2 Defence per level in New Hampshire tiles",
     "FULL PASS — culturalState='NH'"),
    ("State Guard", "State/RI", "Rhode Island Guard",
     "commanderMod, culturalMod", "RI tiles", "+2 Attack +2 Defence per level in Rhode Island tiles",
     "FULL PASS — culturalState='RI'"),
    ("State Guard", "State/VT", "Vermont Guard",
     "commanderMod, culturalMod", "VT tiles", "+2 Attack +2 Defence per level in Vermont tiles",
     "FULL PASS — culturalState='VT'"),
    ("State Guard", "State/ME", "Maine Guard",
     "commanderMod, culturalMod", "ME tiles", "+2 Attack +2 Defence per level in Maine tiles",
     "FULL PASS — culturalState='ME'"),
    ("State Guard", "State/TN", "Tennessee Guard",
     "commanderMod, culturalMod", "TN tiles", "+2 Attack +2 Defence per level in Tennessee tiles",
     "FULL PASS — culturalState='TN'"),
    ("State Guard", "State/AL", "Alabama Guard",
     "commanderMod, culturalMod", "AL tiles", "+2 Attack +2 Defence per level in Alabama tiles",
     "FULL PASS — culturalState='AL'"),
    ("State Guard", "State/FL", "Florida Guard",
     "commanderMod, culturalMod", "FL tiles", "+2 Attack +2 Defence per level in Florida tiles",
     "FULL PASS — culturalState='FL'"),
    ("State Guard", "State/WV", "West Virginia Guard",
     "commanderMod, culturalMod", "WV tiles", "+2 Attack +2 Defence per level in West Virginia tiles",
     "FULL PASS — culturalState='WV'"),
    ("State Guard", "State/DC", "DC Guard",
     "commanderMod, culturalMod", "DC tiles", "+2 Attack +2 Defence per level in DC tiles",
     "FULL PASS — culturalState='DC'"),
    ("State Guard", "Canadian/QB", "Quebec Guard",
     "commanderMod, culturalMod", "CA-QB tiles", "+2 Attack +2 Defence per level in Quebec tiles",
     "FULL PASS — culturalState='CA - QB'"),
    ("State Guard", "Canadian/OT", "Ontario Guard",
     "commanderMod, culturalMod", "CA-OT tiles", "+2 Attack +2 Defence per level in Ontario tiles",
     "FULL PASS — culturalState='CA - OT'"),
    ("State Guard", "Canadian/NS", "Nova Scotia Guard",
     "commanderMod, culturalMod", "CA-NS tiles", "+2 Attack +2 Defence per level in Nova Scotia tiles",
     "FULL PASS — culturalState='CA - NS'"),
    ("State Guard", "Canadian/NB", "New Brunswick Guard",
     "commanderMod, culturalMod", "CA-NB tiles", "+2 Attack +2 Defence per level in New Brunswick tiles",
     "FULL PASS — culturalState='CA - NB'"),
    ("State Guard", "Canadian/PEI", "Prince Edward Island Guard",
     "commanderMod, culturalMod", "CA-PEI tiles", "+2 Attack +2 Defence per level in PEI tiles",
     "FULL PASS — culturalState='CA - PEI'"),
    ("State Guard", "Bahamas/BA", "Bahamas Guard",
     "commanderMod, culturalMod", "BA tiles", "+2 Attack +2 Defence per level in Bahamas tiles",
     "FULL PASS — culturalState='BA'"),

    # ── PROTECTOR BUFF MODS (USA) ───────────────────────────────────────────────
    ("Protector", "Protector/USA", "Mothman Presence",
     "commanderMod", "Magic upkeep", "+20 Ranged Attack, +15 Ranged Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats(); magic upkeep in _tick_status_effects()"),
    ("Protector", "Protector/USA", "Jersey Devil's Fury",
     "commanderMod", "Magic upkeep", "+25 Melee Attack, +10 Ranged, +10 Block while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Bigfoot's Solidarity",
     "commanderMod", "Magic upkeep", "+30 Block, +15 Melee Attack while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Thunderbird's Sovereignty",
     "commanderMod", "Magic upkeep", "+25 Ranged Attack, +10 Melee Attack while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Headless Terror",
     "commanderMod", "Magic upkeep", "+20 Attack, +10 Block; attackers become Terrified 2t while active",
     "FIRST PASS — retaliation in _apply_combat_status_effects(); stat bonus in surveySelf()"),
    ("Protector", "Protector/USA", "Chessie's Blessing",
     "commanderMod", "Magic upkeep", "+20 Block, +15 Ranged Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Bell Witch's Harassment",
     "commanderMod", "Magic upkeep", "+15 Attack, +20 Defence; attackers become Demoralized 2t while active",
     "FIRST PASS — retaliation in _apply_combat_status_effects(); stat bonus in surveySelf()"),
    ("Protector", "Protector/USA", "Old Ironsides' Hull",
     "commanderMod", "Magic upkeep", "+30 Shield, +20 Block while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Valley Forge's Will",
     "commanderMod", "Magic upkeep", "+10 Attack, +25 Block, +20 Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Snallygaster's Claim",
     "commanderMod", "Magic upkeep", "+20 Attack, +10 Ranged, +10 Block while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Paul Revere's Ride",
     "commanderMod", "Magic upkeep", "+15 Ranged, +10 Attack, +3 Movement while active",
     "FIRST PASS — movement via _apply_status_flags(); stats in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Liberty Bell's Resonance",
     "commanderMod", "Magic upkeep", "+25 Block, +15 Ranged Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Green Mountain Haunting",
     "commanderMod", "Magic upkeep", "+20 Block, +15 Ranged Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Presidential Decree",
     "commanderMod", "Magic upkeep", "+20 Attack, +20 Block, +15 Ranged, +15 Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Skunk Ape's Domain",
     "commanderMod", "Magic upkeep", "+20 Attack, +15 Block while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Eternal Vigilance",
     "commanderMod", "Magic upkeep", "+25 Block, +10 Attack while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/USA", "Lincoln's Mandate",
     "commanderMod", "Magic upkeep", "+15 Attack, +15 Block, +15 Ranged, +10 Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),

    # ── PROTECTOR BUFF MODS (CANADIAN) ─────────────────────────────────────────
    ("Protector", "Protector/CA", "Le Wendigo's Hunger",
     "commanderMod", "Magic upkeep", "+30 Melee Attack while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "Loup-Garou's Frenzy",
     "commanderMod", "Magic upkeep", "+25 Attack, +15 Block, +10 Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "Feux Follets' Misdirection",
     "commanderMod", "Magic upkeep", "+25 Ranged Defence, +15 Block while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "Mishepeshu's Depths",
     "commanderMod", "Magic upkeep", "+20 Block, +20 Ranged Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "La Corriveau's Cage",
     "commanderMod", "Magic upkeep", "+20 Ranged Attack, +15 Melee Attack while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "Le Carcajou's Tenacity",
     "commanderMod", "Magic upkeep", "+20 Attack, +15 Block, +10 Defence while active",
     "FIRST PASS — stats applied in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "La Chasse-Galerie",
     "commanderMod", "Magic upkeep", "+15 Attack, +15 Ranged, +4 Movement while active",
     "FIRST PASS — movement via _apply_status_flags(); stats in _apply_status_effects_to_stats()"),
    ("Protector", "Protector/CA", "Le Gougou's Terror",
     "commanderMod", "Magic upkeep", "+15 Attack, +20 Defence; attackers become Terrified 3t while active",
     "FIRST PASS — retaliation in _apply_combat_status_effects(); stat bonus in surveySelf()"),

    # ── NEGATIVE STATUS EFFECTS ─────────────────────────────────────────────────
    ("Negative", "Negative/Combat", "Stunned",
     "isNegative", "—", "Cannot make melee attacks this turn",
     "FULL PASS — attackBlocked set in _apply_status_flags(); meleePressed() checks attackBlocked"),
    ("Negative", "Negative/Combat", "Suppressed",
     "isNegative", "—", "Cannot fire ranged attacks this turn",
     "FULL PASS — attackBlocked set in _apply_status_flags(); rangedPressed() checks attackBlocked"),
    ("Negative", "Negative/Combat", "Shaken",
     "isNegative", "—", "Melee block halved",
     "FULL PASS — armyBlock ×0.5 in _apply_status_effects_to_stats()"),
    ("Negative", "Negative/Combat", "Terrified",
     "isNegative", "—", "Melee attack −50%, melee block −30%",
     "FULL PASS — armyPunch ×0.5, armyBlock ×0.7 in _apply_status_effects_to_stats()"),
    ("Negative", "Negative/Combat", "Routed",
     "isNegative", "—", "Cannot attack; melee attack zeroed; ranged defence halved",
     "FULL PASS — attackBlocked in flags; armyPunch=0, armyDefence ×0.5 in stats; triggered by >40% loss in battle"),
    ("Negative", "Negative/DoT", "Burning",
     "isNegative", "—", "Loses 5×unitCount manpower per turn (DoT)",
     "FULL PASS — DoT applied in _tick_status_effects(); triggered by Rocket Artillery weapon"),
    ("Negative", "Negative/Combat", "Blinded",
     "isNegative", "—", "Ranged attack and ranged defence halved",
     "FULL PASS — armyLaunch ×0.5, armyDefence ×0.5 in _apply_status_effects_to_stats()"),
    ("Negative", "Negative/Magic", "Hexed",
     "isNegative", "—", "Magic defence eliminated (armyMagicDefense = 0)",
     "FULL PASS — armyMagicDefense=0 in _apply_status_effects_to_stats()"),
    ("Negative", "Negative/DoT", "Diseased",
     "isNegative", "—", "Loses 3×unitCount manpower per turn (DoT); caused by tile corruption chance",
     "FULL PASS — DoT in _tick_status_effects(); corruption check in onTurnEnd(); Park Ranger immunity"),
    ("Negative", "Negative/Storm", "Waterlogged",
     "isNegative, stormMod", "Storm: Thunderstorm/Hurricane", "Melee and ranged attack −20%",
     "FULL PASS — applied by _apply_storm_debuffs(); armyPunch/armyLaunch ×0.8 in stats"),
    ("Negative", "Negative/Storm", "Frostbitten",
     "isNegative, stormMod", "Storm: Blizzard/Nor'easter", "Melee and ranged attack −30%",
     "FULL PASS — applied by _apply_storm_debuffs(); armyPunch/armyLaunch ×0.7 in stats"),
    ("Negative", "Negative/Combat", "Demoralized",
     "isNegative", "—", "Melee −20%, block −20%, all mil mods disabled (calculateMilMods skipped)",
     "FULL PASS — armyDemoralized flag set in surveySelf(); calculateMilMods() early return if set"),
    ("Negative", "Negative/Movement", "Exhausted",
     "isNegative", "—", "Movement reduced to 1 this turn",
     "FULL PASS — currentMovementPoints clamped to max(1,val) in _apply_status_flags()"),
    ("Negative", "Negative/Movement", "Bogged Down",
     "isNegative", "—", "Cannot move this turn (movement = 0)",
     "FULL PASS — currentMovementPoints = 0 in _apply_status_flags(); triggered by Tornado storm"),
    ("Negative", "Negative/Combat", "Pacified",
     "isNegative", "—", "Cannot initiate attacks this turn",
     "FULL PASS — attackBlocked in _apply_status_flags()"),
    ("Negative", "Negative/Supply", "Supply Cut",
     "isNegative", "—", "Cannot reinforce or resupply this turn",
     "FULL PASS — reinforcementBlocked in _apply_status_flags(); onTurnEnd() skips refill if set"),
    ("Negative", "Negative/DoT", "Quarantined",
     "isNegative", "—", "No reinforcement; loses 1×unitCount manpower per turn",
     "FULL PASS — reinforcementBlocked in flags; light DoT in _tick_status_effects()"),
    ("Negative", "Negative/Funny", "Seduced",
     "isNegative", "—", "Commander distracted: melee zeroed; cannot attack",
     "FULL PASS — armyPunch=0; attackBlocked in _apply_status_effects_to_stats() / _apply_status_flags()"),
    ("Negative", "Negative/Funny", "Starstruck",
     "isNegative", "—", "All stats −30% (armyPunch/Launch/Block/Defence ×0.7)",
     "FULL PASS — all four scores scaled in _apply_status_effects_to_stats()"),
    ("Negative", "Negative/Funny", "Hangover",
     "isNegative", "—", "All stats −50% (armyPunch/Launch/Block/Defence ×0.5)",
     "FULL PASS — all four scores scaled in _apply_status_effects_to_stats()"),
    ("Negative", "Negative/Funny", "Love-Struck",
     "isNegative", "—", "Melee −70%, block −70%; cannot attack",
     "FULL PASS — armyPunch ×0.3, armyBlock ×0.3; attackBlocked in stats/flags"),
    ("Negative", "Negative/Funny", "Mutinous",
     "isNegative", "—", "Melee −40%; 50% chance army refuses orders each turn",
     "FULL PASS — armyPunch ×0.6; 50% random attackBlocked in _apply_status_flags()"),
]


def build_milmods_sheet(wb):
    ws = wb.create_sheet("Mil Mods")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("TIER",         12),
        ("CATEGORY",     20),
        ("MOD TYPE",     22),
        ("FLAGS",        28),
        ("RESOURCE",     14),
        ("EFFECT",       50),
        ("STATUS / NOTES", 46),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    current_tier = None
    row = 2
    for (tier, cat, mod_type, flags, resource, effect, notes) in MIL_MODS:
        if tier != current_tier:
            tier_bg = {
                "T1 (123)":    "2E5C8A",
                "T2 (23)":     "7A5C1E",
                "T3 (3)":      "8A2E2E",
                "Resource":    "5A5A5A",
                "Legacy":      "3A5A3A",
                "Weapon":      "2E6A5A",
                "Armor":       "5A2E6A",
                "Storm":       "1A4A6A",
                "Cultural":    "6A4A1A",
                "Tool":        "1A6A4A",
                "Special":     "5A1A7A",
                "State Guard": "1A6A3A",
                "Protector":   "7A5A00",
                "Negative":    "8A1A1A",
            }.get(tier, "2F4F6F")
            divider(ws, row, len(COLS), tier + " MODS", bg=tier_bg)
            row += 1
            current_tier = tier

        tbg = TIER_COLORS.get(tier, "FFFFFF")
        # Derive status color from notes
        if notes.startswith("FULL PASS"):
            sbg = STATUS_COLORS["FULL PASS"]
            sfg = STATUS_FG["FULL PASS"]
        elif notes.startswith("FIRST PASS"):
            sbg = STATUS_COLORS["FIRST PASS"]
            sfg = STATUS_FG["FIRST PASS"]
        elif notes.startswith("FIRST DRAFT"):
            sbg = STATUS_COLORS["FIRST DRAFT"]
            sfg = STATUS_FG["FIRST DRAFT"]
        elif notes.startswith("IDEA"):
            sbg = STATUS_COLORS["IDEA"]
            sfg = STATUS_FG["IDEA"]
        elif notes.startswith("CIVILIAN"):
            sbg = "EDD6F7"
            sfg = "3A1A6A"
        else:
            sbg = "FFFFFF"
            sfg = NORMAL_FG

        pcell(ws, row, 1, tier,      bg=tbg, bold=True, align="center")
        pcell(ws, row, 2, cat,       bg=tbg)
        pcell(ws, row, 3, mod_type,  bg=tbg, bold=True)
        pcell(ws, row, 4, flags,     bg="FAFAFA")
        pcell(ws, row, 5, resource,  bg="FAFAFA", align="center")
        pcell(ws, row, 6, effect,    bg="FAFAFA")
        pcell(ws, row, 7, notes,     bg=sbg, fg=sfg)
        ws.row_dimensions[row].height = 36
        row += 1

    ws.freeze_panes = "A2"


# ── SHEET 3: WEAPONS ──────────────────────────────────────────────────────────
# class, type, level, melee_off, melee_def, ranged_off, ranged_def, reload, ammo/level, charge_cost, weapon_mods, notes

WEAPONS = [
    # Sabers
    ("Saber", "Cutlass",        1,  3,  2,  0,  0,  0,  0, "10%", "SaberCharge",                     "Basic melee. FULL PASS."),
    ("Saber", "Cavalry Saber",  2,  5,  3,  0,  0,  0,  0, "10%", "SaberCharge, CavalryMorale",       "CavalryMorale unimplemented. FIRST PASS."),
    ("Saber", "Light Saber",    3,  7,  4,  0,  0,  0,  0, "10%", "SaberCharge",                     "FULL PASS."),
    ("Saber", "Heavy Saber",    4,  9,  5,  0,  0,  0,  0, "10%", "SaberCharge, HeavyCharge",        "HeavyCharge unimplemented. FIRST PASS."),
    # Muskets
    ("Musket", "Flintlock",     1,  1,  1,  4,  1,  2,  1, "—",  "Bayonet",                          "Bayonet melee at 50% penalty. FULL PASS."),
    ("Musket", "Brown Bess",    2,  2,  1,  6,  1,  2,  1, "—",  "Bayonet, VolleyFire",              "VolleyFire unimplemented. FIRST PASS."),
    ("Musket", "Percussion Cap",3,  2,  2,  8,  2,  1,  1, "—",  "Bayonet",                          "Reload 1 (faster). FULL PASS."),
    ("Musket", "Lever Repeater",4,  3,  2, 10,  2,  0,  2, "—",  "Bayonet, RapidFire",              "No reload (fires every round). Burns 2 ammo. FULL PASS."),
    # Artillery
    ("Artillery", "Falconet",   1,  0,  0, 12,  0,  3,  3, "—",  "CannonBlast",                     "No melee. FIRST PASS."),
    ("Artillery", "Field Gun",  2,  0,  0, 18,  0,  3,  3, "—",  "CannonBlast, AreaDamage",          "AreaDamage unimplemented. FIRST PASS."),
    ("Artillery", "Howitzer",   3,  0,  0, 25,  0,  2,  4, "—",  "CannonBlast, AreaDamage",          "FIRST PASS."),
    ("Artillery", "Mortar",     4,  0,  0, 35,  0,  2,  5, "—",  "CannonBlast, AreaDamage, Siege",   "Siege unimplemented. 5 ammo/level. FIRST PASS."),
    # Legacy
    ("Legacy", "Atlatl",        1,  0,  0,  2,  2,  0,  0, "—",  "AtlatlPierce",                    "DODK compat."),
    ("Legacy", "Club",          1,  2,  2,  0,  0,  0,  0, "—",  "ClubBleed",                       "DODK compat."),
    ("Legacy", "Double Axe",    1,  4,  0,  0,  0,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Flail",         1,  1,  2,  0,  1,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Longsword",     1,  3,  0,  0,  1,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Mace",          1,  2,  1,  0,  1,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Machete",       1,  2,  0,  0,  2,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Macuahuitl",    1,  3,  1,  0,  0,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Pike",          1,  1,  3,  0,  0,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Shortsword",    1,  0,  2,  0,  2,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Tomahawk",      1,  1,  0,  3,  0,  0,  0, "—",  "—",                                "DODK compat."),
    ("Legacy", "Spear",         1,  1,  1,  1,  1,  0,  0, "—",  "—",                                "DODK compat."),
    # Mythic — Easter egg event-unlockable weapons
    ("Mythic", "Baseball Bat",        1,  6,  3,  0,  0,  0,  0, "5%",  "BatSweep",         "Unlocked via events only. +10% ATK; first hit ignores shields."),
    ("Mythic", "Trident",             2,  8,  5,  4,  0,  1,  1, "5%",  "TridentPierce",    "Throwable; pierces shields; +3 ATK near naval tiles."),
    ("Mythic", "Mythic Atlatl",       2,  3,  2, 10,  3,  1,  1, "—",   "MythicAtlatl",     "Upgraded Atlatl; +2 ranged/level; stuns on crit."),
    ("Mythic", "Sharps Carbine",      3,  2,  2, 14,  2,  1,  1, "—",   "SharpShot",        "Civil War sniper rifle; ignores 4 DEF/level; terrain cover ignored."),
    ("Mythic", "Blackbeard's Pistols",2,  5,  2,  8,  0,  2,  2, "5%",  "PirateVolley",     "Dual pistols; fires twice per round; +5 terror morale."),
    ("Mythic", "Colt Revolver",       3,  3,  2, 12,  1,  0,  1, "—",   "CylinderFire",     "6-shot cylinder; no reload for 3 turns then 2-turn reload."),
    ("Mythic", "Rocket Artillery",    4,  0,  0, 45,  0,  3,  6, "—",   "RocketBarrage",    "Massive AoE; leaves fire modifier on tile 2 turns."),
    ("Mythic", "Trebuchet",           3,  0,  2, 30,  0,  4,  4, "—",   "TrebuchetLaunch",  "+50% vs Fortress; stuns defenders 1 round."),
    ("Mythic", "Wright Flyer",        4,  0,  5, 20, 10,  2,  3, "—",   "AerialBombing",    "Ignores all ground defense bonuses; terrifies enemy."),
]


def build_weapons_sheet(wb):
    ws = wb.create_sheet("Weapons")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("CLASS",       12),
        ("NAME",        18),
        ("LEVEL",        7),
        ("MELEE ATK",    9),
        ("MELEE DEF",    9),
        ("RANGED ATK",   9),
        ("RANGED DEF",   9),
        ("RELOAD",       8),
        ("AMMO/LVL",     9),
        ("CHARGE COST", 11),
        ("WEAPON MODS", 36),
        ("NOTES",       36),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    row = 2
    current_cls = None
    for (cls, name, level, m_atk, m_def, r_atk, r_def, reload, ammo, charge, mods, notes) in WEAPONS:
        if cls != current_cls:
            cls_bg = {"Saber": "2E5C8A", "Musket": "5A3A1E", "Artillery": "8A2E2E", "Legacy": "5A5A5A", "Mythic": "6A1A6A"}.get(cls, "2F4F6F")
            divider(ws, row, len(COLS), cls + " CLASS", bg=cls_bg)
            row += 1
            current_cls = cls

        cls_colors = {"Saber": "D6E4F7", "Musket": "FFF3CC", "Artillery": "FCE0D6", "Legacy": "DCDCDC", "Mythic": "F5D6F7"}
        bg = cls_colors.get(cls, "FAFAFA")

        pcell(ws, row,  1, cls,    bg=bg, bold=True, align="center")
        pcell(ws, row,  2, name,   bg=bg, bold=True)
        pcell(ws, row,  3, level,  bg=bg, align="center")
        pcell(ws, row,  4, m_atk,  bg=bg, align="center")
        pcell(ws, row,  5, m_def,  bg=bg, align="center")
        pcell(ws, row,  6, r_atk,  bg=bg, align="center")
        pcell(ws, row,  7, r_def,  bg=bg, align="center")
        pcell(ws, row,  8, reload, bg=bg, align="center")
        pcell(ws, row,  9, ammo,   bg=bg, align="center")
        pcell(ws, row, 10, charge, bg=bg, align="center")
        pcell(ws, row, 11, mods,   bg="FAFAFA")
        pcell(ws, row, 12, notes,  bg="FFFBE6")
        ws.row_dimensions[row].height = 28
        row += 1

    ws.freeze_panes = "A2"


# ── SHEET 4: ARMOR & UNIFORMS ─────────────────────────────────────────────────

ARMORS = [
    # Uniforms
    ("Uniform", "Continental",     15, 20, 10, 10, "DrillFormation",           "Balanced all-round. Formation bonus unimplemented. FIRST PASS."),
    ("Uniform", "Redcoat",         20, 15, 10, 10, "LineFormation",            "Strong melee. Line bonus unimplemented. FIRST PASS."),
    ("Uniform", "Militia",         10, 10,  5,  5, "—",                        "Cheap, basic. No mods. FULL PASS."),
    ("Uniform", "Light Infantry",  10, 30, 10, 15, "Skirmish",                 "Strong ranged defense. Skirmish unimplemented. FIRST PASS."),
    ("Uniform", "Heavy Infantry",  35, 10, 10,  8, "ShockTroop",               "Strong melee block. ShockTroop unimplemented. FIRST PASS."),
    ("Uniform", "Cavalry",         25,  5, 15,  5, "MountedCharge, CavalryMorale", "Both mods blocked by disabled check bug. FIRST PASS."),
    ("Uniform", "Artillery Corps",  5,  5, 20, 20, "GunCrewEfficiency",        "GunCrewEfficiency WORKS (bypasses disabled check). FULL PASS."),
    ("Uniform", "Minuteman",       12, 18,  8,  8, "MinutemanSpirit",          "MinutemanSpirit unimplemented. FIRST PASS."),
    ("Uniform", "Tombstone Cap",   10, 20,  5, 12, "QuickDraw",               "Frontier hat; first ranged attack each battle +5 bonus damage. FIRST DRAFT."),
    ("Uniform", "Hardee Hat",      15, 25, 10, 10, "HardeeDisc",              "Civil War dress hat; +3 DEF/level when adjacent friendly unit present. FIRST DRAFT."),
    # Legacy
    ("Legacy", "Canine",           15,  5, 30,  1, "—",                        "DODK compat."),
    ("Legacy", "Cast",             15, 10, 25, 20, "—",                        "DODK compat."),
    ("Legacy", "Chain",             5, 40,  5, 15, "Chain (weapon mod)",        "Chain mod adds to rangedOffence (inverted bug). DODK compat."),
    ("Legacy", "Archer",            0, 50,  0,  1, "—",                        "DODK compat."),
    ("Legacy", "Plate",            50,  0,  0,  1, "—",                        "DODK compat."),
    ("Legacy", "Padded",           10, 30, 10,  1, "—",                        "DODK compat."),
    ("Legacy", "Scout",            16, 16, 16, 90, "—",                        "DODK compat."),
    ("Legacy", "Scale",            28,  8, 14,  6, "Scale (ore mod — +7 shield/level)", "DODK compat."),
    ("Legacy", "Shell",             0,  0, 50,  2, "Shell (weapon mod)",        "DODK compat."),
    ("Legacy", "Point",            25, 25,  0, 55, "—",                        "DODK compat."),
    ("Legacy", "Feline",           15,  5, 30,  1, "—",                        "DODK compat."),
    ("Legacy", "Otter",            15,  5, 30,  1, "—",                        "DODK compat."),
]


def build_armors_sheet(wb):
    ws = wb.create_sheet("Armor & Uniforms")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("CLASS",         10),
        ("NAME",          18),
        ("MELEE BLOCK %", 13),
        ("RANGED BLOCK %",13),
        ("SPELL BLOCK %", 12),
        ("WEAPONS/LVL",   12),
        ("EMBEDDED MODS", 36),
        ("NOTES",         40),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    row = 2
    current_cls = None
    for (cls, name, m_blk, r_blk, s_blk, wpn_lvl, mods, notes) in ARMORS:
        if cls != current_cls:
            cls_bg = {"Uniform": "2E5C3A", "Legacy": "5A5A5A"}.get(cls, "2F4F6F")
            divider(ws, row, len(COLS), cls, bg=cls_bg)
            row += 1
            current_cls = cls

        bg = "D6F0D6" if cls == "Uniform" else "DCDCDC"
        pcell(ws, row, 1, cls,      bg=bg, bold=True, align="center")
        pcell(ws, row, 2, name,     bg=bg, bold=True)
        pcell(ws, row, 3, m_blk,    bg=bg, align="center")
        pcell(ws, row, 4, r_blk,    bg=bg, align="center")
        pcell(ws, row, 5, s_blk,    bg=bg, align="center")
        pcell(ws, row, 6, wpn_lvl,  bg=bg, align="center")
        pcell(ws, row, 7, mods,     bg="FAFAFA")
        pcell(ws, row, 8, notes,    bg="FFFBE6")
        ws.row_dimensions[row].height = 28
        row += 1

    ws.freeze_panes = "A2"


# ── SHEET 5: ARCHETYPES ───────────────────────────────────────────────────────
# id, name, position, terrain, t1_mod1, t1_mod2, t2_mod, t3_mod, narrative hook

ARCHETYPES = [
    ("ARC_01","Wetlands Fisher",      "SCOUT",      "Wetlands",
     "Swamp Legs",       "Coastal Watch",    "Marine",           "Naval Supremacy",
     "Fish the inlets, fight the tides; knows every marsh and cove"),
    ("ARC_02","Appalachian Miner",    "ENGINEER",   "Foothills",
     "Hill Runner",      "Powder & Shot",    "Siege Line",       "Rampart",
     "Blasting rock face taught them how to blow walls too"),
    ("ARC_03","Ivy League Dropout",   "SCHOLAR",    "Metro",
     "Steady Line",      "Street Tough",     "Flanking Drill",   "Continental Line",
     "Book-learned and street-fought; theory meets cobblestones"),
    ("ARC_04","Seminole Fighter",     "WARRIOR",    "Wetlands / Farmlands",
     "Swamp Legs",       "Saber Drill",      "Guerrilla Tactics","Last Stand",
     "Fought the swamps before the British. Will fight them again."),
    ("ARC_05","Green Mountain Farmer","FARMER",     "Foothills / Farmlands",
     "Farmhand",         "Hill Runner",      "Corrupted Ground", "Liberator's Will",
     "The land is theirs by deed and by rifle"),
    ("ARC_06","Chesapeake Shipwright","ENGINEER",   "Wetlands",
     "Coastal Watch",    "Powder & Shot",    "Double Shot",      "Double Cannonade",
     "Built the ships, now mans the guns"),
    ("ARC_07","Loyalist Turncoat",    "SPY",        "Metro / Suburbs",
     "Street Tough",     "Marksman",         "Night Raider",     "Ghost March",
     "Knows the enemy's plans because they used to write them"),
    ("ARC_08","Tobacco Belt Drifter", "SCOUT",      "Farmlands",
     "Farmhand",         "Quick Reload",     "Corrupted Ground", "The Long March",
     "Covered a thousand miles on foot; a thousand more to go"),
    ("ARC_09","War Widow",            "DIPLOMAT",   "Suburbs / Metro",
     "Steady Line",      "Fortified Position","Rallying Voice",  "Undaunted",
     "Lost everything once; will not lose again"),
    ("ARC_10","Indigenous Scout",     "SCOUT",      "Woods / Wetlands",
     "Woodsman",         "Swamp Legs",       "Guerrilla Tactics","Ghost March",
     "The forest is a home, not an obstacle"),
    ("ARC_11","Boston Rabble-Rouser", "ORATOR",     "Metro",
     "Street Tough",     "Saber Drill",      "Rallying Voice",   "Terror",
     "Their voice is a weapon; their fists the punctuation"),
    ("ARC_12","Continental Surgeon",  "HEALER",     "Farmlands / Foothills",
     "Steady Line",      "Fortified Position","Cleaner",         "Liberator's Will",
     "Saves the wounded so they can fight again tomorrow"),
    ("ARC_13","Nantucket Sailor",     "ADMIRAL",    "Wetlands",
     "Coastal Watch",    "Marksman",         "Marine",           "Naval Supremacy",
     "Home is a deck; the shore is just waiting for wind"),
    ("ARC_14","Frontier Preacher",    "ORATOR",     "Woods / Foothills",
     "Woodsman",         "Hill Runner",      "Rallying Voice",   "Undaunted",
     "God and country; in that order"),
    ("ARC_15","DC Bureaucrat",        "BUREAUCRAT", "Metro",
     "Fortified Position","Street Tough",   "Flanking Drill",    "Iron Wall",
     "Knows every regulation and which ones to ignore"),
    ("ARC_16","Rust Belt Steelworker","ENGINEER",   "Suburbs",
     "Powder & Shot",    "Fortified Position","Siege Line",      "Double Cannonade",
     "Built the walls, knows where they crack"),
    ("ARC_17","Plantation Deserter",  "SOLDIER",    "Farmlands",
     "Farmhand",         "Saber Drill",      "Guerrilla Tactics","Last Stand",
     "Left the plantation at midnight; never looked back"),
    ("ARC_18","Swamp Witch",          "MAGE",       "Wetlands",
     "Swamp Legs",       "Coastal Watch",    "Cleaner",          "Terror",
     "The bayou has its own army and she commands it"),
    ("ARC_19","Caribbean Privateer",  "ADMIRAL",    "Wetlands / Suburbs",
     "Coastal Watch",    "Saber Drill",      "Marine",           "Double Cannonade",
     "No flag but profit; for now, the Republic pays better"),
    ("ARC_20","Hawaiian Refugee",     "DIPLOMAT",   "Wetlands / Metro",
     "Farmhand",         "Coastal Watch",    "Marine",           "Naval Supremacy",
     "Crossed the Pacific once. The Atlantic is nothing."),
    ("ARC_21","Border Mercenary",     "SOLDIER",    "Suburbs / Farmlands",
     "Hill Runner",      "Saber Drill",      "Iron Bayonet",     "Last Stand",
     "Fought for whoever paid; now fighting for something real"),
    ("ARC_22","Acadian Forest Ranger","SCOUT",      "Woods / Wetlands",
     "Woodsman",         "Quick Reload",     "Guerrilla Tactics","Ghost March",
     "The British burned their home; the forest remembers"),
    ("ARC_23","Gettysburg Descendant","SOLDIER",    "Farmlands / Foothills",
     "Steady Line",      "Saber Drill",      "Iron Bayonet",     "Entrenched",
     "A family name written in battlefield soil"),
    ("ARC_24","LGBTQ+ Organizer",     "DIPLOMAT",   "Metro / Suburbs",
     "Street Tough",     "Steady Line",      "Rallying Voice",   "Continental Line",
     "Built coalitions the old guard said were impossible"),
    ("ARC_25","Carnival Barker",      "ORATOR",     "Wetlands / Suburbs",
     "Street Tough",     "Marksman",         "Night Raider",     "The Long March",
     "Every performance is a battle; every crowd a campaign"),
]


def build_archetypes_sheet(wb):
    ws = wb.create_sheet("Archetypes")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("ID",           8),
        ("ARCHETYPE",   26),
        ("POSITION",    13),
        ("TERRAIN",     22),
        ("T1 MOD A",    20),
        ("T1 MOD B",    20),
        ("T2 MOD",      20),
        ("T3 MOD",      20),
        ("NARRATIVE HOOK", 52),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    row = 2
    for arc in ARCHETYPES:
        (arc_id, name, pos, terrain, t1a, t1b, t2, t3, hook) = arc

        pcell(ws, row, 1, arc_id,  bg="D6E4F7", bold=True, align="center")
        pcell(ws, row, 2, name,    bg="D6E4F7", bold=True)
        pcell(ws, row, 3, pos,     bg="FFF3CC", align="center")
        pcell(ws, row, 4, terrain, bg="D6F0D6")
        pcell(ws, row, 5, t1a,     bg="D6E4F7")
        pcell(ws, row, 6, t1b,     bg="D6E4F7")
        pcell(ws, row, 7, t2,      bg="FFF3CC")
        pcell(ws, row, 8, t3,      bg="FCE0D6")
        pcell(ws, row, 9, hook,    bg="FAFAFA")
        ws.row_dimensions[row].height = 32
        row += 1

    ws.freeze_panes = "A2"


# ── SHEET 6: BATTLE MECHANICS ─────────────────────────────────────────────────
# section, formula/rule, status, notes

BATTLE_MECHANICS = [
    # ── STAT PIPELINE ──────────────────────────────────────────────────────────
    ("Stat Pipeline", "Unit base stats",
     "FULL PASS",
     "offScore = level × weaponOffence × meleePenalty × effectMultiplier. "
     "effectMultiplier = (man% + wpn%) / 2, or wpn% only for artillery.",
     ""),
    ("Stat Pipeline", "Effect multiplier",
     "FULL PASS",
     "Both manpower and weapons must be adequate for full effect. "
     "Unit at 50% manpower + 100% weapons = 75% effectiveness (except artillery).",
     ""),
    ("Stat Pipeline", "Melee penalty",
     "FULL PASS",
     "Muskets fight at 50% offensive score in melee (weapon.get_melee_penalty() = 0.5). "
     "Sabers at 100%. Artillery cannot melee at all.",
     ""),
    ("Stat Pipeline", "Commander morale multiplier",
     "FULL PASS",
     "mm = 1.0 + (commander.morale / 100) × 0.25. Applied to armyPunch and armyDefence. "
     "Range: ×1.0 (morale 0) to ×1.25 (morale 100).",
     "governor.loyalty range is −20 to +10, not 0–100. Morale input to formula may be loyalty × some factor. Needs clarification."),

    # ── MELEE ROUND ────────────────────────────────────────────────────────────
    ("Melee Round", "Base attack",
     "FULL PASS",
     "raw_attack = attacker.armyPunch (sum of all unit offensiveScores).",
     ""),
    ("Melee Round", "Block reduction",
     "FULL PASS",
     "block_ratio = clamp(armyBlock / unitCount, 0.0, 0.9). "
     "net_damage = raw_attack × (1 − block_ratio).",
     "Max 90% block cap prevents total immunity."),
    ("Melee Round", "Shield absorption",
     "FULL PASS",
     "Shield absorbs before manpower. defenderShieldLoss = min(shield, net_damage). "
     "remainder = net_damage − shieldLoss → applied to manpower.",
     ""),
    ("Melee Round", "Saber charge cost",
     "FULL PASS",
     "Each saber unit pays 10% of its currentManpower on every melee attack. "
     "Cost applied in battle.gd _apply_charge_costs() after battle resolves.",
     ""),
    ("Melee Round", "Simultaneous counter",
     "FULL PASS",
     "Defender counter-attacks simultaneously in melee. "
     "Same formula: defender.armyPunch × (1 − attacker.block_ratio).",
     ""),

    # ── RANGED ROUND ───────────────────────────────────────────────────────────
    ("Ranged Round", "Effective launch",
     "FULL PASS",
     "effective_launch = sum of get_effective_ranged_offence() per unit. "
     "Returns 0 for reloading units, full unitRangedOffence if ready.",
     ""),
    ("Ranged Round", "Ranged block",
     "FULL PASS",
     "ranged_block_ratio = clamp(armyDefence / unitCount, 0.0, 0.9). "
     "Same shield-then-manpower absorption as melee.",
     ""),
    ("Ranged Round", "Ranged counter",
     "FULL PASS",
     "Defender may return fire if they have ready ranged units. "
     "Pure artillery attackers receive no defensive counter (can't melee, no counter).",
     "Artillery special case comment in code but not fully enforced."),
    ("Ranged Round", "Reload",
     "FULL PASS",
     "After firing: unit.start_reload() sets reloadCounter = reloadTurns. "
     "GunCrewEfficiency reduces reloadTurns by 1 at start_reload(). "
     "tick_reload() called each round. Unit fires again when reloadCounter reaches 0.",
     "Reload only ticks during ranged battle rounds. Melee rounds do not advance reload."),

    # ── MORALE & ARMY DEATH ────────────────────────────────────────────────────
    ("Army Death", "Civ-style army destruction",
     "FULL PASS",
     "When manpowerInArmy ≤ 0 after battle, army emits signal armyDestroyed(self). "
     "country._on_army_destroyed() erases from countryArmyList, nulls tile.stationedArmy, queue_free(). "
     "No retreat system — armies die in place, instantly.",
     "inRetreat/retreatTarget vars still declared on Army but no longer emitted."),

    # ── DOUBLE SHOT ───────────────────────────────────────────────────────────
    ("Double Shot", "Second volley (Double Shot T2)",
     "FIRST PASS",
     "After normal ranged calculation, _army_has_siege_mod() checks for Double Shot. "
     "second_multiplier = 0.5 (50% power). Loops artillery units only. "
     "Second net applied to remaining shield then to manpower.",
     "Only applies to artillery weapon class, not muskets."),
    ("Double Shot", "Second volley (Double Cannonade T3)",
     "FIRST PASS",
     "Same as Double Shot but second_multiplier = 1.0 (full power). "
     "Additional second_bonus = artillery unit count × 3. "
     "Overrides Double Shot if both mods somehow present.",
     ""),

    # ── SIEGE ──────────────────────────────────────────────────────────────────
    ("Siege", "Siege score",
     "FIRST DRAFT",
     "armySiegeScore = unitCount × 0.1 × tile.get_siege_difficulty(). "
     "tile.get_siege_difficulty() calculates based on tile buildings.",
     "Siege score never read in battle.gd. No siege battle type implemented."),
    ("Siege", "Fortification defense",
     "FIRST DRAFT",
     "Siege Line mod designed to bypass fortification penalties. "
     "No fortification penalty currently exists in _calculate_ranged_damage.",
     ""),

    # ── BATTLE UI ──────────────────────────────────────────────────────────────
    ("Battle UI", "Projected damage display",
     "FULL PASS",
     "Battle UI shows current manpower bars and projected post-battle values "
     "for both armies before player clicks Attack.",
     ""),
    ("Battle UI", "Reload indicator",
     "FULL PASS",
     "In ranged mode, UI shows '(N reloading)' suffix on attacker's launch stat.",
     ""),
    ("Battle UI", "Battle type display",
     "FULL PASS",
     "Melee shows: Shield / Punch / Block% for each side. "
     "Ranged shows: Shield / Launch / Defence% for each side.",
     ""),
]


def build_battle_sheet(wb):
    ws = wb.create_sheet("Battle Mechanics")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("SECTION",         16),
        ("MECHANIC",        28),
        ("STATUS",          13),
        ("FORMULA / RULE",  58),
        ("GAPS / NOTES",    40),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    row = 2
    current_sec = None
    for (sec, mech, status, rule, notes) in BATTLE_MECHANICS:
        if sec != current_sec:
            divider(ws, row, len(COLS), sec)
            row += 1
            current_sec = sec

        sbg = STATUS_COLORS.get(status, "FFFFFF")
        sfg = STATUS_FG.get(status, NORMAL_FG)

        pcell(ws, row, 1, sec,    bg="EEF2F7")
        pcell(ws, row, 2, mech,   bg="EEF2F7", bold=True)
        pcell(ws, row, 3, status, bg=sbg, fg=sfg, bold=True, align="center")
        pcell(ws, row, 4, rule,   bg="FAFAFA")
        pcell(ws, row, 5, notes,  bg="FFFBE6")
        ws.row_dimensions[row].height = 42
        row += 1

    ws.freeze_panes = "A2"


# ── SHEET 7: KNOWN BUGS ───────────────────────────────────────────────────────

BUGS = [
    ("RESOLVED", "calculateMilMods() inverted disabled check",
     "unit.gd",
     "FIXED: Changed `if MilMod.disabled != false` to `if not MilMod.disabled`. "
     "All mod effects in calculateMilMods() now apply correctly when mods are enabled.",
     "RESOLVED — all terrain, storm, stat, and special mods now active."),

    ("RESOLVED", "addMilMod() skips buildSelf()",
     "governor.gd + mil_mod.gd",
     "FIXED: Added `if has_node('Sprite2D'):` guard to last 5 lines of buildSelf() "
     "so MilMod.new() instances (no scene children) can call buildSelf() safely. "
     "Uncommented `newMM.buildSelf(type)` in governor.gd addMilMod().",
     "RESOLVED — all governor mil mods now have proper flags set."),

    ("RESOLVED", "deleteMode + inRetreat: army not removed on death",
     "army.gd + country.gd",
     "FIXED: Replaced deleteMode/inRetreat with `signal armyDestroyed`. "
     "Emitted when manpowerInArmy ≤ 0 in calculateAttackerResults/calculateDefenderResults. "
     "country.gd addArmy() connects signal to _on_army_destroyed() which erases from "
     "countryArmyList, nulls tile.stationedArmy, and calls queue_free().",
     "RESOLVED — Civ-style army death fully wired."),

    ("MEDIUM", "Weapon cost matching uses legacy names only",
     "army.gd surveySelf() lines 357-365",
     "The match block that calculates armyWeaponsCost uses old weapon names "
     "(Spear, Club, Atlatl etc). No case handles Cutlass, Flintlock, Field Gun, etc. "
     "New weapon armies show 0 weapons cost.",
     "Add Saber/Musket/Artillery weapon cost cases to the match block."),

    ("MEDIUM", "morale vs loyalty confusion in armyPunch multiplier",
     "army.gd surveySelf() line 371",
     "Code uses `commander.morale` but governor.gd has `loyalty` (−20 to +10) "
     "and the morale var may be a separate 0-100 field not clearly initialized. "
     "If morale is always 0, the multiplier is always 1.0 (no bonus ever).",
     "Audit governor morale initialization; confirm formula uses correct field."),

    ("MEDIUM", "commanderCheck() adds mods to units on every updateArmyUI call",
     "army.gd commanderCheck()",
     "commanderCheck() appends mods to each unit via Unit.addMilMod(MilMod) on every call. "
     "Units clear militaryModifierList in getUnitAttributes() via removeMilMod, "
     "but commanderCheck runs AFTER getUnitAttributes in the chain — units accumulate "
     "duplicate mods if updateArmyUI is called multiple times in quick succession.",
     "Verify that getUnitAttributes() fully clears militaryModifierList before "
     "commanderCheck runs. Audit the call order in updateArmyUI()."),

    ("LOW", "GunCrewEfficiency is the only mod that bypasses the disabled check",
     "unit.gd start_reload() line 220-223",
     "GunCrewEfficiency is checked via a separate milModType string comparison in start_reload(), "
     "bypassing the broken calculateMilMods() disabled check. This is the only mod that "
     "actually works. All other mod effects are dead until the disabled check is fixed.",
     "After fixing the disabled check, ensure GunCrewEfficiency is not double-applied."),

    ("LOW", "armyShield vs armyMaxShield display inconsistency",
     "army.gd updateFinalTotals()",
     "updateFinalTotals() passes armyShield for both current and max values in armyCostUI3. "
     "Should pass (armyShield, armyMaxShield) to show depletion.",
     "Change: armyCostUI3.updateSelf(armyShield, armyMaxShield)"),
]


def build_bugs_sheet(wb):
    ws = wb.create_sheet("Known Bugs")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("SEVERITY",   12),
        ("BUG",        34),
        ("LOCATION",   28),
        ("DESCRIPTION",58),
        ("SUGGESTED FIX", 46),
    ]
    for ci, (h, w) in enumerate(COLS, 1):
        hcell(ws, 1, ci, h, w)

    SEV_COLORS = {
        "CRITICAL": ("FF0000", "FFFFFF"),
        "HIGH":     ("FF6B35", "FFFFFF"),
        "MEDIUM":   ("FFEB9C", "7A5A00"),
        "LOW":      ("D6E4F7", "1A3A6A"),
        "RESOLVED": ("00B050", "FFFFFF"),
    }

    row = 2
    for (sev, bug, loc, desc, fix) in BUGS:
        sbg, sfg = SEV_COLORS.get(sev, ("FFFFFF", NORMAL_FG))
        pcell(ws, row, 1, sev,  bg=sbg, fg=sfg, bold=True, align="center")
        pcell(ws, row, 2, bug,  bg="FAFAFA", bold=True)
        pcell(ws, row, 3, loc,  bg="FAFAFA")
        pcell(ws, row, 4, desc, bg="FAFAFA")
        pcell(ws, row, 5, fix,  bg="FFFBE6")
        ws.row_dimensions[row].height = 54
        row += 1

    ws.freeze_panes = "A2"


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    build_systems_sheet(wb)
    build_milmods_sheet(wb)
    build_weapons_sheet(wb)
    build_armors_sheet(wb)
    build_archetypes_sheet(wb)
    build_battle_sheet(wb)
    build_bugs_sheet(wb)

    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH}")
    print(f"  Sheet 1 — Systems Overview:  {len(SYSTEMS)} entries")
    print(f"  Sheet 2 — Mil Mods:          {len(MIL_MODS)} mods")
    print(f"  Sheet 3 — Weapons:           {len(WEAPONS)} weapon types")
    print(f"  Sheet 4 — Armor & Uniforms:  {len(ARMORS)} armor types")
    print(f"  Sheet 5 — Archetypes:        {len(ARCHETYPES)} archetypes")
    print(f"  Sheet 6 — Battle Mechanics:  {len(BATTLE_MECHANICS)} mechanics")
    print(f"  Sheet 7 — Known Bugs:        {len(BUGS)} bugs")


if __name__ == "__main__":
    main()
