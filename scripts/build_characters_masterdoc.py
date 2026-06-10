#!/usr/bin/env python3
"""
Generate character_masterdoc.xlsx — physical + personality design bible for all
named characters in Farsword (excluding procedurally generated commanders).

Sheets:
  1. OVERVIEW     — character count by country/status, design completeness
  2. PLAYABLE     — USA named governors (hire-able characters)
  3. CANADIAN     — Canadian named governors and leaders
  4. NPC          — Non-playable named characters (event figures, antagonists, etc.)
  5. DESIGN NOTES — Art direction and consistency rules
  6. UALANI MODEL — Full production bible: construction, palette, expressions,
                    costume layers, scene-type direction (SFW → explicit)
  7. JESSICA MODEL — Same production bible format for Jessica Commanda Odjick

Status scale:
  FINAL     — physical + personality fully locked, art approved
  DESIGNED  — fully described, not yet in final art
  PARTIAL   — some info defined, gaps remain
  STUB      — character exists in code, not yet designed

Run from repo root:  python3 scripts/build_characters_masterdoc.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "character_masterdoc.xlsx"

# ── PALETTE ──────────────────────────────────────────────────────────────────
HDR_BG, HDR_FG     = "1F2D3D", "FFFFFF"
USA_BG, USA_FG     = "1F3864", "FFFFFF"
CA_BG,  CA_FG      = "9B1C1C", "FFFFFF"
NPC_BG, NPC_FG     = "4A3728", "FFFFFF"
ALT_BG             = "EEF2F7"
WHITE              = "FFFFFF"

FINAL_BG,   FINAL_FG   = "00B050", "FFFFFF"
DESIGN_BG,  DESIGN_FG  = "92D050", "1A3A00"
PARTIAL_BG, PARTIAL_FG = "FFEB9C", "7A5A00"
STUB_BG,    STUB_FG    = "C9B1E8", "3A1A6A"

STATUS_BG = {
    "FINAL":    FINAL_BG,
    "DESIGNED": DESIGN_BG,
    "PARTIAL":  PARTIAL_BG,
    "STUB":     STUB_BG,
}
STATUS_FG = {
    "FINAL":    FINAL_FG,
    "DESIGNED": DESIGN_FG,
    "PARTIAL":  PARTIAL_FG,
    "STUB":     STUB_FG,
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def _fill(h):
    return PatternFill("solid", fgColor=h)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, italic=False, color="000000", size=10):
    return Font(bold=bold, italic=italic, color=color, size=size, name="Calibri")

def _hdr(ws, row, col, text, bg=HDR_BG, fg=HDR_FG, bold=True, size=10, wrap=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=bold, color=fg, size=size)
    c.fill = _fill(bg)
    c.border = _border()
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def _cell(ws, row, col, text, bg=WHITE, bold=False, italic=False, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=bold, italic=italic)
    c.fill = _fill(bg)
    c.border = _border()
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    return c

def _status_cell(ws, row, col, status):
    bg = STATUS_BG.get(status, "FFFFFF")
    fg = STATUS_FG.get(status, "000000")
    c = ws.cell(row=row, column=col, value=status)
    c.font = _font(bold=True, color=fg)
    c.fill = _fill(bg)
    c.border = _border()
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ── CHARACTER DATA ─────────────────────────────────────────────────────────────
#
# Columns: name, position, faction, ethnicity_background, physical_description,
#          personality_traits, narrative_hook, relationships, loyalty_base, status
#
# Physical descriptions are the ART BIBLE — what artists draw, consistent across
# all appearances (cards, events, portraits, cutscenes).
#
# Personality traits are comma-separated keywords for writers (tone of voice,
# behavioral tendencies, dialogue beats).

USA_CHARACTERS = [
    {
        "name": "Ualani Carlisle",
        "position": "PRESIDENT & COMMANDER",
        "faction": "Federal",
        "ethnicity": "Hawaiian / Filipina",
        "physical": (
            "Athletic build, medium height. Warm brown skin. Belly-length straight black hair, "
            "usually worn loose or in a single thick braid during field command. Gold hoop earrings "
            "always present — a personal constant across every context. Sharp dark eyes with "
            "a commander's flat affect. Presidential uniform: deep blue with gold braid and "
            "campaign medals. In the field: same uniform, sleeves rolled, hair braided back. "
            "No powder, no wig. The face of a republic that stopped asking for permission."
        ),
        "personality": "Decisive, unflappable, dry wit, physically present in danger, reads rooms instantly, allergic to pomp, loyal to institutions over individuals",
        "hook": "Commands the APF personally. Security detail has filed 17 formal objections. She has read none of them.",
        "relationships": "Marc Penoit: alliance partner, mutual respect; Jessica Commanda Odjick: peer leader, growing trust; Benjamin Tallmadge: relies on his intel, trusts his discretion",
        "loyalty_base": 20,
        "status": "DESIGNED",
    },
    {
        "name": "Patrick Henry",
        "position": "ORATOR",
        "faction": "Patriot",
        "ethnicity": "Anglo-Virginian",
        "physical": (
            "Lean, angular frame — a man who eats when there's time. Sharp jaw, sunken cheeks, "
            "intense dark eyes that catch light when he speaks. Mid-50s. Dark riding coat, no "
            "powder — he considers it aristocratic affectation. Long fingers, built for pointing "
            "at things accusingly. Often seen mid-gesture, jaw set. The kind of face that looks "
            "like it is always about to say something that will get everyone in trouble."
        ),
        "personality": "Incendiary, rhetorically brilliant, suspicious of centralized power (including his own side), principle over pragmatism, crowd-reader, easily roused, slow to forgive",
        "hook": "'Give me liberty or give me death' — and he means it about every policy disagreement.",
        "relationships": "Ualani Carlisle: respects her directness, uneasy about Federal authority; Thomas Paine: ideological kinship, competitive",
        "loyalty_base": 8,
        "status": "DESIGNED",
    },
    {
        "name": "Abigail Adams",
        "position": "DIPLOMAT",
        "faction": "Moderate",
        "ethnicity": "Anglo-New England",
        "physical": (
            "Composed and precise. Oval face with an expression of patient, slightly disappointed "
            "intelligence. Brown eyes that miss nothing. Dark hair pinned beneath a white cap in "
            "public; loose when writing at her desk — which is where she does her real work. "
            "Colonial dress: practical wool in dark greens and blues, no unnecessary ornament. "
            "The posture of someone who has spent thirty years editing other people's correspondence "
            "and making it better than theirs."
        ),
        "personality": "Methodical, principled, politically sharp, diplomatically patient, holds grudges gracefully, pushes institutions toward their own stated ideals, unfazed by powerful men",
        "hook": "'Remember the ladies' — she said it once and has spent the rest of her career proving she meant it structurally.",
        "relationships": "Patrick Henry: ideological friction; Thomas Paine: finds him too chaotic; Governor Carleton: has his measure completely",
        "loyalty_base": 7,
        "status": "DESIGNED",
    },
    {
        "name": "Thomas Paine",
        "position": "SCHOLAR",
        "faction": "Radical",
        "ethnicity": "Anglo (English immigrant)",
        "physical": (
            "Heavy-browed, broad forehead, quick restless eyes. Mid-40s but looks older from "
            "travel and argument. Ink-stained fingers — always. Worn greatcoat, the collar "
            "turned up against a wind that isn't there. Doesn't own a wig and would never. "
            "Often holding something he's reading. The kind of face that looks like it arrived "
            "from somewhere worse and found this place merely disappointing by comparison."
        ),
        "personality": "Principled to the point of impracticality, believes in people absolutely, distrust of all governments including good ones, tireless pamphleteer, combustible under bad-faith argument, genuinely egalitarian",
        "hook": "His pamphlets lit the fire. He's here to make sure it doesn't burn the wrong things.",
        "relationships": "Patrick Henry: mutual intensity, different targets; Abigail Adams: respects her, exasperates her; Daniel Shays: understands his grievance instinctively",
        "loyalty_base": 4,
        "status": "DESIGNED",
    },
    {
        "name": "Mercy Otis Warren",
        "position": "SCHOLAR",
        "faction": "Patriot",
        "ethnicity": "Anglo-New England",
        "physical": (
            "Sharp-featured, dark quick eyes, an expression of focused appraisal. Mid-50s. "
            "Dark wool dress, simple — she dresses like someone who doesn't need clothes to "
            "make a statement because her words do it. Quill in hand or within reach always. "
            "The posture of a playwright: watching the scene from the wings even when seated "
            "in the center of it."
        ),
        "personality": "Observational, politically precise, satirically sharp, keeps the movement honest, resistant to propaganda from her own side, long memory for hypocrisy",
        "hook": "Her pen is a scalpel. The revolution is her patient. She is not optimistic about the prognosis but keeps working.",
        "relationships": "Abigail Adams: close intellectual partnership; Patrick Henry: respects his fire, documents his contradictions; Thomas Paine: admires his clarity",
        "loyalty_base": 7,
        "status": "DESIGNED",
    },
    {
        "name": "Daniel Shays",
        "position": "FARMER",
        "faction": "Populist",
        "ethnicity": "Anglo-Yankee (Massachusetts)",
        "physical": (
            "Broad-shouldered, weathered. Early 40s but looks 50 — winters, debt, marching. "
            "Continental Army jacket, worn through at the elbows, buttons mismatched where "
            "they've been replaced. Calloused hands. Sun-cracked face with a jaw that's been "
            "set since 1786. Doesn't carry a sword. Carries an axe. The kind of man who "
            "looks exactly like what he is: someone who did everything asked of him and then "
            "found out what it cost."
        ),
        "personality": "Grievance-driven, stubborn, practical anger, distrustful of institutions that have failed him, loyalty earned not given, protective of his people, prone to direct action over deliberation",
        "hook": "He fought for a republic that immediately started taxing him into the ground. He is not impressed by speeches.",
        "relationships": "Thomas Paine: Paine understands him theoretically; Paine has never lost a farm",
        "loyalty_base": 1,
        "status": "DESIGNED",
    },
    {
        "name": "Benjamin Tallmadge",
        "position": "SPYMASTER",
        "faction": "Patriot",
        "ethnicity": "Anglo-Connecticut",
        "physical": (
            "Neat. Precisely neat — the kind of neat that is itself a form of information "
            "control. Mid-30s, trim build, unremarkable in every deliberate way. Brown eyes "
            "that stay still while they assess. Dark coat, clean, no insignia. Hair queued "
            "back — nothing loose. Ink on the inside of his left wrist, below the cuff, "
            "where only he can see it. The face of someone you forgot you met."
        ),
        "personality": "Methodical, deeply loyal, trusts process over instinct, catalogues everything, deeply suspicious of everyone (including himself), calm under duress, keeps ledgers",
        "hook": "The Culper Ring never stopped running. It merely changed names. He is the only one who knows how many names.",
        "relationships": "Ualani Carlisle: serves her completely; Abigail Adams: overlapping intelligence networks; Lord Cornwallis: professional respect across enemy lines",
        "loyalty_base": 9,
        "status": "DESIGNED",
    },
    {
        "name": "Phillis Wheatley",
        "position": "HERALD",
        "faction": "Abolitionist League",
        "ethnicity": "African (Senegalese-born, Boston-raised)",
        "physical": (
            "Small, slight frame with entirely disproportionate presence. Late 20s. "
            "Dark brown skin, large expressive eyes, the careful posture of someone who has "
            "spent a lifetime being watched and learned to use it. Simple but dignified dress — "
            "quality cloth, no excess. Usually holds a quill or has one tucked behind her ear. "
            "The face of someone who has been told, repeatedly, that she cannot do what she "
            "is currently doing."
        ),
        "personality": "Incisive, formally precise, strategically patient, uses language as a weapon, aware of her symbolic weight, refuses to let the revolution forget its contradictions, dignified fury",
        "hook": "She met Washington and wrote him a poem and the poem was better than the war. Crown officers are now confiscating her work. That's how you know it's working.",
        "relationships": "Abigail Adams: mutual regard, different leverage; Francis Asbury: unexpected ideological kinship; Ualani Carlisle: the president reads her work",
        "loyalty_base": 6,
        "status": "DESIGNED",
    },
    {
        "name": "Francis Asbury",
        "position": "CIRCUIT PREACHER",
        "faction": "Common Cause",
        "ethnicity": "Anglo (English immigrant)",
        "physical": (
            "Lean from 300,000 miles of horseback. Mid-50s, permanently wind-burned, "
            "permanently in motion. Simple black Methodist coat, worn smooth at the elbows. "
            "Practical boots. White hair, not powdered — just white. Sharp eyes with "
            "the slightly alarming focus of a man who has decided God wants him specifically "
            "to cover a great deal of ground. His horse looks exhausted. He does not."
        ),
        "personality": "Tireless, democratic, anti-slavery, impossible to stop, converts by sheer relentless presence, more interested in frontier farms than city pulpits, disarming to enemies",
        "hook": "Crown forces tried to arrest him twice. He preached at both arresting officers. One converted.",
        "relationships": "Phillis Wheatley: shared abolition ground; Daniel Shays: knows every farmstead Shays came from; Marc Penoit: has ridden through Habitant territory",
        "loyalty_base": 5,
        "status": "DESIGNED",
    },
]

CANADIAN_CHARACTERS = [
    {
        "name": "Jessica Commanda Odjick",
        "position": "PRIME MINISTER",
        "faction": "Algonquin Nation",
        "ethnicity": "Algonquin (Kitigan Zibi Anishinàbeg)",
        "physical": (
            "Tall for the room she usually walks into. Late 30s. High cheekbones, sharp jaw, "
            "steady dark eyes that don't blink first in any conversation. Long black hair worn "
            "loose or in a single braid threaded with indigo ribbon — the braid appears in "
            "formal settings, the loose hair in the field. Wears a diplomat's dark overcoat "
            "over traditional Algonquin ribbon-work shirt, the beadwork visible at the collar "
            "and cuffs — a deliberate choice. Not a concession to either world; a statement "
            "that she moves through both on her own terms. Carries nothing decorative. "
            "Everything she wears does something."
        ),
        "personality": "Uncompromising, multilingual, historically precise, reads power structures immediately, refused to ask permission before the Governor's Council did, patient with complexity, direct with dishonesty",
        "hook": "The Governor's Council objected to her presence at the table. The Governor's Council is no longer at the table.",
        "relationships": "Marc Penoit: weekly disagreements, zero desertions — that's trust; Ualani Carlisle: two leaders learning each other's language; Governor Carleton: she has his measure, he has not yet taken hers",
        "loyalty_base": 20,
        "status": "DESIGNED",
    },
    {
        "name": "Marc Penoit",
        "position": "DEPUTY GOVERNOR",
        "faction": "French Habitants",
        "ethnicity": "Québécois (French-Canadian)",
        "physical": (
            "Broad-shouldered, built like someone who spent two winters at a siege. Early 40s. "
            "Salt-and-pepper hair worn short and practical. A trimmed dark beard going grey "
            "at the jaw. Weathered face: not from weather, from everything else. Wears a "
            "worn militia coat — French Habitant cut, brass buttons, several replaced — "
            "over a plain wool shirt. The coat has been repaired in four places he can count "
            "and at least two he hasn't found yet. The face of a man who disagrees with the "
            "alliance about once a week and never disagrees enough to leave."
        ),
        "personality": "Pragmatic, militarily precise, protective of his people, frank to the point of bluntness, loyal to Clear-Water's vision if not always her methods, watches the flanks — literal and political",
        "hook": "Spent two winters at the siege of Saint-Georges before anyone knew his name. He doesn't need recognition. He needs the eastern flank watched.",
        "relationships": "Jessica Commanda Odjick: deep trust expressed through productive friction; Ualani Carlisle: allied commander, mutual respect; Governor Carleton: fought him for years, reads him well",
        "loyalty_base": 15,
        "status": "DESIGNED",
    },
    {
        "name": "Governor Guy Carleton",
        "position": "GOVERNOR-GENERAL",
        "faction": "Crown (British)",
        "ethnicity": "Anglo-Irish (British colonial)",
        "physical": (
            "Formal, trim, everything in its place. Late 50s. Powdered wig or short-queued "
            "grey hair depending on context — wig for official functions, hair for field. "
            "Governor's formal coat: red and gold, medals in correct order of precedence. "
            "Sharp light eyes, a thin mouth that is usually slightly disapproving. The "
            "posture of someone who has administered large territories and found the "
            "colonies moderately more difficult than expected. Holds his hands behind "
            "his back when thinking. When he stops doing that, something has gone wrong."
        ),
        "personality": "Administrative rather than military, prefers negotiation to force, privately more pragmatic than his mandate, reads colonial sentiment accurately (and ignores it), underestimates Jessica Commanda Odjick",
        "hook": "He is not a bad administrator. He is administering the wrong empire at the wrong time.",
        "relationships": "Jessica Commanda Odjick: has not yet taken her full measure; Marc Penoit: underestimates his staying power; Abigail Adams: she has taken his measure completely",
        "loyalty_base": None,
        "status": "DESIGNED",
    },
]

NPC_CHARACTERS = [
    {
        "name": "Lord Cornwallis",
        "position": "BRITISH GENERAL",
        "faction": "Crown (British)",
        "ethnicity": "Anglo (British aristocracy)",
        "physical": (
            "Tall, imposing, the full weight of British military tradition in every line. "
            "Mid-40s. Formal scarlet officer's coat with gold epaulettes, the rank "
            "unmistakable from distance. Hawkish features, one eye slightly affected from "
            "a field injury — gives him an unsettlingly focused expression. Clean-shaved, "
            "powdered wig in formal portraits, queued hair in the field. Carries a sword "
            "he knows how to use. The face of a man who has never considered that he might "
            "be on the wrong side of history because the concept had not yet reached him."
        ),
        "personality": "Professionally ruthless, loyal to crown over conscience, genuine military talent, capable of magnanimity in victory, poor at reading asymmetric warfare",
        "hook": "He won Yorktown in this timeline. It didn't help.",
        "relationships": "General Howe: superior officer, occasional friction; Benjamin Tallmadge: professional nemesis",
        "loyalty_base": None,
        "status": "PARTIAL",
    },
    {
        "name": "General Howe",
        "position": "BRITISH COMMANDER",
        "faction": "Crown (British)",
        "ethnicity": "Anglo (British military)",
        "physical": (
            "Heavy build, solid, the physique of a man who campaigns hard and lives well "
            "between campaigns. Mid-50s. Full dress uniform, red and gold. Square face, "
            "confident bearing. Less precise than Cornwallis, more political. "
            "The face of a general who knows how to win battles and is less interested "
            "in the question of whether to fight them."
        ),
        "personality": "Cautious in execution, politically aware, respects formality, occasionally overcautious to the point of strategic failure, convivial away from battle",
        "hook": "Commanded the British forces at a critical juncture. His caution cost them the initiative.",
        "relationships": "Lord Cornwallis: subordinate with his own opinions; Crown command: navigates carefully",
        "loyalty_base": None,
        "status": "PARTIAL",
    },
    {
        "name": "Calico Jack",
        "position": "PIRATE CAPTAIN",
        "faction": "Bahama Free Ports",
        "ethnicity": "Anglo-Caribbean",
        "physical": (
            "The coat is the first thing you see: a long patchwork of calico cloth, "
            "browns and reds and faded blues, sewn from a dozen different sources over "
            "a decade. Mid-30s. Sea-tanned brown skin, a working beard he trims when "
            "he remembers. A single gold earring, left ear. Sharp green eyes and the "
            "slightly amused expression of someone who is always calculating exit routes. "
            "Tricorn hat, battered. Cutlass on the left hip. The kind of man who looks "
            "exactly as dangerous as he is and has decided there's no reason to hide it."
        ),
        "personality": "Opportunist, charismatic, genuinely good at sea warfare, treats his crew fairly by pirate standards, philosophically flexible about law, knows exactly what he is and is comfortable with it",
        "hook": "The Bahamas don't answer to London. Calico Jack is why.",
        "relationships": "Anne Bonny: professional partnership, personal history; Ualani Carlisle: has heard of each other, not yet met",
        "loyalty_base": None,
        "status": "DESIGNED",
    },
    {
        "name": "Anne Bonny",
        "position": "PIRATE FIRST MATE",
        "faction": "Bahama Free Ports",
        "ethnicity": "Anglo-Irish (Caribbean-raised)",
        "physical": (
            "Red-brown hair, salt-stiff, usually tied back with a strip of sail cloth. "
            "Late 20s. Compact, fast, built for deck fighting — not tall but "
            "absolutely present. Sun-burnt nose, freckles, a scar along the left jaw "
            "she got in Nassau and doesn't explain. Sailor's clothing, practical: "
            "loose shirt, breeches, boots. Carries two pistols and is significantly "
            "better with a cutlass. The expression of someone who has made peace with "
            "every decision she has ever made and is prepared to make more of them."
        ),
        "personality": "Direct, impatient with formality, tactically sharp, loyal to chosen crew over all abstractions, fiercely independent, reads people accurately and quickly, dislikes both sides of the war equally",
        "hook": "She left Ireland, then the colonies, then polite society. She stopped at the ocean because there was nowhere left to leave.",
        "relationships": "Calico Jack: complicated, functional, effective; Patricia Eubanks: unlikely alliance of convenience",
        "loyalty_base": None,
        "status": "DESIGNED",
    },
    {
        "name": "Joseph Brant (Thayendanegea)",
        "position": "MOHAWK WAR CHIEF",
        "faction": "Haudenosaunee Confederacy",
        "ethnicity": "Mohawk (Haudenosaunee)",
        "physical": (
            "Striking and deliberate in presentation. Mid-30s. Tall, athletic, "
            "the commanding bearing of a war chief. Long dark hair. Wears a "
            "distinctive combination: British officer's coat (earned, not appropriated) "
            "over Mohawk regalia — gorget, beadwork, clan markings. Both elements "
            "are complete, neither is costume. Face paint in formal contexts. "
            "The appearance of someone navigating two worlds with full knowledge "
            "of what each one costs."
        ),
        "personality": "Strategically sophisticated, bilingual (Mohawk and English), deeply political, prioritizes Haudenosaunee sovereignty over alliance with either side, has been betrayed by British promises before and remembers",
        "hook": "He fights for his people. Everyone else's war is a tool toward that end.",
        "relationships": "Jessica Commanda Odjick: respect across tribal lines, different nations different strategies; Governor Carleton: wary alliance; Lord Cornwallis: uses each other",
        "loyalty_base": None,
        "status": "DESIGNED",
    },
    {
        "name": "Patricia Eubanks",
        "position": "AGENT / UNKNOWN",
        "faction": "Unknown",
        "ethnicity": "Unknown",
        "physical": (
            "Deliberately unremarkable in public — plain dress, hair covered, "
            "nothing that catches the eye. When she wants to be noticed, "
            "she chooses the moment carefully. Eyes that stay still while "
            "assessing. Probably mid-30s. "
            "[DESIGN NOTE: Physical description intentionally incomplete — "
            "her appearance should remain ambiguous until her event arc resolves.]"
        ),
        "personality": "Operationally careful, motives unclear, appears in unexpected contexts, may be working for multiple parties simultaneously",
        "hook": "She appears in event data. Nobody is entirely sure on whose behalf.",
        "relationships": "Anne Bonny: crosses paths in the Bahamas; Benjamin Tallmadge: has noticed her",
        "loyalty_base": None,
        "status": "STUB",
    },
]

# ── ART DIRECTION NOTES ───────────────────────────────────────────────────────
DESIGN_NOTES = [
    ("UNIVERSAL RULES", "Color temperature", "Each faction has a dominant palette: Federal = deep blue/gold; Patriot = navy/cream; Crown = scarlet/gold; Algonquin Nation = forest green/indigo/copper; French Habitants = grey-blue/brown; Bahama Free Ports = sun-bleached warm tones."),
    ("UNIVERSAL RULES", "Era consistency", "Clothing is 1770s-1790s American/British/Canadian colonial, but the game's alt-history allows for slight anachronisms in vocabulary and attitude — not in silhouette. No modern cuts."),
    ("UNIVERSAL RULES", "No placeholder art rule", "Every named character should eventually have a bespoke portrait. The current placeholder (4-22-Ikra-Colors) is identical for all characters — this is known and tracked here."),
    ("UNIVERSAL RULES", "Gold earrings = Ualani", "Gold hoop earrings appear on Ualani Carlisle in ALL contexts: card art, event scenes, portraits. They are a visual signature. No other character wears gold hoop earrings."),
    ("UALANI CARLISLE", "Hair", "Belly-length straight black hair. In field/combat scenes: single thick braid. In diplomatic/presidential scenes: loose OR braided — artist's choice but must reach belly minimum."),
    ("UALANI CARLISLE", "Uniform color", "Deep blue with gold braid. NOT the British scarlet. The blue is deliberate — republic colors."),
    ("JESSICA COMMANDA ODJICK", "Signature element", "Indigo ribbon threaded through braid in formal settings. Algonquin ribbon-work (geometric floral, multi-color) visible at collar and cuffs. These two elements appear in every portrait."),
    ("JESSICA COMMANDA ODJICK", "Clothing philosophy", "Neither full Western diplomatic dress nor full traditional regalia — always both simultaneously. The combination is intentional and should read as authority, not as compromise."),
    ("MARC PENOIT", "Signature element", "The repaired militia coat. Four visible repairs minimum. Should look functional and worn, not derelict."),
    ("MARC PENOIT", "Hair/beard", "Short salt-and-pepper hair, trimmed dark beard going grey at jaw. Practical. No wig."),
    ("CALICO JACK", "Signature element", "The calico coat — patchwork of browns, reds, faded blues. It is the character's most recognizable feature. Should appear in all depictions."),
    ("ANNE BONNY", "Signature element", "Red-brown hair tied with sail cloth. Left jaw scar. Two pistols visible on her person."),
    ("JOSEPH BRANT", "Clothing", "British officer's coat + Mohawk regalia simultaneously. Both elements are complete and intentional. Not costume; not appropriation. Historical accuracy is important here — Brant was a real historical figure who navigated exactly this presentation."),
    ("BENJAMIN TALLMADGE", "Presence", "He is the most deliberately unremarkable-looking person in any scene he is in. This is intentional. His portrait should feel like you almost didn't notice him."),
    ("PHILLIS WHEATLEY", "Signature element", "Quill in hand or tucked behind ear in every depiction. The quill is the weapon."),
]

# ── BUILD FUNCTIONS ────────────────────────────────────────────────────────────

def _build_overview(wb, usa, canadian, npc):
    ws = wb.create_sheet("OVERVIEW")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    _hdr(ws, 1, 1, "CHARACTER MASTERDOC — FARSWORD DESIGN BIBLE", HDR_BG, HDR_FG, bold=True, size=13)

    # Stats block
    stats = [
        ("Total Named Characters", len(usa) + len(canadian) + len(npc)),
        ("USA / Playable Governors", len(usa)),
        ("Canadian Governors / Leaders", len(canadian)),
        ("NPC / Event Characters", len(npc)),
        ("", ""),
        ("Design Complete (FINAL)", sum(1 for c in usa+canadian+npc if c["status"] == "FINAL")),
        ("Described (DESIGNED)", sum(1 for c in usa+canadian+npc if c["status"] == "DESIGNED")),
        ("Partial (PARTIAL)", sum(1 for c in usa+canadian+npc if c["status"] == "PARTIAL")),
        ("Stub Only (STUB)", sum(1 for c in usa+canadian+npc if c["status"] == "STUB")),
    ]
    row = 3
    _hdr(ws, row, 1, "METRIC", HDR_BG, HDR_FG)
    _hdr(ws, row, 2, "COUNT", HDR_BG, HDR_FG)
    row += 1
    alt = False
    for label, val in stats:
        bg = ALT_BG if alt else WHITE
        _cell(ws, row, 1, label, bg=bg, bold=bool(label))
        _cell(ws, row, 2, val if val != "" else "", bg=bg, align="center")
        if label:
            alt = not alt
        row += 1

    # Character quick-list
    row += 1
    _hdr(ws, row, 1, "NAME", USA_BG, USA_FG)
    _hdr(ws, row, 2, "POSITION", USA_BG, USA_FG)
    _hdr(ws, row, 3, "FACTION", USA_BG, USA_FG)
    _hdr(ws, row, 4, "COUNTRY", USA_BG, USA_FG)
    _hdr(ws, row, 5, "STATUS", USA_BG, USA_FG)
    row += 1

    groups = [("USA", usa, "D6E4F7"), ("CANADA", canadian, "FCE0D6"), ("NPC", npc, "EEF2F7")]
    for group_name, chars, bg in groups:
        for i, c in enumerate(chars):
            row_bg = bg if i % 2 == 0 else WHITE
            _cell(ws, row, 1, c["name"], bg=row_bg, bold=True)
            _cell(ws, row, 2, c["position"], bg=row_bg)
            _cell(ws, row, 3, c["faction"], bg=row_bg)
            _cell(ws, row, 4, group_name, bg=row_bg)
            _status_cell(ws, row, 5, c["status"])
            row += 1

    _set_col_widths(ws, [30, 28, 25, 12, 12])
    _freeze(ws, "A2")


def _build_character_sheet(wb, sheet_name, characters, hdr_bg, hdr_fg):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    cols = [
        "NAME", "POSITION", "FACTION", "ETHNICITY / BACKGROUND",
        "PHYSICAL DESCRIPTION (ART BIBLE)",
        "PERSONALITY TRAITS",
        "NARRATIVE HOOK",
        "RELATIONSHIPS",
        "LOYALTY BASE",
        "STATUS",
    ]
    widths = [22, 22, 22, 24, 60, 45, 45, 45, 12, 10]

    for col, label in enumerate(cols, 1):
        _hdr(ws, 1, col, label, hdr_bg, hdr_fg)

    for i, c in enumerate(characters):
        row = i + 2
        bg = ALT_BG if i % 2 == 0 else WHITE
        _cell(ws, row, 1, c["name"], bg=bg, bold=True)
        _cell(ws, row, 2, c["position"], bg=bg)
        _cell(ws, row, 3, c["faction"], bg=bg)
        _cell(ws, row, 4, c["ethnicity"], bg=bg)
        _cell(ws, row, 5, c["physical"], bg=bg)
        _cell(ws, row, 6, c["personality"], bg=bg)
        _cell(ws, row, 7, c["hook"], bg=bg)
        _cell(ws, row, 8, c["relationships"], bg=bg)
        lb = c["loyalty_base"]
        _cell(ws, row, 9, str(lb) if lb is not None else "N/A", bg=bg, align="center")
        _status_cell(ws, row, 10, c["status"])
        ws.row_dimensions[row].height = 120

    _set_col_widths(ws, widths)
    _freeze(ws, "A2")


def _build_design_notes(wb, notes):
    ws = wb.create_sheet("DESIGN NOTES")
    ws.sheet_view.showGridLines = False

    _hdr(ws, 1, 1, "CHARACTER", HDR_BG, HDR_FG)
    _hdr(ws, 1, 2, "ELEMENT", HDR_BG, HDR_FG)
    _hdr(ws, 1, 3, "NOTE", HDR_BG, HDR_FG)

    char_colors = {
        "UNIVERSAL RULES": "1F2D3D",
        "UALANI CARLISLE": "1F3864",
        "JESSICA COMMANDA ODJICK": "2D6A4F",
        "MARC PENOIT": "5C4033",
        "CALICO JACK": "7A5C00",
        "ANNE BONNY": "7A1A1A",
        "JOSEPH BRANT": "4A3060",
        "BENJAMIN TALLMADGE": "2E4057",
        "PHILLIS WHEATLEY": "1A5C3A",
    }

    last_char = None
    for i, (character, element, note) in enumerate(notes):
        row = i + 2
        bg = ALT_BG if i % 2 == 0 else WHITE
        char_bg = char_colors.get(character, HDR_BG)
        if character != last_char:
            _hdr(ws, row, 1, character, char_bg, "FFFFFF", bold=True)
            last_char = character
        else:
            _cell(ws, row, 1, "", bg=bg)
        _cell(ws, row, 2, element, bg=bg, bold=True)
        _cell(ws, row, 3, note, bg=bg)

    _set_col_widths(ws, [28, 28, 80])
    _freeze(ws, "A2")


# ── MODEL SHEET DATA ─────────────────────────────────────────────────────────
#
# Production bible format used by animation studios and game art teams.
# Each section maps to a named block in the model sheet tab.
# "Invariants" = rules that survive EVERY scene type, including explicit.
# Scene ratings: SFW | SENSUAL | EXPLICIT
#

UALANI_MODEL = {
    "name": "Ualani Carlisle",
    "title": "PRESIDENT & COMMANDER — APF / Federal Republic",
    "tab_color": "1F3864",

    # ── CONSTRUCTION ─────────────────────────────────────────────────────────
    "construction": [
        ("Body type", "Athletic, medium-lean. Not lanky — she carries muscle from active command. Visible definition in shoulders, arms, and core without exaggeration. Think swimmer-soldier hybrid."),
        ("Height", "5'7\" (170 cm). Medium-tall. Commands vertical presence without looming."),
        ("Head height ratio", "7.5 heads tall. Slightly idealized but grounded. Not anime-elongated, not naturalistic-compressed."),
        ("Shoulder width", "Wider than average for her height — 2.2 head-widths. This is the first visual cue of her military authority even out of uniform."),
        ("Hip-to-waist ratio", "Moderate hourglass. Hips 1.15× shoulder width. Waist well-defined but not corseted."),
        ("Leg length", "Long relative to torso. 4 heads from hip to floor. Gives the 'walks like she knows where she's going' silhouette."),
        ("Face shape", "Oval with a slightly squared jaw — not soft, not sharp. Filipino jaw structure meeting Hawaiian broadness. Defined cheekbones, not extreme."),
        ("Eyes", "Dark brown, almost black. Almond-shaped, slightly uptilted outer corner. Large but not oversized. The flat affect is in the lids — she rarely blinks fast."),
        ("Nose", "Medium width, slightly rounded tip — Hawaiian-Filipino blend. Avoid making it narrow/Eurocentric. The nose shape is part of her ethnic authenticity."),
        ("Mouth", "Full lips. Natural rest expression is a slight flat-line — not a frown, not a smile. Neutral authority. When she does smile it lands hard."),
        ("Ears", "Standard, slightly small relative to face. ALWAYS show earrings — gold hoops, medium size, not chandelier not tiny."),
        ("Hair construction", "Straight, fine-to-medium texture, very black (not dark brown — full black). No natural wave. Parted loosely near center. In loose mode: falls belly-length minimum, slight weight-pull visible at crown. In braid mode: single thick three-strand braid starts at nape, runs full length of back and past waist, tail tip at hip or lower."),
        ("Hair volume", "NOT flat. The straight black hair has body from its own weight — slight outward curve as it falls past shoulder, then falls straight. The braid is thick (thumb-width at the nape, slightly thinning to fingertip-width at tail)."),
        ("Hands", "Medium-large for her build. Capable hands. Not aristocratic. Short unpolished nails. The hands of someone who has handled weapons."),
        ("Skin tone", "Warm medium-brown. Not golden, not red-brown — the specific warm brown of Hawaiian-Filipino Pacific Islander heritage. Hex reference: #8B5E3C warm undertone, lightest skin highlight #C8906A, deepest shadow #5C3320. Consistent across all lighting."),
    ],

    # ── COLOR PALETTE ─────────────────────────────────────────────────────────
    "palette": [
        ("Skin — base",         "#9B6B4A", "Primary skin tone, flat-lit areas"),
        ("Skin — highlight",    "#C89070", "Brow ridge, nose bridge, shoulder, collarbone"),
        ("Skin — midtone",      "#7A4F35", "Shadow side of face, underarms, stomach crease"),
        ("Skin — deep shadow",  "#5C3320", "Under jaw, inner arms, between legs"),
        ("Skin — warm accent",  "#C07040", "Cheek warmth, sun-caught shoulders, knuckles"),
        ("Hair",                "#0A0A0A", "True near-black. Not dark brown. No highlights unless strong direct light (then #1A1A2A)."),
        ("Hair highlight",      "#1C1C2E", "Only under direct overhead light source. Keep sparse."),
        ("Eyes",                "#1A0D00", "Iris. Very dark warm brown. Pupil blends into iris at most scales."),
        ("Eye sclera",          "#F5EDE0", "Warm white, not pure white. Cold sclera looks wrong on warm skin tones."),
        ("Earrings",            "#D4A017", "Matte gold — not chrome, not brass. Hoop earrings, medium gauge."),
        ("Uniform — primary",   "#1C2E5C", "Deep federal blue. Main body of presidential/field uniform."),
        ("Uniform — secondary", "#D4A017", "Gold braid, buttons, rank insignia. Same gold as earrings."),
        ("Uniform — lining",    "#0F1A3A", "Inside facing of coat, shadow areas of uniform."),
        ("Lips",                "#7A3A2A", "Natural deep nude-brown. Not lipstick red. Consistent in all scenes."),
        ("Eyebrows",            "#150800", "Full, slightly thick, natural arch. Same near-black as hair."),
    ],

    # ── INVARIANTS (never change regardless of scene) ─────────────────────────
    "invariants": [
        ("Gold hoop earrings", "PRESENT IN EVERY DEPICTION WITHOUT EXCEPTION. Both ears. Medium gold hoops, #D4A017. If a scene has her fully undressed, the earrings are still there. They are not removed. This is absolute."),
        ("Hair length", "MINIMUM belly-button length in ALL scenes. Can be longer. Cannot be shorter. Even in bed/undressed scenes — hair is present at full length, loose or braided."),
        ("Hair color", "True black (#0A0A0A). Never brown, never highlighted, never tied with anything that obscures its full length at rest."),
        ("Skin tone", "Consistent warm medium-brown across all scene types. Do not lighten in explicit scenes (common error). Do not darken dramatically in shadow — keep undertone warm."),
        ("Eye color", "Very dark warm brown. Not black irises, not hazel. The darkness is warm, not cold."),
        ("Body type", "Athletic build maintained in ALL scenes. She does not become softer or curvier in explicit scenes. The shoulders are always present. The core definition is always present."),
        ("Age", "Late 30s presentation. She is an experienced head of state. Do not juvenilize her facial features in explicit scenes (common error). The commanding face stays commanding."),
        ("Facial expression authority", "Even in vulnerable/explicit scenes, her default expression has a baseline level of awareness — she is never passive. Ecstasy is an active expression for her, not a passive one."),
        ("Nose authenticity", "Pacific Islander nose shape — do not Europeanize it in explicit scenes. This is a consistency error that destroys ethnic authenticity."),
    ],

    # ── EXPRESSION CHART ──────────────────────────────────────────────────────
    "expressions": [
        ("COMMAND (default)",    "Flat neutral. Lids slightly heavy. Jaw level. Mouth closed or barely parted. This is her baseline — reads as watchful authority, not blank."),
        ("FOCUSED",              "Brows draw very slightly inward and downward — not furrowed, just closer. Eyes narrow 10%. The rest of the face stays still. This is her battle-plan face."),
        ("DRY AMUSEMENT",        "One corner of mouth lifts 3mm. Eyes don't change. This is her laugh. It's not a smile — it's the acknowledgment that something was funny and she is not going to make a big deal of it."),
        ("GENUINE SMILE",        "Rare. Cheeks lift, eyes crinkle, mouth opens to show upper teeth. When this hits, it's worth something because it's rare. Do not deploy it casually."),
        ("ANGER (cold)",         "Brows descend without furrowing — flat and low. Eyes open wider. Mouth stays closed. She gets quieter when angry, not louder. This is the expression that precedes consequences."),
        ("CONCERN",              "Brows rise slightly at inner corner only. Eyes soften — lids lift. Mouth stays neutral. She doesn't broadcast concern widely."),
        ("PHYSICAL EFFORT",      "Jaw tightens. Brows pull together slightly. Lips press. Used in combat, exertion, field command. Neck tendons may be visible."),
        ("SENSUAL INTEREST",     "Eyes soften, lids drop 15%. Lips part slightly. Chin drops fractionally — she looks up through lashes even at eye level. Not coy. Direct. Deliberate."),
        ("PLEASURE (sensual)",   "Eyes close or half-close. Brow smooth. Lips part. Head may tilt slightly. Breathing visible through parted lips. Still has her baseline dignity — not slack-jawed."),
        ("PLEASURE (explicit)",  "Eyes closed or rolled back slightly. Brows raised at center. Mouth open, teeth possibly visible. Chin raised. Hair disheveled but still present at full length. Even at peak she looks like she's in charge of what's happening."),
        ("RELEASE / CLIMAX",     "Eyes fully closed, head thrown back slightly. Brows raised, forehead creased. Mouth fully open. This is the only moment her control expression fully leaves — brief. Then it returns."),
    ],

    # ── COSTUME LAYERS (head to toe, outer to inner) ─────────────────────────
    "costume_layers": [
        ("LAYER 0 — FULL FORMAL", "Presidential reception. Deep blue (#1C2E5C) long military coat with gold (#D4A017) braid at shoulders and cuffs. High collar. Rank insignia. Campaign medals left breast. Matching trousers, dark. Tall dress boots. Hair loose, full length, or formal braid. Earrings."),
        ("LAYER 1 — FIELD COMMAND", "Same blue coat, sleeves rolled to forearm (forearms visible — draw them). Collar open one button. Medal bar removed or single ribbon only. Combat trousers tucked into field boots. Hair always braided in this state. Earrings. This is the most common scene state."),
        ("LAYER 2 — INFORMAL AUTHORITY", "Blue military shirt (not coat), tucked. Gold stripe on collar only. Dark trousers. Boots or low boots. Hair loose. Earrings. Common in planning scenes, private meeting rooms."),
        ("LAYER 3 — CIVILIAN (rare)", "Colonial-era quality clothing — she rarely fully abandons her role. Dark dress or high-waisted skirt and blouse, quality fabric, minimal color. Hair loose. Gold earrings. She does not disappear into civilian dress — she wears it like a costume."),
        ("LAYER 4 — UNDERSHIRT", "Simple white linen shirt, collarless, slightly oversized. Loose at arms. Short sleeves or sleeveless. Tucked or untucked. The first 'off-duty' layer. Hair loose."),
        ("LAYER 5 — BARE TORSO / PARTIAL UNDRESS", "SFW context: shows only in bathing or injury scenes, clinical framing. SENSUAL: deliberate, controlled — she is the one making choices. She wears a simple front-clasp brassiere or nothing depending on scene. EXPLICIT: see explicit section."),
        ("LAYER 6 — UNDRESSED LOWER BODY", "Simple colonial-era underpinnings — loose linen drawers, then nothing. No anachronistic modern underwear unless style choice is deliberate for game tone. For full explicit: bare."),
        ("LAYER 7 — FULLY BARE", "See explicit section. Invariants still apply. Earrings. Hair."),
    ],

    # ── SCENE TYPE DIRECTION ─────────────────────────────────────────────────
    "scene_sfw": [
        ("Portrait / card art",  "Bust or 3/4 shot. Layer 0 or 1. Face in COMMAND expression or slight DRY AMUSEMENT. Hair loose flowing or field braid. Gold earrings prominent — they catch light. Lighting: warm from one side, cool ambient fill. Slight atmospheric haze for depth. Background: flag, map, or plain dark."),
        ("Dialogue scene",       "Standing or seated, Layer 1 or 2. Full figure visible. Expression ranges from FOCUSED to CONCERN depending on context. Both hands visible — she gestures minimally but precisely. Camera at character eye level, never looking down at her."),
        ("Field/combat",         "Layer 1. PHYSICAL EFFORT or COMMAND expression. Hair always braided. Earrings caught by ambient light. Posture: center of gravity low, weight forward slightly. She is never passive in an action frame. Even standing still she looks like a coiled action."),
        ("Map room / planning",  "Layer 2 or informal. Leaning over a table, hands on surface or holding something. FOCUSED expression. Light from map table or candles — shadows up from below, unusual lighting direction that suits her face structure. Slight shadow under brow ridge."),
        ("Diplomatic reception", "Layer 0 full formal. Neutral-to-pleasant expression — controlled affect. She is reading the room. Eyes moving. Posture upright, hands at sides or clasped behind back. Not stiff — disciplined."),
        ("Casual/off-duty",      "Layer 3 or 4. Rare scene type — should feel like a glimpse. Hair fully loose, belly length. Earrings. Expression: DRY AMUSEMENT or GENUINE SMILE (use the genuine smile sparingly — this is one of the few appropriate scenes). Light casual and warm."),
        ("Injury / recovery",    "Layer 4 or partial 5. Clinical dignity — this is not sexualized. CONCERN or suppressed PHYSICAL EFFORT expression. Her body reads as a soldier's body: functional damage, not ornamental. Show the competence, not the vulnerability."),
    ],

    "scene_sensual": [
        ("Rating definition",    "SENSUAL = charged but not explicit. Clothing displaced but key areas covered or at the edge. Physical proximity and tension visible. Designed to imply rather than show. Often more effective than explicit for character integrity."),
        ("Entry state",          "Transition from Layer 2 or 3 into partial Layer 5. Shirt open but held or fallen to one side. Her agency is visible — she is not happening to. She is choosing. SENSUAL INTEREST expression."),
        ("Framing rule",         "Camera stays at or above mid-torso in pure sensual framing. What the frame doesn't show is as important as what it does. Bare shoulder, collarbone, lower throat = highly effective. The earrings in this frame read as charged detail — use them."),
        ("Hair in sensual",      "Loose and full-length. Slightly disheveled — not tidy formal braid. Some strands across shoulder or chest. The hair is a sensual element but not the centerpiece. It frames."),
        ("Touch and proximity",  "Her hand on someone's jaw or neck reads as dominant. Someone's hands on her shoulders or at her waist reads as invitation accepted. She does not appear startled or passive in physical contact. She has decided this."),
        ("Lighting",             "Warmer, lower contrast than SFW scenes. Single warm source (candle, fire, lamp) from side or below. Longer shadows. Skin highlights (#C89070) catch the warm light prominently — the warm medium-brown skin reads beautifully in warm light. Lean into it."),
        ("Partial undress rules", "Brassiere or binding if present: simple linen, colonial-appropriate. Earrings always visible. Hair always present. Skin tone consistent — warm medium-brown, do not wash out in candlelight. Her torso shows the athletic muscle — do not omit the shoulder/arm definition in favor of a softer look."),
        ("Expression in sensual", "SENSUAL INTEREST transitioning toward PLEASURE (sensual). The authority baseline remains — she is enjoying herself on her own terms. Not overwhelmed. Present and directive."),
    ],

    "scene_explicit": [
        ("Rating definition",    "EXPLICIT = full nudity and/or explicit sexual content. Full production rules apply. All invariants are MANDATORY. This is not a different character — it is the same woman with the same face, same hair, same earrings."),
        ("Body canon — full nude", "Athletic body with visible muscle definition: shoulder caps, bicep curve, oblique/core lines visible, quadriceps defined. Not bodybuilder extreme — functional athlete. She does not have the soft undifferentiated body of SFW 'cute' character design. Breasts: medium, natural hang, consistent with her build. Not exaggerated upward or outward. They sit naturally on a muscled chest — slight lateral outward angle when arms are raised. Nipples: medium, warm-toned (#9B5040). Consistent across scenes."),
        ("Pubic area",           "Present and consistent. Not completely bare (anachronistic for era, and inconsistent with her character's un-curated presentation). Dark, trimmed naturally. Not elaborate. Colonial era authenticity matters."),
        ("Explicit positions — her dominant/active role", "She is the initiating party in dominant scenes. Posture: upright, center of gravity controlled. Hands placed with intent. Face: PLEASURE expression that retains authority — her brows do not go fully passive. Eyes maintain direction. She looks AT what she is doing or AT who she is with."),
        ("Explicit positions — her receptive role", "When receptive, she is still active — weight shifted, hands gripping, back arched with intention. She does not become a passive object. PLEASURE (explicit) or RELEASE expression. Even at RELEASE she is the main subject of the frame, not a prop."),
        ("Oral — giving",        "Face: focused, deliberate. Eyes may be closed or upward. Not servile — purposeful. The authority in her posture must be maintained even in submissive-presenting positions. She owns the act."),
        ("Oral — receiving",     "Head thrown back or tilted, PLEASURE or RELEASE expression. One hand in partner's hair or on their shoulder — she is directing even when receiving. Hair falls full-length around her."),
        ("Penetrative — general", "All positions must maintain: hair visible and full-length (loose or disheveled braid), gold earrings visible or catching light, skin tone warm and consistent, body definition present. No position flattens her athleticism into softness."),
        ("Penetrative — missionary/equivalent", "Her face always has access to the frame — do not obscure it. PLEASURE or RELEASE. Legs show the athletic build — inner thigh definition, not just soft shape."),
        ("Penetrative — from behind", "Profile or 3/4 back view. Back muscles visible — she has a back, not just a back surface. Spine visible. Hair falls forward over shoulder or is swept aside. Earrings visible in profile. PLEASURE expression in whatever view of face is available."),
        ("Penetrative — riding/dominant", "She is upright. Back straight or slightly arched. Both hands on partner's chest or shoulders. PLEASURE expression that is visibly directive — she sets the pace. This is her most 'in character' explicit position."),
        ("Penetrative — against wall/furniture", "Layer 1 partially on — open coat, uniform displaced. She is not fully undressed. Partially clothed explicit is both more era-appropriate and more character-consistent. Boots may remain. Earrings must."),
        ("Multiple partners",    "Her focal dominance is maintained regardless of partner count. She is never lost in the scene — she is the organizing presence. One partner at a time gets her full gaze."),
        ("Post-scene state",     "Hair disheveled but still long. Earrings still present. Expression returning to baseline authority — the DRY AMUSEMENT or COMMAND returning. She does not look spent in an undignified way. She looks like someone who has concluded an engagement successfully."),
        ("Lighting in explicit", "Warm, low. Same as sensual but deeper shadows. The skin highlights (#C89070) are prominent. The gold earrings catch the warm light — they are a visual anchor in every explicit scene, confirming identity at a glance."),
        ("Common errors to avoid", "1) Lightening skin tone in explicit scenes. 2) Making hair brown under warm light — it stays near-black even in candlelight. 3) Removing earrings. 4) Softening the shoulder/arm definition. 5) Giving her a passive or confused expression. 6) Europeanizing the nose. 7) Shortening the hair. 8) Adding anachronistic modern elements (nail polish, modern lingerie, contemporary hairstyles)."),
    ],
}

JESSICA_MODEL = {
    "name": "Jessica Commanda Odjick",
    "title": "PRIME MINISTER — Algonquin Nation / Continental Republic of Canada",
    "tab_color": "2D6A4F",

    # ── CONSTRUCTION ─────────────────────────────────────────────────────────
    "construction": [
        ("Body type", "Tall, lean-strong. Not visibly athletic in the soldier sense — her strength is posture and stillness. Long limbs. The kind of tall that fills a doorway with presence rather than size."),
        ("Height", "5'9\" (175 cm). Noticeably tall, especially given her political context. This height is part of her visual authority."),
        ("Head height ratio", "8 heads tall — the full idealized figure proportion. She is drawn to be imposing in a frame."),
        ("Shoulder width", "Narrower than Ualani's but with straight horizontal line — carried shoulders, not hunched. 2.0 head-widths. The straightness is more striking than the width."),
        ("Hip-to-waist ratio", "Lean hourglass. Less pronounced than Ualani. Her figure is elongated, not curvaceous. Hips 1.1× shoulder width."),
        ("Leg length", "Long. 4.2 heads from hip to floor — more than standard idealized figure. The visual impression of someone who covers ground efficiently."),
        ("Face shape", "Oval tending toward long — the elongation of her face matches her height. High, prominent cheekbones (Algonquin bone structure — wide across the cheekbone, tapering sharply to a defined jaw). The cheekbones are the defining feature of her face."),
        ("Eyes", "Dark brown to near-black, almond-shaped, straight — not uptilted like Ualani's. Horizontal eye line. Heavy upper lids. The stillness in her expression comes from these eyes staying level and open without blinking often."),
        ("Nose", "Distinctive Algonquin nose structure: straight bridge, slightly broader at base than a European nose, nostrils that flare slightly. Authentically Indigenous. Do not narrow this in any scene type."),
        ("Mouth", "Medium-full lips, slightly narrower than Ualani's. The dominant expression is a flat closed line — not hostile, simply waiting. When she speaks, the mouth becomes very precise. When she smiles, it's with the full weight of her face behind it."),
        ("Ears", "Partly hidden by hair in most states. When visible, plain — she does not wear ear jewelry. Her ornamentation is at the collar and cuffs."),
        ("Hair construction", "Long, straight, very dark brown-black (slightly warmer than Ualani's — a hint of dark brown undertone is correct). Natural Algonquin/Anishinaabe texture: smooth, straight, fine. In loose mode: falls past mid-back, approximately to lower lumbar. In braid mode (formal): single braid threaded with indigo ribbon (#2E1760), starting at crown, running to mid-back. The braid is the formal state. The loose hair is the relaxed/field state."),
        ("Hair ornamentation", "In formal braid: a single length of indigo ribbon (#2E1760) woven into the braid from root to tip. The ribbon is narrow (6mm), runs the full braid length. This is always present in formal/diplomatic scenes, always absent in field/combat scenes."),
        ("Hands", "Long-fingered, elegant. The hands of someone who uses language as her primary tool. Nails short and clean. No nail color. The hands are often the most expressive part of her body in SFW scenes."),
        ("Skin tone", "Medium-warm brown, slightly cooler undertone than Ualani — less orange-warm, more neutral-warm. Algonquin/Anishinaabe complexion. Base: #7A5238, highlight: #B07850, midtone: #5C3A22, deep shadow: #3D2210, warm accent #9A6040 (cheeks, hands)."),
    ],

    # ── COLOR PALETTE ─────────────────────────────────────────────────────────
    "palette": [
        ("Skin — base",         "#7A5238", "Primary skin tone, flat-lit areas"),
        ("Skin — highlight",    "#B07850", "Cheekbone, brow ridge, nose bridge, collarbone"),
        ("Skin — midtone",      "#5C3A22", "Shadow side of face, underarms"),
        ("Skin — deep shadow",  "#3D2210", "Under jaw, under brows, inner arms"),
        ("Skin — warm accent",  "#9A6040", "Cheek warmth, knuckles"),
        ("Hair",                "#120800", "Very dark with warm brown undertone. Not full black like Ualani."),
        ("Hair highlight",      "#2A1800", "Only in strong direct light. Very subtle."),
        ("Eyes",                "#140800", "Very dark warm brown-black iris."),
        ("Eye sclera",          "#F0E8DC", "Warm white. Slightly more warm than Ualani's due to cooler skin contrast."),
        ("Ribbon (braid)",      "#2E1760", "Indigo — a specific deep blue-purple. NOT navy, NOT black. This color is invariant."),
        ("Ribbon highlight",    "#4A2B96", "Sheen on ribbon silk in direct light."),
        ("Ribbon-work embroidery", "#C84B18", "Dominant warm red in Algonquin floral ribbon-work at collar/cuffs."),
        ("Ribbon-work accent 1","#D4A017", "Gold-yellow in ribbon-work geometric border."),
        ("Ribbon-work accent 2","#2E6B3A", "Forest green in ribbon-work floral centers."),
        ("Ribbon-work accent 3","#1F3090", "Deep blue in ribbon-work outer border."),
        ("Diplomat coat",       "#1C2A1C", "Very dark forest green. The overcoat base color."),
        ("Coat interior lining","#6B1414", "Deep burgundy lining visible at lapels — the only warm element on the coat."),
        ("Lips",                "#6B3A28", "Natural dark nude-rose. Consistent across scenes."),
        ("Eyebrows",            "#180A00", "Full, straight — not highly arched. Strong horizontal line."),
    ],

    # ── INVARIANTS ───────────────────────────────────────────────────────────
    "invariants": [
        ("Ribbon-work at collar/cuffs", "MANDATORY in all formal and diplomatic scenes. The geometric-floral Algonquin ribbon-work in warm red, gold, green, and blue appears at the collar edge and cuff edge of whatever she is wearing in those contexts. In undressed/explicit scenes, the garment is removed but the NEXT layer beneath may have a simpler version. This element is her signature of dual-world authority — it does not vanish."),
        ("Indigo ribbon in formal braid", "When in braid mode, the indigo ribbon (#2E1760) is always threaded through. It cannot be a different color. It cannot be absent while she is braided. In loose-hair scenes, no ribbon is present."),
        ("Cheekbone structure", "The high Algonquin cheekbones are the defining feature of her face and must remain prominent in all scene types, all lighting conditions, all expressions. Do not soften this structure in explicit scenes."),
        ("Skin tone consistency", "Warm medium-brown, slightly cooler than Ualani. Do not lighten in explicit scenes. The specific tone is part of her Indigenous identity and must not be adjusted for 'softness' or 'glow'."),
        ("Height", "She must read as tall relative to other characters in multi-character scenes. She is the tallest named female character."),
        ("Hair length", "Lower lumbar minimum length. Can be longer. Cannot be shorter."),
        ("Nose authenticity", "Algonquin nose structure. Do not narrow or Europeanize. This is as important as any other invariant."),
        ("Age", "Late 30s. She is a mature political leader. Juvenile facial features in explicit scenes break her characterization completely."),
        ("Eyes — directness", "She makes direct eye contact in every scene type including explicit. Her gaze does not go coy or soft. She looks at things — including the viewer — with full deliberate awareness."),
        ("No ear jewelry", "She does not wear earrings or ear jewelry of any kind. This is a contrast with Ualani. Do not add ear jewelry to distinguish her from other characters in multi-character scenes — use the ribbon-work and face structure instead."),
    ],

    # ── EXPRESSION CHART ──────────────────────────────────────────────────────
    "expressions": [
        ("DEFAULT (waiting)",    "Still and level. Eyes open, level, unblinking. Mouth closed, horizontal. No tension in brow. This expression does not read as blank — it reads as a person who is receiving information and processing it before speaking. The room waits for her."),
        ("CONSIDERING",          "The brow descends fractionally over one eye (usually right). The other eye stays level. This asymmetry means she has encountered something worth examining. Mouth stays closed."),
        ("POLITICAL PRECISION",  "Lips part very slightly — she is about to speak. Eyes sharpen. The brow descends evenly, not furrowed but lowered. This is her 'I am about to say the exact right thing' face. It is followed by silence for one beat before she speaks."),
        ("DISMISSAL",            "Chin rises fractionally. Eyes stay level but the lids lower slightly — not sleepy, imperial. The mouth corners pull back neutrally. This is the expression that ended the Governor's Council's participation."),
        ("DRY SATISFACTION",     "The left corner of her mouth lifts 2mm. Not a smile — a notation. Her version of acknowledging that something went correctly."),
        ("GENUINE WARMTH",       "Rare. Both corners of mouth lift evenly. Cheeks lift — the cheekbones become even more prominent. Eyes soften. This reads as completely transformative on a face that rarely deploys it. Reserve for Marc Penoit scenes of hard-won trust and moments of actual victory."),
        ("CONCERN",              "Inner brow corners lift slightly. Eyes widen fractionally. Mouth stays closed. She does not broadcast concern — this is the smallest possible signal."),
        ("ANGER",                "Flat and cold. Eyes open fully — wider than default. Brow descends entirely, evenly. Jaw sets. Mouth closes hard. She becomes more still when angry. The expression that precedes a sentence that will change how someone's week goes."),
        ("SENSUAL INTEREST",     "The eyes do the work. Lids drop 20%. The gaze becomes focused rather than wide. Her chin drops fractionally — she looks up through her lashes even when at height. Mouth opens slightly, lips part. This is chosen, deliberate, unhurried."),
        ("PLEASURE (sensual)",   "Eyes half-closed, gaze soft and inward. Brow smooth. Mouth open, breath visible. Head may tilt back or to side. Her long neck becomes prominent in this expression. Still. She processes pleasure with stillness, not movement."),
        ("PLEASURE (explicit)",  "Eyes closed, brow lifted at center. Mouth open and active — this is the loudest her face gets. Head thrown back or pressed forward depending on position. The cheekbones catch side light beautifully in this expression. Her neck is long and exposed."),
        ("RELEASE / CLIMAX",     "Full opening of expression — brows raised, eyes closed, mouth wide. The normally controlled face fully releases. This is the most readable moment of vulnerability she offers, and it should feel earned and significant for that reason."),
    ],

    # ── COSTUME LAYERS ────────────────────────────────────────────────────────
    "costume_layers": [
        ("LAYER 0 — FULL DIPLOMATIC FORMAL", "Dark forest green overcoat (#1C2A1C), floor-length or ankle-length, deep burgundy interior lining visible at open lapels. Algonquin ribbon-work at collar and cuffs (red/gold/green/blue geometric floral). Hair in formal braid with indigo ribbon. No ear jewelry. Severe, composed, absolutely authoritative."),
        ("LAYER 1 — WORKING FORMAL", "Same coat, less volume — slightly shorter, more practical cut. Ribbon-work still present at collar/cuffs. Underneath: dark fitted shirt or tunic with some ribbon-work detail. Hair may be braid or loose depending on the scene energy. This is her most common state."),
        ("LAYER 2 — FIELD / TRAVEL", "Coat removed. Dark fitted jacket or heavy shirt, simpler — ribbon-work detail only at collar, minimal. Practical trousers, good boots. Hair loose. No ribbon. She reads as active and direct in this state."),
        ("LAYER 3 — INFORMAL", "Dark simple shirt, loose or tucked. No coat. Trousers. Hair loose, full length. The ribbon-work is absent. This is her 'not performing authority' state — she still has it, it just isn't costumed. This layer feels like a reveal."),
        ("LAYER 4 — UNDERSHIRT", "Simple linen undershirt — slightly warm-toned, off-white or pale tan. Loose. Collarless. The first private layer. Hair fully loose, long."),
        ("LAYER 5 — PARTIAL UNDRESS", "Undershirt open or removed. SFW: shown in injury or recovery context, not sexualized. SENSUAL: deliberate entry. Her torso is lean and long — defined but not muscular in the athletic-soldier way. The length of her torso and the prominence of her hip bones read differently than Ualani's build. Colonial-era underpinnings if present."),
        ("LAYER 6 — BASE/UNDERPINNINGS", "Colonial linen or simple cloth. Loose. Then nothing."),
        ("LAYER 7 — FULLY BARE", "See explicit section. No ear jewelry in any undressed scene — this is already the invariant for this character."),
    ],

    # ── SCENE TYPE DIRECTION ──────────────────────────────────────────────────
    "scene_sfw": [
        ("Portrait / card art",  "Bust or 3/4. Layer 0 or 1. DEFAULT or POLITICAL PRECISION expression. Formal braid with indigo ribbon. Ribbon-work visible at collar. Lighting: cool-neutral ambient with one warm directional source — this plays against her cooler skin tone and creates beautiful cheekbone contrast. Background: forest, waterway, or abstract dark with Algonquin geometric motif."),
        ("Dialogue scene",       "Full figure, Layer 1 or 2. Expression depends on context — most commonly CONSIDERING or POLITICAL PRECISION. Her hands are often at rest at her sides or one hand raised to make a point, fingers precise. The long figure reads well in dialogue scenes — she takes up vertical space."),
        ("Diplomatic/council",   "Layer 0 full formal. DISMISSAL or DEFAULT. She is the tallest person at the table and she knows it. Camera angle: slightly below eye level to emphasize her height and cheekbone structure from below."),
        ("Field scene",          "Layer 2. Hair loose. DEFAULT or CONSIDERING. She is not a combat character — her field presence is intelligence gathering and authority projection, not fighting. She should be still in field scenes while others move."),
        ("Alliance signing",     "This is the EtherTask scene (CAN_ALLIANCE_SIGNED). Layer 0 full formal. Next to Ualani (Layer 0) and Marc Penoit (his coat). The visual contrast: Ualani's gold earrings catching light, Jessica's ribbon-work at collar, Marc's worn coat. This trio should be visually distinctive from a distance."),
        ("Private/informal",     "Layer 3 or 4. Hair fully loose. GENUINE WARMTH expression — one of its few appropriate uses. Candle or fire lighting. The cool skin tone reads warm in firelight. This is the rare 'she has allowed this moment' scene."),
    ],

    "scene_sensual": [
        ("Rating definition",    "SENSUAL = charged implication. Clothing displaced, key areas at frame edge. Her agency is absolute — she permits, she decides, she chooses the moment."),
        ("Entry state",          "Layer 2 to Layer 4 transition. Coat or jacket removed, undershirt loosened or open. The ribbon-work of the removed coat is OFF — she has stepped out of formal identity. This is significant. The long throat and collarbone are the focal sensual element at this stage."),
        ("Her specific sensual register", "She is deliberate and unhurried. She does not respond to urgency — she sets pace. If there is a scene partner, she is the one who decided this was happening. SENSUAL INTEREST expression: eyes doing the work."),
        ("The neck", "Her long neck is the primary sensual focus before and during undress. It is a significant visual element unique to her proportions — Ualani does not share this. Tilted head exposing the throat is her most vulnerable visual language and should be used with that weight."),
        ("Hair",                 "Fully loose in sensual scenes. The length (lower lumbar) creates a visual curtain element — hair falling forward across one shoulder, or falling around her in a reclining position. No ribbon."),
        ("Partial undress framing", "The long torso reads differently than Ualani's — narrower, longer. The hip bones become prominent at partial undress and are a characteristic element of her specific body. The lean build means the ribcage is visible under skin in some poses — authentic and specific to this character."),
        ("Lighting",             "Cool ambient with one warm source. The cheekbones read dramatically under side lighting. Her cooler skin tone (#7A5238 base) takes warm light differently than Ualani — slightly more golden rather than orange. Firelight is excellent for this character."),
        ("Expression",           "SENSUAL INTEREST transitioning to PLEASURE (sensual). The stillness of her default expression makes the opening of pleasure expression significantly charged — the contrast is the content."),
    ],

    "scene_explicit": [
        ("Rating definition",    "EXPLICIT = full nudity and explicit sexual content. All invariants mandatory. No ear jewelry (already her standard). Ribbon-work gone (she has removed the formal layer). Indigo ribbon gone (braid would be loose). She is fully herself without armor."),
        ("Body canon — full nude", "Lean and long. The opposite of soft curves — elongated lines, visible structure. Hip bones prominent and angular. Waist long. Ribcage structure visible under skin in extended/arched positions. Breasts: medium-small, natural shape consistent with lean build — sitting slightly wide and lateral on a flat chest wall. No exaggerated volume. Nipples: medium, warm-toned (#8A4030). Consistent. The overall impression: a body that moves with precision and economy, not one designed for soft display."),
        ("Pubic area",           "Present and natural. Dark, consistent with natural hair color. Colonial-era authentic. Not manicured into a modern shape."),
        ("Explicit positions — her dominant/active role", "She is deliberate and economical in movement. No excess motion. Her dominant posture: upright, weight controlled, pace fully hers. Hands — her long fingers are prominent and expressive in explicit scenes. Her face: PLEASURE expression that returns quickly to baseline — she does not stay at maximum expression for long. The control returning is part of her character."),
        ("Explicit positions — her receptive role", "She receives attention with stillness. This is the defining feature of her receptive scenes — she is very still except for specific reactive moments (the PLEASURE or RELEASE expression). Her long body in a reclining or bent position is a different silhouette than Ualani's — the length is the visual content."),
        ("The neck in explicit", "Her neck is the most important body element in explicit scenes after her face. Exposed throat — head thrown back, neck arched — is her primary body language of pleasure. Every explicit scene should have at least one moment where the long neck is fully visible and extended."),
        ("Oral — giving",        "Face: eyes closed or partially open, expression focused and chosen. Her long fingers may be visible. Deliberate, unhurried, precise — she approaches this the way she approaches everything. Not servile."),
        ("Oral — receiving",     "Head thrown fully back — neck extended, prominent. PLEASURE expression. One hand on partner's head or at her own thigh, fingers splayed. The hair falls loose around and behind her. This position most clearly shows her body's elongated proportions."),
        ("Penetrative — general", "All positions maintain: hair loose and visible at full lower-lumbar length, no ear jewelry, skin tone consistent (#7A5238 base), cheekbone structure prominent in any face angle, long body proportions visible."),
        ("Penetrative — missionary/equivalent", "The length of her body is the visual content — full figure visible to emphasize the elongated proportions. Her face always accessible to the frame. PLEASURE expression. Legs: long, straight line or wrapped, showing the 4.2-head leg proportion."),
        ("Penetrative — from behind", "Her back is long and straight. The spine is visible as a line from neck to lumbar. Hair falls forward or to one side. The prominent hip bones visible in this position. PLEASURE expression in whatever face view is available — the neck is always relevant here."),
        ("Penetrative — riding/dominant", "Her most character-consistent explicit position. Upright, back straight, weight fully controlled. The long torso creates a dramatic upright line. Hands on partner's chest or shoulders. PLEASURE expression with control — she sets the pace, she reads as knowing exactly what she is doing. The hip bones and long waist are prominent in this position."),
        ("Penetrative — against wall/furniture", "Layer 1 or 2 partially displaced — coat against a surface, arms pinned or braced. The ribbon-work of the coat visible in the displaced clothing. This is the most charged visual contrast: the diplomatic symbols of her authority physically moved aside but still present in frame."),
        ("Multiple partners",    "Her stillness as the organizing center of a scene. She is the axis around which things happen. Her gaze controls the scene even in receptive positions. She looks at each partner with full deliberate focus when directing them."),
        ("Post-scene state",     "Hair loose and full, possibly more disheveled than sensual scenes — the long hair becomes a visual element in the aftermath. No ribbon. The POLITICAL PRECISION or DEFAULT expression returning to her face. She looks as if she has concluded something that went exactly as intended."),
        ("Lighting",             "Cool ambient, single warm point source. The cheekbones are the light-catching feature — in warm light they glow at the highlight (#B07850). The long neck catches the warm source along one side while the other falls into cool shadow. This lighting setup is specific and beautiful for her face structure."),
        ("Contrast with Ualani", "In multi-character explicit scenes: Ualani is warm, muscular, gold earrings catching light. Jessica is cool, elongated, no jewelry, ribbon-work gone. They are visually distinct from a distance at every level. The contrast in body type (Ualani: athlete; Jessica: elongated lean) should be fully realized."),
        ("Common errors to avoid", "1) Shortening her proportions — she must remain the 8-head tall figure. 2) Softening or rounding the hip bones — they are angular and prominent. 3) Giving her ear jewelry — she never wears any. 4) Narrowing the nose. 5) Softening the cheekbones. 6) Making the breasts larger than canon. 7) Changing the hair to loose waves — it is straight. 8) Making her expression passive or confused — she is always present and choosing. 9) Lightening skin tone. 10) Shortening the hair below lower lumbar."),
    ],
}


# ── MODEL SHEET BUILD FUNCTION ─────────────────────────────────────────────────

def _build_model_sheet(wb, model):
    name_short = model["name"].split()[0]  # "Ualani" or "Jessica"
    ws = wb.create_sheet(f"{name_short.upper()} MODEL")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = model["tab_color"]

    # Column widths: section label | element | content
    _set_col_widths(ws, [26, 28, 90])

    row = 1

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    _hdr(ws, row, 1, f"{model['name'].upper()} — PRODUCTION MODEL SHEET", model["tab_color"], "FFFFFF", bold=True, size=13)
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    _hdr(ws, row, 1, model["title"], HDR_BG, "CCCCCC", bold=False, size=10)
    row += 2

    # ── Section renderer ──────────────────────────────────────────────────────
    def _section(title, bg, fg, items, col1_label="ELEMENT", col2_label="SPECIFICATION"):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        t = ws.cell(row=row, column=1, value=f"▶  {title}")
        t.font = Font(bold=True, color=fg, size=11, name="Calibri")
        t.fill = _fill(bg)
        t.border = _border()
        t.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        row += 1
        _hdr(ws, row, 1, "SECTION", bg, fg, size=9)
        _hdr(ws, row, 2, col1_label, bg, fg, size=9)
        _hdr(ws, row, 3, col2_label, bg, fg, size=9)
        row += 1
        for i, (el, spec) in enumerate(items):
            alt = ALT_BG if i % 2 == 0 else WHITE
            _cell(ws, row, 1, title, bg=alt)
            _cell(ws, row, 2, el, bg=alt, bold=True)
            _cell(ws, row, 3, spec, bg=alt)
            ws.row_dimensions[row].height = max(60, min(180, len(spec) // 2))
            row += 1
        row += 1  # spacer

    # ── Sections ──────────────────────────────────────────────────────────────
    _section("CONSTRUCTION",          "2E4057", "FFFFFF", model["construction"])
    _section("COLOR PALETTE",         "1A3A2A", "FFFFFF", [(row[0], f"{row[1]}  —  {row[2]}") for row in model["palette"]])
    _section("INVARIANTS",            "5C2020", "FFFFFF", model["invariants"], col1_label="RULE", col2_label="DESCRIPTION — ABSOLUTE, NO EXCEPTIONS")
    _section("EXPRESSION CHART",      "3A3A1A", "FFFFFF", model["expressions"], col1_label="EXPRESSION", col2_label="DIRECTION")
    _section("COSTUME LAYERS",        "2A2A4A", "FFFFFF", model["costume_layers"], col1_label="LAYER", col2_label="DESCRIPTION")
    _section("SCENE DIRECTION — SFW", "1A4A1A", "FFFFFF", model["scene_sfw"],     col1_label="SCENE TYPE",    col2_label="DIRECTION")
    _section("SCENE DIRECTION — SENSUAL", "5C3A00", "FFFFFF", model["scene_sensual"], col1_label="ELEMENT", col2_label="DIRECTION")
    _section("SCENE DIRECTION — EXPLICIT", "7A0000", "FFFFFF", model["scene_explicit"], col1_label="SCENE / RULE", col2_label="DIRECTION — FULL PRODUCTION SPEC")

    _freeze(ws, "A3")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_overview(wb, USA_CHARACTERS, CANADIAN_CHARACTERS, NPC_CHARACTERS)
    _build_character_sheet(wb, "PLAYABLE (USA)", USA_CHARACTERS, USA_BG, USA_FG)
    _build_character_sheet(wb, "CANADIAN", CANADIAN_CHARACTERS, CA_BG, CA_FG)
    _build_character_sheet(wb, "NPC", NPC_CHARACTERS, NPC_BG, NPC_FG)
    _build_design_notes(wb, DESIGN_NOTES)
    _build_model_sheet(wb, UALANI_MODEL)
    _build_model_sheet(wb, JESSICA_MODEL)

    wb.save(OUT_PATH)
    total = len(USA_CHARACTERS) + len(CANADIAN_CHARACTERS) + len(NPC_CHARACTERS)
    designed = sum(1 for c in USA_CHARACTERS + CANADIAN_CHARACTERS + NPC_CHARACTERS if c["status"] in ("DESIGNED", "FINAL"))
    print(f"Wrote {OUT_PATH}")
    print(f"  {total} total characters: {len(USA_CHARACTERS)} USA, {len(CANADIAN_CHARACTERS)} Canadian, {len(NPC_CHARACTERS)} NPC")
    print(f"  {designed}/{total} fully described  |  {total-designed} partial/stub")
    print(f"  2 full model sheets: Ualani Carlisle, Jessica Commanda Odjick")


if __name__ == "__main__":
    main()
